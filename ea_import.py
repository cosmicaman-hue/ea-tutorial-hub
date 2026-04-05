#!/usr/bin/env python3
"""
EA Historical Import: Aug 2024 - Jan 2026
- Full replace of historical months from Excel
- Removes duplicate student IDs (future roll entries)
- Adds frozen_months for JS hardening
- Writes to both data file paths
"""
import json, datetime, re, openpyxl, time
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\sujit\OneDrive\Desktop\EA STUDENT SCORE TALLY  v5.5.xlsx"
LIVE_PATH  = r"C:\var\data\ea_tutorial_hub\offline_scoreboard_data.json"
INST_PATH  = r"c:\Users\sujit\OneDrive\Desktop\Project EA\instance\offline_scoreboard_data.json"

TARGET_MONTHS = {
    'Aug 24':'2024-08','Sep 24':'2024-09','Oct 24':'2024-10',
    'Nov 24':'2024-11','Dec 24':'2024-12','Jan 25':'2025-01',
    'Feb 25':'2025-02','Mar 25':'2025-03','Apr 25':'2025-04',
    'May 25':'2025-05','Jun 25':'2025-06','Jul 25':'2025-07',
    'Aug 25':'2025-08','Sep 25':'2025-09','Oct 25':'2025-10',
    'Nov 25':'2025-11','Dec 25':'2025-12','Jan 26':'2026-01',
}

SKIP_COL_LABELS = {
    'rank','vote power','points used','buffer','teams','activity point',
    'cumulative point','total points','combined power','activity calendar',
    'vision test','fees','column1','column2','column3','column4',
    'column5','column6','column7','column8','column9','column10',
}

def parse_date_value(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        v = v.strip()
        m = re.match(r'^(\d{1,2})-([A-Za-z]{3})-(\d{2,4})$', v)
        if m:
            day, mon_str, yr = int(m.group(1)), m.group(2), int(m.group(3))
            if yr < 100:
                yr += 2000
            months = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                      'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
            mon = months.get(mon_str.lower())
            if mon:
                try:
                    return datetime.date(yr, mon, day).strftime('%Y-%m-%d')
                except Exception:
                    pass
        m2 = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', v)
        if m2:
            return v
    return None

def is_skip_col(label):
    if not label:
        return False
    l = str(label).strip().lower()
    return any(s in l for s in SKIP_COL_LABELS)

def find_header_row(ws):
    for r in range(1, 5):
        for c in range(1, 5):
            v = str(ws.cell(r, c).value or '').lower()
            if 'roll' in v:
                return r
    return None

def inspect_sheet(ws):
    hidden_cols = {cl for cl, cd in ws.column_dimensions.items() if cd.hidden}
    hrow = find_header_row(ws)
    if not hrow:
        return None

    roll_col = name_col = class_col = final_col = awf_col = None
    awf_col_hidden = False
    all_date_cols = []  # (ci, date_str, is_hidden)

    for ci in range(1, ws.max_column + 1):
        v = ws.cell(hrow, ci).value
        cl = get_column_letter(ci)
        is_hidden = cl in hidden_cols

        d = parse_date_value(v)
        if d:
            all_date_cols.append((ci, d, is_hidden))
            continue

        if v is None:
            continue
        vl = str(v).strip().lower()

        if is_skip_col(vl):
            continue
        if 'roll' in vl and roll_col is None:
            roll_col = ci
        elif any(x in vl for x in ['candidate', "student's name", 'student name']) and name_col is None:
            name_col = ci
        elif 'name' in vl and name_col is None and roll_col is not None:
            name_col = ci
        elif any(x in vl for x in ['class', 'grade']) and class_col is None:
            class_col = ci
        elif any(x in vl for x in ['final score', 'total score']):
            final_col = ci
        elif 'awf' in vl and 'buffer' not in vl:
            awf_col = ci
            awf_col_hidden = is_hidden

    return hrow, roll_col, name_col, class_col, all_date_cols, final_col, awf_col, awf_col_hidden

def strip_suffixes(name):
    name = re.sub(r'\((?:L|CoL|LoP|CR|PP|DPP|WCI|ECS|CCAI|CITC|ECJ|DWI|CI|SC|RM|v|Vv|vvv)\)', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\*+', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

now_iso = datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
id_counter = [int(time.time() * 1000) + 10000]

def next_id():
    id_counter[0] += 1
    return id_counter[0]

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data...")
with open(LIVE_PATH, 'r', encoding='utf-8') as f:
    data = json.load(f)

students = data.get('students', [])
print(f"  {len(students)} student entries (before dedup fix)")

# ── Step 1: Fix duplicate student IDs ─────────────────────────────────────────
# Remove future-roll entries that don't exist yet (April 2026 changes)
REMOVE_ROLLS = {'EA24B16', 'EA24B15', 'EA24B14'}

clean_students = []
removed_rolls = []
for s in students:
    roll = str(s.get('roll', '') or '').strip().upper()
    if roll in REMOVE_ROLLS:
        removed_rolls.append(f"id={s.get('id')} roll={roll} name={s.get('name')}")
        continue
    clean_students.append(s)

print(f"  Removed {len(removed_rolls)} future-roll entries: {removed_rolls}")
print(f"  {len(clean_students)} clean student entries")

# Build name → studentId lookup (name-first: names are stable, rolls were recycled)
name_to_id = {}
roll_to_id = {}
for s in clean_students:
    sid = int(s['id'])
    raw_name = strip_suffixes(str(s.get('name') or s.get('base_name') or ''))
    norm = raw_name.lower().strip()
    if norm:
        name_to_id[norm] = sid
    roll = str(s.get('roll', '') or '').strip().upper()
    if roll:
        roll_to_id[roll] = sid

def resolve_sid(roll, name):
    # Name-first: names are stable identifiers; rolls were recycled (e.g. EA24A04 = Sahil then Tanu)
    n = strip_suffixes(str(name or '')).lower().strip()
    if n and n in name_to_id:
        return name_to_id[n]
    r = str(roll or '').strip().upper()
    if r and r in roll_to_id:
        return roll_to_id[r]
    return None

# ── Step 2: Parse Excel ────────────────────────────────────────────────────────
print("\nParsing Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

excel_months = {}

for sheet_name, month_key in TARGET_MONTHS.items():
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP {sheet_name}: sheet not found")
        continue
    ws = wb[sheet_name]
    result = inspect_sheet(ws)
    if result is None:
        print(f"  SKIP {sheet_name}: no header row")
        continue

    hrow, roll_col, name_col, class_col, all_date_cols, final_col, awf_col, awf_col_hidden = result

    if not roll_col or not name_col:
        print(f"  SKIP {sheet_name}: missing roll={roll_col} name={name_col}")
        continue

    month_students = []
    for r in range(hrow + 1, ws.max_row + 1):
        roll = str(ws.cell(r, roll_col).value or '').strip()
        if not roll.startswith('EA'):
            continue
        raw_name = str(ws.cell(r, name_col).value or '').strip()
        if not raw_name:
            continue
        base_name = strip_suffixes(raw_name)
        class_val = ws.cell(r, class_col).value if class_col else None

        # Daily scores — include ALL date cols, hidden ones become 0
        daily = {}
        for ci, date_str, is_hidden in all_date_cols:
            if is_hidden:
                daily[date_str] = 0
            else:
                raw = ws.cell(r, ci).value
                daily[date_str] = int(raw) if isinstance(raw, (int, float)) else 0

        final_score = None
        if final_col:
            raw_fs = ws.cell(r, final_col).value
            if isinstance(raw_fs, (int, float)):
                final_score = int(raw_fs)

        awf_value = None
        if awf_col:
            raw_awf = ws.cell(r, awf_col).value
            if isinstance(raw_awf, (int, float)):
                awf_value = int(raw_awf)

        month_students.append({
            'roll': roll,
            'name': raw_name,
            'base_name': base_name,
            'class': int(class_val) if isinstance(class_val, (int, float)) else None,
            'daily': daily,
            'final_score': final_score,
            'awf_value': awf_value,
        })

    excel_months[month_key] = {
        'students': month_students,
        'has_awf': awf_col is not None,
        'awf_hidden': awf_col_hidden,
    }
    print(f"  {sheet_name} -> {month_key}: {len(month_students)} students, "
          f"{len(all_date_cols)} dates ({sum(1 for _,_,h in all_date_cols if h)} hidden), "
          f"final={'YES' if final_col else 'MISSING'}, "
          f"awf={'hidden' if awf_col_hidden else 'shown' if awf_col else 'none'}")

# ── Step 3: Rebuild historical data ───────────────────────────────────────────
print("\nRebuilding historical data...")
LOCKED_MONTHS = set(TARGET_MONTHS.values())

# Remove all existing historical scores
kept_scores = [
    s for s in data.get('scores', [])
    if (s.get('month') or s.get('date', '')[:7]) not in LOCKED_MONTHS
]
print(f"  Removed {len(data.get('scores',[])) - len(kept_scores)} old historical scores, kept {len(kept_scores)} active")

new_scores = list(kept_scores)
new_profiles = dict(data.get('month_roster_profiles', {}))
new_extras = dict(data.get('month_student_extras', {}))
new_extra_cols = dict(data.get('month_extra_columns', {}))
frozen_months = dict(data.get('frozen_months') or {})
no_sid_count = 0

for month_key, mdata in excel_months.items():
    month_students = mdata['students']
    has_awf = mdata['has_awf']

    # Profile entries (all students in the month, with locked flag)
    profiles = []
    for row in month_students:
        sid = resolve_sid(row['roll'], row['base_name'])
        profiles.append({
            'roll': row['roll'],
            'name': row['name'],
            'base_name': row['base_name'],
            'class': row['class'],
            'studentId': sid,
            'locked': True,
            'month_star_count': 0,
            'month_veto_count': 0,
            'month_designations': [],
        })
    new_profiles[month_key] = profiles

    # Score entries
    extras_this_month = {}
    scores_added = 0
    for row in month_students:
        sid = resolve_sid(row['roll'], row['base_name'])
        if sid is None:
            no_sid_count += 1
            continue

        # Authoritative total
        if row['final_score'] is not None:
            new_scores.append({
                'id': next_id(),
                'studentId': sid,
                'date': f"{month_key}-15",
                'month': month_key,
                'points': row['final_score'],
                'stars': 0, 'vetos': 0,
                'star_usage_normal': 0, 'star_usage_disciplinary': 0,
                'notes': 'excel_total_score',
                'recordedBy': 'admin',
                'created_at': now_iso, 'updated_at': now_iso,
                'locked': True,
            })
            scores_added += 1

        # Daily scores
        for date_str, pts in row['daily'].items():
            new_scores.append({
                'id': next_id(),
                'studentId': sid,
                'date': date_str,
                'month': month_key,
                'points': pts,
                'stars': 0, 'vetos': 0,
                'star_usage_normal': 0, 'star_usage_disciplinary': 0,
                'notes': 'excel_daily_score',
                'recordedBy': 'admin',
                'created_at': now_iso, 'updated_at': now_iso,
                'locked': True,
            })
            scores_added += 1

        # AWF extra
        if has_awf and row['awf_value'] is not None:
            rk = row['roll'].upper()
            if rk not in extras_this_month:
                extras_this_month[rk] = {}
            extras_this_month[rk]['awf'] = str(row['awf_value'])

    new_extras[month_key] = extras_this_month
    new_extra_cols[month_key] = [{'key': 'awf', 'label': 'AWF'}] if has_awf else []
    frozen_months[month_key] = {
        'hardened': True,
        'allow_modifications': False,
        'locked_at': now_iso,
        'source': 'excel_import_v2',
    }
    print(f"  {month_key}: {len(profiles)} profiles, {scores_added} scores")

if no_sid_count:
    print(f"  WARNING: {no_sid_count} students had no matching studentId (profiles kept, scores skipped)")

# ── Step 4: Write ──────────────────────────────────────────────────────────────
data['students'] = clean_students
data['scores'] = new_scores
data['month_roster_profiles'] = new_profiles
data['month_student_extras'] = new_extras
data['month_extra_columns'] = new_extra_cols
data['frozen_months'] = frozen_months
data['locked_months'] = sorted(LOCKED_MONTHS)
data['updated_at'] = now_iso
data['server_updated_at'] = now_iso

out = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
print(f"\nOutput: {len(out)/1024:.1f} KB, {len(new_scores)} total scores")

for path in [LIVE_PATH, INST_PATH]:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(out)
    print(f"  Written: {path}")

print("\nDone. Run Master Update in browser to sync.")
