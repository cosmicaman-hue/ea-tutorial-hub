"""
Nuclear Scoreboard Rebuild from Excel
======================================
Clears all student/score data and rebuilds from scratch using:
  - Excel: Aug 2024 - Feb 2026  (authoritative historical source)
  - Backup: Feb 2026 - Apr 2026 (platform data, matched by canonical name)

Run from project root:
  python scripts/rebuild_from_excel.py
"""

import json
import re
import sys
import os
from datetime import datetime, timezone, timedelta
from calendar import monthrange

EXCEL_PATH = r"C:\Users\sujit\OneDrive\Desktop\EA STUDENT SCORE TALLY  v5.5.xlsx"
# Use shared data path resolver (respects EA_STORAGE_ROOT, RENDER_DISK_PATH, /var/data, or instance/)
_project_root = str(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)
from app.utils.data_paths import get_data_path
DB_PATH = get_data_path()
BACKUP_PATH = os.path.join(os.path.dirname(__file__), "..", "instance", "BACKUP_FEB_APR_2026_20260405_134210.json")

# IST timezone offset
IST = timezone(timedelta(hours=5, minutes=30))
NOW_IST = datetime.now(IST).isoformat()

# Sheet name → month key
SHEET_TO_MONTH = {
    'Aug 24': '2024-08', 'Sep 24': '2024-09',
    'Oct 24': '2024-10', 'Nov 24': '2024-11', 'Dec 24': '2024-12',
    'Jan 25': '2025-01', 'Feb 25': '2025-02', 'Mar 25': '2025-03',
    'Apr 25': '2025-04', 'May 25': '2025-05', 'Jun 25': '2025-06',
    'Jul 25': '2025-07', 'Aug 25': '2025-08', 'Sep 25': '2025-09',
    'Oct 25': '2025-10', 'Nov 25': '2025-11', 'Dec 25': '2025-12',
    'Jan 26': '2026-01', 'Feb 26': '2026-02',
}


# ---------------------------------------------------------------------------
# Name normalization and formatting (mirrors JS reconcilePromotedStudentDuplicates logic)
# ---------------------------------------------------------------------------
def canonical_name(raw):
    """Convert to lowercase, strip ALL decorations for matching.
    Strips: asterisks, ALL parenthetical suffixes (CR, CoL, SC, PP, WCI, RM, etc.),
    v/V decoration markers, trailing slashes, numbers.
    CRITICAL: Only strips v/V when they're standalone decorations, NOT within names.
    """
    if not raw:
        return ""
    name = str(raw)
    name = re.sub(r'\*+', '', name)               # strip asterisks
    name = re.sub(r'\s*\([^)]*\)\s*', ' ', name)  # strip ALL (...) suffixes
    # Only strip v/V when followed by space, end of string, or non-letter char (not within words)
    # This preserves names like "Vishes", "Shiva", "Divya"
    name = re.sub(r'(?<!\w)[vV]+(?!\w)', ' ', name)
    name = re.sub(r'[/\\]+', ' ', name)            # strip slashes (including trailing)
    name = re.sub(r'\d+', '', name)                # strip numbers
    name = re.sub(r'[^\w\s]', '', name)            # strip remaining punctuation (dots etc)
    return re.sub(r'\s+', ' ', name).strip().lower()


def normalize_decoration_format(raw):
    """
    Option B: Convert repeated decorations to counts.
    E.g. "Sahil Yadav******* vvv (CoL)" → "Sahil Yadav *7 V3 (CoL)"

    Handles decoration formats:
    - base name***
    - base name*** vV
    - base name(v)
    - base name *****(vV)

    Base name ends at first * or ( — everything before that is the base.
    This correctly handles abbreviations like "MD. Aamir", single letters like "N Riya Kumari".
    """
    if not raw:
        return ""

    s = str(raw).strip()

    # Find where decorations start: first * or (
    deco_match = re.search(r'[\*\(]', s)

    if not deco_match:
        # No decorations: return as-is (preserve original name like "MD. Aamir")
        return s

    base = s[:deco_match.start()].strip()
    decorations_part = s[deco_match.start():].strip()

    if not base:
        return s

    # Process decorations: count *, v/V; preserve other parenthetical suffixes
    result_parts = [base]
    i = 0

    while i < len(decorations_part):
        char = decorations_part[i]

        if char == '*':
            count = 1
            while i + count < len(decorations_part) and decorations_part[i + count] == '*':
                count += 1
            result_parts.append(f"*{count}")
            i += count
        elif char.lower() == 'v':
            # Only count as veto if outside a word (preceded by space/start)
            count = 1
            while i + count < len(decorations_part) and decorations_part[i + count].lower() == 'v':
                count += 1
            result_parts.append(f"V{count}")
            i += count
        elif char == '(':
            match_paren = re.match(r'\([^)]*\)', decorations_part[i:])
            if match_paren:
                paren_str = match_paren.group(0)
                inner = paren_str[1:-1]

                # Check if purely v/V based: (v), (vvv), (vV)
                if inner and all(c.lower() == 'v' for c in inner):
                    v_count = sum(1 for c in inner if c.lower() == 'v')
                    result_parts.append(f"V{v_count}")
                else:
                    # Keep suffix as-is: (CR), (CoL), (SC), (LoP), (ECS), (PP), etc.
                    result_parts.append(paren_str)
                i += len(paren_str)
            else:
                i += 1
        elif char == ' ':
            i += 1
        else:
            i += 1

    result = ' '.join(result_parts)
    result = re.sub(r'\s+', ' ', result).strip()
    return result


def extract_decoration_counts(formatted_name):
    """Parse Option B formatted name (e.g. 'Vishes Xalxo *3 V1 (CoL)')
    and return (star_count, veto_count) as integers.
    """
    text = str(formatted_name or '')
    star_match = re.search(r'\*(\d+)', text)
    stars = int(star_match.group(1)) if star_match else 0
    veto_match = re.search(r'V(\d+)', text)
    vetos = int(veto_match.group(1)) if veto_match else 0
    return stars, vetos


def is_class_number(val):
    """Check if value is just a class number (int 1-12)."""
    if isinstance(val, int):
        return 1 <= val <= 12
    if isinstance(val, str):
        try:
            num = int(val.strip())
            return 1 <= num <= 12
        except ValueError:
            return False
    return False


def display_name(raw):
    """Format name with Option B decoration style, preserving exact Excel decorations."""
    if not raw:
        return ""
    # Check if this looks like a class number (Jan 26 anomaly)
    if is_class_number(raw):
        return ""
    return normalize_decoration_format(raw)


# ---------------------------------------------------------------------------
# Known same-person name variant pairs to merge after extraction.
# Format: (canonical_name_to_remove, canonical_name_to_keep)
# The student record under "to_keep" wins; scores/profiles are re-pointed.
# ---------------------------------------------------------------------------
MERGE_RULES = [
    ('rashi singh',   'rashi'),          # EA24A05 – keep "Rashi"
    ('sakshi singh',  'sakshi'),         # EA24C06 – keep "Sakshi"
    ('riya kumari',   'n riya kumari'),  # EA25C18 – keep "N Riya Kumari"
    ('aamna',         'aamna khatoon'),  # EA25D13 – keep "Aamna Khatoon"
    ('mahak mahato',  'mahek mahato'),   # EA25D22 – keep "Mahek Mahato"
    ('salman khan',   'hasibul alam'),   # EA24D02 – keep "Hasibul Alam" (Name 1)
]

# Known roll-change effective months that must survive rebuilds exactly.
# These came from manual admin decisions and cannot be inferred safely from
# the Feb-Apr backup alone because the backup contains overlapping variants.
ROLL_CHANGE_EFFECTIVE_MONTH_OVERRIDES = {
    ('ayush gupta', 'EA24B15'): '2026-04',
    ('tanu sinha', 'EA24B16'): '2026-04',
}


# ---------------------------------------------------------------------------
# Auto-detect sheet layout: header row, data start row, metadata col count
# ---------------------------------------------------------------------------
def detect_sheet_layout(ws):
    """
    Returns (header_row, data_start_row, name_col_idx, date_col_start_idx)
    All indices are 1-based (openpyxl convention).
    """
    def row_values(r):
        return [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 60))]

    row1 = row_values(1)
    row2 = row_values(2)

    def has_dates(row):
        return any(isinstance(v, datetime) for v in row)

    def has_string_dates(row):
        # Nov 24 style: "01-Nov-24"
        return any(
            isinstance(v, str) and re.match(r'\d{2}-[A-Za-z]{3}-\d{2}', v)
            for v in row
        )

    def first_date_col(row):
        for i, v in enumerate(row, 1):
            if isinstance(v, datetime) or (isinstance(v, str) and re.match(r'\d{2}-[A-Za-z]{3}-\d{2}', v)):
                return i
        return None

    if has_dates(row1):
        header_row = 1
        # Sep 24 has a blank row 2 between header and data
        row2_has_data = any(v is not None for v in row2[:4])
        data_start = 2 if row2_has_data else 3
        date_start = first_date_col(row1)
    elif has_dates(row2) or has_string_dates(row2):
        header_row = 2
        data_start = 3
        date_start = first_date_col(row2)
    else:
        # Fallback: try row 3
        row3 = row_values(3)
        header_row = 3
        data_start = 4
        date_start = first_date_col(row3)

    if date_start is None:
        return None  # Sheet has no date columns at all

    # Name column is the one just before date columns
    name_col = date_start - 1

    return (header_row, data_start, name_col, date_start)


# ---------------------------------------------------------------------------
# Parse a date value from a header cell
# ---------------------------------------------------------------------------
def parse_date(val, month_key):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        # Nov 24 style: "01-Nov-24"
        try:
            return datetime.strptime(val.strip(), "%d-%b-%y").strftime('%Y-%m-%d')
        except ValueError:
            pass
        try:
            return datetime.strptime(val.strip(), "%d-%b-%Y").strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Main rebuild logic
# ---------------------------------------------------------------------------
def rebuild():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    print("Loading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("Loading current database...")
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)

    # -----------------------------------------------------------------------
    # Step 1: Reset data fields, keep settings
    # -----------------------------------------------------------------------
    print("\n[Step 1] Resetting database data fields...")
    settings_to_keep = [
        'good_conduct_formula', 'points_star_conversion', 'hardened_months',
        'locked_months', 'pending_cr_requests', 'role_veto_applied_month',
        'server_version', 'server_updated_at', 'month_extra_columns',
        'month_student_extras',
    ]
    new_db = {k: db[k] for k in settings_to_keep if k in db}
    new_db['students'] = []
    new_db['scores'] = []
    new_db['months'] = []
    new_db['month_students'] = {}
    new_db['month_roster_profiles'] = {}
    new_db['roll_history'] = []
    new_db['deleted_students'] = []
    # fee_records removed - using simple student.fees field instead
    print("   Done. Settings preserved, data cleared.")

    # -----------------------------------------------------------------------
    # Step 2: Extract from Excel (Aug 2024 – Feb 2026)
    # -----------------------------------------------------------------------
    print("\n[Step 2] Extracting from Excel...")

    name_to_student = {}   # canonical_name → student dict
    roll_to_student = {}   # roll → student dict (for exact roll-based matching)
    student_id_counter = [1]
    score_id_counter = [1]
    all_scores = []
    month_roster_profiles = {}
    month_students = {}
    roll_history = []

    def is_ascending_roll_change(old_roll, new_roll):
        """True if new_roll is an ascending group-letter change from old_roll.
        e.g., EA24A01 → EA24B15 is ascending (A→B). B→A is not.
        Must be same batch prefix (EA24, EA25, etc.).
        """
        if not old_roll or not new_roll or old_roll == new_roll:
            return False
        if old_roll[:4] != new_roll[:4]:  # different batch year
            return False
        if len(old_roll) < 5 or len(new_roll) < 5:
            return False
        return new_roll[4].upper() > old_roll[4].upper()

    def get_effective_month_for_backup_roll_change(cname, new_roll, default_month):
        """Return the safest effective month for a backup-derived roll change."""
        override = ROLL_CHANGE_EFFECTIVE_MONTH_OVERRIDES.get((cname, new_roll))
        if override:
            return override
        return default_month

    def get_roll_for_month(student_id, month_key):
        """Resolve which roll a student should hold in a given month."""
        student = next(
            (s for s in name_to_student.values() if s.get('id') == student_id),
            None
        )
        if not student:
            return ''

        current_roll = student.get('roll', '')
        entries = sorted(
            [
                entry for entry in roll_history
                if isinstance(entry, dict) and entry.get('student_id') == student_id
            ],
            key=lambda entry: str(entry.get('effective_month') or ''),
            reverse=True,
        )

        roll = current_roll
        for entry in entries:
            eff_month = str(entry.get('effective_month') or '').strip()
            if eff_month and eff_month > month_key:
                old_roll = entry.get('old_roll', '')
                if old_roll:
                    roll = old_roll
            else:
                break
        return roll

    def get_or_create_student(roll, raw_name, class_val, month_key, fee_val=0):
        cname = canonical_name(raw_name)
        dname = display_name(raw_name)

        if not cname:
            return None

        if cname in name_to_student:
            s = name_to_student[cname]
            # Track roll change (promotion detection)
            prev_roll = s.get('roll')
            if prev_roll and prev_roll != roll and roll:
                # Check if this roll change is already recorded
                already = any(
                    r['student_id'] == s['id'] and r['new_roll'] == roll
                    for r in roll_history
                )
                if not already:
                    roll_history.append({
                        'student_id': s['id'],
                        'old_roll': prev_roll,
                        'new_roll': roll,
                        'effective_month': month_key,
                        'reason': 'promotion_or_roll_change',
                        'timestamp': NOW_IST,
                    })
                s['roll'] = roll  # update to latest known roll
            # Update class if present (and not a class number from Jan 26 anomaly)
            if class_val and not s.get('class'):
                s['class'] = class_val
            # Update fees if present
            if fee_val is not None and fee_val > 0:
                s['fees'] = fee_val
            # Store or update name: use dname if it's non-empty and more informative
            # (dname has decorations formatted in Option B style)
            if dname:
                # Prefer the version with most decorations/information
                current_name = s.get('name', '')
                # If dname is longer or has decorations, prefer it
                if len(dname) >= len(current_name):
                    s['name'] = dname
                    # base_name is the raw name without Option B formatting, for fallback
                    base_only = re.sub(r'[\*\d\s]*\([A-Za-z\d]+\)', '', dname).strip()
                    base_only = re.sub(r'[\*\d\s]+', ' ', base_only).strip()
                    s['base_name'] = base_only or dname
            s['rolls_by_month'][month_key] = roll
            s['last_month_appeared'] = month_key
            return s
        else:
            s = {
                'id': student_id_counter[0],
                'roll': roll,
                'name': dname or roll,
                'base_name': dname.split()[0] + (' ' + dname.split()[1] if len(dname.split()) > 1 else '') if dname else roll,
                'class': class_val,
                'fees': fee_val,
                'active': True,
                'stars': 0,
                'veto_count': 0,
                'rolls_by_month': {month_key: roll},
                'last_month_appeared': month_key,
                'deactivation_month': None,
                'active_from_month': None,
            }
            student_id_counter[0] += 1
            name_to_student[cname] = s
            if roll:
                roll_to_student[roll] = s
            return s

    for sheet_name, month_key in SHEET_TO_MONTH.items():
        if sheet_name not in wb.sheetnames:
            print(f"   SKIP {sheet_name}: not in workbook")
            continue

        ws = wb[sheet_name]
        layout = detect_sheet_layout(ws)

        if layout is None:
            print(f"   SKIP {sheet_name}: could not detect date columns")
            continue

        header_row, data_start, name_col, date_col_start = layout

        # Build date list from header row
        dates_in_month = []
        for col in range(date_col_start, ws.max_column + 1):
            val = ws.cell(header_row, col).value
            date_str = parse_date(val, month_key)
            if date_str:
                dates_in_month.append((col, date_str))

        if not dates_in_month:
            print(f"   SKIP {sheet_name}: no parseable dates in header row {header_row}")
            continue

        # Find class column (if exists — between roll and name)
        # Col A = roll, then class/grade cols may exist before name col
        # name_col is detected as col just before dates; roll is col 1
        # Anything between col 2 and name_col-1 is metadata (class, fees, etc.)
        # Class is typically the first such column (col 2)
        class_col = 2 if name_col > 2 else None
        # Fees column (if exists) - typically after class column
        fee_col = 3 if name_col > 3 else None

        month_profiles = []
        month_roll_list = []
        scores_this_month = 0

        for row_idx in range(data_start, ws.max_row + 1):
            roll_val = ws.cell(row_idx, 1).value
            if not roll_val or not str(roll_val).startswith('EA'):
                continue

            roll = str(roll_val).strip()
            name_val = ws.cell(row_idx, name_col).value
            if not name_val:
                continue

            class_val = None
            if class_col:
                cv = ws.cell(row_idx, class_col).value
                if isinstance(cv, (int, float)):
                    class_val = int(cv)

            fee_val = 0
            if fee_col:
                fv = ws.cell(row_idx, fee_col).value
                if isinstance(fv, (int, float)):
                    fee_val = int(fv)

            student = get_or_create_student(roll, str(name_val), class_val, month_key, fee_val)
            if not student:
                continue

            # Format the name with Option B decorations for display
            formatted_name = display_name(str(name_val))
            if not formatted_name:
                # If dname is empty (e.g., Jan 26 class number), use canonical name from student
                formatted_name = student.get('name', roll)

            # Add to month roster
            month_roll_list.append(roll)
            prof_stars, prof_vetos = extract_decoration_counts(formatted_name)
            month_profiles.append({
                'roll': roll,
                'name': formatted_name,  # Display name with Option B formatted decorations
                'base_name': student.get('base_name', ''),
                'class': class_val,
                'studentId': student['id'],
                'month_star_count': prof_stars,
                'month_veto_count': prof_vetos,
                'month_designations': [],
            })

            # Extract score for each date
            for col, date_str in dates_in_month:
                val = ws.cell(row_idx, col).value
                if val is None:
                    continue
                try:
                    points = int(val) if isinstance(val, (int, float)) else int(float(str(val)))
                except (ValueError, TypeError):
                    continue

                all_scores.append({
                    'id': score_id_counter[0],
                    'studentId': student['id'],
                    'date': date_str,
                    'month': month_key,
                    'points': points,
                    'stars': 0,
                    'vetos': 0,
                    'star_usage_normal': 0,
                    'star_usage_disciplinary': 0,
                    'notes': 'excel_daily_score',
                    'created_at': NOW_IST,
                    'updated_at': NOW_IST,
                })
                score_id_counter[0] += 1
                scores_this_month += 1

        month_roster_profiles[month_key] = month_profiles
        month_students[month_key] = month_roll_list
        if month_key not in new_db['months']:
            new_db['months'].append(month_key)

        print(f"   {sheet_name} ({month_key}): {len(month_profiles)} students, {scores_this_month} scores")

    print(f"\n   Excel extraction complete: {len(name_to_student)} unique students, {len(all_scores)} scores")

    # -----------------------------------------------------------------------
    # Merge same-person name variants (e.g. "Rashi Singh" → "Rashi")
    # -----------------------------------------------------------------------
    for remove_cname, keep_cname in MERGE_RULES:
        old_s = name_to_student.get(remove_cname)
        new_s = name_to_student.get(keep_cname)
        if not old_s or not new_s:
            continue  # one or both not found – skip silently

        old_id = old_s['id']
        new_id = new_s['id']

        # Merge rolls_by_month (old months not already in new)
        for m, r in old_s.get('rolls_by_month', {}).items():
            new_s['rolls_by_month'].setdefault(m, r)

        # Keep the latest last_month_appeared
        if old_s.get('last_month_appeared', '') > new_s.get('last_month_appeared', ''):
            new_s['last_month_appeared'] = old_s['last_month_appeared']

        # Re-point all scores from old_id → new_id
        for score in all_scores:
            if score['studentId'] == old_id:
                score['studentId'] = new_id

        # Re-point roll_history entries
        for rh in roll_history:
            if rh['student_id'] == old_id:
                rh['student_id'] = new_id

        # Re-point month_roster_profiles
        for profiles in month_roster_profiles.values():
            for p in profiles:
                if p.get('studentId') == old_id:
                    p['studentId'] = new_id

        # Remove old student from lookup dicts
        del name_to_student[remove_cname]
        if old_s.get('roll') in roll_to_student and roll_to_student[old_s['roll']]['id'] == old_id:
            roll_to_student[old_s['roll']] = new_s

        print(f"   Merged: '{remove_cname}' (id={old_id}) → '{keep_cname}' (id={new_id})")

    print(f"   After merge: {len(name_to_student)} unique students")

    # Compute deactivation_month for students who left before Feb 2026
    def next_month(month_key):
        """Return the next month after the given YYYY-MM string."""
        y, m = map(int, month_key.split('-'))
        if m == 12:
            return f"{y + 1:04d}-01"
        return f"{y:04d}-{m + 1:02d}"

    # Use student IDs from Jan 2026 roster (not just rolls) — correctly handles
    # roll reassignments where two different people share the same roll number.
    jan_2026_student_ids = set()
    if '2026-01' in month_roster_profiles:
        jan_2026_student_ids = {p['studentId'] for p in month_roster_profiles['2026-01']}

    deactivated_count = 0
    for student in name_to_student.values():
        last_appeared = student.get('last_month_appeared', '2024-08')

        # If THIS student's ID is not in the Jan 2026 roster, they left before Feb 2026
        if student['id'] not in jan_2026_student_ids and last_appeared < '2026-01':
            student['deactivation_month'] = next_month(last_appeared)
            student['active'] = False
            deactivated_count += 1
            print(f"   Deactivated: {student.get('name', student.get('roll','?'))} (last seen {last_appeared}, inactive from {student['deactivation_month']})")
    
    print(f"   {deactivated_count} students marked as deactivated")

    # -----------------------------------------------------------------------
    # Step 3: Inject Feb–Apr 2026 backup
    # -----------------------------------------------------------------------
    print("\n[Step 3] Injecting Feb–Apr 2026 backup...")
    with open(BACKUP_PATH, 'r', encoding='utf-8') as f:
        backup = json.load(f)

    backup_students = backup.get('students', [])
    backup_scores = backup.get('scores', [])

    # Build roll→canonical_name index for existing students (to detect roll reassignment)
    student_cname_by_id = {
        s['id']: canonical_name(s.get('name', '') or s.get('base_name', ''))
        for s in name_to_student.values()
    }

    # Build mapping: backup_student_id → new_student_id
    # Process ALL backup student records (no dedup) — same-roll same-person variants
    # match via roll+name; roll reassignments are handled by name-only fallback.
    old_id_to_new_id = {}

    for bs in backup_students:
        old_id = bs['id']
        cname = canonical_name(bs.get('name', '') or bs.get('base_name', ''))
        dname = display_name(bs.get('name', '') or bs.get('base_name', ''))
        roll = bs.get('roll', '')
        batch = roll[:4] if roll and len(roll) >= 4 else ''
        is_active = bs.get('active', True)
        s = None

        def same_batch(r1, r2):
            b1 = (r1 or '')[:4]
            b2 = (r2 or '')[:4]
            return not b1 or not b2 or b1 == b2

        def name_match_same_batch():
            """Find existing student by canonical name in same batch."""
            if not cname or cname not in name_to_student:
                return None
            candidate = name_to_student[cname]
            if same_batch(roll, candidate.get('roll', '')):
                return candidate
            return None

        # --- Matching rules ---
        if roll and roll in roll_to_student:
            roll_candidate = roll_to_student[roll]
            roll_cname = student_cname_by_id.get(roll_candidate['id'], '')
            if cname == roll_cname or not cname:
                # Rule 1: Roll + same name → same person, different variant
                s = roll_candidate
            else:
                # Roll matches but different person name
                name_s = name_match_same_batch()
                if name_s:
                    # Person exists elsewhere (e.g. Sahil Yadav on old roll EA24A04
                    # while current roll EA24C02 is already in Excel) → map to them
                    s = name_s
                elif is_active:
                    # Active, different canonical, not in Excel by name
                    # → active roll reassignment (new person on this roll) → new student
                    s = None
                else:
                    # Rule 2 (inactive, roll-diff, name not found):
                    # Could be Aug-Sep 2024 original roll holder → skip
                    s = None
                    old_id_to_new_id[old_id] = roll_candidate['id']  # map scores to current holder
                    continue
        else:
            # No roll match — try by canonical name (same batch)
            s = name_match_same_batch()
            # Rule 3 (inactive, no match at all): skip — old name variant, old roll
            if s is None and not is_active:
                old_id_to_new_id[old_id] = old_id  # drop scores (unmappable inactive)
                continue

        if s:
            new_id = s['id']
            old_roll = s.get('roll', '')
            # Update roll if backup has an ascending promotion
            if roll and roll != old_roll and is_ascending_roll_change(old_roll, roll):
                already = any(
                    r['student_id'] == s['id'] and r['new_roll'] == roll
                    for r in roll_history
                )
                if not already:
                    effective_month = get_effective_month_for_backup_roll_change(
                        cname,
                        roll,
                        '2026-02',
                    )
                    roll_history.append({
                        'student_id': s['id'],
                        'old_roll': old_roll,
                        'new_roll': roll,
                        'effective_month': effective_month,
                        'reason': 'promotion_or_roll_change',
                        'timestamp': NOW_IST,
                    })
                    print(f"   Roll change: {dname} {old_roll} → {roll} (effective {effective_month})")
                s['roll'] = roll
                roll_to_student[roll] = s
                student_cname_by_id[s['id']] = cname or student_cname_by_id.get(s['id'], '')
            # Only apply active/stars/veto from the active variant
            if is_active:
                s['active'] = True
                s['stars'] = bs.get('stars', s.get('stars', 0))
                s['veto_count'] = bs.get('veto_count', s.get('veto_count', 0))
            if bs.get('class') and not s.get('class'):
                s['class'] = bs['class']
        else:
            # Rule 4: Active + no match → genuine new joiner post-Jan 2026
            new_id = student_id_counter[0]
            student_id_counter[0] += 1
            s = {
                'id': new_id,
                'roll': roll,
                'name': dname or roll,
                'base_name': dname or roll,
                'class': bs.get('class'),
                'active': is_active,
                'stars': bs.get('stars', 0),
                'veto_count': bs.get('veto_count', 0),
                'rolls_by_month': {},
            }
            if cname:
                name_to_student[cname] = s
            if roll:
                roll_to_student[roll] = s
            student_cname_by_id[s['id']] = cname
            print(f"   New student from backup: {dname} ({roll})")

        old_id_to_new_id[old_id] = new_id

    # Add backup scores, remapping student IDs
    backup_months_seen = set()
    backup_score_count = 0
    for sc in backup_scores:
        old_sid = sc.get('studentId')
        new_sid = old_id_to_new_id.get(old_sid)
        if new_sid is None:
            continue

        month_key = sc.get('month', '')
        date_str = sc.get('date', '')

        # Skip if this exact score already exists (same student + date from Excel Feb 2026)
        duplicate = any(
            s['studentId'] == new_sid and s['date'] == date_str
            for s in all_scores
        )
        if duplicate:
            continue

        all_scores.append({
            'id': score_id_counter[0],
            'studentId': new_sid,
            'date': date_str,
            'month': month_key,
            'points': sc.get('points', 0),
            'stars': sc.get('stars', 0),
            'vetos': sc.get('vetos', 0),
            'star_usage_normal': sc.get('star_usage_normal', 0),
            'star_usage_disciplinary': sc.get('star_usage_disciplinary', 0),
            'notes': sc.get('notes', ''),
            'created_at': sc.get('created_at', NOW_IST),
            'updated_at': sc.get('updated_at', NOW_IST),
        })
        score_id_counter[0] += 1
        backup_score_count += 1
        backup_months_seen.add(month_key)

    # Build month_roster_profiles and month_students for backup months
    for month_key in sorted(backup_months_seen):
        if month_key not in new_db['months']:
            new_db['months'].append(month_key)

        # Get all students that have scores in this month
        sids_in_month = {s['studentId'] for s in all_scores if s['month'] == month_key}
        id_to_student = {s['id']: s for s in name_to_student.values()}

        profiles = []
        rolls = []
        for sid in sorted(sids_in_month):
            s = id_to_student.get(sid)
            if not s:
                continue
            roll_for_month = get_roll_for_month(sid, month_key) or s.get('roll', '')
            profiles.append({
                'roll': roll_for_month,
                'name': s.get('name', ''),
                'base_name': s.get('base_name', ''),
                'class': s.get('class'),
                'studentId': sid,
                'month_star_count': 0,  # Feb 2026+: governed by live star/veto system
                'month_veto_count': 0,  # (student.stars + score records + role_veto_monthly)
                'month_designations': [],
            })
            rolls.append(roll_for_month)

        month_roster_profiles[month_key] = profiles
        month_students[month_key] = rolls

    print(f"   Backup injection: {backup_score_count} scores across {len(backup_months_seen)} months")

    # -----------------------------------------------------------------------
    # Step 4: Finalize and write database
    # -----------------------------------------------------------------------
    print("\n[Step 4] Finalizing database...")

    # Build final students list
    final_students = []
    seen_ids = set()
    for s in name_to_student.values():
        if s['id'] in seen_ids:
            continue
        seen_ids.add(s['id'])
        entry = {
            'id': s['id'],
            'roll': s.get('roll', ''),
            'name': s.get('name', ''),
            'base_name': s.get('base_name', ''),
            'class': s.get('class'),
            'active': s.get('active', True),
            'stars': s.get('stars', 0),
            'veto_count': s.get('veto_count', 0),
            'last_month_appeared': s.get('last_month_appeared'),
        }
        if s.get('deactivation_month'):
            entry['deactivation_month'] = s['deactivation_month']
        if s.get('active_from_month'):
            entry['active_from_month'] = s['active_from_month']
        final_students.append(entry)
    final_students.sort(key=lambda s: s['id'])

    new_db['students'] = final_students
    new_db['scores'] = all_scores
    new_db['months'] = sorted(set(new_db['months']))
    new_db['month_students'] = month_students
    new_db['month_roster_profiles'] = month_roster_profiles
    new_db['roll_history'] = roll_history
    new_db['_schema_version'] = 23
    new_db['_app_schema_version'] = 23
    new_db['_force_authoritative_pull_once'] = True

    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(new_db, f, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"REBUILD COMPLETE")
    print(f"{'='*60}")
    print(f"  Students : {len(final_students)}")
    print(f"  Scores   : {len(all_scores)}")
    print(f"  Months   : {len(new_db['months'])} → {new_db['months']}")
    print(f"  Roll changes tracked: {len(roll_history)}")
    print(f"  DB written to: {DB_PATH}")
    print(f"\nNEXT STEPS:")
    print(f"  1. Restart Flask server")
    print(f"  2. Clear browser cache (DevTools → Application → Clear site data)")
    print(f"  3. Hard refresh (Ctrl+Shift+R)")


if __name__ == '__main__':
    rebuild()
