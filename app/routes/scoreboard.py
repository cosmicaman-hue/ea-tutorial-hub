from flask import Blueprint, render_template, request, jsonify, current_app, send_file, after_this_request, Response, stream_with_context, g, has_request_context
from flask_login import login_required, current_user
from app import db, csrf, limiter
from app.models import (
    User,
    Proposal,
    ProposalVote,
    ProposalMessage,
    ScoreAdjustmentAction,
    DeviceSession,
    AccountAction,
    JoinCode,
    UserAccessWindow,
    StudentTransfer,
    PublicSiteCredential,
)
from app.utils.syllabus_helpers import merge_syllabus_catalog_superset, merge_syllabus_tracking_superset
from app.utils.file_operations import (
    atomic_write_json as _shared_atomic_write_json,
    ensure_ledger_payload as _ensure_ledger_payload,
)
from app.utils.sync_config import (
    get_sync_peers,
    is_full_ledger_snapshot,
    is_private_peer_url,
    normalize_peer_urls,
    resolve_sync_shared_key,
)
from app.utils.sync_payloads import payload_for_external_replication
from app.config.constants import SCOREBOARD_DEFAULT_LEADERSHIP, SCOREBOARD_DEFAULT_PARTIES, VETO_QUOTAS, VETO_INDIVIDUAL_ALLOCATIONS
import app.utils.score_balance as _score_balance
from app.utils.data_paths import (
    get_storage_root,
    get_data_path as _shared_data_path,
    get_backup_dir as _shared_backup_dir,
    load_json_data_cached as _cached_load_json_data,
    invalidate_data_cache as _invalidate_data_cache,
    prime_data_cache as _prime_data_cache,
    get_serialized_response as _get_serialized_response,
    store_serialized_response as _store_serialized_response,
)
from datetime import datetime, date, timedelta, timezone
from dateutil.relativedelta import relativedelta
import calendar
from werkzeug.utils import secure_filename
import os
import json
import re
import hashlib
import secrets
import tempfile
import shutil
import glob
import subprocess
import urllib.request
import urllib.error
import urllib.parse
from werkzeug.security import generate_password_hash
from zoneinfo import ZoneInfo
from queue import Queue, Empty
import threading
import time
import math
import functools
import logging

_ledger_log = logging.getLogger(__name__)

points_bp = Blueprint('points', __name__, url_prefix='/scoreboard')
_sync_subscribers = []
_sync_lock = threading.Lock()

# Serializes read-merge-write cycles on the offline JSON ledger within this
# process. Waitress serves with multiple threads; without this lock, two
# concurrent mutating requests could each load the ledger, merge independently,
# and the last writer would silently discard the other's merge (lost update).
# RLock so a request already holding the lock can call _save_offline_data
# (which also acquires it) without deadlocking.
_LEDGER_WRITE_LOCK = threading.RLock()


def _ledger_write_guard(view):
    """Route decorator: serialize ledger-mutating requests. Plain GETs skip the lock."""
    @functools.wraps(view)
    def _wrapped(*args, **kwargs):
        if request.method == 'GET':
            return view(*args, **kwargs)
        with _LEDGER_WRITE_LOCK:
            return view(*args, **kwargs)
    return _wrapped

# NOTE: DEFAULT_PARTIES also defined in app/config/constants.py (without 'name' field).
# This version is the operational one used by the scoreboard routes.
DEFAULT_PARTIES = [dict(party) for party in SCOREBOARD_DEFAULT_PARTIES]

# NOTE: DEFAULT_LEADERSHIP also defined in app/config/constants.py (without holder names).
# This version is the operational one with actual holder assignments.
DEFAULT_LEADERSHIP = [dict(post) for post in SCOREBOARD_DEFAULT_LEADERSHIP]


def _storage_root_path():
    """Thin wrapper — logic lives in app.utils.data_paths (single source of truth)."""
    return get_storage_root()


def _legacy_instance_file(name):
    return os.path.join(current_app.instance_path, name)


def _politics_file_path():
    return os.path.join(_storage_root_path(), 'scoreboard_politics.json')


def _offline_data_path():
    """Thin wrapper — logic lives in app.utils.data_paths (single source of truth)."""
    return _shared_data_path()


def _offline_backup_dir():
    """Thin wrapper — logic lives in app.utils.data_paths (single source of truth)."""
    return _shared_backup_dir()

def _offline_hourly_backup_dir():
    return os.path.join(_storage_root_path(), 'offline_scoreboard_hourly_backups')

def _offline_startup_restore_dir():
    return os.path.join(_storage_root_path(), 'startup_restore_points')

def _restore_points_meta_path():
    return os.path.join(_storage_root_path(), 'restore_points_meta.json')


def _device_log_path():
    return os.path.join(_storage_root_path(), 'device_log.json')

def _load_restore_points_meta():
    path = _restore_points_meta_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _save_restore_points_meta(meta):
    _atomic_write_json(_restore_points_meta_path(), meta if isinstance(meta, dict) else {})


def _get_server_timezone():
    return os.getenv('EA_TIMEZONE', 'Asia/Kolkata').strip() or 'Asia/Kolkata'


def _server_now_iso():
    tz_name = _get_server_timezone()
    try:
        return datetime.now(ZoneInfo(tz_name)).isoformat()
    except Exception:
        if tz_name.lower() == 'asia/kolkata':
            return datetime.now(timezone(timedelta(hours=5, minutes=30))).isoformat()
        return datetime.utcnow().isoformat()


def _roll_key(value):
    return str(value or '').strip().upper()


def _name_key(value):
    text = str(value or '').strip().lower()
    text = re.sub(r'\s*\([^)]*\)', ' ', text)
    text = re.sub(r'[^a-z0-9]+', '', text)
    return text


def _safe_int(value, default=0):
    try:
        if value in (None, ''):
            return default
        return int(value)
    except Exception:
        return default


def _safe_float(value, default=0.0):
    try:
        if value in (None, ''):
            return default
        return float(value)
    except Exception:
        return default


def _parse_stamp(value):
    raw = str(value or '').strip()
    if not raw:
        return 0.0
    try:
        return datetime.fromisoformat(raw.replace('Z', '+00:00')).timestamp()
    except Exception:
        return 0.0


def _norm_attendance_status(value):
    status = str(value or '').strip().lower()
    return status if status in {'present', 'absent', 'late', 'leave'} else 'present'


def _attendance_penalty(status):
    if status == 'absent':
        return -20
    if status == 'late':
        return -5
    return 0


def _clean_public_name(value):
    raw = str(value or '').strip()
    if not raw:
        return '-'
    # Remove bracketed annotations like "Name (CR)" for a cleaner public name.
    cleaned = re.sub(r'\s*\([^)]*\)', '', raw)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned or '-'


def _clean_public_photo_path(value):
    """Allow only photo paths generated by the student portrait uploader."""
    raw = str(value or '').strip()
    if re.fullmatch(r'/static/uploads/students/[A-Za-z0-9_-]+\\.(?:jpg|jpeg|png|webp)', raw, re.IGNORECASE):
        return raw
    return ''


def _public_student_photo_path(student):
    if not isinstance(student, dict):
        return ''
    profile = student.get('profile_data') if isinstance(student.get('profile_data'), dict) else {}
    return _clean_public_photo_path(profile.get('photo_path'))


def _recent_public_month_window():
    try:
        now = datetime.now(ZoneInfo(_get_server_timezone()))
    except Exception:
        now = datetime.now()
    out = []
    year = now.year
    month = now.month
    for _ in range(3):
        out.append(f'{year:04d}-{month:02d}')
        month -= 1
        if month <= 0:
            month = 12
            year -= 1
    return out


def _public_information_item_id(value):
    if value in (None, ''):
        return int(time.time() * 1000)
    if isinstance(value, bool):
        return int(time.time() * 1000)
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return int(time.time() * 1000)
    if text.isdigit():
        try:
            return int(text)
        except Exception:
            pass
    return text


def _sanitize_public_planner_data(raw):
    if not isinstance(raw, dict):
        return None
    planner_type = str(raw.get('type') or 'monthly').strip().lower()
    if planner_type not in {'monthly', 'weekly'}:
        planner_type = 'monthly'
    data = raw.get('data') if isinstance(raw.get('data'), dict) else {}
    safe_data = {}
    for key, value in data.items():
        key_text = str(key or '').strip()
        if not key_text:
            continue
        if isinstance(value, dict):
            safe_data[key_text] = {
                str(slot or '').strip(): str(text or '').strip()
                for slot, text in value.items()
                if str(slot or '').strip()
            }
        else:
            safe_data[key_text] = str(value or '').strip()
    out = {
        'type': planner_type,
        'month': _safe_int(raw.get('month'), 0),
        'year': _safe_int(raw.get('year'), 0),
        'weekStart': str(raw.get('weekStart') or '').strip(),
        'data': safe_data,
        'goals': str(raw.get('goals') or '').strip(),
        'notes': str(raw.get('notes') or '').strip(),
    }
    return out


def _public_information_payload_from_rows(rows):
    if not isinstance(rows, list):
        rows = []
    normalized = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        layout = str(item.get('layout') or 'cards').strip().lower()
        if layout not in {'cards', 'table', 'planner'}:
            layout = 'cards'
        title = str(item.get('title') or '').strip()
        content = str(item.get('content') or '').strip()
        planner_data = _sanitize_public_planner_data(item.get('planner_data')) if layout == 'planner' else None
        if layout == 'planner' and not planner_data:
            continue
        if not title and not content and not item.get('table_rows') and not planner_data:
            continue
        entry = {
            'id': _public_information_item_id(item.get('id')),
            'title': title or 'Information',
            'layout': layout,
            'content': content,
            'category': str(item.get('category') or 'general').strip().lower() or 'general',
            'audience': str(item.get('audience') or '').strip(),
            'period_type': str(item.get('period_type') or '').strip().lower(),
            'period_label': str(item.get('period_label') or '').strip(),
            'updated_at': str(item.get('updated_at') or item.get('created_at') or ''),
        }
        if layout == 'table':
            cols = item.get('table_columns', []) or []
            rows_data = item.get('table_rows', []) or []
            cols = [str(col or '').strip() for col in cols if str(col or '').strip()]
            safe_rows = []
            if isinstance(rows_data, list):
                for row in rows_data:
                    if not isinstance(row, list):
                        continue
                    safe_row = [str(cell or '').strip() for cell in row]
                    if any(cell for cell in safe_row):
                        safe_rows.append(safe_row)
            entry['table_columns'] = cols
            entry['table_rows'] = safe_rows
        if layout == 'planner' and planner_data:
            entry['planner_data'] = planner_data
        attachment = item.get('attachment')
        if isinstance(attachment, dict):
            data_url = str(attachment.get('data_url') or '').strip()
            mime = str(attachment.get('mime') or '').strip().lower()
            name = str(attachment.get('name') or '').strip()
            if data_url.startswith('data:') and mime in {'application/pdf', 'image/jpeg', 'image/jpg', 'image/png'}:
                entry['attachment'] = {
                    'name': name or 'attachment',
                    'mime': mime,
                    'data_url': data_url,
                }
        normalized.append(entry)
    normalized.sort(key=lambda x: str(x.get('updated_at') or ''), reverse=True)
    return normalized


def _public_information_payload(payload):
    rows = payload.get('public_information', []) or []
    return _public_information_payload_from_rows(rows)


def _merge_public_information_snapshot(client_rows, server_rows, client_updated_at='', server_updated_at=''):
    """Prefer server ledger for published information unless the client is unsynced.

    Mirrors the LAN ``applyRemoteSnapshot`` guard: stale browser snapshots must
    not resurrect deleted information items during Force Publish.
    """
    client_stamp = _parse_stamp(client_updated_at)
    server_stamp = _parse_stamp(server_updated_at)
    if client_stamp > server_stamp:
        source = client_rows if isinstance(client_rows, list) else []
    else:
        source = server_rows if isinstance(server_rows, list) else []
    return _public_information_payload_from_rows(source)

def _merge_public_chess_champion_snapshot(client_chess, payload):
    """Publish the richest safe chess snapshot available.

    Force Publish sends a browser-built chess snapshot, but sanitization used to
    discard it and rebuild only from the ledger. That dropped Admin-held boards
    whenever the ledger merge lagged the LAN UI. Keep a match-id superset from
    both sources and prefer newer stage/round metadata from the client.
    """
    server_chess = _build_public_chess_champion_payload(payload if isinstance(payload, dict) else {})
    if not isinstance(client_chess, dict):
        return server_chess

    client_matches = client_chess.get('matches') if isinstance(client_chess.get('matches'), list) else []
    server_matches = server_chess.get('matches') if isinstance(server_chess.get('matches'), list) else []
    by_id = {}

    def match_key(match):
        match_id = str(match.get('id') or '').strip()
        if match_id:
            return match_id
        white = match.get('white') if isinstance(match.get('white'), dict) else {}
        black = match.get('black') if isinstance(match.get('black'), dict) else {}
        return '|'.join([
            str(match.get('stage') or ''),
            str(match.get('round') or ''),
            str(match.get('board') or ''),
            str(white.get('roll') or ''),
            str(black.get('roll') or ''),
        ])

    for source in (server_matches, client_matches):
        for match in source:
            if not isinstance(match, dict):
                continue
            if not (match.get('white') or match.get('black')):
                continue
            key = match_key(match)
            previous = by_id.get(key)
            if not previous:
                by_id[key] = dict(match)
                continue
            # Prefer the row that has a recorded result / newer stamp / explicit origin.
            prev_complete = bool(previous.get('result'))
            next_complete = bool(match.get('result'))
            if next_complete and not prev_complete:
                by_id[key] = dict(match)
                continue
            if str(match.get('updatedAt') or '') > str(previous.get('updatedAt') or ''):
                merged_row = dict(previous)
                merged_row.update(match)
                by_id[key] = merged_row
                continue
            if match.get('origin') in {'manual', 'historical'} and previous.get('origin') not in {'manual', 'historical'}:
                previous = dict(previous)
                previous['origin'] = match.get('origin')
                by_id[key] = previous

    merged = dict(server_chess)
    client_stamp = str(client_chess.get('updatedAt') or '')
    server_stamp = str(server_chess.get('updatedAt') or '')
    if client_stamp >= server_stamp:
        for key in (
            'eventName', 'stage', 'stageLabel', 'round', 'basicRoundCount',
            'qualificationCount', 'enrolledCount', 'qualifiedCount', 'champion',
            'series', 'standings', 'scoring', 'updatedAt'
        ):
            if key in client_chess:
                merged[key] = client_chess.get(key)
    merged_matches = list(by_id.values())
    merged_matches.sort(key=lambda row: (
        str(row.get('stage') or ''),
        _safe_int(row.get('round'), 0),
        _safe_int(row.get('board'), 0),
        str(row.get('id') or ''),
    ))
    merged['matches'] = merged_matches
    merged['completedMatches'] = sum(
        1 for row in merged_matches
        if str(row.get('status') or '').lower() == 'completed' or str(row.get('result') or '').strip()
    )
    return merged


def _sanitize_client_public_snapshot(snapshot, payload):
    recent_window = _recent_public_month_window()
    allowed_months = set(recent_window)
    months = [str(m or '').strip() for m in (snapshot.get('months') or []) if re.match(r'^\d{4}-\d{2}$', str(m or '').strip())]
    months = [m for m in months if m in allowed_months]
    months = sorted(dict.fromkeys(months), reverse=True)[:3]
    scoreboard_src = snapshot.get('scoreboard') if isinstance(snapshot.get('scoreboard'), dict) else {}
    scoreboard = {}
    top_full = 15
    photo_by_roll = {}
    for student in payload.get('students', []) or []:
        if not isinstance(student, dict):
            continue
        roll = _roll_key(student.get('roll'))
        photo_path = _public_student_photo_path(student)
        if roll and photo_path:
            photo_by_roll[roll] = photo_path
    for month in months:
        rows = scoreboard_src.get(month) if isinstance(scoreboard_src.get(month), list) else []
        safe_rows = []
        for idx, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            roll = _roll_key(row.get('roll') or '-')
            safe_rows.append({
                'roll': roll,
                'name': _clean_public_name(row.get('name') or ''),
                'class': str(row.get('class') or '').strip(),
                'total': _safe_int(row.get('total')),
                'photo_path': _clean_public_photo_path(row.get('photo_path')) or photo_by_roll.get(roll, ''),
            })
        # Always re-sort by total descending (client rank field is unreliable when
        # the cache was stale at publish time — totals are the ground truth).
        safe_rows.sort(key=lambda item: (-_safe_int(item.get('total'), 0), str(item.get('roll') or '')))
        for i, item in enumerate(safe_rows):
            rank = i + 1
            masked = rank > top_full
            item['rank'] = rank
            item['masked'] = masked
            if masked:
                item['name'] = ''
                item['class'] = ''
                item['photo_path'] = ''
        scoreboard[month] = safe_rows
    client_parties = _sanitize_client_public_parties(snapshot.get('parties'), months)
    # Older LAN browser bundles did not send a `parties` field. Do not publish
    # an empty party snapshot just because the scoreboard portion is valid;
    # rebuild the missing months from the authoritative server payload instead.
    server_parties = _sanitize_client_public_parties(
        _build_public_parties_payload(payload, months), months
    )
    for month in months:
        if not isinstance(client_parties.get(month), list) or not client_parties.get(month):
            client_parties[month] = server_parties.get(month, [])

    return {
        'updated_at': str(snapshot.get('updated_at') or payload.get('server_updated_at') or _server_now_iso()),
        'top_full_count': top_full,
        'months': months,
        'scoreboard': scoreboard,
        'parties': client_parties,
        'chess_champion': _merge_public_chess_champion_snapshot(snapshot.get('chess_champion'), payload),
        'public_information': _merge_public_information_snapshot(
            snapshot.get('public_information') if isinstance(snapshot.get('public_information'), list) else [],
            payload.get('public_information') if isinstance(payload, dict) else [],
            str(snapshot.get('updated_at') or ''),
            str((payload or {}).get('server_updated_at') or (payload or {}).get('updated_at') or ''),
        ),
    }


# Vibrant public-site party accents matched to LAN Party Standings ledger hues.
# Neon/near-white values are clamped for contrast while keeping each party distinct.
_PUBLIC_PARTY_COLORS = {
    'MAP': '#f0a400',  # gold
    'BWP': '#ef4444',  # vivid red
    'ESP': '#3b82f6',  # blue (LAN near-white has no usable hue)
    'MRP': '#22c55e',  # lime green
    'SSP': '#06b6d4',  # cyan (LAN #14ffef)
    'NJP': '#f97316',  # orange
    'EYP': '#d946ef',  # magenta (LAN #f500ed)
}


def _hex_to_rgb(value):
    text = str(value or '').strip().lstrip('#')
    if len(text) == 3:
        text = ''.join(ch * 2 for ch in text)
    if len(text) != 6 or any(ch not in '0123456789abcdefABCDEF' for ch in text):
        return None
    try:
        return tuple(int(text[i:i + 2], 16) for i in (0, 2, 4))
    except Exception:
        return None


def _relative_luminance(hex_color):
    rgb = _hex_to_rgb(hex_color)
    if not rgb:
        return 0.5

    def channel(c):
        x = c / 255.0
        return x / 12.92 if x <= 0.03928 else ((x + 0.055) / 1.055) ** 2.4

    r, g, b = (channel(v) for v in rgb)
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def _darken_hex(hex_color, factor=0.55):
    rgb = _hex_to_rgb(hex_color)
    if not rgb:
        return '#64748b'
    factor = max(0.15, min(1.0, float(factor)))
    darkened = tuple(max(0, min(255, int(round(c * factor)))) for c in rgb)
    return '#{:02x}{:02x}{:02x}'.format(*darkened)


def _normalize_hex_color(value):
    color = str(value or '').strip()
    if not re.match(r'^#[0-9A-Fa-f]{3}([0-9A-Fa-f]{3})?$', color):
        return ''
    if len(color) == 4:
        color = '#' + ''.join(ch * 2 for ch in color[1:])
    return color.lower()


def _hex_chroma(hex_color):
    rgb = _hex_to_rgb(hex_color)
    if not rgb:
        return 0
    return max(rgb) - min(rgb)


def _clamp_public_party_color(hex_color, target_max_luminance=0.48):
    """Darken while preserving hue until white-on-color text stays readable."""
    out = _normalize_hex_color(hex_color) or '#64748b'
    for _ in range(10):
        if _relative_luminance(out) <= target_max_luminance:
            return out
        out = _darken_hex(out, 0.84)
    return out


def _public_party_color(code, raw_color=None):
    """Prefer LAN Party Standings colors; clamp only when contrast would fail."""
    key = str(code or '').strip().upper()
    raw = _normalize_hex_color(raw_color)
    # Near-white / washed colors (e.g. ESP #f0f4ff) have no usable hue — fall back.
    if raw and not (_relative_luminance(raw) > 0.82 and _hex_chroma(raw) < 40):
        return _clamp_public_party_color(raw)
    if key in _PUBLIC_PARTY_COLORS:
        return _clamp_public_party_color(_PUBLIC_PARTY_COLORS[key])
    return '#64748b'


def _sanitize_client_public_parties(raw_parties, allowed_months):
    """Sanitize the client-built per-month party standings snapshot.

    Only months present in ``allowed_months`` are kept. Each party row is
    reduced to a safe, serializable subset (no student IDs or internal state).
    Member names are NOT masked here — party membership is a public governance
    fact, not a ranked-score disclosure, so the top-15 masking rule that applies
    to the scoreboard does not apply to party rosters.
    """
    if not isinstance(raw_parties, dict):
        return {}
    allowed = set(allowed_months)
    out = {}
    for month, rows in raw_parties.items():
        m = str(month or '').strip()
        if m not in allowed:
            continue
        if not isinstance(rows, list):
            continue
        safe_rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            members = []
            for mem in (row.get('members') or []):
                if not isinstance(mem, dict):
                    continue
                members.append({
                    'roll': str(mem.get('roll') or '-').strip().upper(),
                    'name': _clean_public_name(mem.get('name') or ''),
                    'class': str(mem.get('class') or '').strip(),
                    'role': str(mem.get('role') or 'Member').strip(),
                    'score': _safe_int(mem.get('score')),
                    'votePower': max(1, min(25, _safe_int(mem.get('votePower'), 1))),
                })
            code = str(row.get('code') or '').strip().upper()[:10]
            safe_rows.append({
                'code': code,
                'fullName': str(row.get('fullName') or '').strip()[:120],
                'symbol': str(row.get('symbol') or '').strip()[:8],
                'color': _public_party_color(code, row.get('color')),
                'memberCount': _safe_int(row.get('memberCount')),
                'totalScore': _safe_int(row.get('totalScore')),
                'avgScore': _safe_int(row.get('avgScore')),
                'votePower': max(0, _safe_int(row.get('votePower'))),
                'rank': _safe_int(row.get('rank')),
                'members': members,
            })
        out[m] = safe_rows
    return out


def _public_snapshot_has_useful_rows(snapshot):
    if not isinstance(snapshot, dict):
        return False
    months = snapshot.get('months') if isinstance(snapshot.get('months'), list) else []
    scoreboard = snapshot.get('scoreboard') if isinstance(snapshot.get('scoreboard'), dict) else {}
    if not months or not scoreboard:
        return False
    for month in months:
        rows = scoreboard.get(month)
        if isinstance(rows, list) and any(isinstance(row, dict) for row in rows):
            return True
    return False


def _public_month_keys(payload):
    months = set()
    for month in payload.get('months', []) or []:
        item = str(month or '').strip()
        if re.match(r'^\d{4}-\d{2}$', item):
            months.add(item)
    month_students = payload.get('month_students', {}) if isinstance(payload.get('month_students'), dict) else {}
    for month in month_students.keys():
        item = str(month or '').strip()
        if re.match(r'^\d{4}-\d{2}$', item):
            months.add(item)
    month_profiles = payload.get('month_roster_profiles', {}) if isinstance(payload.get('month_roster_profiles'), dict) else {}
    for month in month_profiles.keys():
        item = str(month or '').strip()
        if re.match(r'^\d{4}-\d{2}$', item):
            months.add(item)
    for score in payload.get('scores', []) or []:
        item = str((score or {}).get('month') or '').strip()
        if re.match(r'^\d{4}-\d{2}$', item):
            months.add(item)
    for attendance in payload.get('attendance', []) or []:
        item = str((attendance or {}).get('date') or '').strip()
        if len(item) >= 7:
            month = item[:7]
            if re.match(r'^\d{4}-\d{2}$', month):
                months.add(month)
    return sorted(months, reverse=True)


def _get_month_roster_rolls(payload, month):
    rolls = set()
    for value in payload.get('month_students', {}).get(month, []) or []:
        roll = _roll_key(value)
        if roll:
            rolls.add(roll)
    for profile in payload.get('month_roster_profiles', {}).get(month, []) or []:
        roll = _roll_key((profile or {}).get('roll'))
        if roll:
            rolls.add(roll)
    return rolls


def _build_public_month_rows(payload, month):
    # month_roster_profiles is the authoritative historical snapshot:
    # it records which student ID held which roll in each specific month.
    # Using it (rather than the current students list) ensures that months
    # before a roll reassignment show the original holder with their correct
    # name and scores, not the current holder of that roll number.
    hist_profiles = (payload.get('month_roster_profiles', {}) or {}).get(month, []) or []

    hist_profiles_have_ids = False
    if hist_profiles:
        by_id = {}
        for profile in hist_profiles:
            if not isinstance(profile, dict):
                continue
            sid = _safe_int(profile.get('studentId'))
            if sid > 0 and sid not in by_id:
                by_id[sid] = profile
                hist_profiles_have_ids = True
        totals = {sid: 0 for sid in by_id}
    if not hist_profiles or not hist_profiles_have_ids:
        # Fallback for months with no roster snapshot: use current student list.
        roster_rolls = _get_month_roster_rolls(payload, month)
        base_students = [
            student for student in (payload.get('students', []) or [])
            if isinstance(student, dict) and student.get('active', True) is not False
        ]
        if roster_rolls:
            visible_students = [
                student for student in base_students
                if _roll_key(student.get('roll')) in roster_rolls
            ]
        else:
            visible_students = base_students

        dedup_by_roll = {}
        for student in visible_students:
            roll = _roll_key(student.get('roll'))
            if roll and roll not in dedup_by_roll:
                dedup_by_roll[roll] = student

        by_id = {}
        totals = {}
        for student in dedup_by_roll.values():
            sid = _safe_int(student.get('id'))
            if sid <= 0:
                continue
            row_student = dict(student)
            if hist_profiles:
                profile_by_roll = {
                    _roll_key((profile or {}).get('roll')): profile
                    for profile in hist_profiles
                    if isinstance(profile, dict) and _roll_key((profile or {}).get('roll'))
                }
                matched_profile = profile_by_roll.get(_roll_key(student.get('roll')))
                if isinstance(matched_profile, dict):
                    row_student['roll'] = matched_profile.get('roll') or row_student.get('roll')
                    row_student['base_name'] = matched_profile.get('base_name') or matched_profile.get('name') or row_student.get('base_name') or row_student.get('name')
                    row_student['name'] = matched_profile.get('name') or matched_profile.get('base_name') or row_student.get('name') or row_student.get('base_name')
                    row_student['class'] = matched_profile.get('class') or row_student.get('class')
            by_id[sid] = row_student
            totals[sid] = 0

    month_scores = []
    for score in payload.get('scores', []) or []:
        if not isinstance(score, dict):
            continue
        if str(score.get('month') or '').strip() != month:
            continue
        month_scores.append(score)

    # Historical imports can contain canonical total-column rows. When present,
    # use those totals directly to avoid day-wise re-summing mismatches.
    excel_total_by_sid = {}
    for score in month_scores:
        note = str(score.get('notes') or '').strip().lower()
        if note.startswith('excel_total_score') or note.startswith('excel_total_from_dates'):
            sid = _safe_int(score.get('studentId'))
            if sid > 0 and sid in totals:
                excel_total_by_sid[sid] = excel_total_by_sid.get(sid, 0) + _safe_float(score.get('points'))
    if excel_total_by_sid:
        for sid in totals.keys():
            totals[sid] = excel_total_by_sid.get(sid, 0)
        rows = []
        for sid, total in totals.items():
            student = by_id.get(sid, {})
            rows.append({
                'roll': _roll_key(student.get('roll')),
                'name': str(student.get('base_name') or student.get('name') or '').strip() or '-',
                'classVal': student.get('class') or '-',
                'total': round(total, 2) or 0,
                'photo_path': _public_student_photo_path(student),
            })
        return sorted(rows, key=lambda item: (-item['total'], item['roll']))

    is_historical = month < '2026-02'
    hist_totals = {}
    hist_has = set()
    per_date_score = {}
    per_date_star_normal = {}
    per_date_star_disciplinary = {}
    apply_star_bonus = month >= '2026-02'

    for score in month_scores:
        sid = _safe_int(score.get('studentId'))
        if sid <= 0 or sid not in totals:
            continue

        note = str(score.get('notes') or '').strip().lower()
        points = _safe_float(score.get('points'))
        if is_historical:
            is_excel_total = note.startswith('excel_total_score') or note.startswith('excel_total_from_dates')
            is_excel_daily = note.startswith('excel_daily_score')
            is_excel_star = note.startswith('excel_star_usage')
            if not (is_excel_total or is_excel_daily or is_excel_star):
                continue
            if is_excel_total:
                hist_totals[sid] = hist_totals.get(sid, 0) + points
                hist_has.add(sid)
                continue

        score_date = str(score.get('date') or '').strip()
        totals[sid] = totals.get(sid, 0) + points

        if apply_star_bonus and score_date:
            date_scores = per_date_score.setdefault(sid, {})
            date_scores[score_date] = date_scores.get(score_date, 0) + points

            star_delta = _safe_int(score.get('stars'))
            if star_delta < 0:
                usage_abs = abs(star_delta)
                normal = max(0, _safe_int(score.get('star_usage_normal')))
                disciplinary = max(0, _safe_int(score.get('star_usage_disciplinary')))
                if normal + disciplinary <= 0:
                    is_transfer = bool(score.get('star_transfer_out') or score.get('star_transfer_in') or '[star transfer' in note)
                    if is_transfer:
                        normal = 0
                        disciplinary = 0
                    elif 'disciplinary' in note:
                        disciplinary = usage_abs
                        normal = 0
                    else:
                        normal = usage_abs
                        disciplinary = 0
                if normal > 0:
                    date_norm = per_date_star_normal.setdefault(sid, {})
                    date_norm[score_date] = date_norm.get(score_date, 0) + normal
                if disciplinary > 0:
                    date_disc = per_date_star_disciplinary.setdefault(sid, {})
                    date_disc[score_date] = date_disc.get(score_date, 0) + disciplinary

    if apply_star_bonus:
        for sid, date_map in per_date_star_disciplinary.items():
            if sid not in totals:
                continue
            date_scores = per_date_score.get(sid, {})
            normal_map = per_date_star_normal.get(sid, {})
            for score_date, disc_uses in date_map.items():
                normal_uses = max(0, _safe_int(normal_map.get(score_date)))
                if normal_uses > 0:
                    continue
                day_score = _safe_int(date_scores.get(score_date))
                if day_score < 0:
                    halved = day_score
                    for _ in range(disc_uses):
                        halved = math.floor(halved / 2)
                    totals[sid] = totals.get(sid, 0) - (day_score - halved)
                    date_scores[score_date] = halved

    if apply_star_bonus:
        for sid, date_map in per_date_star_normal.items():
            if sid not in totals:
                continue
            date_scores = per_date_score.get(sid, {})
            for score_date, normal_uses in date_map.items():
                if normal_uses <= 0:
                    continue
                day_score = _safe_int(date_scores.get(score_date))
                if day_score < 0:
                    totals[sid] = totals.get(sid, 0) - day_score
                    date_scores[score_date] = 0
                if day_score > -50:
                    totals[sid] = totals.get(sid, 0) + (100 * normal_uses)

    attendance_latest = {}
    for attendance in payload.get('attendance', []) or []:
        if not isinstance(attendance, dict):
            continue
        sid = _safe_int(attendance.get('studentId'))
        day = str(attendance.get('date') or '').strip()
        if sid <= 0 or not day.startswith(month):
            continue
        key = f'{sid}|{day}'
        prev = attendance_latest.get(key)
        prev_ts = _parse_stamp((prev or {}).get('updated_at') or (prev or {}).get('created_at'))
        next_ts = _parse_stamp(attendance.get('updated_at') or attendance.get('created_at'))
        if prev is None or next_ts >= prev_ts:
            attendance_latest[key] = attendance

    for attendance in attendance_latest.values():
        sid = _safe_int(attendance.get('studentId'))
        if sid in totals:
            # GCB-immune students are not subject to absence/late penalties.
            if not (by_id.get(sid) or {}).get('gcb'):
                totals[sid] = totals.get(sid, 0) + _attendance_penalty(_norm_attendance_status(attendance.get('status')))

    if is_historical:
        for sid in hist_has:
            if sid in totals:
                totals[sid] = hist_totals.get(sid, 0)

    rows = []
    for sid, total in totals.items():
        student = by_id.get(sid, {})
        rows.append({
            'roll': _roll_key(student.get('roll')),
            'name': _clean_public_name(student.get('base_name') or student.get('name') or ''),
            'classVal': student.get('class') or '-',
            'total': total or 0,
            'photo_path': _public_student_photo_path(student),
        })
    return sorted(rows, key=lambda item: (-item['total'], item['roll']))


def _build_public_site_payload(payload):
    recent_window = _recent_public_month_window()
    months = list(recent_window)
    scoreboard = {}
    top_full = 15
    for month in months:
        rows = _build_public_month_rows(payload, month)
        public_rows = []
        for idx, row in enumerate(rows):
            rank = idx + 1
            masked = rank > top_full
            public_rows.append({
                'rank': rank,
                'roll': row.get('roll') or '-',
                'name': row.get('name') if not masked else '',
                'class': row.get('classVal') if not masked else '',
                'total': row.get('total') or 0,
                'photo_path': _clean_public_photo_path(row.get('photo_path')) if not masked else '',
                'masked': masked,
            })
        scoreboard[month] = public_rows
    return {
        'updated_at': str(payload.get('server_updated_at') or _server_now_iso()),
        'top_full_count': top_full,
        'months': months,
        'scoreboard': scoreboard,
        'parties': _build_public_parties_payload(payload, months),
        'chess_champion': _build_public_chess_champion_payload(payload),
        'public_information': _public_information_payload(payload),
    }


def _build_public_parties_payload(payload, months):
    """Server-side fallback for the per-month party standings snapshot.

    Used only when the client-provided public_snapshot is missing or incomplete.
    Mirrors the JS _buildPublicPartyStandingsForMonth logic: active members only
    (member.status active AND student not explicitly deactivated for the month),
    per-member month total from scores, vote power = clamp(floor(total/20)+1, 1, 25).
    """
    parties_raw = payload.get('parties') if isinstance(payload.get('parties'), list) else []
    students_raw = payload.get('students') if isinstance(payload.get('students'), list) else []
    student_by_id = {}
    for student in students_raw:
        if not isinstance(student, dict):
            continue
        sid = _safe_int(student.get('id'))
        if sid > 0:
            student_by_id[sid] = student

    def _is_deactivated(student, month):
        if not student:
            return True
        if student.get('active') is False:
            return True
        dm = str(student.get('deactivation_month') or '').strip()
        if dm and month and month >= dm:
            return True
        return False

    def _status_active(value):
        text = str(value or 'active').strip().lower()
        return text not in ('suspended', 'vacant', 'ended')

    out = {}
    for month in months:
        # Pre-compute per-student month totals from scores for this month.
        month_totals = {}
        for score in payload.get('scores', []) or []:
            if not isinstance(score, dict):
                continue
            if str(score.get('month') or '').strip() != month:
                continue
            sid = _safe_int(score.get('studentId'))
            if sid > 0:
                month_totals[sid] = month_totals.get(sid, 0) + _safe_float(score.get('points'))

        rows = []
        for party in parties_raw:
            if not isinstance(party, dict):
                continue
            code = str(party.get('code') or '').strip().upper()[:10]
            if not code:
                continue
            members_in = party.get('members') if isinstance(party.get('members'), list) else []
            member_rows = []
            total_score = 0
            vote_power = 0
            for mem in members_in:
                if not isinstance(mem, dict):
                    continue
                if not _status_active(mem.get('status')):
                    continue
                sid = _safe_int(mem.get('studentId'))
                if sid <= 0:
                    continue
                student = student_by_id.get(sid)
                if not student or _is_deactivated(student, month):
                    continue
                member_score = round(month_totals.get(sid, 0))
                total_score += member_score
                member_vp = max(1, min(25, (member_score // 20) + 1))
                vote_power += member_vp
                designation = str(mem.get('designation') or '').strip().lower()
                role = 'President' if designation in ('president', 'party president', 'pp') else (
                    'Deputy President' if designation in ('deputy_president', 'deputy party president', 'dpp') else 'Member'
                )
                member_rows.append({
                    'roll': _roll_key(student.get('roll')) or '-',
                    'name': _clean_public_name(student.get('base_name') or student.get('name') or ''),
                    'class': str(student.get('class') or '').strip(),
                    'role': role,
                    'score': member_score,
                    'votePower': member_vp,
                })
            member_rows.sort(key=lambda m: -m['score'])
            rows.append({
                'code': code,
                'fullName': str(party.get('fullName') or party.get('full_name') or code).strip()[:120],
                'symbol': str(party.get('symbol') or '').strip()[:8],
                'color': _public_party_color(code, party.get('color')),
                'memberCount': len(member_rows),
                'totalScore': round(total_score),
                'avgScore': round(total_score / len(member_rows)) if member_rows else 0,
                'votePower': vote_power,
                'rank': 0,
                'members': member_rows,
            })
        rows.sort(key=lambda r: -r['totalScore'])
        for i, r in enumerate(rows):
            r['rank'] = i + 1
        out[month] = rows
    return out


def _build_public_chess_champion_payload(payload):
    """Build a safe, read-only tournament snapshot for the Cloudflare site.

    The canonical tournament ledger stores only student IDs in match records. This
    transform resolves those IDs to the small public identity subset needed by the
    bracket and standings, without exposing the full offline ledger.
    """
    raw = payload.get('chess_champion') if isinstance(payload, dict) else None
    state = _normalize_chess_champion(raw)
    students = payload.get('students', []) if isinstance(payload, dict) else []
    by_id = {}
    for student in students if isinstance(students, list) else []:
        if not isinstance(student, dict):
            continue
        sid = _safe_int(student.get('id'))
        if sid > 0:
            by_id[sid] = student

    def public_student(student_id):
        student = by_id.get(_safe_int(student_id))
        if not student:
            return None
        return {
            'roll': _roll_key(student.get('roll')) or '-',
            'name': _clean_public_name(student.get('base_name') or student.get('name') or ''),
            'class': str(student.get('class') or student.get('class_name') or '').strip() or '-',
            'group': str(student.get('group') or student.get('group_name') or '').strip().upper(),
        }

    enrolled_ids = []
    for value in state.get('enrolledStudentIds', []) if isinstance(state.get('enrolledStudentIds'), list) else []:
        sid = _safe_int(value)
        if sid > 0 and sid not in enrolled_ids:
            enrolled_ids.append(sid)

    standings = {}
    for sid in enrolled_ids:
        standings[sid] = {
            'id': sid,
            'student': public_student(sid),
            'points': 0,
            'matches': 0,
            'wins': 0,
            'draws': 0,
            'losses': 0,
            'killPoints': 0,
            'bonusPoints': 0,
        }

    safe_matches = []
    matches = state.get('matches', []) if isinstance(state.get('matches'), list) else []
    for raw_match in matches:
        if not isinstance(raw_match, dict):
            continue
        white_id = _safe_int(raw_match.get('whiteId'))
        black_id = _safe_int(raw_match.get('blackId'))
        result = str(raw_match.get('result') or '').strip().lower()
        status = str(raw_match.get('status') or ('completed' if result else 'scheduled')).strip().lower()
        complete = status == 'completed' and bool(result)
        advancing_id = _safe_int(raw_match.get('advancingId') or raw_match.get('winnerId'))
        advancing_side = 'white' if advancing_id == white_id and white_id > 0 else ('black' if advancing_id == black_id and black_id > 0 else '')
        awards = raw_match.get('awards') if isinstance(raw_match.get('awards'), dict) else {}
        breakdown = raw_match.get('breakdown') if isinstance(raw_match.get('breakdown'), dict) else {}

        if complete:
            for sid, side in ((white_id, 'white'), (black_id, 'black')):
                if sid <= 0:
                    continue
                if sid not in standings:
                    standings[sid] = {
                        'id': sid,
                        'student': public_student(sid),
                        'points': 0,
                        'matches': 0,
                        'wins': 0,
                        'draws': 0,
                        'losses': 0,
                        'killPoints': 0,
                        'bonusPoints': 0,
                    }
                row = standings[sid]
                row['matches'] += 1
                award = _safe_float(awards.get(str(sid), awards.get(sid, 0)))
                detail = breakdown.get(side) if isinstance(breakdown.get(side), dict) else {}
                row['points'] += award
                row['killPoints'] += _safe_float(detail.get('killPoints'))
                row['bonusPoints'] += (
                    _safe_float(detail.get('classBonus'))
                    + _safe_float(detail.get('opponentBonus'))
                    + _safe_float(detail.get('timeBonus'))
                    + _safe_float(detail.get('killBonus'))
                )
                if result == 'draw':
                    row['draws'] += 1
                elif result == 'bye' or _safe_int(raw_match.get('winnerId')) == sid:
                    row['wins'] += 1
                else:
                    row['losses'] += 1

        origin = str(raw_match.get('origin') or 'system').strip().lower()
        if origin not in {'system', 'historical', 'manual'}:
            origin = 'system'
        safe_matches.append({
            'id': str(raw_match.get('id') or ''),
            'stage': str(raw_match.get('stage') or 'basic').strip(),
            'round': max(1, _safe_int(raw_match.get('round'), 1)),
            'board': max(1, _safe_int(raw_match.get('board'), 1)),
            'status': status,
            'result': result,
            'origin': origin,
            'winnerSide': result if result in {'white', 'black'} else '',
            'advancingSide': advancing_side,
            'whiteSeed': max(0, _safe_int(raw_match.get('whiteSeed'))),
            'blackSeed': max(0, _safe_int(raw_match.get('blackSeed'))),
            'white': public_student(white_id),
            'black': public_student(black_id),
            'durationSeconds': max(0, _safe_int(raw_match.get('durationSeconds'))),
            'killPoints': max(0, _safe_float(raw_match.get('killPoints'))),
            'whitePoints': _safe_float(awards.get(str(white_id), awards.get(white_id, 0))),
            'blackPoints': _safe_float(awards.get(str(black_id), awards.get(black_id, 0))),
            'updatedAt': str(raw_match.get('updatedAt') or raw_match.get('createdAt') or '').strip(),
        })

    sorted_standings = sorted(
        [row for row in standings.values() if row.get('student')],
        key=lambda row: (
            -_safe_float(row.get('points')),
            -_safe_int(row.get('wins')),
            -_safe_float(row.get('killPoints')),
            -_safe_float(row.get('bonusPoints')),
            str((row.get('student') or {}).get('name') or ''),
        ),
    )
    for index, row in enumerate(sorted_standings, 1):
        row['rank'] = index
        row['points'] = round(_safe_float(row.get('points')), 2)
        row['killPoints'] = round(_safe_float(row.get('killPoints')), 2)
        row['bonusPoints'] = round(_safe_float(row.get('bonusPoints')), 2)
        identity = row.pop('student', None) or {}
        row.update({
            'roll': str(identity.get('roll') or '-'),
            'name': str(identity.get('name') or '').strip(),
            'class': str(identity.get('class') or '-').strip() or '-',
            'group': str(identity.get('group') or '').strip().upper(),
        })
        row.pop('id', None)

    stage = str(state.get('stage') or 'registration').strip()
    round_number = max(0, _safe_int(state.get('round')))
    stage_labels = {
        'registration': 'Registration',
        'basic': f'Basic rounds · Round {round_number or 1}',
        'elimination': f'Elimination · Round {round_number or 1}',
        'semifinal': 'Semifinals',
        'finale': 'Finale',
        'complete': 'Champion crowned',
    }
    scoring = state.get('scoring') if isinstance(state.get('scoring'), dict) else {}
    qualified_ids = state.get('qualifiedStudentIds', []) if isinstance(state.get('qualifiedStudentIds'), list) else []
    champion_id = _safe_int(state.get('championStudentId'))
    if champion_id <= 0 and stage == 'complete':
        for raw_match in reversed(matches):
            if not isinstance(raw_match, dict) or str(raw_match.get('stage') or '').strip().lower() != 'finale':
                continue
            if str(raw_match.get('result') or '').strip().lower() not in {'white', 'black', 'draw', 'bye'}:
                continue
            champion_id = _safe_int(raw_match.get('advancingId') or raw_match.get('winnerId'))
            if champion_id > 0:
                break
    completed_count = sum(1 for item in safe_matches if item.get('status') == 'completed' or item.get('result') in {'white', 'black', 'draw', 'bye'})
    public_series = []
    for item in state.get('seriesArchive', []) if isinstance(state.get('seriesArchive'), list) else []:
        champion = public_student(item.get('championStudentId'))
        item_matches = item.get('matches') if isinstance(item.get('matches'), list) else []
        item_completed = sum(1 for match in item_matches if isinstance(match, dict) and (match.get('status') == 'completed' or str(match.get('result') or '').lower() in {'white', 'black', 'draw', 'bye'}))
        public_series.append({
            'id': str(item.get('id') or ''),
            'eventName': str(item.get('eventName') or 'Excel Chess Champion').strip()[:80],
            'stage': str(item.get('stage') or 'registration'),
            'stageLabel': stage_labels.get(str(item.get('stage') or ''), 'Registration'),
            'round': max(0, _safe_int(item.get('round'))),
            'enrolledCount': len(item.get('enrolledStudentIds') or []),
            'matchCount': len(item_matches),
            'completedMatches': item_completed,
            'champion': champion,
            'archivedAt': str(item.get('archivedAt') or '').strip(),
        })
    return {
        'eventName': str(state.get('eventName') or 'Excel Chess Champion').strip()[:80],
        'stage': stage if stage in stage_labels else 'registration',
        'stageLabel': stage_labels.get(stage, 'Registration'),
        'round': round_number,
        'basicRoundCount': 3,
        'qualificationCount': 8,
        'eliminationRoundCount': 3,
        'semifinalRoundCount': 3,
        'enrolledCount': len(enrolled_ids),
        'qualifiedCount': len([value for value in qualified_ids if _safe_int(value) > 0]),
        'champion': public_student(champion_id),
        'series': public_series,
        'completedMatches': completed_count,
        'updatedAt': str(state.get('updatedAt') or payload.get('server_updated_at') or payload.get('updated_at') or '').strip(),
        'scoring': {
            'winPoints': _safe_float(scoring.get('winPoints'), 3),
            'drawPoints': _safe_float(scoring.get('drawPoints'), 1),
            'lossPoints': _safe_float(scoring.get('lossPoints'), 0),
            'classBonusPerLevel': _safe_float(scoring.get('classBonusPerLevel'), 1),
            'opponentBonusPerFivePoints': _safe_float(scoring.get('opponentBonusPerFivePoints'), 1),
            'fastWinSeconds': max(60, _safe_int(scoring.get('fastWinSeconds'), 600)),
            'standardWinSeconds': max(60, _safe_int(scoring.get('standardWinSeconds'), 1200)),
            'fastTimeBonus': _safe_float(scoring.get('fastTimeBonus'), 3),
            'standardTimeBonus': _safe_float(scoring.get('standardTimeBonus'), 1),
            'killPointMultiplier': _safe_float(scoring.get('killPointMultiplier'), 1),
        },
        'standings': sorted_standings,
        'matches': [item for item in safe_matches if item.get('white') or item.get('black')],
    }


_CHESS_STAGE_VALUES = {'registration', 'basic', 'elimination', 'semifinal', 'finale', 'complete'}
_CHESS_MATCH_STAGES = {'basic', 'elimination', 'semifinal', 'finale'}
_CHESS_RESULT_VALUES = {'', 'white', 'black', 'draw', 'bye'}


def _normalize_chess_champion_archive(raw, index=0):
    source = raw if isinstance(raw, dict) else {}
    normalized = _normalize_chess_champion({**source, 'seriesArchive': []})
    return {
        'id': str(source.get('id') or f'series-{index}')[:120],
        'eventName': normalized['eventName'],
        'stage': normalized['stage'],
        'status': normalized['status'],
        'round': normalized['round'],
        'basicRoundCount': normalized['basicRoundCount'],
        'qualificationCount': normalized['qualificationCount'],
        'enrolledStudentIds': normalized['enrolledStudentIds'],
        'qualifiedStudentIds': normalized['qualifiedStudentIds'],
        'championStudentId': normalized['championStudentId'],
        'scoring': normalized['scoring'],
        'matches': normalized['matches'],
        'archivedAt': str(source.get('archivedAt') or source.get('updatedAt') or '').strip()[:80],
    }


def _normalize_chess_champion(raw):
    """Return a bounded, merge-safe Chess Champion state.

    The browser owns the tournament workflow, but the JSON ledger is shared by
    multiple clients and peers. Keep a defensive server-side representation so
    malformed or partial snapshots cannot erase a populated tournament.
    """
    source = raw if isinstance(raw, dict) else {}
    scoring = source.get('scoring') if isinstance(source.get('scoring'), dict) else {}

    def bounded(value, minimum, maximum, fallback):
        number = _safe_float(value, fallback)
        return max(minimum, min(maximum, number))

    def bounded_int(value, minimum, maximum, fallback):
        return int(round(bounded(value, minimum, maximum, fallback)))

    def normalize_breakdown(raw_breakdown):
        if not isinstance(raw_breakdown, dict):
            return {}
        output = {}
        for side in ('white', 'black'):
            row = raw_breakdown.get(side)
            if not isinstance(row, dict):
                continue
            output[side] = {
                key: bounded(row.get(key), 0, 10000, 0)
                for key in ('base', 'classBonus', 'opponentBonus', 'timeBonus', 'iplTotal', 'killPoints', 'killBonus', 'total')
                if key in row
            }
        return output

    matches = []
    raw_matches = source.get('matches') if isinstance(source.get('matches'), list) else []
    for index, raw_match in enumerate(raw_matches[:5000]):
        if not isinstance(raw_match, dict):
            continue
        stage = str(raw_match.get('stage') or 'basic').strip().lower()
        if stage not in _CHESS_MATCH_STAGES:
            stage = 'basic'
        result = str(raw_match.get('result') or '').strip().lower()
        if result not in _CHESS_RESULT_VALUES:
            result = ''
        white_id = max(0, _safe_int(raw_match.get('whiteId'), 0))
        black_id = max(0, _safe_int(raw_match.get('blackId'), 0))
        winner_id = max(0, _safe_int(raw_match.get('winnerId'), 0))
        if result == 'white':
            winner_id = white_id
        elif result == 'black':
            winner_id = black_id
        elif result in {'draw', 'bye'}:
            winner_id = 0
        white_seed = max(0, min(1000, _safe_int(raw_match.get('whiteSeed'), 0)))
        black_seed = max(0, min(1000, _safe_int(raw_match.get('blackSeed'), 0)))
        advancing_id = max(0, _safe_int(raw_match.get('advancingId'), 0))
        if not advancing_id and result in {'white', 'black'}:
            advancing_id = winner_id
        if not advancing_id and result == 'draw' and stage in _CHESS_MATCH_STAGES - {'basic'} and white_seed and black_seed:
            advancing_id = white_id if white_seed < black_seed else black_id
        raw_kills = raw_match.get('killPointsBySide') if isinstance(raw_match.get('killPointsBySide'), dict) else {}
        legacy_kills = bounded(raw_match.get('killPoints'), 0, 1000, 0)
        kill_points_by_side = {
            'white': bounded(raw_kills.get('white'), 0, 1000, legacy_kills if result == 'white' else 0),
            'black': bounded(raw_kills.get('black'), 0, 1000, legacy_kills if result == 'black' else 0),
        }
        awards_raw = raw_match.get('awards') if isinstance(raw_match.get('awards'), dict) else {}
        awards = {
            str(key): bounded(value, 0, 10000, 0)
            for key, value in list(awards_raw.items())[:20]
            if str(key).strip()
        }
        history = raw_match.get('revisionHistory') if isinstance(raw_match.get('revisionHistory'), list) else []
        matches.append({
            'id': str(raw_match.get('id') or f'match-{index}')[:120],
            'stage': stage,
            'round': max(1, min(12, _safe_int(raw_match.get('round'), 1))),
            'board': max(1, min(1000, _safe_int(raw_match.get('board'), index + 1))),
            'whiteId': white_id,
            'blackId': black_id,
            'whiteSeed': white_seed,
            'blackSeed': black_seed,
            'status': 'completed' if result else 'scheduled',
            'result': result,
            'winnerId': winner_id,
            'advancingId': advancing_id,
            'durationSeconds': bounded_int(raw_match.get('durationSeconds'), 0, 36000, 0),
            'killPoints': max(kill_points_by_side['white'], kill_points_by_side['black']),
            'killPointsBySide': kill_points_by_side,
            'awards': awards,
            'breakdown': normalize_breakdown(raw_match.get('breakdown')),
            'revisionHistory': [item for item in history[-5:] if isinstance(item, dict)],
            'notes': str(raw_match.get('notes') or '').strip()[:240],
            'createdAt': str(raw_match.get('createdAt') or '').strip()[:80],
            'origin': str(raw_match.get('origin') or 'system').strip().lower() if str(raw_match.get('origin') or 'system').strip().lower() in {'system', 'historical', 'manual'} else 'system',
            'updatedAt': str(raw_match.get('updatedAt') or '').strip()[:80],
        })

    enrolled_values = source.get('enrolledStudentIds') if isinstance(source.get('enrolledStudentIds'), list) else []
    qualified_values = source.get('qualifiedStudentIds') if isinstance(source.get('qualifiedStudentIds'), list) else []
    stage = str(source.get('stage') or 'registration').strip().lower()
    if stage not in _CHESS_STAGE_VALUES:
        stage = 'registration'
    status = str(source.get('status') or stage).strip().lower()
    if status not in _CHESS_STAGE_VALUES:
        status = stage
    state = {
        'version': max(2, _safe_int(source.get('version'), 2)),
        'eventEpoch': max(0, _safe_int(source.get('eventEpoch'), 0)),
        'eventName': str(source.get('eventName') or 'Excel Chess Champion').strip()[:80] or 'Excel Chess Champion',
        'stage': stage,
        'status': status,
        'round': max(0, min(12, _safe_int(source.get('round'), 0))),
        'basicRoundCount': 3,
        'qualificationCount': 8,
        'eliminationRoundCount': 3,
        'semifinalRoundCount': 3,
        'enrolledStudentIds': sorted({sid for sid in (_safe_int(value, 0) for value in enrolled_values) if sid > 0}),
        'qualifiedStudentIds': [sid for sid in (_safe_int(value, 0) for value in qualified_values) if sid > 0],
        'championStudentId': max(0, _safe_int(source.get('championStudentId'), 0)),
        'seriesArchive': [],
        'scoring': {
            'winPoints': bounded(scoring.get('winPoints'), 0, 100, 3),
            'drawPoints': bounded(scoring.get('drawPoints'), 0, 100, 1),
            'lossPoints': bounded(scoring.get('lossPoints'), 0, 100, 0),
            'classBonusPerLevel': bounded(scoring.get('classBonusPerLevel'), 0, 20, 1),
            'opponentBonusPerFivePoints': bounded(scoring.get('opponentBonusPerFivePoints'), 0, 20, 1),
            'fastWinSeconds': bounded_int(scoring.get('fastWinSeconds'), 60, 10800, 600),
            'standardWinSeconds': bounded_int(scoring.get('standardWinSeconds'), 60, 14400, 1200),
            'fastTimeBonus': bounded(scoring.get('fastTimeBonus'), 0, 50, 3),
            'standardTimeBonus': bounded(scoring.get('standardTimeBonus'), 0, 50, 1),
            'killPointMultiplier': bounded(scoring.get('killPointMultiplier'), 0, 20, 1),
        },
        'matches': matches,
        'updatedAt': str(source.get('updatedAt') or '').strip()[:80],
    }
    archive_values = source.get('seriesArchive') if isinstance(source.get('seriesArchive'), list) else []
    state['seriesArchive'] = [_normalize_chess_champion_archive(item, index) for index, item in enumerate(archive_values[-20:]) if isinstance(item, dict)]
    if state['scoring']['standardWinSeconds'] < state['scoring']['fastWinSeconds']:
        state['scoring']['standardWinSeconds'] = state['scoring']['fastWinSeconds']
    return state


def _merge_chess_champion_archives(existing, incoming):
    archive_by_id = {}
    for item in (existing.get('seriesArchive') or []) + (incoming.get('seriesArchive') or []):
        archive_id = str(item.get('id') or '').strip()
        if archive_id:
            archive_by_id[archive_id] = item
    return list(archive_by_id.values())[-20:]


def _merge_chess_champion_superset(existing_state, incoming_state):
    """Merge tournament state without allowing partial/stale snapshots to erase results."""
    existing_raw = existing_state if isinstance(existing_state, dict) else {}
    incoming_raw = incoming_state if isinstance(incoming_state, dict) else {}
    existing = _normalize_chess_champion(existing_raw)
    incoming = _normalize_chess_champion(incoming_raw)
    if not incoming_raw:
        return existing
    if not existing_raw or (not existing.get('matches') and incoming.get('matches') and incoming.get('eventEpoch', 0) >= existing.get('eventEpoch', 0)):
        return incoming

    existing_stamp = _parse_sync_stamp(existing.get('updatedAt'))
    incoming_stamp = _parse_sync_stamp(incoming.get('updatedAt'))
    existing_epoch = max(0, _safe_int(existing.get('eventEpoch'), 0))
    incoming_epoch = max(0, _safe_int(incoming.get('eventEpoch'), 0))
    # A higher eventEpoch is an intentional Admin reset/new event and must win
    # even when the incoming snapshot has empty matches/enrollments.
    if incoming_epoch > existing_epoch:
        reset_state = dict(incoming)
        reset_state['seriesArchive'] = _merge_chess_champion_archives(existing, incoming)
        return _normalize_chess_champion(reset_state)
    if existing_epoch > incoming_epoch:
        keep_state = dict(existing)
        keep_state['seriesArchive'] = _merge_chess_champion_archives(existing, incoming)
        return _normalize_chess_champion(keep_state)

    explicit_reset = (
        str(incoming_raw.get('stage') or '').strip().lower() == 'registration'
        and isinstance(incoming_raw.get('matches'), list)
        and not incoming_raw.get('matches')
        and incoming_stamp > existing_stamp
    )
    if explicit_reset:
        reset_state = dict(incoming)
        reset_state['seriesArchive'] = _merge_chess_champion_archives(existing, incoming)
        return _normalize_chess_champion(reset_state)
    if existing.get('matches') and not incoming.get('matches'):
        return existing

    incoming_newer = incoming_stamp >= existing_stamp and incoming_stamp > 0
    base = dict(existing if not incoming_newer else incoming)
    if incoming_newer:
        base.update({key: value for key, value in incoming.items() if key not in {'matches', 'scoring'}})
        base['scoring'] = incoming.get('scoring') or existing.get('scoring')
    else:
        base.update({key: value for key, value in existing.items() if key not in {'matches', 'scoring'}})
        base['scoring'] = existing.get('scoring') or incoming.get('scoring')

    base['seriesArchive'] = _merge_chess_champion_archives(existing, incoming)

    by_id = {}
    for item in existing.get('matches', []) + incoming.get('matches', []):
        match_id = str(item.get('id') or '').strip()
        if not match_id:
            continue
        previous = by_id.get(match_id)
        if not previous:
            by_id[match_id] = dict(item)
            continue
        previous_stamp = _parse_sync_stamp(previous.get('updatedAt') or previous.get('createdAt'))
        item_stamp = _parse_sync_stamp(item.get('updatedAt') or item.get('createdAt'))
        chosen = item if item_stamp > previous_stamp or (item_stamp == previous_stamp and incoming_newer) else previous
        merged_match = dict(previous)
        merged_match.update(chosen)
        merged_match['revisionHistory'] = (previous.get('revisionHistory') or []) + (item.get('revisionHistory') or [])
        merged_match['revisionHistory'] = merged_match['revisionHistory'][-5:]
        by_id[match_id] = merged_match
    base['matches'] = list(by_id.values())
    # Allow intentional clears (newer empty enrollment/qualifier lists) instead of
    # always resurrecting the older side's roster selection.
    intentional_clear_enrolled = incoming_newer and not incoming.get('enrolledStudentIds')
    intentional_clear_qualified = incoming_newer and not incoming.get('qualifiedStudentIds')
    if existing.get('enrolledStudentIds') and not base.get('enrolledStudentIds') and not intentional_clear_enrolled:
        base['enrolledStudentIds'] = existing['enrolledStudentIds']
    if existing.get('qualifiedStudentIds') and not base.get('qualifiedStudentIds') and not intentional_clear_qualified:
        base['qualifiedStudentIds'] = existing['qualifiedStudentIds']
    return _normalize_chess_champion(base)


def _project_root_path():
    return os.path.abspath(os.path.join(current_app.root_path, os.pardir))


def _public_site_dir():
    explicit = str(os.getenv('EA_PUBLIC_SITE_DIR', '') or '').strip()
    if explicit:
        return os.path.abspath(explicit)
    return os.path.join(_project_root_path(), 'public_site')


def _sync_spa_to_public_site(site_dir):
    """Copy offline_scoreboard.html and CSS into public_site/ for Cloudflare Pages hosting.
    Injects the <meta name="ea-backend-url"> tag with the tunnel origin so the SPA
    operates in cross-origin mode when served from Cloudflare Pages."""
    import shutil
    repo_root = _project_root_path()
    spa_src = os.path.join(repo_root, 'app', 'static', 'offline_scoreboard.html')
    css_src = os.path.join(repo_root, 'app', 'static', 'css', 'offline-scoreboard.css')
    spa_dst = os.path.join(site_dir, 'offline_scoreboard.html')
    css_dst_dir = os.path.join(site_dir, 'static', 'css')
    css_dst = os.path.join(css_dst_dir, 'offline-scoreboard.css')
    copied = []

    if os.path.isfile(spa_src):
        with open(spa_src, 'r', encoding='utf-8') as f:
            html = f.read()
        tunnel_origin = str(os.getenv('EA_TUNNEL_ORIGIN', '') or '').strip()
        if tunnel_origin:
            tunnel_url = tunnel_origin if tunnel_origin.startswith('http') else f'https://{tunnel_origin}'
            meta_tag = f'<meta name="ea-backend-url" content="{tunnel_url}">'
            if 'ea-backend-url' not in html:
                html = html.replace('<head>', f'<head>\n    {meta_tag}', 1)
            else:
                import re
                html = re.sub(
                    r'<meta\s+name="ea-backend-url"\s+content="[^"]*"\s*/?>',
                    meta_tag, html, count=1
                )
        os.makedirs(site_dir, exist_ok=True)
        with open(spa_dst, 'w', encoding='utf-8') as f:
            f.write(html)
        copied.append(spa_dst)

    if os.path.isfile(css_src):
        os.makedirs(css_dst_dir, exist_ok=True)
        shutil.copy2(css_src, css_dst)
        copied.append(css_dst)

    # Mirror student portrait photos to public_site/static/uploads/students/
    # so they appear on the Cloudflare Pages site (static hosting).
    uploads_src = os.path.join(repo_root, 'app', 'static', 'uploads', 'students')
    uploads_dst = os.path.join(site_dir, 'static', 'uploads', 'students')
    if os.path.isdir(uploads_src):
        os.makedirs(uploads_dst, exist_ok=True)
        for fname in os.listdir(uploads_src):
            src_file = os.path.join(uploads_src, fname)
            if os.path.isfile(src_file):
                try:
                    shutil.copy2(src_file, os.path.join(uploads_dst, fname))
                    copied.append(os.path.join(uploads_dst, fname))
                except Exception as photo_err:
                    current_app.logger.warning(f'Failed to mirror student photo {fname}: {photo_err}')

    return copied


def _public_site_scores_path():
    return os.path.join(_public_site_dir(), 'scores.json')


def _public_site_credentials_path():
    return os.path.join(_public_site_dir(), 'credentials.json')


def _hash_public_credential(password):
    """Salted SHA-256 of a password (UTF-8), matching the browser's Web Crypto
    SHA-256(salt + password) and scripts/generate_credentials.py exactly.
    Returns (salt_hex, hash_hex)."""
    salt = secrets.token_hex(8)
    digest = hashlib.sha256((salt + str(password)).encode('utf-8')).hexdigest()
    return salt, digest


def _normalize_roll(roll):
    return str(roll or '').strip().upper()


def _student_rolls_from_ledger():
    """Best-effort extraction of student roll numbers from the offline ledger,
    used to populate the admin datalist and to drive bulk-set operations."""
    try:
        data = _load_offline_data() or {}
    except Exception:
        data = {}
    rolls = []
    seen = set()
    students = data.get('students') if isinstance(data, dict) else None
    if isinstance(students, list):
        for s in students:
            if isinstance(s, dict):
                roll = str(s.get('roll') or s.get('roll_number') or '').strip().upper()
                if roll and roll not in seen:
                    seen.add(roll)
                    rolls.append(roll)
    return rolls


def _build_public_credentials_payload():
    """Read active PublicSiteCredential rows and build the
    public_site/credentials.json payload. DB is the single source of truth."""
    try:
        rows = (
            PublicSiteCredential.query
            .filter(PublicSiteCredential.active.is_(True))
            .order_by(PublicSiteCredential.roll.asc())
            .all()
        )
    except Exception as exc:
        current_app.logger.warning('Failed to load public credentials: %s', exc)
        return {'credentials': []}
    return {
        'credentials': [
            {
                'roll': str(r.roll or '').upper(),
                'salt': str(r.salt or ''),
                'hash': str(r.hash or ''),
            }
            for r in rows
        ]
    }


def _auto_push_public_site_enabled():
    flag = str(os.getenv('EA_PUBLIC_SITE_AUTO_PUSH', '1') or '').strip().lower()
    return flag not in {'0', 'false', 'no', 'off'}


def _clear_stale_git_index_lock(repo_root, max_age_seconds=90):
    """Remove a leftover index.lock from a crashed or interrupted git command."""
    lock_path = os.path.join(repo_root, '.git', 'index.lock')
    if not os.path.exists(lock_path):
        return False
    try:
        age_seconds = time.time() - os.path.getmtime(lock_path)
        if age_seconds < max_age_seconds:
            return False
        os.remove(lock_path)
        return True
    except Exception:
        return False


def _run_git(repo_root, args):
    def _invoke():
        return subprocess.run(
            ['git', '-C', repo_root, *args],
            check=False,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace'
        )

    completed = _invoke()
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or '').strip()
        if 'index.lock' in detail and _clear_stale_git_index_lock(repo_root, max_age_seconds=0):
            completed = _invoke()
            if completed.returncode == 0:
                return completed
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or '').strip()
            raise RuntimeError(detail or f"git {' '.join(args)} failed")
    return completed


def _safe_commit_stamp():
    try:
        return datetime.now(ZoneInfo(_get_server_timezone())).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _resync_students_from_veto_tracking(payload):
    """
    Before publishing, sync VETO counts on each student record from the
    authoritative veto_tracking.students map. This prevents a stale client-side
    students[] array from overwriting the server's VETO ledger in the published
    public_site/scores.json (which is then committed to git and seen by peers).
    Stars are intentionally not touched — student.stars is authoritative and
    should not be modified by publish.
    """
    if not isinstance(payload, dict):
        return
    veto_tracking = payload.get('veto_tracking') or {}
    tracked = veto_tracking.get('students') or {}
    if not isinstance(tracked, dict) or not tracked:
        return
    current_month = _server_now_iso()[:7]
    roll_to_sid = {
        s.get('roll'): _parse_int_safe(s.get('id'), 0)
        for s in (payload.get('students') or [])
        if s.get('roll') and _parse_int_safe(s.get('id'), 0) > 0
    }
    for student in payload.get('students') or []:
        if not isinstance(student, dict):
            continue
        roll = str(student.get('roll') or '').strip()
        if not roll or roll not in tracked:
            continue
        entry = tracked[roll] or {}
        try:
            ind = int(entry.get('individual_vetos', 0) or 0)
            role = int(entry.get('role_vetos', 0) or 0)
        except (TypeError, ValueError):
            continue

        sid = roll_to_sid.get(roll, 0)
        # Individual VETOs are PERMANENT (total usage deducted);
        # Role VETOs are MONTHLY (only current-month usage deducted).
        ind_rem, role_rem, total_used = _compute_veto_remaining_counters(
            payload, sid, ind, role, current_month
        )

        student['veto_count'] = ind_rem
        student['role_veto_count'] = role_rem
        student['used_veto_count'] = total_used


def _publish_public_site_snapshot(payload, push=None, public_snapshot=None):
    site_dir = _public_site_dir()
    scores_path = _public_site_scores_path()
    # Align students[] VETO counts with authoritative veto_tracking before any
    # downstream transform — prevents stale client state from leaking into the
    # published snapshot.
    try:
        _resync_students_from_veto_tracking(payload)
    except Exception as _resync_err:
        current_app.logger.warning(f"Pre-publish VETO resync failed: {_resync_err}")
    # Prefer the browser-generated public snapshot when it is structurally sane.
    # The LAN UI already computes month-aware totals and historical identity
    # resolution, so this keeps the public site aligned with what admins see
    # locally. Fall back to the server-side rebuild if the client snapshot is
    # missing or looks incomplete.
    public_payload = None
    if isinstance(public_snapshot, dict) and isinstance(public_snapshot.get('scoreboard'), dict):
        sanitized = _sanitize_client_public_snapshot(public_snapshot, payload if isinstance(payload, dict) else {})
        if _public_snapshot_has_useful_rows(sanitized):
            public_payload = sanitized
        else:
            current_app.logger.warning(
                "Client-provided public_snapshot was empty or incomplete; rebuilding public site from server payload instead"
            )
    if public_payload is None:
        public_payload = _build_public_site_payload(payload if isinstance(payload, dict) else {})

    os.makedirs(site_dir, exist_ok=True)
    _atomic_write_json(scores_path, public_payload)

    # Credentials.json: DB is the single source of truth. Always (re)write from
    # active PublicSiteCredential rows so revocations propagate on publish.
    # If no active credentials exist, write an empty list (site locks until the
    # admin configures credentials via the Admin Control Panel).
    credentials_path = _public_site_credentials_path()
    credentials_payload = _build_public_credentials_payload()
    try:
        _atomic_write_json(credentials_path, credentials_payload)
    except Exception as creds_err:
        current_app.logger.warning(f"credentials.json write failed: {creds_err}")
    credentials_count = len(credentials_payload.get('credentials') or [])

    try:
        _sync_spa_to_public_site(site_dir)
    except Exception as spa_err:
        current_app.logger.warning(f"SPA sync to public_site failed: {spa_err}")

    result = {
        'status': 'ok',
        'site_dir': site_dir,
        'scores_path': scores_path,
        'credentials_path': credentials_path,
        'credentials_count': credentials_count,
        'updated_at': public_payload.get('updated_at'),
        'pushed': False,
    }

    should_push = _auto_push_public_site_enabled() if push is None else bool(push)
    if not should_push:
        result['status'] = 'written'
        return result

    repo_root = _project_root_path()
    git_dir = os.path.join(repo_root, '.git')
    if not os.path.isdir(git_dir):
        result.update({
            'status': 'write_only',
            'error': f'Git repository not found at {repo_root}',
        })
        return result

    rel_scores = os.path.relpath(scores_path, repo_root)
    rel_index = os.path.relpath(os.path.join(site_dir, 'index.html'), repo_root)
    rel_readme = os.path.relpath(os.path.join(site_dir, 'README.md'), repo_root)
    rel_headers = os.path.relpath(os.path.join(site_dir, '_headers'), repo_root)
    rel_creds = os.path.relpath(os.path.join(site_dir, 'credentials.json'), repo_root)
    rel_spa = os.path.relpath(os.path.join(site_dir, 'offline_scoreboard.html'), repo_root)
    rel_css = os.path.relpath(os.path.join(site_dir, 'static', 'css', 'offline-scoreboard.css'), repo_root)
    candidate_paths = [rel_index, rel_readme, rel_headers, rel_scores, rel_creds, rel_spa, rel_css]
    # Add all student portrait photos in public_site/static/uploads/students/
    uploads_dst_dir = os.path.join(site_dir, 'static', 'uploads', 'students')
    if os.path.isdir(uploads_dst_dir):
        for fname in os.listdir(uploads_dst_dir):
            fpath = os.path.join(uploads_dst_dir, fname)
            if os.path.isfile(fpath):
                candidate_paths.append(os.path.relpath(fpath, repo_root))
    tracked_paths = [path for path in candidate_paths if os.path.exists(os.path.join(repo_root, path))]
    commit_stamp = _safe_commit_stamp()

    try:
        _clear_stale_git_index_lock(repo_root)
        _run_git(repo_root, ['add', '--', *tracked_paths])
        diff = subprocess.run(
            ['git', '-C', repo_root, 'diff', '--cached', '--quiet', '--', *tracked_paths],
            check=False,
            capture_output=True
        )
        if diff.returncode == 0:
            result['status'] = 'up_to_date'
            return result
        _run_git(repo_root, ['commit', '-m', f'Publish public scoreboard ({commit_stamp})', '--', *tracked_paths])
        push_output = _run_git(repo_root, ['push', 'origin', 'HEAD'])
        result.update({
            'status': 'pushed',
            'pushed': True,
            'push_output': (push_output.stdout or '').strip(),
        })
    except Exception as exc:
        result.update({
            'status': 'write_only',
            'error': str(exc),
        })
    return result


def _normalize_peer_list(raw_values):
    if not isinstance(raw_values, list):
        return []
    return normalize_peer_urls(raw_values)


def _forward_offline_data_to_peers(payload, extra_peers=None):
    peers = get_sync_peers() + _normalize_peer_list(extra_peers or [])
    peers = list(dict.fromkeys(peers))
    if not peers:
        return
    # request.host_url is only available inside a Flask request context.
    # This function is also called from background daemon threads (spawned
    # during route handlers) where the request proxy is no longer valid.
    # Swallow the RuntimeError and default to '' so the origin-skip check
    # simply never matches any peer URL.
    try:
        current_origin = (request.host_url or '').rstrip('/')
    except RuntimeError:
        current_origin = ''
    is_master = str(os.getenv('EA_MASTER_MODE', '')).strip() == '1'
    if is_master:
        # Master publishes authoritative snapshots to peers (Render/backup),
        # so receivers can accept them even when peer timestamps drift ahead.
        base_flags = {
            'authoritative_master_push': True,
            'force_replace': True,
        }
    else:
        base_flags = {}
    shared_key = resolve_sync_shared_key()
    for peer in peers:
        if peer.rstrip('/') == current_origin:
            continue
        # Preserve the fees module on LAN peers so offline/backup nodes stay
        # consistent; only strip it for non-private WAN/cloud mirrors.
        peer_payload = payload if is_private_peer_url(peer) else payload_for_external_replication(payload)
        body = json.dumps({'data': peer_payload, **base_flags}).encode('utf-8')
        target_url = f'{peer}/scoreboard/offline-data'
        req = urllib.request.Request(
            target_url,
            data=body,
            method='POST',
            headers={
                'Content-Type': 'application/json',
                'X-EA-Replicated': '1',
                'X-EA-Sync-Key': shared_key
            }
        )
        try:
            # WAN peers (Render) cold-start can take 30-50 s on free tier.
            # This function always runs in a background daemon thread so blocking is fine.
            with urllib.request.urlopen(req, timeout=45):
                pass
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError):
            continue


def _forward_offline_data_to_peers_async(payload, extra_peers=None):
    threading.Thread(
        target=_forward_offline_data_to_peers,
        args=(payload, extra_peers),
        daemon=True
    ).start()


# ── Background peer-sync thread ──────────────────────────────────────────────
# Runs every 30 s on the local master server only (EA_MASTER_MODE=1).
# Keeps Render awake (free-tier dynos sleep after 15 min of inactivity),
# pushes local data when local is newer, and pulls Render data when Render is
# newer — giving true bidirectional sync without any browser interaction.

_peer_sync_thread_started = False
_NON_ADMIN_EDIT_WINDOW_DAYS = 2
_supabase_push_lock = threading.Lock()
_supabase_last_pushed_stamp = ''
_SYNC_OPS_MAX = 4000


def _fees_module_enabled():
    raw = str(os.getenv('EA_ENABLE_FEES_MODULE', '1') or '1').strip().lower()
    return raw in ('1', 'true', 'yes', 'on')


def _supabase_enabled():
    raw = str(os.getenv('EA_DISABLE_SUPABASE', '0') or '0').strip().lower()
    return raw not in ('1', 'true', 'yes', 'on')


def _supabase_snapshot_config():
    if not _supabase_enabled():
        return {
            'enabled_read': False,
            'enabled_write': False,
            'url': '',
            'read_key': '',
            'write_key': '',
            'table': '',
            'row_id': '',
        }
    url = str(os.getenv('SUPABASE_URL', '') or '').strip().rstrip('/')
    service_key = str(os.getenv('SUPABASE_SERVICE_ROLE_KEY', '') or '').strip()
    anon_key = str(os.getenv('SUPABASE_ANON_KEY', '') or '').strip()
    read_key = service_key or anon_key
    write_key = service_key
    table = str(os.getenv('SUPABASE_SNAPSHOT_TABLE', 'offline_snapshots') or 'offline_snapshots').strip()
    row_id = str(os.getenv('SUPABASE_SNAPSHOT_ROW_ID', 'main') or 'main').strip()
    enabled_read = bool(url and read_key and table and row_id)
    enabled_write = bool(url and write_key and table and row_id)
    return {
        'enabled_read': enabled_read,
        'enabled_write': enabled_write,
        'url': url,
        'read_key': read_key,
        'write_key': write_key,
        'table': table,
        'row_id': row_id,
    }


def _supabase_headers(cfg, key_name='read_key', for_write=False):
    key = cfg.get(key_name, '')
    headers = {
        'apikey': key,
        'Authorization': f"Bearer {key}",
    }
    if for_write:
        headers['Content-Type'] = 'application/json'
        headers['Prefer'] = 'resolution=merge-duplicates,return=minimal'
    return headers


def _supabase_fetch_snapshot(timeout_sec=12):
    cfg = _supabase_snapshot_config()
    if not cfg.get('enabled_read'):
        return None, 'not-configured'
    endpoint = f"{cfg['url']}/rest/v1/{cfg['table']}"
    try:
        params = urllib.parse.urlencode({
            'select': 'data,updated_at,source',
            'id': f"eq.{cfg['row_id']}",
            'limit': 1
        })
        req = urllib.request.Request(
            f"{endpoint}?{params}",
            headers=_supabase_headers(cfg, key_name='read_key', for_write=False),
            method='GET'
        )
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            raw = resp.read()
        rows = json.loads(raw.decode('utf-8', errors='replace')) if raw else []
        if not isinstance(rows, list) or not rows:
            return None, 'empty'
        row = rows[0] if isinstance(rows[0], dict) else {}
        data = row.get('data')
        if isinstance(data, dict):
            return data, 'supabase'
    except Exception as exc:
        current_app.logger.warning('Supabase fetch snapshot exception: %s', exc)
    return None, 'error'


def _supabase_push_snapshot(payload, reason='snapshot_save', timeout_sec=15):
    cfg = _supabase_snapshot_config()
    if not cfg.get('enabled_write'):
        return False
    if not isinstance(payload, dict):
        return False

    # Exclude fees module data from Supabase mirror by requirement.
    mirror_payload = payload_for_external_replication(payload)

    stamp = str(mirror_payload.get('server_updated_at') or mirror_payload.get('updated_at') or '').strip()
    global _supabase_last_pushed_stamp
    with _supabase_push_lock:
        if stamp and stamp == _supabase_last_pushed_stamp:
            return True

    endpoint = f"{cfg['url']}/rest/v1/{cfg['table']}"
    body = [{
        'id': cfg['row_id'],
        'data': mirror_payload,
        'source': str(reason or 'snapshot_save')[:80],
        'updated_at': _server_now_iso(),
    }]
    try:
        params = urllib.parse.urlencode({'on_conflict': 'id'})
        req = urllib.request.Request(
            f"{endpoint}?{params}",
            data=json.dumps(body).encode('utf-8'),
            headers=_supabase_headers(cfg, key_name='write_key', for_write=True),
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=timeout_sec):
            pass
        if stamp:
            with _supabase_push_lock:
                _supabase_last_pushed_stamp = stamp
        return True
    except Exception as exc:
        current_app.logger.warning('Supabase push snapshot exception: %s', exc)
        return False


def _supabase_push_snapshot_async(payload, reason='snapshot_save'):
    cfg = _supabase_snapshot_config()
    if not cfg.get('enabled_write'):
        return
    threading.Thread(
        target=_supabase_push_snapshot,
        args=(payload, reason),
        daemon=True
    ).start()


# ── GitHub Gist cloud backup (Supabase-free alternative) ─────────────────────
# Set GITHUB_GIST_TOKEN (personal access token, gist scope) and
# GITHUB_GIST_ID (the Gist ID) on both master PC and Render.
# Create the Gist once at https://gist.github.com — can be secret/private.
_gist_push_lock = threading.Lock()
_gist_last_pushed_stamp = ''


def _gist_config():
    token   = str(os.getenv('GITHUB_GIST_TOKEN', '') or '').strip()
    gist_id = str(os.getenv('GITHUB_GIST_ID', '') or '').strip()
    filename = str(os.getenv('GITHUB_GIST_FILENAME', 'ea_snapshot.json') or 'ea_snapshot.json').strip()
    return {
        'enabled_read':  bool(gist_id),           # public gists readable without token
        'enabled_write': bool(token and gist_id),
        'token':   token,
        'gist_id': gist_id,
        'filename': filename,
    }


def _gist_fetch_snapshot(timeout_sec=15):
    cfg = _gist_config()
    if not cfg.get('enabled_read'):
        return None, 'not-configured'
    headers = {
        'Accept': 'application/vnd.github.v3+json',
        'User-Agent': 'EA-Scoreboard/1.0',
    }
    if cfg['token']:
        headers['Authorization'] = f"token {cfg['token']}"
    try:
        req = urllib.request.Request(
            f"https://api.github.com/gists/{cfg['gist_id']}",
            headers=headers, method='GET'
        )
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            meta = json.loads(resp.read().decode('utf-8', errors='replace'))
        file_info = (meta.get('files') or {}).get(cfg['filename'], {})
        # Large files are truncated in the API response; follow raw_url instead.
        raw_url = file_info.get('raw_url', '')
        content = None
        if raw_url:
            req2 = urllib.request.Request(raw_url, headers=headers, method='GET')
            with urllib.request.urlopen(req2, timeout=timeout_sec) as r2:
                content = r2.read().decode('utf-8', errors='replace')
        else:
            content = file_info.get('content') or ''
        if content:
            data = json.loads(content)
            if isinstance(data, dict):
                return data, 'gist'
    except Exception as exc:
        current_app.logger.warning('Gist fetch snapshot exception: %s', exc)
    return None, 'error'


def _gist_push_snapshot(payload, reason='snapshot_save', timeout_sec=20):
    cfg = _gist_config()
    if not cfg.get('enabled_write'):
        return False
    if not isinstance(payload, dict):
        return False
    mirror_payload = payload_for_external_replication(payload)
    stamp = str(mirror_payload.get('server_updated_at') or mirror_payload.get('updated_at') or '').strip()
    global _gist_last_pushed_stamp
    with _gist_push_lock:
        if stamp and stamp == _gist_last_pushed_stamp:
            return True
    content = json.dumps(mirror_payload, ensure_ascii=False)
    body = json.dumps({'files': {cfg['filename']: {'content': content}}}).encode('utf-8')
    headers = {
        'Authorization': f"token {cfg['token']}",
        'Content-Type': 'application/json',
        'Accept': 'application/vnd.github.v3+json',
        'User-Agent': 'EA-Scoreboard/1.0',
    }
    try:
        req = urllib.request.Request(
            f"https://api.github.com/gists/{cfg['gist_id']}",
            data=body, headers=headers, method='PATCH'
        )
        with urllib.request.urlopen(req, timeout=timeout_sec):
            pass
        if stamp:
            with _gist_push_lock:
                _gist_last_pushed_stamp = stamp
        current_app.logger.info('Gist snapshot pushed (%s)', reason)
        return True
    except Exception as exc:
        current_app.logger.warning('Gist push snapshot exception: %s', exc)
        return False


def _gist_push_snapshot_async(payload, reason='snapshot_save'):
    cfg = _gist_config()
    if not cfg.get('enabled_write'):
        return
    threading.Thread(
        target=_gist_push_snapshot,
        args=(payload, reason),
        daemon=True
    ).start()


def _do_peer_sync_cycle(app):
    """One bidirectional sync cycle with all configured peers."""
    peers = get_sync_peers()
    shared_key = resolve_sync_shared_key()
    if not peers or not shared_key:
        return

    with app.app_context():
        local_data = _load_offline_data() or {}
        local_stamp = _payload_sync_stamp(local_data) or 0.0
        local_count = _student_count(local_data)

        for peer in peers:
            try:
                # ── Pull: fetch peer's current snapshot ─────────────────────
                get_req = urllib.request.Request(
                    f'{peer}/scoreboard/offline-data',
                    method='GET',
                    headers={
                        'Cache-Control': 'no-store',
                        'X-EA-Replicated': '1',
                        'X-EA-Sync-Key': shared_key,
                    }
                )
                with urllib.request.urlopen(get_req, timeout=55) as resp:
                    peer_body = resp.read()
                peer_parsed = json.loads(peer_body.decode('utf-8', errors='replace'))
                peer_data = peer_parsed.get('data') if isinstance(peer_parsed, dict) else None
                if not isinstance(peer_data, dict):
                    continue
                # Hard safety: never merge a sanitized/month-clipped view (served to
                # unauthenticated callers) into a full local ledger — it would drop
                # historical months and private collections.
                if not is_full_ledger_snapshot(peer_data):
                    app.logger.warning('[BgSync] Peer %s returned a sanitized snapshot (key mismatch?) — skipping.', peer)
                    continue

                peer_stamp = _payload_sync_stamp(peer_data) or 0.0
                peer_count = _student_count(peer_data)

                is_master = str(os.getenv('EA_MASTER_MODE', '')).strip() == '1'
                min_students = _min_safe_student_roster()

                # Master is authoritative: never pull peer snapshots into master.
                # This prevents stale/inflated peer data from overriding local canonical data.
                if is_master:
                    if local_count >= min_students and not _is_suspicious_student_shrink(peer_data, local_data):
                        # Push whenever snapshots differ by stamp/count.
                        if (abs(local_stamp - peer_stamp) > 1) or (local_count != peer_count):
                            body = json.dumps({
                                'data': payload_for_external_replication(local_data),
                                'authoritative_master_push': True,
                                'force_replace': True
                            }).encode('utf-8')
                            post_req = urllib.request.Request(
                                f'{peer}/scoreboard/offline-data',
                                data=body,
                                method='POST',
                                headers={
                                    'Content-Type': 'application/json',
                                    'X-EA-Replicated': '1',
                                    'X-EA-Sync-Key': shared_key,
                                }
                            )
                            with urllib.request.urlopen(post_req, timeout=55):
                                pass
                            app.logger.info(
                                "[BgSync] Master pushed authoritative snapshot to %s (%s students, stamp=%s)",
                                peer, local_count, local_data.get('server_updated_at', '')
                            )
                    continue

                # ── Backup/slave mode: bidirectional with skew margin ───────
                # Margin of 30 s avoids flip-flopping on tiny clock skew.
                if peer_stamp > local_stamp + 30 and peer_count >= min_students:
                    if not _is_suspicious_student_shrink(local_data, peer_data):
                        # CRITICAL FIX: Merge instead of overwrite to preserve locally-added students
                        merged = dict(local_data)
                        merged['chess_champion'] = _merge_chess_champion_superset(
                            local_data.get('chess_champion', {}),
                            peer_data.get('chess_champion', {})
                        )

                        # Superset merge key tables (never shrink)
                        merged['students'] = _merge_students_preserve_active(
                            local_data.get('students', []),
                            peer_data.get('students', [])
                        )
                        merged['scores'] = _merge_scores_superset(
                            local_data.get('scores', []),
                            peer_data.get('scores', [])
                        )
                        merged['attendance'] = _merge_attendance_superset(local_data, peer_data)
                        merged['appeals'] = _merge_appeals_superset(
                            local_data.get('appeals', []),
                            peer_data.get('appeals', [])
                        )
                        merged['postholder_tickets'] = _merge_postholder_tickets(
                            local_data.get('postholder_tickets', {}),
                            peer_data.get('postholder_tickets', {})
                        )
                        merged['postholder_ticket_log'] = _merge_postholder_ticket_log(
                            local_data.get('postholder_ticket_log', []),
                            peer_data.get('postholder_ticket_log', [])
                        )

                        # Update timestamp only if peer is genuinely newer
                        if peer_stamp > local_stamp:
                            merged['server_updated_at'] = peer_data.get('server_updated_at', local_data.get('server_updated_at'))

                        merged_count = _student_count(merged)
                        _save_offline_data(merged)
                        _broadcast_sync_event(
                            merged.get('server_updated_at', ''),
                            source='bg-peer-pull'
                        )

                        # Alert if merge detected students not in peer (locally added)
                        local_only = local_count - peer_count
                        preserved = merged_count - peer_count
                        if preserved > 0:
                            app.logger.warning(
                                "[BgSync] PRESERVED %d locally-added students during peer pull from %s "
                                "(local: %s, peer: %s, merged: %s)",
                                preserved, peer, local_count, peer_count, merged_count
                            )

                        app.logger.info(
                            "[BgSync] Merged & pulled newer snapshot from %s (local: %s students, peer: %s students, result: %s students, stamp=%s)",
                            peer, local_count, peer_count, merged_count, merged.get('server_updated_at', '')
                        )

                elif local_stamp > peer_stamp + 30 and local_count >= min_students:
                    # Local is newer -> push to peer
                    body = json.dumps({'data': payload_for_external_replication(local_data)}).encode('utf-8')
                    post_req = urllib.request.Request(
                        f'{peer}/scoreboard/offline-data',
                        data=body,
                        method='POST',
                        headers={
                            'Content-Type': 'application/json',
                            'X-EA-Replicated': '1',
                            'X-EA-Sync-Key': shared_key,
                        }
                    )
                    with urllib.request.urlopen(post_req, timeout=55):
                        pass
                    app.logger.info(
                        "[BgSync] Pushed local snapshot to %s (%s students, stamp=%s)",
                        peer, local_count, local_data.get('server_updated_at', '')
                    )

            except Exception as exc:
                app.logger.debug("[BgSync] Peer %s unreachable: %s", peer, exc)


def start_peer_sync_background(app):
    """Start a single persistent background thread that syncs with peers every 30 s.
    Safe to call multiple times — only one thread is ever started.
    Only active when EA_MASTER_MODE=1 and SYNC_PEERS is configured."""
    global _peer_sync_thread_started
    if _peer_sync_thread_started:
        return
    if str(os.getenv('EA_MASTER_MODE', '')).strip() != '1':
        return
    if not get_sync_peers():
        return
    _peer_sync_thread_started = True

    def _loop():
        import time as _time
        # Initial delay so the server finishes booting before the first sync.
        _time.sleep(15)
        last_error_repr = None
        while True:
            try:
                _do_peer_sync_cycle(app)
                last_error_repr = None
            except Exception as exc:
                if repr(exc) != last_error_repr:
                    last_error_repr = repr(exc)
                    app.logger.exception('[BgSync] Peer sync cycle failed')
                else:
                    app.logger.debug('[BgSync] Peer sync cycle failed (repeat): %s', exc)
            _time.sleep(30)

    t = threading.Thread(target=_loop, daemon=True, name='ea-peer-sync')
    t.start()
    app.logger.info("[BgSync] Background peer-sync thread started (interval=30s)")


def _atomic_write_json(path, payload):
    # Use dir=target_dir so the temp file is on the same filesystem as the
    # target.  Without this, tempfile.mkstemp() uses /tmp which is a separate
    # mount on Render/Docker, causing os.replace() to raise:
    #   [Errno 18] Invalid cross-device link
    kwargs = {'separators': (',', ':')}
    if os.path.abspath(os.fspath(path)) == os.path.abspath(_shared_data_path()):
        kwargs['validator'] = _ensure_ledger_payload
    _shared_atomic_write_json(path, payload, **kwargs)


def _verify_backup_copy(source_path, backup_path):
    """Verify a freshly copied backup has the same size as its source."""
    try:
        if os.path.getsize(backup_path) == os.path.getsize(source_path):
            return True
        _ledger_log.error(
            'Backup verification failed (size mismatch): %s — removing bad copy', backup_path
        )
    except OSError:
        _ledger_log.exception('Backup verification failed for %s', backup_path)
    try:
        os.remove(backup_path)
    except OSError:
        pass
    return False


def _backup_offline_file(path, keep=50):
    if not os.path.exists(path):
        return
    # Skip per-save backup when an hourly backup for the current hour already
    # exists — the hourly backup provides the same safety net without the
    # overhead of shutil.copy2 + directory listing on every save.
    hour_key = datetime.now().strftime('%Y%m%d_%H')
    hourly_path = os.path.join(_offline_hourly_backup_dir(), f'offline_scoreboard_hourly_{hour_key}.json')
    if os.path.exists(hourly_path):
        return
    os.makedirs(_offline_backup_dir(), exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f'offline_scoreboard_{timestamp}.json'
    backup_path = os.path.join(_offline_backup_dir(), backup_name)
    shutil.copy2(path, backup_path)
    _verify_backup_copy(path, backup_path)

    backups = sorted(
        [os.path.join(_offline_backup_dir(), f) for f in os.listdir(_offline_backup_dir()) if f.endswith('.json')],
        key=os.path.getmtime,
        reverse=True
    )
    for old in backups[keep:]:
        try:
            os.remove(old)
        except Exception:
            pass


def _backup_offline_hourly_immutable(payload, keep=24 * 30):
    """
    Create one immutable snapshot per hour (local server time).
    This is append-only per hour and protects against rapid accidental overwrites.
    Uses shutil.copy2 from the just-written main file (22ms) instead of
    _atomic_write_json (1.1s) since the main file is already atomically written.
    """
    os.makedirs(_offline_hourly_backup_dir(), exist_ok=True)
    hour_key = datetime.now().strftime('%Y%m%d_%H')
    backup_name = f'offline_scoreboard_hourly_{hour_key}.json'
    backup_path = os.path.join(_offline_hourly_backup_dir(), backup_name)
    if not os.path.exists(backup_path):
        main_path = _offline_data_path()
        if os.path.exists(main_path):
            shutil.copy2(main_path, backup_path)
            _verify_backup_copy(main_path, backup_path)
        else:
            _atomic_write_json(backup_path, payload)

    backups = sorted(
        [os.path.join(_offline_hourly_backup_dir(), f) for f in os.listdir(_offline_hourly_backup_dir()) if f.endswith('.json')],
        key=os.path.getmtime,
        reverse=True
    )
    for old in backups[keep:]:
        try:
            os.remove(old)
        except Exception:
            pass


def _load_latest_offline_backup():
    backup_dir = _offline_backup_dir()
    if not os.path.isdir(backup_dir):
        return None
    backup_files = sorted(
        [os.path.join(backup_dir, f) for f in os.listdir(backup_dir) if f.endswith('.json')],
        key=os.path.getmtime,
        reverse=True
    )
    for backup_path in backup_files:
        try:
            with open(backup_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as exc:
            _ledger_log.warning('Skipping unreadable backup %s: %s', backup_path, exc)
            continue
    return None



def _load_offline_data():
    """
    Load the offline scoreboard JSON, using an in-memory mtime cache.

    PERFORMANCE: the offline data file is typically 4+ MB. Re-parsing it on
    every request is the biggest source of server latency. The cache is keyed
    on (path, mtime_ns, size) and invalidated automatically when the file
    changes on disk. Every mutation path goes through _save_offline_data(),
    which re-primes the cache with the new payload.

    CAVEAT: the returned dict is the shared cached object. Callers that mutate
    it MUST eventually call _save_offline_data(data) — the save primes the
    cache with the mutated dict, keeping cache == disk. Don't mutate and then
    discard; that will leave stale state in the cache until the next write.
    """
    path = _offline_data_path()
    legacy_path = _legacy_instance_file('offline_scoreboard_data.json')
    if (not os.path.exists(path)) and os.path.exists(legacy_path):
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            shutil.copy2(legacy_path, path)
        except Exception:
            pass
    # Fallback seed stamp: fixed old date so any real pushed data always wins the
    # timestamp comparison.  Using datetime.now() here caused Render to reject local
    # pushes with 409 because the seed appeared "newer" than real data.
    _SEED_STAMP = "2026-02-26T00:00:00+00:00"

    # ── FAST PATH: mtime-cached load (avoids re-parsing 4+ MB per request) ──
    if os.path.exists(path):
        cached = _cached_load_json_data()
        if cached is not None and _student_count(cached) > 0:
            return cached
        # Cache returned None (read/parse failure) or snapshot is empty.
        # Fall through to recovery logic below.

    if not os.path.exists(path):
        _ledger_log.warning('Offline ledger missing at %s — attempting recovery from backups/gist', path)
        data = _load_latest_offline_backup()
        if not data:
            gist_data, _ = _gist_fetch_snapshot()
            if isinstance(gist_data, dict):
                data = gist_data
        if data:
            try:
                _atomic_write_json(path, data)
                _invalidate_data_cache()
            except Exception:
                # Recovered data is served but NOT persisted — recovery will
                # re-run on every request until the write succeeds.
                _ledger_log.exception('Failed to persist recovered ledger to %s', path)
            return data
        # Last resort: hardcoded seed so the UI is never completely blank.
        try:
            _ledger_log.error('No recoverable ledger found — falling back to FEB26 seed data')
            payload = json.loads(json.dumps(FEB26_SEED))
            payload['server_updated_at'] = _SEED_STAMP
            payload['updated_at'] = _SEED_STAMP
            try:
                _atomic_write_json(path, payload)
                _invalidate_data_cache()
            except Exception:
                _ledger_log.exception('Failed to persist seed ledger to %s', path)
            return payload
        except Exception:
            return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if _student_count(data) == 0:
                raise ValueError('Empty offline snapshot')
            return data
    except Exception:
        _ledger_log.exception('Offline ledger at %s is corrupted/empty — attempting recovery', path)
        data = _load_latest_offline_backup()
        if not data:
            gist_data, _ = _gist_fetch_snapshot()
            if isinstance(gist_data, dict):
                data = gist_data
        if data:
            try:
                _atomic_write_json(path, data)
                _invalidate_data_cache()
            except Exception:
                _ledger_log.exception('Failed to persist recovered ledger to %s', path)
            return data
        try:
            _ledger_log.error('No recoverable ledger found — falling back to FEB26 seed data')
            payload = json.loads(json.dumps(FEB26_SEED))
            payload['server_updated_at'] = _SEED_STAMP
            payload['updated_at'] = _SEED_STAMP
            try:
                _atomic_write_json(path, payload)
                _invalidate_data_cache()
            except Exception:
                _ledger_log.exception('Failed to persist seed ledger to %s', path)
            return payload
        except Exception:
            return None


def _subscribe_sync_events():
    queue = Queue(maxsize=128)
    with _sync_lock:
        _sync_subscribers.append(queue)
    return queue


def _unsubscribe_sync_events(queue):
    with _sync_lock:
        if queue in _sync_subscribers:
            _sync_subscribers.remove(queue)


def _broadcast_sync_event(updated_at, source='server'):
    payload = json.dumps({
        'updated_at': updated_at,
        'source': source
    })
    stale = []
    with _sync_lock:
        for queue in _sync_subscribers:
            try:
                queue.put_nowait(payload)
            except Exception:
                stale.append(queue)
        for queue in stale:
            if queue in _sync_subscribers:
                _sync_subscribers.remove(queue)


def _ledger_journal_mode():
    mode = str(os.getenv('EA_LEDGER_JOURNAL_MODE', 'shadow') or 'shadow').strip().lower()
    return mode if mode in {'off', 'shadow', 'strict'} else 'shadow'


def _ledger_journal_context(next_version):
    op_id = ''
    base_version = 0
    actor_login_id = ''
    actor_role = ''
    source = 'server'
    if has_request_context():
        source = str(request.endpoint or 'request').strip()[:80] or 'request'
        try:
            body = request.get_json(silent=True) or {}
        except Exception:
            body = {}
        if isinstance(body, dict):
            op_id = str(body.get('op_id') or '').strip()
            base_version = _parse_int_safe(body.get('base_version'), 0)
        try:
            actor_login_id = str(getattr(current_user, 'login_id', '') or '').strip()[:80]
            actor_role = str(getattr(current_user, 'role', '') or '').strip()[:20]
        except Exception:
            pass
        op_id = str(getattr(g, 'ea_ledger_op_id', '') or op_id).strip()
    if not op_id:
        op_id = f'auto:{source}:{next_version}:{secrets.token_hex(12)}'
    return {
        'op_id': op_id[:160],
        'base_version': base_version or None,
        'actor_login_id': actor_login_id,
        'actor_role': actor_role,
        'source': source,
    }


def _save_offline_data(payload):
    with _LEDGER_WRITE_LOCK:
        return _save_offline_data_locked(payload)


def _save_offline_data_locked(payload):
    path = _offline_data_path()
    if isinstance(payload, dict):
        # Never persist client-view markers into the canonical ledger. These mark
        # sanitized/clipped GET responses; if a client merged one into its local
        # data and pushed it back, replication peers would refuse the server
        # snapshot forever (they treat these keys as "not a full ledger").
        payload.pop('sync_scope', None)
        payload.pop('allowed_months', None)
        # Monotonic server-side version for optimistic sync checks.
        # Read server_version from the shared cache directly instead of calling
        # _load_offline_data() (which has recovery/seed fallback logic that's
        # unnecessary here and adds overhead).
        current = _cached_load_json_data() or {}
        prev_ver = _parse_int_safe(current.get('server_version'), 0)
        next_ver = max(prev_ver + 1, _parse_int_safe(payload.get('server_version'), 0) or 0)
        payload['server_version'] = next_ver if next_ver > 0 else 1
        if not payload.get('updated_at'):
            payload['updated_at'] = payload.get('server_updated_at') or _server_now_iso()

    journal_mode = _ledger_journal_mode()
    journal_state = None
    journal_context = None
    if isinstance(payload, dict):
        _ensure_ledger_payload(payload)
        if journal_mode != 'off':
            journal_context = _ledger_journal_context(payload.get('server_version', 0))
            try:
                from app.utils.ledger_repository import prepare_revision
                journal_state = prepare_revision(
                    revision=_parse_int_safe(payload.get('server_version'), 0),
                    payload=payload,
                    op_id=journal_context['op_id'],
                    source=journal_context['source'],
                    actor_login_id=journal_context['actor_login_id'],
                    actor_role=journal_context['actor_role'],
                    base_version=journal_context['base_version'],
                )
                if journal_state.get('duplicate'):
                    payload['server_version'] = journal_state['revision']
                    return payload
            except Exception as journal_error:
                if journal_mode == 'strict':
                    raise
                _ledger_log.exception('Ledger journal prepare failed; continuing in shadow mode: %s', journal_error)

    try:
        _backup_offline_file(path)
    except Exception:
        _ledger_log.exception('Pre-save backup failed for %s', path)
    try:
        _atomic_write_json(path, payload)
    except Exception as save_error:
        if journal_state and journal_context:
            try:
                from app.utils.ledger_repository import mark_revision_failed
                mark_revision_failed(
                    revision=journal_state['revision'],
                    op_id=journal_context['op_id'],
                    error=str(save_error),
                )
            except Exception:
                _ledger_log.exception('Failed to mark ledger journal revision as failed')
        raise
    if journal_state and journal_context:
        try:
            from app.utils.ledger_repository import mark_revision_committed
            mark_revision_committed(
                revision=journal_state['revision'],
                op_id=journal_context['op_id'],
                payload_sha256=journal_state['payload_sha256'],
            )
        except Exception as journal_error:
            if journal_mode == 'strict':
                raise
            _ledger_log.exception('Ledger journal commit failed in shadow mode: %s', journal_error)
    # Prime the shared cache with the just-saved payload so the next read
    # (immediate refetch from frontend after save) is a cache hit.
    if isinstance(payload, dict):
        _prime_data_cache(payload)
    else:
        _invalidate_data_cache()
    try:
        _backup_offline_hourly_immutable(payload)
    except Exception:
        _ledger_log.exception('Hourly immutable backup failed for %s', path)
    _gist_push_snapshot_async(payload, reason='save_offline_data')
    return payload


def _is_duplicate_sync_op(payload, op_id):
    if not isinstance(payload, dict):
        return False
    op_key = str(op_id or '').strip()
    if not op_key:
        return False
    rows = payload.get('_sync_ops') or []
    if not isinstance(rows, list):
        return False
    for item in rows:
        if isinstance(item, dict) and str(item.get('id') or '').strip() == op_key:
            return True
    return False


def _record_sync_op(payload, op_id, actor=''):
    if not isinstance(payload, dict):
        return
    op_key = str(op_id or '').strip()
    if not op_key:
        return
    rows = payload.get('_sync_ops')
    if not isinstance(rows, list):
        rows = []
    rows.append({
        'id': op_key,
        'at': _server_now_iso(),
        'actor': str(actor or '').strip()[:80]
    })
    if len(rows) > _SYNC_OPS_MAX:
        rows = rows[-_SYNC_OPS_MAX:]
    payload['_sync_ops'] = rows


def _parse_sync_stamp(value):
    if not value:
        return 0.0
    try:
        text = str(value).strip().replace('Z', '+00:00')
        return datetime.fromisoformat(text).timestamp()
    except Exception:
        return 0.0


def _get_last_history_stamp(row):
    """Return the most recent history entry timestamp for a score row."""
    if not isinstance(row, dict):
        return 0.0
    history = row.get('history')
    if not isinstance(history, list) or not history:
        return 0.0
    max_stamp = 0.0
    for entry in history:
        if isinstance(entry, dict):
            ts = _parse_sync_stamp(entry.get('timestamp', ''))
            if ts > max_stamp:
                max_stamp = ts
    return max_stamp


def _payload_sync_stamp(payload):
    if not isinstance(payload, dict):
        return 0.0
    return max(
        _parse_sync_stamp(payload.get('server_updated_at')),
        _parse_sync_stamp(payload.get('updated_at'))
    )


def _is_suspicious_student_shrink(existing_payload, incoming_payload):
    """Detect stale snapshots that would silently shrink student master data."""
    if not isinstance(existing_payload, dict) or not isinstance(incoming_payload, dict):
        return False

    existing_students = existing_payload.get('students', []) or []
    incoming_students = incoming_payload.get('students', []) or []
    if not existing_students or not incoming_students:
        return False

    existing_rolls = {
        _normalize_roll_value(s.get('roll'))
        for s in existing_students
        if isinstance(s, dict) and _normalize_roll_value(s.get('roll'))
    }
    incoming_rolls = {
        _normalize_roll_value(s.get('roll'))
        for s in incoming_students
        if isinstance(s, dict) and _normalize_roll_value(s.get('roll'))
    }
    if not existing_rolls or not incoming_rolls:
        return False

    removed_rolls = existing_rolls - incoming_rolls
    hard_drop = len(incoming_rolls) + 5 < len(existing_rolls)
    large_removed_set = len(removed_rolls) >= 8
    return hard_drop and large_removed_set


def _min_safe_student_roster():
    raw = str(os.getenv('EA_MIN_SAFE_STUDENT_ROSTER', '')).strip()
    if not raw:
        return 25
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return 25
    return max(1, min(value, 10000))


def _student_count(payload):
    if not isinstance(payload, dict):
        return 0
    students = payload.get('students') or []
    return len(students) if isinstance(students, list) else 0


def _is_tiny_roster(payload, min_count=25):
    count = _student_count(payload)
    # Only treat non-empty snapshots as corrupt. Empty snapshot may be intentional for fresh bootstraps.
    return count > 0 and count < max(1, int(min_count or 25))


def _load_json_file(path):
    if not path or not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def _merge_fee_records_from_local_sources(payload):
    if not isinstance(payload, dict):
        return payload

    current_fee_records = payload.get('fee_records', [])
    if isinstance(current_fee_records, list) and current_fee_records:
        return payload

    merged_fee_records = []
    candidate_paths = []
    live_path = _offline_data_path()
    legacy_path = _legacy_instance_file('offline_scoreboard_data.json')
    for path in [live_path, legacy_path]:
        if path:
            candidate_paths.append(path)
    latest_backup = _load_latest_offline_backup()
    if isinstance(latest_backup, dict):
        merged_fee_records = _merge_fee_records_superset(
            merged_fee_records,
            latest_backup.get('fee_records', []),
        )

    seen_paths = set()
    for path in candidate_paths:
        if not path or path in seen_paths:
            continue
        seen_paths.add(path)
        candidate = _load_json_file(path)
        if not isinstance(candidate, dict):
            continue
        merged_fee_records = _merge_fee_records_superset(
            merged_fee_records,
            candidate.get('fee_records', []),
        )

    merged_fee_records = _merge_fee_records_superset(
        merged_fee_records,
        payload.get('fee_records', []),
    )
    if not merged_fee_records:
        return payload

    next_payload = dict(payload)
    next_payload['fee_records'] = merged_fee_records
    return next_payload


def _iter_offline_recovery_candidate_paths():
    """
    Yield local snapshot files that can be used to recover from a corrupt/tiny roster.
    This never deletes any files; it only reads candidates.
    """
    storage_dir = _storage_root_path()
    legacy_instance_dir = current_app.instance_path
    paths = set()
    # Prefer explicitly marked stable backups when available.
    patterns = [
        os.path.join(storage_dir, 'offline_scoreboard_data.STABLE_BACKUP*.json'),
        os.path.join(storage_dir, 'offline_scoreboard_data.pre_*.json'),
        os.path.join(_offline_backup_dir(), '*.json'),
        os.path.join(_offline_hourly_backup_dir(), '*.json'),
        os.path.join(_offline_startup_restore_dir(), '*.json'),
        # Legacy instance-path snapshots (for migration/recovery only)
        os.path.join(legacy_instance_dir, 'offline_scoreboard_data.STABLE_BACKUP*.json'),
        os.path.join(legacy_instance_dir, 'offline_scoreboard_data.pre_*.json'),
    ]
    for pattern in patterns:
        try:
            for match in glob.glob(pattern):
                if match and os.path.isfile(match):
                    paths.add(match)
        except Exception:
            continue

    live_path = _offline_data_path()
    if live_path in paths:
        paths.remove(live_path)

    for path in sorted(paths, key=os.path.getmtime, reverse=True):
        yield path


def _best_local_snapshot(min_students=25, candidate_limit=80):
    best = None
    best_stamp = 0.0
    best_mtime = 0.0
    best_count = 0
    best_src = ''
    considered = 0
    for path in _iter_offline_recovery_candidate_paths():
        considered += 1
        if considered > candidate_limit:
            break
        payload = _load_json_file(path)
        if not payload:
            continue
        count = _student_count(payload)
        if count < min_students:
            continue
        stamp = _payload_sync_stamp(payload)
        mtime = 0.0
        try:
            mtime = float(os.path.getmtime(path))
        except Exception:
            mtime = 0.0
        # Prefer higher sync stamp; fall back to mtime for payloads missing stamps.
        rank_stamp = stamp if stamp else mtime
        if not best or (rank_stamp, mtime, count) > (best_stamp, best_mtime, best_count):
            best = payload
            best_stamp = rank_stamp
            best_mtime = mtime
            best_count = count
            best_src = path
    return best, best_src


def _fetch_peer_offline_payload(base_url, timeout_sec=2.5):
    if not base_url:
        return None
    peer = str(base_url).rstrip('/')
    url = f'{peer}/scoreboard/offline-data'
    # Authenticate as a replication peer so the remote serves the FULL snapshot
    # (unauthenticated GETs now receive a sanitized public view — see
    # _sanitize_anonymous_snapshot — which must never be persisted as a backup).
    headers = {'Cache-Control': 'no-store', 'X-EA-Replicated': '1'}
    shared_key = resolve_sync_shared_key()
    if shared_key:
        headers['X-EA-Sync-Key'] = shared_key
    req = urllib.request.Request(url, method='GET', headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            body = resp.read()
    except Exception:
        return None
    try:
        if not body:
            return None
        parsed = json.loads(body.decode('utf-8', errors='replace'))
    except Exception:
        return None
    if not isinstance(parsed, dict):
        return None
    data = parsed.get('data')
    if not isinstance(data, dict):
        return None
    # Never accept sanitized/clipped payloads as recovery snapshots — they are
    # month-filtered views, not the full ledger. Persisting one would destroy data.
    if not is_full_ledger_snapshot(data):
        return None
    return data


def _best_peer_snapshot(min_students=25, timeout_sec=2.5):
    peers = get_sync_peers()
    best = None
    best_stamp = 0.0
    best_count = 0
    best_src = ''
    for peer in peers:
        payload = _fetch_peer_offline_payload(peer, timeout_sec=timeout_sec)
        if not payload:
            continue
        count = _student_count(payload)
        if count < min_students:
            continue
        stamp = _payload_sync_stamp(payload) or 0.0
        if not best or (stamp, count) > (best_stamp, best_count):
            best = payload
            best_stamp = stamp
            best_count = count
            best_src = peer
    return best, best_src


def _recover_tiny_roster_if_needed(payload, min_students=25):
    """
    If the current payload appears corrupt (tiny roster), recover from the best peer or local snapshot.
    This avoids ever serving or persisting the known "20 students" stale snapshot.
    """
    if not _is_tiny_roster(payload, min_students):
        return payload, ''

    # Use a longer timeout during startup/tiny-roster recovery so Render has enough
    # time to reach the local master before falling back to the seed.
    recovered, src = _best_peer_snapshot(min_students=min_students, timeout_sec=20)
    if not recovered:
        recovered, src = _best_local_snapshot(min_students=min_students)

    if recovered and not _is_tiny_roster(recovered, min_students):
        recovered = dict(recovered)
        # Prefer local fee_records if recovered snapshot has none (protects against stale cloud copy).
        if not recovered.get('fee_records') and payload.get('fee_records'):
            recovered['fee_records'] = payload['fee_records']
        # CRITICAL: Merge students instead of raw overwrite so structural visibility
        # fields (active_from_month, deactivation_month) set locally are never dropped.
        if isinstance(payload.get('students'), list) and isinstance(recovered.get('students'), list):
            recovered['students'] = _merge_students_preserve_active(
                payload.get('students', []),
                recovered.get('students', [])
            )
        if isinstance(payload.get('month_roster_profiles'), dict) and isinstance(recovered.get('month_roster_profiles'), dict):
            recovered['month_roster_profiles'] = _merge_month_roster_profiles_superset(
                payload.get('month_roster_profiles', {}),
                recovered.get('month_roster_profiles', {})
            )
        recovered['chess_champion'] = _merge_chess_champion_superset(
            payload.get('chess_champion', {}),
            recovered.get('chess_champion', {})
        )
        try:
            _save_offline_data(recovered)
        except Exception:
            current_app.logger.exception("Failed to persist recovered roster snapshot from %s", src or 'unknown source')
        current_app.logger.warning(
            "Recovered tiny roster (%s students) using %s (%s students).",
            _student_count(payload),
            src or 'unknown source',
            _student_count(recovered),
        )
        return recovered, src

    return payload, ''


def _recover_stale_snapshot_if_needed(payload, min_students=25, min_newer_seconds=30, allow_local_scan=False):
    """
    If a peer has a clearly newer healthy snapshot, adopt it locally.
    This heals nodes that remain stuck on an older server_updated_at.
    """
    if not isinstance(payload, dict):
        return payload, ''

    local_stamp = _payload_sync_stamp(payload) or 0.0
    best_payload = None
    best_src = ''
    best_stamp = local_stamp

    # Prefer a newer peer snapshot when peers are configured.
    # On master, still fetch peer data but protect local deactivation decisions.
    is_master = str(os.getenv('EA_MASTER_MODE', '')).strip() == '1'
    peer_payload, peer_src = _best_peer_snapshot(min_students=min_students)
    peer_stamp = _payload_sync_stamp(peer_payload) if peer_payload else 0.0
    if peer_payload and peer_stamp > best_stamp and not _is_suspicious_student_shrink(payload, peer_payload):
        if is_master:
            # Preserve local deactivation decisions — never let a peer re-activate
            # a student that admin explicitly deactivated on the master.
            local_inactive_rolls = {
                str(s.get('roll', '')).strip().upper()
                for s in payload.get('students', [])
                if s.get('active') is False
            }
            guarded = dict(peer_payload)
            for s in guarded.get('students', []):
                if str(s.get('roll', '')).strip().upper() in local_inactive_rolls:
                    s['active'] = False
            best_payload = guarded
        else:
            best_payload = peer_payload
        best_src = peer_src
        best_stamp = peer_stamp

    # Optional local backup scan (expensive): disabled on hot sync paths by default.
    if allow_local_scan:
        local_payload, local_src = _best_local_snapshot(min_students=min_students)
        local_best_stamp = _payload_sync_stamp(local_payload) if local_payload else 0.0
        if local_payload and local_best_stamp > best_stamp and not _is_suspicious_student_shrink(payload, local_payload):
            best_payload = local_payload
            best_src = local_src
            best_stamp = local_best_stamp

    if best_payload and best_stamp >= (local_stamp + float(min_newer_seconds or 0)):
        best_payload = dict(best_payload)
        # Prefer local fee_records if recovered snapshot has none (protects against stale cloud copy).
        if not best_payload.get('fee_records') and payload.get('fee_records'):
            best_payload['fee_records'] = payload['fee_records']
        # CRITICAL: Merge students instead of raw overwrite so structural visibility
        # fields (active_from_month, deactivation_month) set locally are never
        # dropped by a peer snapshot that lacks them.
        if isinstance(payload.get('students'), list) and isinstance(best_payload.get('students'), list):
            best_payload['students'] = _merge_students_preserve_active(
                payload.get('students', []),
                best_payload.get('students', [])
            )
        # Superset-merge month rosters so local roster additions are preserved.
        if isinstance(payload.get('month_roster_profiles'), dict) and isinstance(best_payload.get('month_roster_profiles'), dict):
            best_payload['month_roster_profiles'] = _merge_month_roster_profiles_superset(
                payload.get('month_roster_profiles', {}),
                best_payload.get('month_roster_profiles', {})
            )
        best_payload['chess_champion'] = _merge_chess_champion_superset(
            payload.get('chess_champion', {}),
            best_payload.get('chess_champion', {})
        )
        try:
            _save_offline_data(best_payload)
        except Exception:
            current_app.logger.exception("Failed to persist stale-recovery snapshot from %s", best_src or 'unknown source')
        current_app.logger.warning(
            "Recovered stale snapshot from %s (local=%s, recovered=%s).",
            best_src or 'unknown source',
            payload.get('server_updated_at') or payload.get('updated_at') or '',
            best_payload.get('server_updated_at') or best_payload.get('updated_at') or '',
        )
        return best_payload, best_src

    return payload, ''


def _ensure_score_timestamps(payload):
    """
    Backfill missing score timestamps for legacy rows so client-side merge logic
    can consistently prefer the newest snapshot.
    """
    if not isinstance(payload, dict):
        return False
    scores = payload.get('scores')
    if not isinstance(scores, list):
        return False

    changed = False
    fallback = str(payload.get('server_updated_at') or payload.get('updated_at') or _server_now_iso()).strip() or _server_now_iso()
    for row in scores:
        if not isinstance(row, dict):
            continue
        updated_at = str(row.get('updated_at') or '').strip()
        created_at = str(row.get('created_at') or '').strip()
        if not updated_at:
            row['updated_at'] = created_at or fallback
            updated_at = str(row.get('updated_at') or '').strip()
            changed = True
        if not created_at:
            row['created_at'] = updated_at or fallback
            changed = True
    return changed


def _normalize_roll_value(value):
    return str(value or '').strip().upper()


def _get_roll_for_month(data, student_id, month_key):
    """Thin wrapper — logic lives in app.utils.score_balance (Step 5 module split)."""
    return _score_balance.get_roll_for_month(data, student_id, month_key)


def _parse_int_safe(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_float_safe(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_holder_status(value):
    text = str(value or '').strip().lower()
    if text == 'suspended':
        return 'suspended'
    if text == 'vacant':
        return 'vacant'
    return 'active'


def _normalize_post_text(value):
    return str(value or '').strip().lower()


def _leadership_role_type(post_name):
    text = _normalize_post_text(post_name)
    if not text:
        return ''
    # Check for opposition / lop
    if re.search(r'\b(leader of opposition|opposition leader|lop)\b', text) or '(lop)' in text:
        return 'lop'
    # Check for co-leader / co leader / col
    if re.search(r'\b(co-leader|co-leaders|co leader|co leaders|col)\b', text) or '(col)' in text:
        return 'co_leader'
    # Check for leader / l
    if (re.search(r'\b(leader|l)\b', text) or '(l)' in text) and not re.search(r'\b(opposition|co-leader|co-leaders|co leader|co leaders|col)\b', text):
        return 'leader'
    return ''


def _leadership_veto_quota(post_name):
    """Role VETO quota for a leadership post — sourced from VETO_QUOTAS (H2 fix).

    Previously hardcoded 5/3/2/1 here while constants.py and the JS defaults
    maintained their own copies, causing divergence. Now all consumers read
    from the same constants.VETO_QUOTAS dict.
    """
    role_type = _leadership_role_type(post_name)
    if role_type == 'leader':
        return VETO_QUOTAS.get('LEADER', 0)
    if role_type == 'co_leader':
        return VETO_QUOTAS.get('CO_LEADER', 0)
    if role_type == 'lop':
        return VETO_QUOTAS.get('LOP', 0)

    text = str(post_name or '').strip().lower()
    # Check for discipline and welfare or dwi
    if ('discipline' in text and 'welfare' in text) or re.search(r'\bdwi\b', text) or '(dwi)' in text:
        return VETO_QUOTAS.get('DWI', 0)
    return VETO_QUOTAS.get('OTHER', 0)


def _tenure_months_for_assignment(source, post_name=''):
    source_key = str(source or '').strip().lower()
    if source_key in ('class_rep', 'group_cr'):
        return 1
    if source_key == 'leadership' and _leadership_role_type(post_name) == 'co_leader':
        return 1
    return 2


def _parse_date_key(value):
    text = str(value or '').strip()
    if not text:
        return None
    try:
        if len(text) >= 10:
            return datetime.fromisoformat(text[:10]).date()
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def _is_assignment_active_by_tenure(elected_on, tenure_months=2, extension_months=0, on_date=None):
    start_date = _parse_date_key(elected_on)
    if not start_date:
        return False
    total_months = max(0, _parse_int_safe(tenure_months) + _parse_int_safe(extension_months))
    end_date = start_date + relativedelta(months=total_months)
    check_date = on_date if isinstance(on_date, date) else _parse_date_key(on_date)
    if not check_date:
        check_date = _parse_date_key(_server_now_iso())
    if not check_date:
        check_date = date.today()
    return start_date <= check_date <= end_date


def _build_student_lookups(data):
    students = data.get('students', []) or []
    by_id = {}
    by_roll = {}
    for student in students:
        sid = _parse_int_safe(student.get('id'), 0)
        if sid <= 0:
            continue
        by_id[sid] = student
        roll = _normalize_roll_value(student.get('roll'))
        if roll and roll not in by_roll:
            by_roll[roll] = sid
    return by_id, by_roll


def _compute_active_role_veto_quotas(data, date_key=None):
    by_id, by_roll = _build_student_lookups(data)
    quotas = {}
    check_date = _parse_date_key(date_key) if date_key else _parse_date_key(_server_now_iso())
    if not check_date:
        check_date = date.today()

    def _add_quota(student_id, amount):
        sid = _parse_int_safe(student_id, 0)
        quota = _parse_int_safe(amount, 0)
        if sid <= 0 or quota <= 0:
            return
        quotas[sid] = quotas.get(sid, 0) + quota

    for post in data.get('leadership', []) or []:
        if _normalize_holder_status(post.get('status')) != 'active':
            continue
        tenure_months = _tenure_months_for_assignment('leadership', post.get('post'))
        extension_months = _parse_int_safe(post.get('tenure_extension_months'), 0)
        if not _is_assignment_active_by_tenure(post.get('elected_on'), tenure_months, extension_months, check_date):
            continue
        quota = _leadership_veto_quota(post.get('post'))
        if quota <= 0:
            continue
        sid = _parse_int_safe(post.get('studentId'), 0)
        if sid <= 0:
            roll = _normalize_roll_value(post.get('roll'))
            sid = by_roll.get(roll, 0) if roll else 0
        if sid in by_id:
            _add_quota(sid, quota)

    # CR quota (+2) once per student if either class CR or group CR is active.
    # class_reps and group_crs are explicitly managed by admins via status field
    # (active/ended), so we trust status alone — no tenure-expiry gate applied.
    # This ensures vetoes are granted instantly on election win, direct appointment,
    # or post resume (the reconciliation runs immediately after each data sync).
    cr_students = set()
    for rep in data.get('class_reps', []) or []:
        if _normalize_holder_status(rep.get('status') or 'active') != 'active':
            continue
        sid = _parse_int_safe(rep.get('studentId'), 0)
        if sid in by_id:
            cr_students.add(sid)

    for rep in data.get('group_crs', []) or []:
        if _normalize_holder_status(rep.get('status') or 'active') != 'active':
            continue
        sid = _parse_int_safe(rep.get('studentId'), 0)
        if sid in by_id:
            cr_students.add(sid)

    for sid in cr_students:
        _add_quota(sid, VETO_QUOTAS.get('CR', 0))

    return quotas


def _reconcile_role_veto_monthly(data, month_key=None, date_key=None):
    if not isinstance(data, dict):
        return
    by_id, _ = _build_student_lookups(data)
    if 'role_veto_monthly' not in data or not isinstance(data.get('role_veto_monthly'), dict):
        data['role_veto_monthly'] = {}

    resolved_month = str(month_key or _server_now_iso()[:7]).strip()
    if not re.match(r'^\d{4}-\d{2}$', resolved_month):
        resolved_month = _server_now_iso()[:7]

    # Compute current role-grant quotas for active postholders.
    target = _compute_active_role_veto_quotas(data, date_key=date_key)
    next_state = {str(sid): grant for sid, grant in target.items() if grant > 0}
    data['role_veto_monthly'][resolved_month] = next_state
    data['role_veto_applied_month'] = resolved_month

    # Role-grant VETOs are stored in role_veto_count only — veto_count (individual
    # carry) is never touched here so admin corrections are permanently preserved.
    students = data.get('students', []) or []
    for student in students:
        sid = _parse_int_safe(student.get('id'), 0)
        if sid <= 0:
            continue
        grant = max(0, _parse_int_safe(target.get(sid), 0))
        student['role_veto_count'] = grant


def _get_student_month_consumed_vetos(data, student_id, month):
    sid = _parse_int_safe(student_id, 0)
    if sid <= 0 or not isinstance(data, dict):
        return 0
    scores = data.get('scores', []) or []
    consumed = 0
    for score in scores:
        if not isinstance(score, dict):
            continue
        if _parse_int_safe(score.get('studentId'), 0) != sid:
            continue
        if score.get('month') != month:
            continue
        
        # 1. Negative vetos delta in scores
        v_val = _parse_int_safe(score.get('vetos'), 0)
        if v_val < 0:
            consumed += abs(v_val)
            
        # 2. Check notes for display-only veto usage
        notes = str(score.get('notes', '') or '')
        parts = [p.strip() for p in notes.split('|') if p.strip()]
        display_only = 0
        for part in parts:
            if re.match(r'^1V used to select (.+?) post for (.+)$', part, re.IGNORECASE):
                display_only += 1
        if '[veto shield]' in notes.lower():
            display_only += 1
            
        # Add display-only if it was not already accounted for by negative vetos field
        if v_val >= 0:
            consumed += display_only
            
    return consumed


def _get_student_total_used_vetos_from_scores(data, student_id):
    sid = _parse_int_safe(student_id, 0)
    if sid <= 0 or not isinstance(data, dict):
        return 0
    scores = data.get('scores', []) or []
    consumed = 0
    for score in scores:
        if not isinstance(score, dict):
            continue
        if _parse_int_safe(score.get('studentId'), 0) != sid:
            continue
        
        # 1. Negative vetos delta in scores
        v_val = _parse_int_safe(score.get('vetos'), 0)
        if v_val < 0:
            consumed += abs(v_val)
            
        # 2. Check notes for display-only veto usage
        notes = str(score.get('notes', '') or '')
        parts = [p.strip() for p in notes.split('|') if p.strip()]
        display_only = 0
        for part in parts:
            if re.match(r'^1V used to select (.+?) post for (.+)$', part, re.IGNORECASE):
                display_only += 1
        if '[veto shield]' in notes.lower():
            display_only += 1
            
        # Add display-only if it was not already accounted for by negative vetos field
        if v_val >= 0:
            consumed += display_only
            
    return consumed


def _compute_veto_remaining_counters(data, sid, ind_alloc, role_alloc, month_key):
    """Compute live remaining counters for veto_count and role_veto_count.

    Individual VETOs are PERMANENT — usage from any month permanently reduces
    the pool. Role VETOs are MONTHLY — they expire at month end, so only
    current-month usage is deducted from the current grant.

    Spend priority: individual first, then role (matches spendStudentVetoPower
    in the JS and adjustStudentVetoCount).

    Returns (ind_remaining, role_remaining, total_used).
    """
    total_used = _get_student_total_used_vetos_from_scores(data, sid) if sid else 0
    month_used = _get_student_month_consumed_vetos(data, sid, month_key) if sid else 0

    # Individual VETOs: permanently deduct total usage (all months)
    ind_used_total = min(ind_alloc, total_used)
    ind_rem = ind_alloc - ind_used_total

    # Role VETOs: deduct only this month's role usage.
    # Role is used after individual is exhausted, so we need to figure out
    # how much of this month's usage hit role vs individual.
    prior_used = max(0, total_used - month_used)
    ind_available_at_month_start = max(0, ind_alloc - prior_used)
    ind_used_this_month = min(month_used, ind_available_at_month_start)
    role_used_this_month = max(0, month_used - ind_used_this_month)
    role_rem = max(0, role_alloc - role_used_this_month)

    return ind_rem, role_rem, total_used


def _reconcile_veto_tracking_from_data(data):
    """
    Reconciles the veto_tracking ledger with the actual scores and log entries.
    - Adds star-to-veto conversions to student's individual_vetos.
    - Updates used_vetos and remaining_vetos based on both ledger and scores.
    """
    if not isinstance(data, dict):
        return
        
    if 'veto_tracking' not in data or not isinstance(data.get('veto_tracking'), dict):
        data['veto_tracking'] = {
            'hardened': True,
            'initialized_at': _server_now_iso(),
            'version': 2,
            'students': {},
            'usage_log': [],
            'last_reset': _server_now_iso()
        }
        
    veto_tracking = data['veto_tracking']
    students_map = veto_tracking.setdefault('students', {})
    usage_log = veto_tracking.setdefault('usage_log', [])
    
    # 1. Individual veto allocations — single source of truth in constants.py
    initial_allocations = VETO_INDIVIDUAL_ALLOCATIONS
    
    # 1b. Compute conversion gains from usage_log (deduplicated by timestamp)
    #     This is computed BEFORE writing students_map so individual_vetos can be
    #     set as initial_allocations + conversions in a single pass, making the
    #     function fully idempotent (no accumulation across calls).
    seen_conversion_ts = set()
    conversion_gains = {}
    # 1c. Compute manual admin grants/revoke from usage_log (deduplicated by
    #     entry_id, falling back to timestamp). These let an admin durably add
    #     or remove individual VETOs via the Adjust VETO dialog — without this,
    #     the server reconcile would overwrite student.veto_count from the
    #     constant allocations and revert the admin's manual change.
    seen_manual_keys = set()
    manual_grants = {}
    for entry in usage_log:
        if not isinstance(entry, dict):
            continue
        action = entry.get('action')
        if action == 'star_to_veto_converted':
            ts = entry.get('timestamp')
            if ts and ts in seen_conversion_ts:
                continue
            roll = entry.get('roll')
            gained = _parse_int_safe(entry.get('vetos_gained'), 0)
            if roll and gained > 0:
                conversion_gains[roll] = conversion_gains.get(roll, 0) + gained
            if ts:
                seen_conversion_ts.add(ts)
        elif action in ('manual_veto_grant', 'manual_veto_revoke'):
            key = entry.get('entry_id') or entry.get('timestamp')
            if key and key in seen_manual_keys:
                continue
            roll = entry.get('roll')
            if not roll:
                continue
            if action == 'manual_veto_grant':
                n = _parse_int_safe(entry.get('vetos_granted'), 0)
                if n > 0:
                    manual_grants[roll] = manual_grants.get(roll, 0) + n
            else:
                n = _parse_int_safe(entry.get('vetos_revoked'), 0)
                if n > 0:
                    manual_grants[roll] = manual_grants.get(roll, 0) - n
            if key:
                seen_manual_keys.add(key)

    # Reset/ensure everyone in veto_tracking has their base allocations and current role allocations
    # ALWAYS reset individual_vetos = initial_allocations + conversion_gains + manual_grants.
    # This makes the function idempotent: running it N times produces the same
    # result as running it once. Previous versions preserved existing
    # individual_vetos which caused unbounded inflation on repeated calls.
    # Compute active gross role veto quotas for office holders
    role_grant_map = _compute_active_role_veto_quotas(data)

    for student in data.get('students', []):
        roll = student.get('roll')
        if not roll:
            continue
        sid = _parse_int_safe(student.get('id'), 0)
        role_vc = max(0, _parse_int_safe(role_grant_map.get(sid), 0))
        base_alloc = initial_allocations.get(roll, 0)
        conv_gain = conversion_gains.get(roll, 0)
        manual_gain = manual_grants.get(roll, 0)
        ind_vetos = max(0, base_alloc + conv_gain + manual_gain)
        students_map[roll] = {
            'name': student.get('name', 'Unknown'),
            'individual_vetos': ind_vetos,
            'role_vetos': role_vc,
            'total_vetos': ind_vetos + role_vc,
            'used_vetos': 0,
            'remaining_vetos': ind_vetos + role_vc
        }
            
    # 2b. Remove stale roll entries from students_map that are no longer in the
    #     current student list (e.g. after a roll change). This prevents old
    #     inflated entries from persisting and being looked up by stale rolls.
    current_rolls = {s.get('roll') for s in data.get('students', []) if s.get('roll')}
    stale_rolls = [r for r in list(students_map.keys()) if r not in current_rolls]
    for r in stale_rolls:
        del students_map[r]

    # 3. Update used_vetos and remaining_vetos from actual score entries
    roll_to_sid = {s.get('roll'): s.get('id') for s in data.get('students', []) if s.get('roll') and s.get('id')}
    for roll, s_data in students_map.items():
        sid = roll_to_sid.get(roll)
        used_from_scores = _get_student_total_used_vetos_from_scores(data, sid) if sid else 0
        used_from_ledger = _parse_int_safe(s_data.get('used_vetos'), 0)
        
        total_used = max(used_from_ledger, used_from_scores)
        s_data['used_vetos'] = total_used
        s_data['total_vetos'] = s_data.get('individual_vetos', 0) + s_data.get('role_vetos', 0)
        s_data['remaining_vetos'] = max(0, s_data['total_vetos'] - total_used)


def _reconcile_veto_counters_from_scores(data, month_key=None):
    if not isinstance(data, dict):
        return
    month = str(month_key or _server_now_iso()[:7]).strip()
    if not re.match(r'^\d{4}-\d{2}$', month):
        month = _server_now_iso()[:7]

    # Reconcile veto_tracking ledger first to ensure it's up to date
    try:
        _reconcile_veto_tracking_from_data(data)
    except Exception as e:
        current_app.logger.warning(f"Failed to reconcile veto_tracking from data: {e}")

    grants = {}
    if isinstance(data.get('role_veto_monthly'), dict):
        grants = data['role_veto_monthly'].get(month, {}) or {}
    if not isinstance(grants, dict):
        grants = {}

    students = data.get('students', []) or []
    by_id = { _parse_int_safe(s.get('id'), 0): s for s in students if _parse_int_safe(s.get('id'), 0) > 0 }
    carry_by_student = {}
    month_profiles = {}
    carry_month = month
    try:
        if month == _server_now_iso()[:7]:
            dt = datetime.strptime(month + '-01', '%Y-%m-%d')
            prev_year = dt.year if dt.month > 1 else dt.year - 1
            prev_month = dt.month - 1 if dt.month > 1 else 12
            carry_month = f"{prev_year:04d}-{prev_month:02d}"
    except Exception:
        carry_month = month
    if isinstance(data.get('month_roster_profiles'), dict):
        month_profiles = data.get('month_roster_profiles', {}).get(carry_month, {}) or {}
    profile_rows = month_profiles if isinstance(month_profiles, list) else list((month_profiles or {}).values())
    roll_to_sid = {
        _normalize_roll_value(s.get('roll')): _parse_int_safe(s.get('id'), 0)
        for s in students
        if _parse_int_safe(s.get('id'), 0) > 0 and _normalize_roll_value(s.get('roll'))
    }
    for profile in profile_rows:
        if not isinstance(profile, dict):
            continue
        roll = _normalize_roll_value(profile.get('roll'))
        if not roll:
            continue
        sid = roll_to_sid.get(roll, 0)
        if sid <= 0:
            continue
        carry_by_student[sid] = max(0, _parse_int_safe(profile.get('month_veto_count'), 0))

    # Role-grant VETOs (expire monthly) are stored separately in role_veto_count.
    # individual veto_count is never touched here — it is set by admin and carries
    # forward via month_roster_profiles; role grants must not pollute it.
    for k, v in grants.items():
        sid = _parse_int_safe(k, 0)
        if sid <= 0:
            continue
        student = by_id.get(sid)
        if not student:
            continue
        student['role_veto_count'] = max(0, _parse_int_safe(v, 0))
    # Zero out role_veto_count for any student no longer in the grant list
    grant_sids = {_parse_int_safe(k, 0) for k in grants.keys()}
    for student in students:
        sid = _parse_int_safe(student.get('id'), 0)
        if sid > 0 and sid not in grant_sids:
            if student.get('role_veto_count', 0) != 0:
                student['role_veto_count'] = 0

    # Deduct spent vetoes from veto_count and role_veto_count.
    # Priority: individual veto_count first, then role_veto_count (matching spendStudentVetoPower).
    # Individual VETOs are PERMANENT — total usage across all months is deducted.
    # Role VETOs are MONTHLY — only current-month usage is deducted (they expire).
    veto_tracking = data.get('veto_tracking', {})
    tracked = veto_tracking.get('students') or {}
    roll_to_sid_for_month = {
        s.get('roll'): _parse_int_safe(s.get('id'), 0)
        for s in students if s.get('roll') and _parse_int_safe(s.get('id'), 0) > 0
    }
    for student in students:
        roll = student.get('roll')
        if not roll or roll not in tracked:
            continue
        entry = tracked[roll]
        sid = roll_to_sid_for_month.get(roll, 0)

        ind_alloc = _parse_int_safe(entry.get('individual_vetos'), 0)
        role_alloc = _parse_int_safe(entry.get('role_vetos'), 0)
        ind_rem, role_rem, total_used = _compute_veto_remaining_counters(
            data, sid, ind_alloc, role_alloc, month
        )

        old_vc = _parse_int_safe(student.get('veto_count'), 0)
        old_rvc = _parse_int_safe(student.get('role_veto_count'), 0)
        old_used = _parse_int_safe(student.get('used_veto_count'), 0)
        if old_vc != ind_rem or old_rvc != role_rem or old_used != total_used:
            student['veto_count'] = ind_rem
            student['role_veto_count'] = role_rem
            student['used_veto_count'] = total_used
            # Update timestamp so client merge logic (timestamp-based) accepts
            # the server's corrected veto counts over locally inflated values.
            student['updated_at'] = _server_now_iso()


def _compute_student_star_balance(data, student_id, month_key):
    """Thin wrapper — logic lives in app.utils.score_balance (Step 5 module split)."""
    return _score_balance.compute_star_balance(data, student_id, month_key)


def _compute_student_veto_balance(data, student_id, month_key):
    """Thin wrapper — logic lives in app.utils.score_balance (Step 5 module split)."""
    return _score_balance.compute_veto_balance(data, student_id, month_key, _server_now_iso()[:7])


def _merge_teacher_scores(existing_data, incoming_data):
    """Merge teacher score payload safely across devices with different local IDs."""
    existing_scores = list(existing_data.get('scores', []) or [])
    incoming_scores = incoming_data.get('scores', []) or []
    if not incoming_scores:
        return existing_scores
    now_iso = _server_now_iso()

    existing_students = existing_data.get('students', []) or []
    incoming_students = incoming_data.get('students', []) or []

    # Build set of student IDs that have GCB immunity (score floor -20).
    gcb_student_ids = {
        _parse_int_safe(s.get('id'), 0)
        for s in existing_students
        if isinstance(s, dict) and s.get('gcb')
    }

    existing_id_by_roll = {}
    existing_id_set = set()
    for student in existing_students:
        sid = student.get('id')
        if sid is None:
            continue
        existing_id_set.add(sid)
        roll = _normalize_roll_value(student.get('roll'))
        if roll and roll not in existing_id_by_roll:
            existing_id_by_roll[roll] = sid

    incoming_roll_by_id = {}
    for student in incoming_students:
        sid = student.get('id')
        if sid is None:
            continue
        roll = _normalize_roll_value(student.get('roll'))
        if roll:
            incoming_roll_by_id[str(sid)] = roll

    score_index = {}
    max_score_id = 0
    for idx, score in enumerate(existing_scores):
        sid = score.get('studentId')
        date_key = str(score.get('date') or '').strip()
        if sid is None or not date_key:
            continue
        score_index[(str(sid), date_key)] = idx
        max_score_id = max(max_score_id, _parse_int_safe(score.get('id')))

    for incoming in incoming_scores:
        if not isinstance(incoming, dict):
            continue
        recorded_by = str(incoming.get('recordedBy') or '').strip().lower()
        if recorded_by and recorded_by != 'teacher':
            continue
        if not recorded_by:
            continue
        incoming_sid = incoming.get('studentId')
        date_key = str(incoming.get('date') or '').strip()
        if incoming_sid is None or not date_key:
            continue

        incoming_roll = incoming_roll_by_id.get(str(incoming_sid), '')
        target_sid = existing_id_by_roll.get(incoming_roll)
        if target_sid is None and incoming_sid in existing_id_set:
            target_sid = incoming_sid
        if target_sid is None:
            continue

        index_key = (str(target_sid), date_key)
        existing_idx = score_index.get(index_key)
        existing_score = existing_scores[existing_idx] if existing_idx is not None else None
        # Teachers are not allowed to change stars/vetos directly. Preserve whatever is on record.
        approved_stars = _parse_int_safe(existing_score.get('stars')) if isinstance(existing_score, dict) else 0
        approved_vetos = _parse_int_safe(existing_score.get('vetos')) if isinstance(existing_score, dict) else 0
        month_key = str(incoming.get('month') or '').strip() or date_key[:7]
        raw_points = _parse_float_safe(incoming.get('points'))
        
        clamped_points = raw_points
                
        # Teacher can input a maximum of 50 points
        if clamped_points > 50:
            clamped_points = 50
                
        # Apply GCB student immunity minimum score floor (-20)
        if target_sid in gcb_student_ids:
            clamped_points = max(-20, clamped_points)

        normalized_score = {
            'studentId': target_sid,
            'date': date_key,
            'points': clamped_points,
            'stars': approved_stars,
            'vetos': approved_vetos,
            'month': month_key,
            'notes': str(incoming.get('notes') or ''),
            'recordedBy': 'teacher',
            # Critical for client convergence: without updated_at, stale local rows can
            # win tie-breaks and keep showing old points after a valid teacher update.
            'updated_at': str(incoming.get('updated_at') or now_iso).strip() or now_iso
        }
        if existing_idx is not None:
            # Keep the earliest known created_at when updating an existing record.
            existing_created = str(existing_score.get('created_at') or '').strip() if isinstance(existing_score, dict) else ''
            incoming_created = str(incoming.get('created_at') or '').strip()
            normalized_score['created_at'] = existing_created or incoming_created or normalized_score['updated_at']
            existing_score.update(normalized_score)
        else:
            max_score_id += 1
            normalized_score['id'] = max_score_id
            normalized_score['created_at'] = str(incoming.get('created_at') or normalized_score['updated_at']).strip() or normalized_score['updated_at']
            score_index[index_key] = len(existing_scores)
            existing_scores.append(normalized_score)

    return existing_scores


def _merge_appeals_superset(existing_appeals, incoming_appeals):
    """Merge appeals by id and keep the latest updated entry."""
    merged = {}

    def _appeal_key(item):
        if not isinstance(item, dict):
            return ''
        appeal_id = item.get('id')
        if appeal_id is None:
            return ''
        key = str(appeal_id).strip()
        return key

    for source in (existing_appeals or []), (incoming_appeals or []):
        for item in source:
            if not isinstance(item, dict):
                continue
            key = _appeal_key(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = dict(item)
                continue
            prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
            next_stamp = _parse_sync_stamp(item.get('updated_at') or item.get('created_at'))
            if next_stamp >= prev_stamp:
                merged[key] = dict(item)

    return list(merged.values())


def _non_admin_edit_window_bounds(today=None):
    base = today or datetime.now().date()
    start = base - timedelta(days=_NON_ADMIN_EDIT_WINDOW_DAYS)
    return start, base


def _is_date_within_non_admin_window(date_value, today=None):
    d = _parse_date_key(date_value)
    if not d:
        return False
    start, end = _non_admin_edit_window_bounds(today=today)
    return start <= d <= end


def _filter_teacher_payload_to_edit_window(incoming_data, teacher_login_id='Teacher'):
    """
    Teachers can only modify:
    - scores within edit window (today and previous N days) recordedBy=teacher
    - attendance within edit window
    - their own appeals created in the same edit window
    """
    if not isinstance(incoming_data, dict):
        return {}
    today = datetime.now().date()
    filtered = dict(incoming_data)

    filtered_scores = []
    for row in incoming_data.get('scores', []) or []:
        if not isinstance(row, dict):
            continue
        recorded_by = str(row.get('recordedBy') or '').strip().lower()
        if recorded_by != 'teacher':
            continue
        date_key = str(row.get('date') or '').strip()
        if not _is_date_within_non_admin_window(date_key, today=today):
            continue
        filtered_scores.append(row)
    filtered['scores'] = filtered_scores

    filtered_attendance = []
    for item in incoming_data.get('attendance', []) or []:
        if not isinstance(item, dict):
            continue
        date_key = str(item.get('date') or '').strip()
        if not _is_date_within_non_admin_window(date_key, today=today):
            continue
        filtered_attendance.append(item)
    filtered['attendance'] = filtered_attendance

    # Appeals: allow only teacher-originated, current-month entries.
    filtered_appeals = []
    teacher_key = str(teacher_login_id or 'Teacher').strip().lower()
    for item in incoming_data.get('appeals', []) or []:
        if not isinstance(item, dict):
            continue
        from_role = str(item.get('from_role') or '').strip().lower()
        created_by = str(item.get('created_by') or '').strip().lower()
        if from_role and from_role != 'teacher' and created_by != teacher_key and created_by != 'teacher':
            continue
        score_month = str(item.get('score_month') or '').strip()
        score_date = str(item.get('score_date') or '').strip()
        date_for_window = score_date or (f"{score_month}-01" if re.match(r'^\d{4}-\d{2}$', score_month) else '')
        if not date_for_window:
            created_at = str(item.get('created_at') or '').strip()
            if len(created_at) >= 10:
                date_for_window = created_at[:10]
        if not _is_date_within_non_admin_window(date_for_window, today=today):
            continue
        filtered_appeals.append(item)
    filtered['appeals'] = filtered_appeals

    # CRITICAL FIX: Ensure students are preserved for attendance merge identity lookup
    # The merge function needs student ID->roll mappings to properly identify attendance records
    if 'students' not in filtered and isinstance(incoming_data.get('students'), list):
        filtered['students'] = incoming_data.get('students', [])

    return filtered


def _build_teacher_replication_patch(full_payload, teacher_login_id='Teacher'):
    """Build a narrow replication patch safe to apply on master server."""
    today = datetime.now().date()
    payload = full_payload if isinstance(full_payload, dict) else {}

    students_min = []
    for student in payload.get('students', []) or []:
        if not isinstance(student, dict):
            continue
        sid = student.get('id')
        roll = _normalize_roll_value(student.get('roll'))
        if sid is None or not roll:
            continue
        students_min.append({'id': sid, 'roll': roll})

    scores = []
    for row in payload.get('scores', []) or []:
        if not isinstance(row, dict):
            continue
        recorded_by = str(row.get('recordedBy') or '').strip().lower()
        if recorded_by != 'teacher':
            continue
        date_key = str(row.get('date') or '').strip()
        if not _is_date_within_non_admin_window(date_key, today=today):
            continue
        scores.append(row)

    attendance = []
    for item in payload.get('attendance', []) or []:
        if not isinstance(item, dict):
            continue
        date_key = str(item.get('date') or '').strip()
        if not _is_date_within_non_admin_window(date_key, today=today):
            continue
        attendance.append(item)

    appeals = []
    teacher_key = str(teacher_login_id or 'Teacher').strip().lower()
    for item in payload.get('appeals', []) or []:
        if not isinstance(item, dict):
            continue
        from_role = str(item.get('from_role') or '').strip().lower()
        created_by = str(item.get('created_by') or '').strip().lower()
        if from_role and from_role != 'teacher' and created_by != teacher_key and created_by != 'teacher':
            continue
        score_month = str(item.get('score_month') or '').strip()
        score_date = str(item.get('score_date') or '').strip()
        date_for_window = score_date or (f"{score_month}-01" if re.match(r'^\d{4}-\d{2}$', score_month) else '')
        if not date_for_window:
            created_at = str(item.get('created_at') or '').strip()
            if len(created_at) >= 10:
                date_for_window = created_at[:10]
        if not _is_date_within_non_admin_window(date_for_window, today=today):
            continue
        appeals.append(item)

    syllabus_tracking = []
    for row in payload.get('syllabus_tracking', []) or []:
        if not isinstance(row, dict):
            continue
        key = str(row.get('key') or '').strip()
        if not key:
            continue
        syllabus_tracking.append(row)

    return {
        'actor_role': 'teacher',
        'replica_purpose': 'teacher_patch',
        'students': students_min,
        'scores': scores,
        'attendance': attendance,
        'appeals': appeals,
        'syllabus_tracking': syllabus_tracking,
        'updated_at': payload.get('updated_at')
    }


def _merge_month_students_superset(existing_ms, incoming_ms):
    """Superset merge for month_students: never let an incoming partial roster shrink an existing month's list."""
    if not isinstance(existing_ms, dict):
        existing_ms = {}
    if not isinstance(incoming_ms, dict):
        incoming_ms = {}
    merged = {}
    for month in set(list(existing_ms.keys()) + list(incoming_ms.keys())):
        existing_rolls = list(existing_ms.get(month) or [])
        incoming_rolls = list(incoming_ms.get(month) or [])
        seen = set(str(r or '').strip().upper() for r in existing_rolls if r)
        combined = list(existing_rolls)
        for r in incoming_rolls:
            key = str(r or '').strip().upper()
            if key and key not in seen:
                combined.append(r)
                seen.add(key)
        merged[month] = combined
    return merged


def _merge_month_roster_profiles_superset(existing_rp, incoming_rp):
    """Superset merge for month_roster_profiles: union by roll across months."""
    if not isinstance(existing_rp, dict):
        existing_rp = {}
    if not isinstance(incoming_rp, dict):
        incoming_rp = {}
    merged = {}

    def _class_stamp(profile):
        if not isinstance(profile, dict):
            return 0.0
        class_stamp = _parse_sync_stamp(profile.get('class_updated_at') or '')
        if class_stamp > 0:
            return class_stamp
        return _parse_sync_stamp(profile.get('updated_at') or profile.get('created_at') or '')

    for month in set(list(existing_rp.keys()) + list(incoming_rp.keys())):
        by_roll = {}
        for p in (existing_rp.get(month) or []):
            if not isinstance(p, dict):
                continue
            roll = str(p.get('roll') or '').strip().upper()
            if roll:
                by_roll[roll] = dict(p)
        for p in (incoming_rp.get(month) or []):
            if not isinstance(p, dict):
                continue
            roll = str(p.get('roll') or '').strip().upper()
            if roll:
                existing_profile = by_roll.get(roll) or {}
                merged_profile = {**existing_profile, **p}
                if _class_stamp(existing_profile) >= _class_stamp(p):
                    if 'class' in existing_profile:
                        merged_profile['class'] = existing_profile.get('class')
                    if existing_profile.get('class_updated_at'):
                        merged_profile['class_updated_at'] = existing_profile.get('class_updated_at')
                # Preserve name from the side with the newer name stamp.
                def _rp_name_stamp(row):
                    if not isinstance(row, dict):
                        return 0.0
                    return _parse_sync_stamp(row.get('name_updated_at') or row.get('updated_at') or row.get('created_at') or '')
                if _rp_name_stamp(existing_profile) > _rp_name_stamp(p) and existing_profile.get('base_name'):
                    merged_profile['base_name'] = existing_profile.get('base_name')
                    merged_profile['name'] = existing_profile.get('base_name')
                elif _rp_name_stamp(p) > _rp_name_stamp(existing_profile) and p.get('base_name'):
                    merged_profile['base_name'] = p.get('base_name')
                    merged_profile['name'] = p.get('base_name')
                by_roll[roll] = merged_profile
        merged[month] = list(by_roll.values())
    return merged


def _clone_jsonish(value):
    return json.loads(json.dumps(value))


def _locked_month_keys(payload):
    if not isinstance(payload, dict):
        return set()
    months = set()
    for month in payload.get('locked_months', []) or []:
        mk = str(month or '').strip()
        if re.match(r'^\d{4}-\d{2}$', mk):
            months.add(mk)
    frozen = payload.get('frozen_months', {}) or {}
    if isinstance(frozen, dict):
        for month, info in frozen.items():
            mk = str(month or '').strip()
            if not re.match(r'^\d{4}-\d{2}$', mk):
                continue
            if isinstance(info, dict) and info.get('hardened') is True and info.get('allow_modifications') is False:
                months.add(mk)
    return months


def _preserve_locked_historical_window(existing_payload, incoming_payload):
    """Never let sync/publish overwrite locked historical months.

    Once months are locked/frozen, the live server copy is authoritative for:
    - locked_months / frozen_months metadata
    - month_students / month_roster_profiles / month_extra_columns / month_student_extras / role_veto_monthly
    - score rows inside the locked month window

    Historical rebuilds/restores should use dedicated admin/import endpoints, not generic sync/publish.
    """
    existing = existing_payload if isinstance(existing_payload, dict) else {}
    incoming = incoming_payload if isinstance(incoming_payload, dict) else {}
    locked_months = sorted(_locked_month_keys(existing) | _locked_month_keys(incoming))
    if not locked_months:
        return incoming

    incoming['locked_months'] = locked_months

    existing_frozen = existing.get('frozen_months', {}) if isinstance(existing.get('frozen_months'), dict) else {}
    incoming_frozen = incoming.get('frozen_months', {}) if isinstance(incoming.get('frozen_months'), dict) else {}
    merged_frozen = dict(incoming_frozen)
    for month in locked_months:
        if month in existing_frozen:
            merged_frozen[month] = _clone_jsonish(existing_frozen[month])
    incoming['frozen_months'] = merged_frozen

    month_scoped_keys = [
        'month_students',
        'month_roster_profiles',
        'month_extra_columns',
        'month_student_extras',
        'role_veto_monthly',
    ]
    for key in month_scoped_keys:
        existing_map = existing.get(key, {}) if isinstance(existing.get(key), dict) else {}
        incoming_map = incoming.get(key, {}) if isinstance(incoming.get(key), dict) else {}
        merged_map = dict(incoming_map)
        for month in locked_months:
            if month in existing_map:
                merged_map[month] = _clone_jsonish(existing_map[month])
        incoming[key] = merged_map

    existing_scores = existing.get('scores', []) if isinstance(existing.get('scores'), list) else []
    incoming_scores = incoming.get('scores', []) if isinstance(incoming.get('scores'), list) else []
    preserved_scores = []
    for row in existing_scores:
        month_key = str(row.get('month') or str(row.get('date') or '')[:7]).strip()
        if month_key in locked_months:
            preserved_scores.append(_clone_jsonish(row))
    live_scores = []
    for row in incoming_scores:
        month_key = str(row.get('month') or str(row.get('date') or '')[:7]).strip()
        if month_key in locked_months:
            continue
        live_scores.append(row)
    incoming['scores'] = preserved_scores + live_scores
    return incoming


def _apply_admin_historical_score_ops(payload, ops, actor_login_id='Admin'):
    """Apply explicit admin score ops for locked historical months only.

    This keeps the historical lock guard in place for generic sync payloads while
    allowing controlled Record Score operations (add/edit/delete) to pass through.
    """
    if not isinstance(payload, dict):
        return payload
    if not isinstance(ops, list) or not ops:
        return payload

    locked_months = _locked_month_keys(payload)
    if not locked_months:
        return payload

    # GCB-immune students have a -20 score floor.
    gcb_student_ids = {
        _parse_int_safe(s.get('id'), 0)
        for s in (payload.get('students', []) or [])
        if isinstance(s, dict) and s.get('gcb')
    }

    scores = payload.get('scores', [])
    if not isinstance(scores, list):
        scores = []

    by_key = {}
    max_id = 0
    for row in scores:
        if not isinstance(row, dict):
            continue
        sid = _parse_int_safe(row.get('studentId'), 0)
        date_key = str(row.get('date') or '').strip()[:10]
        month_key = str(row.get('month') or date_key[:7]).strip()
        if sid <= 0 or not re.match(r'^\d{4}-\d{2}-\d{2}$', date_key) or not re.match(r'^\d{4}-\d{2}$', month_key):
            continue
        key = (sid, date_key, month_key)
        by_key[key] = dict(row)
        max_id = max(max_id, _parse_int_safe(row.get('id'), 0))

    now_iso = _server_now_iso()
    for raw in ops[:200]:
        if not isinstance(raw, dict):
            continue
        op_type = str(raw.get('type') or 'upsert').strip().lower()
        sid = _parse_int_safe(raw.get('studentId'), 0)
        date_key = str(raw.get('date') or '').strip()[:10]
        month_key = str(raw.get('month') or date_key[:7]).strip()
        if sid <= 0 or not re.match(r'^\d{4}-\d{2}-\d{2}$', date_key) or not re.match(r'^\d{4}-\d{2}$', month_key):
            continue
        if month_key not in locked_months:
            continue

        key = (sid, date_key, month_key)
        if op_type == 'delete':
            by_key.pop(key, None)
            continue

        existing = by_key.get(key, {})
        row = dict(existing)
        if _parse_int_safe(row.get('id'), 0) <= 0:
            max_id += 1
            row['id'] = max_id
            row['created_at'] = now_iso

        row['studentId'] = sid
        row['date'] = date_key
        row['month'] = month_key
        raw_pts = _parse_int_safe(raw.get('points'), 0)
        row['points'] = max(-20, raw_pts) if sid in gcb_student_ids else raw_pts
        row['stars'] = _parse_int_safe(raw.get('stars'), 0)
        row['vetos'] = _parse_int_safe(raw.get('vetos'), 0)
        row['notes'] = str(raw.get('notes') or '').strip()
        row['recordedBy'] = str(raw.get('recordedBy') or actor_login_id or row.get('recordedBy') or 'Admin').strip()
        row['star_usage_normal'] = max(0, _parse_int_safe(raw.get('star_usage_normal'), _parse_int_safe(row.get('star_usage_normal'), 0)))
        row['star_usage_disciplinary'] = max(0, _parse_int_safe(raw.get('star_usage_disciplinary'), _parse_int_safe(row.get('star_usage_disciplinary'), 0)))
        row['updated_at'] = now_iso

        history_entry = raw.get('history_entry')
        if isinstance(history_entry, dict):
            row_history = row.get('history')
            if not isinstance(row_history, list):
                row_history = []
            row_history.append(dict(history_entry))
            row['history'] = row_history

        by_key[key] = row

    payload['scores'] = list(by_key.values())
    return payload


def _merge_students_preserve_active(existing_students, incoming_students):
    """Merge student lists, never downgrading active:True→False without a genuinely newer timestamp.
    This protects against sync-induced corruption where a peer device pushes stale active:false flags."""
    if not isinstance(existing_students, list):
        existing_students = []
    if not isinstance(incoming_students, list):
        incoming_students = []
    merged = {}
    id_index = {}
    roll_index = {}
    max_id = 0

    def _class_stamp(student_row):
        if not isinstance(student_row, dict):
            return 0.0
        class_stamp = _parse_sync_stamp(student_row.get('class_updated_at') or '')
        if class_stamp > 0:
            return class_stamp
        return _parse_sync_stamp(student_row.get('updated_at') or student_row.get('created_at') or '')

    def _derive_group_from_roll(roll_value):
        match = re.search(r'^EA\d{2}([A-Z])', _normalize_roll_value(roll_value))
        return match.group(1) if match else ''

    def _normalize_student_record(student):
        normalized = dict(student or {})
        normalized['roll'] = _normalize_roll_value(normalized.get('roll'))
        normalized['base_name'] = str(
            normalized.get('base_name') or normalized.get('name') or normalized.get('raw_name') or ''
        ).strip()
        normalized['name'] = str(
            normalized.get('name') or normalized.get('base_name') or normalized.get('raw_name') or normalized.get('roll') or ''
        ).strip()
        normalized['raw_name'] = str(
            normalized.get('raw_name') or normalized.get('name') or normalized.get('base_name') or ''
        ).strip()
        class_raw = normalized.get('class')
        try:
            class_val = int(class_raw)
        except (TypeError, ValueError):
            class_val = None
        normalized['class'] = class_val
        if not normalized.get('group'):
            normalized['group'] = _derive_group_from_roll(normalized['roll'])
        if normalized.get('active') is None:
            normalized['active'] = True
        return normalized

    def _merge_pair(existing, incoming):
        existing_stamp = _parse_sync_stamp(existing.get('updated_at') or existing.get('created_at') or '')
        incoming_stamp = _parse_sync_stamp(incoming.get('updated_at') or incoming.get('created_at') or '')
        if incoming_stamp > existing_stamp:
            merged_s = {**existing, **incoming}
        elif incoming_stamp < existing_stamp:
            merged_s = {**incoming, **existing}
        else:
            merged_s = {**existing, **incoming}
            if existing.get('active') is not False and merged_s.get('active') is False:
                merged_s['active'] = existing.get('active', True)

        stable_id = _parse_int_safe(existing.get('id'), 0) or _parse_int_safe(incoming.get('id'), 0)
        if stable_id > 0:
            merged_s['id'] = stable_id

        if _class_stamp(existing) >= _class_stamp(incoming) and 'class' in existing:
            merged_s['class'] = existing.get('class')
            if existing.get('class_updated_at'):
                merged_s['class_updated_at'] = existing.get('class_updated_at')

        # Preserve name from the record with the newer name timestamp.
        def _name_stamp(row):
            return _parse_sync_stamp(row.get('name_updated_at') or row.get('updated_at') or row.get('created_at') or '')
        if _name_stamp(existing) > _name_stamp(incoming) and existing.get('base_name'):
            merged_s['base_name'] = existing.get('base_name')
            merged_s['name'] = existing.get('base_name')
            merged_s['raw_name'] = existing.get('raw_name') or existing.get('base_name')
            if existing.get('name_updated_at'):
                merged_s['name_updated_at'] = existing.get('name_updated_at')
        elif _name_stamp(incoming) > _name_stamp(existing) and incoming.get('base_name'):
            merged_s['base_name'] = incoming.get('base_name')
            merged_s['name'] = incoming.get('base_name')
            merged_s['raw_name'] = incoming.get('raw_name') or incoming.get('base_name')
            if incoming.get('name_updated_at'):
                merged_s['name_updated_at'] = incoming.get('name_updated_at')

        if existing.get('active') is not False and merged_s.get('active') is False:
            merged_s['active'] = existing.get('active', True)

        # Preserve stars from the record with the newer star-specific timestamp.
        # student.stars is the authoritative current-month balance; a stale sync
        # snapshot with a bumped updated_at (but no actual star mutation) must NOT
        # overwrite a genuinely newer star value. Falls back to updated_at when
        # stars_updated_at is absent (legacy records). Defensive fix for the
        # "should be 11, shows 8" class of stale-sync star regressions.
        def _star_stamp(row):
            return _parse_sync_stamp(row.get('stars_updated_at') or row.get('updated_at') or row.get('created_at') or '')
        if _star_stamp(existing) > _star_stamp(incoming) and existing.get('stars') is not None:
            merged_s['stars'] = max(0, _parse_int_safe(existing.get('stars')))
            if existing.get('stars_updated_at'):
                merged_s['stars_updated_at'] = existing.get('stars_updated_at')
        elif _star_stamp(incoming) > _star_stamp(existing) and incoming.get('stars') is not None:
            merged_s['stars'] = max(0, _parse_int_safe(incoming.get('stars')))
            if incoming.get('stars_updated_at'):
                merged_s['stars_updated_at'] = incoming.get('stars_updated_at')

        # Preserve structural visibility fields: never drop active_from_month or
        # deactivation_month once set on either side.  These control whether a
        # student appears in historical months and roll-change visibility checks.
        # EXCEPTION: an incoming record that carries the key with an explicit
        # null/empty value is a DELIBERATE clear (the reactivation path sets
        # deactivation_month=null). Honor the clear; only an ABSENT key marks a
        # stale/legacy/imported record that never knew about the field.
        for _vis_field in ('active_from_month', 'deactivation_month'):
            _existing_val = str(existing.get(_vis_field) or '').strip()
            _merged_val = str(merged_s.get(_vis_field) or '').strip()
            _explicit_clear = _vis_field in incoming and str(incoming.get(_vis_field) or '').strip() == ''
            if _existing_val and not _merged_val and not _explicit_clear:
                merged_s[_vis_field] = _existing_val

        roll_value = _normalize_roll_value(merged_s.get('roll'))
        if roll_value:
            merged_s['roll'] = roll_value
            group_value = _derive_group_from_roll(roll_value)
            if group_value:
                merged_s['group'] = group_value

        return merged_s

    def _remember_student(student):
        nonlocal max_id
        if not isinstance(student, dict):
            return
        sid = _parse_int_safe(student.get('id'), 0)
        roll = _normalize_roll_value(student.get('roll'))
        if sid > 0 and sid > max_id:
            max_id = sid
        if sid > 0:
            id_index[sid] = None  # populated by caller
        if roll:
            roll_index[roll] = None  # populated by caller

    def _store_student(key, student):
        merged[key] = student
        sid = _parse_int_safe(student.get('id'), 0)
        roll = _normalize_roll_value(student.get('roll'))
        if sid > 0:
            id_index[sid] = key
            if sid > max_id:
                nonlocal_max_id[0] = sid
        if roll:
            roll_index[roll] = key

    nonlocal_max_id = [0]
    for student in existing_students:
        if not isinstance(student, dict):
            continue
        _remember_student(student)
        sid = _parse_int_safe(student.get('id'), 0)
        roll = _normalize_roll_value(student.get('roll'))
        name = str(student.get('base_name') or student.get('name') or student.get('raw_name') or '').strip()
        key = ''
        if sid > 0:
            key = f'id:{sid}'
        elif roll:
            key = f'roll:{roll}'
        elif name:
            key = f'name:{name}'
        if not key:
            continue
        _store_student(key, dict(student))

    max_id = max(max_id, nonlocal_max_id[0])

    for student in incoming_students:
        if not isinstance(student, dict):
            continue
        normalized = _normalize_student_record(student)
        sid = _parse_int_safe(normalized.get('id'), 0)
        roll = _normalize_roll_value(normalized.get('roll'))
        name = _name_key(normalized.get('base_name') or normalized.get('name') or normalized.get('raw_name') or '')

        target_key = None
        if sid > 0 and sid in id_index and id_index[sid]:
            target_key = id_index[sid]
        elif roll and roll in roll_index and roll_index[roll]:
            target_key = roll_index[roll]
        elif sid > 0:
            target_key = f'id:{sid}'
        elif roll:
            target_key = f'roll:{roll}'
        elif name:
            target_key = f'name:{name}'
        if not target_key:
            continue

        existing = merged.get(target_key)
        if existing:
            merged_student = _merge_pair(existing, normalized)
        else:
            merged_student = normalized

        if sid > 0 and _parse_int_safe(merged_student.get('id'), 0) <= 0:
            merged_student['id'] = sid
        elif _parse_int_safe(merged_student.get('id'), 0) <= 0 and target_key.startswith('id:'):
            merged_student['id'] = _parse_int_safe(target_key.split(':', 1)[1], 0)

        sid_final = _parse_int_safe(merged_student.get('id'), 0)
        roll_final = _normalize_roll_value(merged_student.get('roll'))
        if sid_final > 0 and sid_final > max_id:
            max_id = sid_final
        if sid > 0:
            id_index[sid] = target_key
        if sid_final > 0:
            id_index[sid_final] = target_key
        if roll_final:
            roll_index[roll_final] = target_key

        merged[target_key] = merged_student

    result = list(merged.values())
    used_ids = set()
    deduped = []
    for student in result:
        if not isinstance(student, dict):
            continue
        sid = _parse_int_safe(student.get('id'), 0)
        if sid <= 0 or sid in used_ids:
            max_id += 1
            student = {**student, 'id': max_id}
            sid = max_id
        used_ids.add(sid)
        deduped.append(_normalize_student_record(student))
    return deduped


def _merge_scores_superset(existing_scores, incoming_scores):
    """Merge score lists without dropping existing rows when an incoming snapshot is stale."""
    merged = {}
    max_score_id = 0

    def _normalize_score(score):
        if not isinstance(score, dict):
            return None, None
        sid = score.get('studentId')
        date_key = str(score.get('date') or '').strip()
        if sid is None or not date_key:
            return None, None
        month_key = str(score.get('month') or '').strip() or date_key[:7]
        normalized = dict(score)
        normalized['studentId'] = sid
        normalized['date'] = date_key
        normalized['month'] = month_key
        normalized['points'] = round(_parse_float_safe(score.get('points')), 2)
        normalized['stars'] = _parse_int_safe(score.get('stars'))
        normalized['vetos'] = _parse_int_safe(score.get('vetos'))
        key = (str(sid), date_key, month_key)
        return key, normalized

    for score in existing_scores or []:
        key, normalized = _normalize_score(score)
        if key is None:
            continue
        merged[key] = normalized
        max_score_id = max(max_score_id, _parse_int_safe(normalized.get('id')))

    for score in incoming_scores or []:
        key, normalized = _normalize_score(score)
        if key is None:
            continue
        if key in merged:
            prev = merged[key]
            prev_points = _parse_float_safe(prev.get('points'))
            prev_stars = _parse_int_safe(prev.get('stars'))
            prev_vetos = _parse_int_safe(prev.get('vetos'))
            
            # History-aware merge: updated_at can be bumped by sync operations,
            # but history entries record actual user edits. If prev has a more
            # recent history entry than incoming, prev is more authoritative.
            prev_hist_stamp = _get_last_history_stamp(prev)
            next_hist_stamp = _get_last_history_stamp(normalized)
            
            # Perform standard timestamp and ID tiebreaker updates
            prev_stamp = _parse_sync_stamp(prev.get('updated_at', ''))
            next_stamp = _parse_sync_stamp(normalized.get('updated_at', ''))
            if next_stamp > prev_stamp:
                if prev_hist_stamp > next_hist_stamp:
                    pass  # Local has a more recent user edit — keep prev
                else:
                    merged[key] = normalized
            elif next_stamp == prev_stamp and _parse_int_safe(normalized.get('id')) > _parse_int_safe(prev.get('id')):
                # Same age (or both missing updated_at): keep the higher-id record as tiebreaker.
                merged[key] = normalized
            # else prev is same age or newer — keep it
        else:
            merged[key] = normalized
        max_score_id = max(max_score_id, _parse_int_safe(normalized.get('id')))

    result = []
    for record in merged.values():
        if not _parse_int_safe(record.get('id')):
            max_score_id += 1
            record['id'] = max_score_id
        result.append(record)
    return result


def _merge_notification_history(existing_history, incoming_history):
    """Keep all notification history entries across sync peers."""
    merged = {}

    def _entry_key(item):
        if not isinstance(item, dict):
            return ''
        fp = str(item.get('fingerprint') or '').strip().lower()
        if fp:
            return fp
        title = str(item.get('title') or '').strip().lower()
        detail = str(item.get('detail') or '').strip().lower()
        meta = str(item.get('meta') or '').strip().lower()
        return f'{title}||{detail}||{meta}'

    for source in (existing_history or []), (incoming_history or []):
        for item in source:
            if not isinstance(item, dict):
                continue
            key = _entry_key(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = dict(item)
                continue
            prev_stamp = _parse_sync_stamp(prev.get('logged_at'))
            next_stamp = _parse_sync_stamp(item.get('logged_at'))
            if next_stamp >= prev_stamp:
                merged[key] = dict(item)

    return list(merged.values())


def _merge_records_superset(existing_rows, incoming_rows, key_fields, ts_fields=('updated_at', 'created_at')):
    merged = {}

    def _key(item):
        if not isinstance(item, dict):
            return ''
        parts = []
        for field in key_fields:
            parts.append(str(item.get(field) or '').strip().lower())
        return '::'.join(parts)

    def _ts(item):
        if not isinstance(item, dict):
            return 0.0
        stamps = [_parse_sync_stamp(item.get(field)) for field in ts_fields]
        return max(stamps) if stamps else 0.0

    for source in (existing_rows or []), (incoming_rows or []):
        for row in source:
            if not isinstance(row, dict):
                continue
            key = _key(row)
            if not key:
                continue
            prev = merged.get(key)
            if not prev or _ts(row) >= _ts(prev):
                merged[key] = dict(row)

    return list(merged.values())


def _active_leadership_role_for_login(data, login_id):
    roll = _normalize_roll_value(login_id)
    if not roll.startswith('EA'):
        return ''
    check_date = _parse_date_key(_server_now_iso()) or date.today()
    for post in data.get('leadership', []) or []:
        if _normalize_holder_status(post.get('status') or 'active') != 'active':
            continue
        tenure_months = _tenure_months_for_assignment('leadership', post.get('post'))
        extension_months = _parse_int_safe(post.get('tenure_extension_months'), 0)
        if not _is_assignment_active_by_tenure(post.get('elected_on'), tenure_months, extension_months, check_date):
            continue
        post_roll = _normalize_roll_value(post.get('roll'))
        if post_roll and post_roll == roll:
            role_type = _leadership_role_type(post.get('post'))
            if role_type in ('leader', 'co_leader'):
                return role_type
    return ''


def _get_score_row_for_student_date(data, student_id, date_key):
    best = None
    best_stamp = 0.0
    for row in data.get('scores', []) or []:
        if not isinstance(row, dict):
            continue
        sid = _parse_int_safe(row.get('studentId'), 0)
        d = str(row.get('date') or '').strip()
        if sid != student_id or d != date_key:
            continue
        stamp = max(
            _parse_sync_stamp(row.get('updated_at')),
            _parse_sync_stamp(row.get('created_at')),
        )
        if not best or stamp >= best_stamp:
            best = row
            best_stamp = stamp
    return best


def _student_id_by_login(data, login_id):
    roll = _normalize_roll_value(login_id)
    if not roll.startswith('EA'):
        return 0
    for student in data.get('students', []) or []:
        if _normalize_roll_value(student.get('roll')) == roll:
            return _parse_int_safe(student.get('id'), 0)
    return 0


def _is_active_assignment(item, source='leadership', check_date=None):
    if not isinstance(item, dict):
        return False
    if _normalize_holder_status(item.get('status') or 'active') != 'active':
        return False
    post_name = item.get('post') or ''
    tenure_months = _tenure_months_for_assignment(source, post_name)
    extension_months = _parse_int_safe(item.get('tenure_extension_months'), 0)
    return _is_assignment_active_by_tenure(item.get('elected_on'), tenure_months, extension_months, check_date or date.today())


def _is_student_council_member(data, student_id):
    sid = _parse_int_safe(student_id, 0)
    if sid <= 0:
        return False
    check_date = _parse_date_key(_server_now_iso()) or date.today()
    for post in data.get('leadership', []) or []:
        if _parse_int_safe(post.get('studentId'), 0) == sid and _is_active_assignment(post, 'leadership', check_date):
            return True
    for post in data.get('class_reps', []) or []:
        if _parse_int_safe(post.get('studentId'), 0) == sid and _is_active_assignment(post, 'class_rep', check_date):
            return True
    for post in data.get('group_crs', []) or []:
        if _parse_int_safe(post.get('studentId'), 0) == sid and _is_active_assignment(post, 'group_cr', check_date):
            return True
    for party in data.get('parties', []) or []:
        for member in party.get('members', []) or []:
            if _parse_int_safe(member.get('studentId'), 0) != sid:
                continue
            status = str(member.get('status') or 'active').strip().lower()
            if status not in ('active', ''):
                continue
            return True
    return False


def _is_proposal_stakeholder(data, proposal, user):
    role = str(getattr(user, 'role', '') or '').strip().lower()
    if role in ('admin', 'teacher'):
        return True
    if role != 'student':
        return False
    scope = str((proposal or {}).get('scope') or 'student_council').strip().lower()
    if scope == 'all_students':
        return True
    sid = _student_id_by_login(data, getattr(user, 'login_id', ''))
    return _is_student_council_member(data, sid)


def _month_key_from_date_like(value):
    text = str(value or '').strip()
    if re.match(r'^\d{4}-\d{2}$', text):
        return text
    if re.match(r'^\d{4}-\d{2}-\d{2}$', text):
        return text[:7]
    if len(text) >= 7 and re.match(r'^\d{4}-\d{2}', text[:7]):
        return text[:7]
    return ''


def _iter_month_keys_between(start_key, end_key):
    start = _month_key_from_date_like(start_key)
    end = _month_key_from_date_like(end_key)
    if not start or not end:
        return []
    try:
        sy, sm = [int(x) for x in start.split('-')]
        ey, em = [int(x) for x in end.split('-')]
    except Exception:
        return []
    s = date(sy, sm, 1)
    e = date(ey, em, 1)
    if s > e:
        s, e = e, s
    out = []
    cur = s
    while cur <= e:
        out.append(f"{cur.year:04d}-{cur.month:02d}")
        year = cur.year + (1 if cur.month == 12 else 0)
        month = 1 if cur.month == 12 else cur.month + 1
        cur = date(year, month, 1)
    return out


def _student_allowed_months_from_roster(data, login_id):
    roll = _normalize_roll_value(login_id)
    if not roll.startswith('EA'):
        return set()
    allowed = set()
    month_students = data.get('month_students', {}) if isinstance(data, dict) else {}
    if isinstance(month_students, dict):
        for month, rows in month_students.items():
            mk = _month_key_from_date_like(month)
            if not mk:
                continue
            for value in rows or []:
                if _normalize_roll_value(value) == roll:
                    allowed.add(mk)
                    break
    month_profiles = data.get('month_roster_profiles', {}) if isinstance(data, dict) else {}
    if isinstance(month_profiles, dict):
        for month, rows in month_profiles.items():
            mk = _month_key_from_date_like(month)
            if not mk:
                continue
            for profile in rows or []:
                if not isinstance(profile, dict):
                    continue
                if _normalize_roll_value(profile.get('roll')) == roll:
                    allowed.add(mk)
                    break
    return allowed


def _teacher_allowed_months_from_windows(user_id):
    allowed = set()
    try:
        rows = (
            UserAccessWindow.query
            .filter_by(user_id=_parse_int_safe(user_id, 0))
            .order_by(UserAccessWindow.updated_at.desc(), UserAccessWindow.id.desc())
            .all()
        )
    except Exception:
        db.session.rollback()
        rows = []
    for row in rows:
        for mk in _iter_month_keys_between(row.month_from, row.month_to):
            allowed.add(mk)
    return allowed


def _allowed_months_for_user(data, user):
    role = str(getattr(user, 'role', '') or '').strip().lower()
    if role == 'admin':
        return None  # unrestricted
    if role == 'teacher':
        months = _teacher_allowed_months_from_windows(getattr(user, 'id', 0))
        # Ensure teachers can always at least see the current active month
        # so they are not locked out when a new calendar month begins.
        if months:
            months.add(_server_now_iso()[:7])
            return months
        # Backward compatibility: until admin configures windows, keep current behaviour.
        return None
    if role == 'student':
        months = _student_allowed_months_from_roster(data, getattr(user, 'login_id', ''))
        return months or {(_server_now_iso()[:7])}
    return {(_server_now_iso()[:7])}


def _is_month_allowed_for_user(data, user, month_key):
    allowed = _allowed_months_for_user(data, user)
    if allowed is None:
        return True
    mk = _month_key_from_date_like(month_key)
    if not mk:
        return False
    return mk in allowed


def _clip_payload_to_allowed_months(payload, allowed_months):
    if allowed_months is None or not isinstance(payload, dict):
        return payload

    allowed = {m for m in (allowed_months or set()) if _month_key_from_date_like(m)}
    data = dict(payload)

    def _in_allowed(mk):
        return bool(mk and mk in allowed)

    data['scores'] = [
        row for row in (payload.get('scores') or [])
        if isinstance(row, dict) and _in_allowed(_month_key_from_date_like(row.get('month') or row.get('date')))
    ]
    data['attendance'] = [
        row for row in (payload.get('attendance') or [])
        if isinstance(row, dict) and _in_allowed(_month_key_from_date_like(row.get('month') or row.get('date')))
    ]
    data['appeals'] = [
        row for row in (payload.get('appeals') or [])
        if isinstance(row, dict) and _in_allowed(
            _month_key_from_date_like(row.get('score_month') or row.get('score_date') or row.get('created_at'))
        )
    ]
    data['resource_requests'] = [
        row for row in (payload.get('resource_requests') or [])
        if isinstance(row, dict) and _in_allowed(
            _month_key_from_date_like(row.get('month') or row.get('request_date') or row.get('created_at'))
        )
    ]
    data['resource_transactions'] = [
        row for row in (payload.get('resource_transactions') or [])
        if isinstance(row, dict) and _in_allowed(
            _month_key_from_date_like(row.get('month') or row.get('date') or row.get('created_at'))
        )
    ]
    # fee_records are per-student (not per-month), pass through unfiltered.
    data['fee_records'] = list(payload.get('fee_records') or [])

    month_students = payload.get('month_students', {}) if isinstance(payload.get('month_students'), dict) else {}
    month_profiles = payload.get('month_roster_profiles', {}) if isinstance(payload.get('month_roster_profiles'), dict) else {}
    data['month_students'] = {
        mk: rows for mk, rows in month_students.items()
        if _month_key_from_date_like(mk) in allowed
    }
    data['month_roster_profiles'] = {
        mk: rows for mk, rows in month_profiles.items()
        if _month_key_from_date_like(mk) in allowed
    }
    data['allowed_months'] = sorted(list(allowed))
    return data


def _sanitize_anonymous_snapshot(payload):
    """
    Public/display-safe view of the ledger for UNAUTHENTICATED GET /offline-data.
    Keeps the recent scoreboard months renderable (students, scores, parties,
    leadership) but strips private data: fees, appeals, resource money trails,
    activity log, notifications, proposals and per-student personal profile
    fields. The payload is marked with sync_scope='anonymous-public' so peers
    and backup bootstraps refuse to persist it as a full snapshot.
    """
    months = set(_recent_public_month_window())
    data = _clip_payload_to_allowed_months(payload, months)
    data['fee_records'] = []
    data['appeals'] = []
    data['resource_requests'] = []
    data['resource_transactions'] = []
    data['resource_advantage_deductions'] = []
    data.pop('activity_log', None)
    data['notification_history'] = []
    data['proposals'] = []
    data['proposal_votes'] = []
    data['proposal_messages'] = []
    data['score_adjustment_actions'] = []
    data['_sync_ops'] = []
    safe_students = []
    for s in (data.get('students') or []):
        if not isinstance(s, dict):
            continue
        s = dict(s)
        s.pop('profile_data', None)
        s.pop('remarks', None)
        safe_students.append(s)
    data['students'] = safe_students
    data['sync_scope'] = 'anonymous-public'
    return data


def _sum_points_for_student_month(snapshot, student_id, month_key):
    sid = _parse_int_safe(student_id, 0)
    if sid <= 0:
        return 0
    month = _month_key_from_date_like(month_key)
    total = 0
    for row in snapshot.get('scores', []) or []:
        if not isinstance(row, dict):
            continue
        if _parse_int_safe(row.get('studentId'), 0) != sid:
            continue
        mk = _month_key_from_date_like(row.get('month') or row.get('date'))
        if mk != month:
            continue
        total += _parse_int_safe(row.get('points'), 0)
    return total


def _upsert_score_delta(snapshot, student_id, date_key, month_key, delta_points=0, delta_stars=0, note=''):
    sid = _parse_int_safe(student_id, 0)
    if sid <= 0:
        return None
    date_text = str(date_key or '').strip()
    month_text = _month_key_from_date_like(month_key) or _month_key_from_date_like(date_text)
    if not date_text or not month_text:
        return None

    scores = snapshot.get('scores', [])
    if not isinstance(scores, list):
        scores = []
        snapshot['scores'] = scores

    target = None
    best_stamp = 0.0
    max_id = 0
    for row in scores:
        if not isinstance(row, dict):
            continue
        max_id = max(max_id, _parse_int_safe(row.get('id'), 0))
        if _parse_int_safe(row.get('studentId'), 0) != sid:
            continue
        if str(row.get('date') or '').strip() != date_text:
            continue
        if _month_key_from_date_like(row.get('month') or row.get('date')) != month_text:
            continue
        stamp = max(_parse_sync_stamp(row.get('updated_at')), _parse_sync_stamp(row.get('created_at')))
        if not target or stamp >= best_stamp:
            target = row
            best_stamp = stamp

    now_iso = _server_now_iso()
    if target is None:
        target = {
            'id': max_id + 1,
            'studentId': sid,
            'date': date_text,
            'month': month_text,
            'points': 0,
            'stars': 0,
            'vetos': 0,
            'notes': '',
            'recordedBy': 'transfer',
            'created_at': now_iso,
        }
        scores.append(target)

    target['points'] = round(_parse_float_safe(target.get('points'), 0) + _parse_float_safe(delta_points, 0), 2)
    target['stars'] = _parse_int_safe(target.get('stars'), 0) + _parse_int_safe(delta_stars, 0)
    existing_note = str(target.get('notes') or '').strip()
    if note:
        target['notes'] = f"{existing_note} | {note}" if existing_note else note
    target['updated_at'] = now_iso
    # Propagate star-specific timestamp to the student ledger so the merge
    # strategy can distinguish a genuine star mutation from a sync-bumped
    # updated_at on a stale snapshot. Only touch the student record when stars
    # actually changed — avoids false "newer star" signals on points-only deltas.
    if _parse_int_safe(delta_stars, 0) != 0:
        students = snapshot.get('students', []) or []
        for s in students:
            if isinstance(s, dict) and _parse_int_safe(s.get('id'), 0) == sid:
                s['stars_updated_at'] = now_iso
                break
    return target


def _calculate_election_results(election_data, student_votes, teacher_votes):
    """Calculates the winner of an election.

    Uses the same weighted-vote formula as the frontend:
    - Student votes are weighted by each voter's votePower (stored on the vote record).
    - Teacher vote power = (highest_student_vote_power + bonus) / 2,
      where bonus comes from app_settings.teacher_vote_bonus (default 5).
    - A winner must pass three thresholds: average weight, clear lead, and majority weight.

    Args:
        election_data (dict): The full offline data dict (must contain 'students'
            and optionally 'app_settings').
        student_votes (list): A list of student vote records for a single post.
        teacher_votes (list): A list of teacher vote records for a single post.

    Returns:
        dict: A dictionary containing the winner, tie flag, candidate vote totals,
              and threshold metadata.  Returns a null result if there is a tie or
              thresholds are not met.
    """
    # --- Read configurable bonus from app_settings (matches frontend default) ---
    app_settings = election_data.get('app_settings') or {}
    bonus = 0
    try:
        bonus = int(app_settings.get('teacher_vote_bonus', 5))
    except (TypeError, ValueError):
        bonus = 5
    bonus = max(0, bonus)

    # --- Compute highest student vote power ---
    # The frontend computes vote power from total scores via computeVotePower.
    # Vote records store votePower at cast time, so we use the stored value.
    highest_student_vote_power = 0
    for vote in student_votes:
        vp = _parse_float_safe(vote.get('votePower'), 0)
        if vp > highest_student_vote_power:
            highest_student_vote_power = vp

    # Also check student objects as a fallback (for cases where votes don't have votePower)
    for student in election_data.get('students', []):
        vp = _parse_float_safe(student.get('vote_power'), 0)
        if vp > highest_student_vote_power:
            highest_student_vote_power = vp

    # --- Teacher vote power: (highest + bonus) / 2 (matches frontend getTeacherVotePower) ---
    teacher_vote_power = (highest_student_vote_power + bonus) / 2.0

    # --- Calculate weighted votes for each candidate ---
    candidate_votes = {}  # candidate_id -> total composite weight
    for vote in student_votes:
        candidate_id = vote.get('candidateId')
        if not candidate_id:
            continue
        # Use stored votePower, fall back to 1 if missing
        weight = _parse_float_safe(vote.get('votePower'), 1)
        if weight <= 0:
            weight = 1
        candidate_votes[candidate_id] = candidate_votes.get(candidate_id, 0) + weight

    for vote in teacher_votes:
        candidate_id = vote.get('candidateId')
        if not candidate_id:
            continue
        candidate_votes[candidate_id] = candidate_votes.get(candidate_id, 0) + teacher_vote_power

    if not candidate_votes:
        return {'winner': None, 'tie': False, 'candidate_votes': {}, 'nullified': True,
                'runoff': False, 'reason': 'No candidate votes'}

    # --- Sort candidates by composite weight (descending) ---
    sorted_candidates = sorted(candidate_votes.items(), key=lambda x: x[1], reverse=True)
    winner_id, top_weight = sorted_candidates[0]
    second_weight = sorted_candidates[1][1] if len(sorted_candidates) > 1 else 0

    # --- Check for tie (no clear lead) ---
    eps = 0.0001
    tie = len(sorted_candidates) > 1 and abs(top_weight - second_weight) < eps

    if tie:
        return {'winner': None, 'tie': True, 'candidate_votes': candidate_votes,
                'nullified': False, 'runoff': False, 'reason': 'Tie — no clear lead'}

    # --- Threshold checks (matches frontend _concludeStudentAndReveal logic) ---
    # Count eligible student voters for threshold computation
    eligible_voter_ids = set()
    for vote in student_votes:
        voter_id = vote.get('voterStudentId')
        if voter_id:
            eligible_voter_ids.add(voter_id)

    # Total eligible weight = sum of votePower across all eligible voters who cast a vote
    # (frontend uses all eligible voters, but backend only has vote records)
    total_eligible_weight = 0
    seen_voters = set()
    for vote in student_votes:
        voter_id = vote.get('voterStudentId')
        if voter_id and voter_id not in seen_voters:
            seen_voters.add(voter_id)
            total_eligible_weight += _parse_float_safe(vote.get('votePower'), 1)

    # Add teacher eligible weight
    teacher_ids = set()
    for vote in teacher_votes:
        tid = vote.get('teacherId')
        if tid:
            teacher_ids.add(tid)
    total_teacher_weight = len(teacher_ids) * teacher_vote_power
    total_eligible_weight += total_teacher_weight

    avg_threshold = total_eligible_weight / 2.0 if (len(seen_voters) + len(teacher_ids)) > 0 else 0
    majority_threshold = total_eligible_weight / 2.0

    passes_avg = top_weight >= avg_threshold
    passes_majority = top_weight >= majority_threshold
    clear_lead = top_weight > second_weight

    if not passes_avg or not passes_majority:
        if clear_lead and passes_avg:
            return {'winner': None, 'tie': False, 'candidate_votes': candidate_votes,
                    'nullified': False, 'runoff': True,
                    'reason': 'Runoff required — clear lead but no majority',
                    'top_weight': top_weight, 'second_weight': second_weight,
                    'avg_threshold': avg_threshold,
                    'majority_threshold': majority_threshold}
        reason = 'Did not meet average threshold' if not passes_avg else 'Combined majority not reached'
        return {'winner': None, 'tie': False, 'candidate_votes': candidate_votes,
                'nullified': True, 'runoff': False, 'reason': reason,
                'top_weight': top_weight, 'avg_threshold': avg_threshold,
                'majority_threshold': majority_threshold}

    return {'winner': winner_id, 'tie': False, 'candidate_votes': candidate_votes,
            'top_weight': top_weight, 'second_weight': second_weight,
            'avg_threshold': avg_threshold, 'majority_threshold': majority_threshold,
            'teacher_vote_power': teacher_vote_power}


def _merge_election_votes_superset(existing_votes, incoming_votes, mode='party'):
    """Merge election votes by voter key and keep latest timestamped entry."""
    merged = {}

    def _normalize_vote(item):
        if not isinstance(item, dict):
            return None, None
        post = str(item.get('post') or '').strip()
        if not post:
            return None, None
        normalized = dict(item)
        normalized['post'] = post
        if mode == 'party':
            party_id = _parse_int_safe(item.get('partyId'), 0)
            if party_id <= 0:
                return None, None
            normalized['partyId'] = party_id
            candidate_id = _parse_int_safe(item.get('candidateId'), 0)
            if candidate_id <= 0:
                return None, None
            normalized['candidateId'] = candidate_id
            key = f'{post}::party::{party_id}'
        elif mode == 'teacher':
            teacher_id = _parse_int_safe(item.get('teacherId'), 0)
            if teacher_id <= 0:
                return None, None
            # Do not allow admin to vote
            if teacher_id == 1:
                return None, None
            normalized['teacherId'] = teacher_id
            candidate_id = _parse_int_safe(item.get('candidateId'), 0)
            if candidate_id <= 0:
                return None, None
            normalized['candidateId'] = candidate_id
            key = f'{post}::teacher::{teacher_id}'
        else:
            voter_id = _parse_int_safe(item.get('voterStudentId'), 0)
            if voter_id <= 0:
                return None, None
            normalized['voterStudentId'] = voter_id
            vote_type = str(item.get('voteType') or 'candidate').strip().lower()
            if vote_type not in ('candidate', 'abstain', 'nota'):
                vote_type = 'candidate'
            normalized['voteType'] = vote_type
            if vote_type == 'candidate':
                candidate_id = _parse_int_safe(item.get('candidateId'), 0)
                if candidate_id <= 0:
                    return None, None
                normalized['candidateId'] = candidate_id
            else:
                normalized['candidateId'] = None
            key = f'{post}::student::{voter_id}'
        return key, normalized

    for source in (existing_votes or []), (incoming_votes or []):
        for item in source:
            # Skip admin votes.
            if item.get('teacherId') == 1:
                continue
            key, normalized = _normalize_vote(item)
            if key is None:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            prev_stamp = _parse_sync_stamp(prev.get('timestamp') or prev.get('updated_at') or prev.get('created_at'))
            next_stamp = _parse_sync_stamp(normalized.get('timestamp') or normalized.get('updated_at') or normalized.get('created_at'))
            if next_stamp >= prev_stamp:
                current_app.logger.warning(
                    'Duplicate vote key overwritten during sync: key=%s mode=%s '
                    'prev_candidate=%s next_candidate=%s prev_stamp=%s next_stamp=%s',
                    key, mode,
                    prev.get('candidateId'), normalized.get('candidateId'),
                    prev_stamp, next_stamp
                )
                merged[key] = normalized

    return list(merged.values())


def _merge_pending_results_superset(existing_results, incoming_results):
    """Merge pending election results by post/source and keep latest record."""
    merged = {}

    def _entry_key(item):
        if not isinstance(item, dict):
            return ''
        post = str(item.get('post') or '').strip().lower()
        source = str(item.get('source') or '').strip().lower()
        if not post or not source:
            return ''
        return f'{post}::{source}'

    def _timestamp(item):
        if not isinstance(item, dict):
            return 0.0
        return max(
            _parse_sync_stamp(item.get('decided_at')),
            _parse_sync_stamp(item.get('updated_at')),
            _parse_sync_stamp(item.get('created_at'))
        )

    for source in (existing_results or []), (incoming_results or []):
        for item in source:
            if not isinstance(item, dict):
                continue
            key = _entry_key(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev or _timestamp(item) >= _timestamp(prev):
                merged[key] = dict(item)

    return list(merged.values())


def _normalize_attendance_status(value):
    status = str(value or '').strip().lower()
    if status in ('absent', 'late', 'leave', 'present'):
        return status
    return 'present'


def _merge_attendance_superset(existing_data, incoming_data):
    """Merge attendance by latest updated_at, keyed by date + roll (fallback studentId)."""
    existing_attendance = existing_data.get('attendance', []) if isinstance(existing_data, dict) else []
    incoming_attendance = incoming_data.get('attendance', []) if isinstance(incoming_data, dict) else []
    existing_students = existing_data.get('students', []) if isinstance(existing_data, dict) else []
    incoming_students = incoming_data.get('students', []) if isinstance(incoming_data, dict) else []

    def _normalize_att_roll(value):
        # Attendance payloads from older clients may contain formatted rolls
        # (spaces or separators). Normalize aggressively for stable identity mapping.
        return re.sub(r'[^A-Z0-9]', '', _normalize_roll_value(value))

    existing_id_by_roll = {}
    for student in existing_students or []:
        sid = _parse_int_safe(student.get('id'), 0)
        roll = _normalize_att_roll(student.get('roll'))
        if sid > 0 and roll and roll not in existing_id_by_roll:
            existing_id_by_roll[roll] = sid

    incoming_roll_by_id = {}
    for student in incoming_students or []:
        sid = _parse_int_safe(student.get('id'), 0)
        roll = _normalize_att_roll(student.get('roll'))
        if sid > 0 and roll:
            incoming_roll_by_id[str(sid)] = roll

    merged = {}

    def _normalize_item(item):
        if not isinstance(item, dict):
            return None, None
        date_key = str(item.get('date') or '').strip()
        if not date_key:
            return None, None
        roll = _normalize_att_roll(item.get('roll'))
        sid = _parse_int_safe(item.get('studentId'), 0)
        if not roll and sid > 0:
            roll = incoming_roll_by_id.get(str(sid), '')
        if roll and roll in existing_id_by_roll:
            sid = existing_id_by_roll[roll]
        identity = roll or (str(sid) if sid > 0 else '')
        if not identity:
            return None, None
        normalized = dict(item)
        normalized['date'] = date_key
        normalized['status'] = _normalize_attendance_status(item.get('status'))
        normalized['remarks'] = str(item.get('remarks') or '')
        if roll:
            normalized['roll'] = roll
        if sid > 0:
            normalized['studentId'] = sid
        key = f'{date_key}::{identity}'
        return key, normalized

    for source in (existing_attendance or []), (incoming_attendance or []):
        for item in source:
            key, normalized = _normalize_item(item)
            if key is None:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
            next_stamp = _parse_sync_stamp(normalized.get('updated_at') or normalized.get('created_at'))
            if next_stamp >= prev_stamp:
                merged[key] = normalized

    # SAFEGUARD: Ensure all merged attendance records have timestamps for proper sync ordering
    now_iso = _server_now_iso()
    for item in merged.values():
        if isinstance(item, dict):
            if not item.get('updated_at'):
                item['updated_at'] = item.get('created_at', now_iso)
            if not item.get('created_at'):
                item['created_at'] = now_iso

    return list(merged.values())


def _merge_fee_records_superset(existing_records, incoming_records):
    """
    Merge fee records by studentId.

    Safety: never lose evidence of payments (last_paid_date/payment_history) even if a stale device
    has a newer updated_at due to clock skew.
    """
    merged = {}

    def _normalize_record(item):
        if not isinstance(item, dict):
            return None, None
        sid = _parse_int_safe(item.get('studentId'), 0)
        if sid <= 0:
            return None, None
        normalized = dict(item)
        normalized['studentId'] = sid
        return str(sid), normalized

    def _max_date(a, b):
        a = str(a or '').strip()
        b = str(b or '').strip()
        if not a:
            return b
        if not b:
            return a
        # YYYY-MM-DD lexicographic compare works.
        return a if a >= b else b

    def _merge_payment_history(prev_list, next_list):
        merged_list = []
        seen = set()
        for src in (prev_list or []), (next_list or []):
            if not isinstance(src, list):
                continue
            for item in src:
                if not isinstance(item, dict):
                    continue
                date_key = str(item.get('date') or item.get('paid_on') or item.get('paidAt') or '').strip()
                amount_key = str(item.get('amount') or '').strip()
                note_key = str(item.get('note') or item.get('remarks') or '').strip().lower()
                fp = f'{date_key}::{amount_key}::{note_key}'
                if fp in seen:
                    continue
                seen.add(fp)
                merged_list.append(dict(item))
        # Keep stable ordering by (date, amount) when possible.
        def _sort_key(x):
            d = str(x.get('date') or x.get('paid_on') or x.get('paidAt') or '')
            a = _parse_int_safe(x.get('amount'), 0)
            return (d, a)
        try:
            merged_list.sort(key=_sort_key)
        except Exception:
            pass
        return merged_list

    def _parse_float(value):
        try:
            v = float(value)
            return v if v == v else None
        except Exception:
            return None

    def _choose_text(prev_val, next_val, prefer_next=False):
        prev_text = str(prev_val or '').strip()
        next_text = str(next_val or '').strip()
        if not prev_text:
            return next_text
        if not next_text:
            return prev_text
        return next_text if prefer_next else prev_text

    def _merge_pair(prev, nxt):
        prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
        next_stamp = _parse_sync_stamp(nxt.get('updated_at') or nxt.get('created_at'))
        prefer_next = next_stamp >= prev_stamp

        result = dict(prev if not prefer_next else nxt)
        # Always preserve payment proof.
        result['payment_history'] = _merge_payment_history(prev.get('payment_history'), nxt.get('payment_history'))
        result['last_paid_date'] = _max_date(prev.get('last_paid_date'), nxt.get('last_paid_date'))

        # Prefer the latest cycle anchor so paid records don't revert to an older due date.
        result['start_date'] = _max_date(prev.get('start_date'), nxt.get('start_date'))

        # Numeric fields: prefer whichever record matches the chosen start_date (current cycle).
        chosen_cycle = str(result.get('start_date') or '').strip()
        prev_cycle = str(prev.get('start_date') or '').strip()
        next_cycle = str(nxt.get('start_date') or '').strip()
        cycle_source = None
        if chosen_cycle and chosen_cycle == next_cycle:
            cycle_source = nxt
        elif chosen_cycle and chosen_cycle == prev_cycle:
            cycle_source = prev
        else:
            cycle_source = nxt if prefer_next else prev

        amount = _parse_float(cycle_source.get('amount'))
        pending = _parse_float(cycle_source.get('pending_amount'))
        if amount is not None:
            result['amount'] = amount
        if pending is not None:
            result['pending_amount'] = pending

        # Period months: keep a sane integer.
        period = _parse_int_safe(cycle_source.get('period_months'), 0)
        if period > 0:
            result['period_months'] = period

        # Remarks: keep non-empty; prefer whichever record is newer.
        result['remarks'] = _choose_text(prev.get('remarks'), nxt.get('remarks'), prefer_next=prefer_next)

        # Timestamps: keep created_at earliest, updated_at latest-ish.
        created = _choose_text(prev.get('created_at'), nxt.get('created_at'), prefer_next=False)
        if created:
            result['created_at'] = created
        updated = _choose_text(prev.get('updated_at'), nxt.get('updated_at'), prefer_next=prefer_next)
        if updated:
            result['updated_at'] = updated

        # Ensure studentId preserved.
        result['studentId'] = _parse_int_safe(result.get('studentId'), _parse_int_safe(prev.get('studentId'), 0))
        return result

    for source in (existing_records or []), (incoming_records or []):
        for item in source:
            key, normalized = _normalize_record(item)
            if key is None:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            merged[key] = _merge_pair(prev, normalized)

    return list(merged.values())


def _merge_resource_cabinet_superset(existing_items, incoming_items):
    """Merge resource cabinet items by id (int)."""
    merged = {}

    def _normalize(item):
        if not isinstance(item, dict):
            return None, None
        item_id = _parse_int_safe(item.get('id'), 0)
        if item_id <= 0:
            return None, None
        normalized = dict(item)
        normalized['id'] = item_id
        normalized['name'] = str(item.get('name') or '').strip()
        normalized['unit'] = str(item.get('unit') or '').strip()
        normalized['price_per_unit'] = float(item.get('price_per_unit') or 0) if str(item.get('price_per_unit') or '').strip() else float(item.get('price_per_unit') or 0)
        normalized['total_held'] = max(0, _parse_int_safe(item.get('total_held'), 0))
        normalized['updated_at'] = str(item.get('updated_at') or normalized.get('updated_at') or '').strip()
        normalized['created_at'] = str(item.get('created_at') or normalized.get('created_at') or normalized['updated_at'] or '').strip()
        return str(item_id), normalized

    for src in (existing_items or []), (incoming_items or []):
        if not isinstance(src, list):
            continue
        for item in src:
            key, normalized = _normalize(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
            next_stamp = _parse_sync_stamp(normalized.get('updated_at') or normalized.get('created_at'))
            merged[key] = normalized if next_stamp >= prev_stamp else prev

    return list(merged.values())


def _resource_status_rank(value):
    text = str(value or '').strip().lower()
    ranks = {
        'draft': 0,
        'pending_teacher': 1,
        'recommended': 2,
        'not_recommended': 2,
        'pending_admin': 3,
        'approved': 4,
        'rejected': 4,
        'fulfilled': 5,
        'cancelled': 5
    }
    return ranks.get(text, 0)


def _merge_resource_requests_superset(existing_requests, incoming_requests):
    """
    Merge resource requests by id.
    Safety: never lose recommendation/approval decisions once present.
    """
    merged = {}

    def _normalize(item):
        if not isinstance(item, dict):
            return None, None
        rid = _parse_int_safe(item.get('id'), 0)
        if rid <= 0:
            return None, None
        normalized = dict(item)
        normalized['id'] = rid
        sid = _parse_int_safe(item.get('studentId'), 0)
        if sid > 0:
            normalized['studentId'] = sid
        normalized['month'] = str(item.get('month') or '').strip()
        normalized['status'] = str(item.get('status') or '').strip().lower()
        normalized['updated_at'] = str(item.get('updated_at') or '').strip()
        normalized['created_at'] = str(item.get('created_at') or normalized['updated_at'] or '').strip()
        return str(rid), normalized

    def _merge_pair(prev, nxt):
        prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
        next_stamp = _parse_sync_stamp(nxt.get('updated_at') or nxt.get('created_at'))
        prefer_next = next_stamp >= prev_stamp
        base = dict(prev if not prefer_next else nxt)

        prev_status = str(prev.get('status') or '').strip().lower()
        next_status = str(nxt.get('status') or '').strip().lower()
        resolved_statuses = {'approved', 'rejected', 'fulfilled', 'cancelled'}
        prev_resolved = prev_status in resolved_statuses
        next_resolved = next_status in resolved_statuses

        # Preserve decisions/proof.
        for key in ('teacher_decision', 'teacher_remark', 'teacher_login_id', 'teacher_updated_at',
                    'admin_decision', 'admin_remark', 'admin_login_id', 'admin_updated_at',
                    'approved_at', 'fulfilled_at', 'urgent'):
            prev_val = prev.get(key)
            next_val = nxt.get(key)
            if str(next_val or '').strip() and prefer_next:
                base[key] = next_val
            elif str(prev_val or '').strip() and not str(base.get(key) or '').strip():
                base[key] = prev_val

        # Never downgrade a resolved request back to a pending/in-progress state.
        # This protects against stale client snapshots re-opening already approved/rejected rows.
        if prev_resolved and not next_resolved:
            base['status'] = prev_status
            for key in ('admin_decision', 'admin_remark', 'admin_login_id', 'admin_updated_at',
                        'approved_at', 'fulfilled_at', 'teacher_decision', 'teacher_remark'):
                prev_val = prev.get(key)
                if str(prev_val or '').strip():
                    base[key] = prev_val
        # If both are resolved, keep the furthest status; on tie, prefer newer stamp.
        elif prev_resolved and next_resolved:
            prev_rank = _resource_status_rank(prev_status)
            next_rank = _resource_status_rank(next_status)
            if next_rank > prev_rank:
                base['status'] = next_status
            elif prev_rank > next_rank:
                base['status'] = prev_status
            else:
                base['status'] = next_status if prefer_next else prev_status
        elif _resource_status_rank(next_status) >= _resource_status_rank(prev_status):
            base['status'] = next_status or prev_status
        else:
            base['status'] = prev_status or next_status

        # Keep created_at earliest, updated_at latest (by chosen record timestamp semantics).
        prev_created = str(prev.get('created_at') or '').strip()
        next_created = str(nxt.get('created_at') or '').strip()
        if prev_created and next_created:
            base['created_at'] = prev_created if prev_created <= next_created else next_created
        elif prev_created or next_created:
            base['created_at'] = prev_created or next_created
        base['updated_at'] = str((nxt if prefer_next else prev).get('updated_at') or base.get('updated_at') or '').strip()

        return base

    for src in (existing_requests or []), (incoming_requests or []):
        if not isinstance(src, list):
            continue
        for item in src:
            key, normalized = _normalize(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            merged[key] = _merge_pair(prev, normalized)

    return list(merged.values())


def _merge_leadership_superset(existing_posts, incoming_posts):
    """Merge leadership posts by id; never overwrite a populated holder with an empty one.
    An 'ended' status is preserved: once a post is ended, a stale client cannot revive it
    unless the incoming entry assigns a *different* holder (intentional reassignment)."""
    merged = {}
    def is_populated(p):
        return bool(str(p.get('holder') or '').strip() or str(p.get('roll') or '').strip())
    for p in (existing_posts or []):
        pid = int(p.get('id') or 0)
        if pid:
            merged[pid] = dict(p)
    for p in (incoming_posts or []):
        pid = int(p.get('id') or 0)
        if not pid:
            continue
        existing = merged.get(pid)
        if not existing:
            merged[pid] = dict(p)
            continue
        if is_populated(existing) and not is_populated(p):
            continue
        new_entry = {**existing, **p}
        # Preserve 'ended' against stale clients that still carry 'active' for the same holder.
        prev_status = str(existing.get('status') or '').strip().lower()
        new_status = str(p.get('status') or '').strip().lower()
        same_holder = (
            str(existing.get('holder') or '').strip().lower() ==
            str(p.get('holder') or '').strip().lower()
        )
        if prev_status == 'ended' and new_status != 'ended' and same_holder:
            new_entry['status'] = 'ended'
        merged[pid] = new_entry
    return list(merged.values())


def _merge_group_crs_superset(existing_crs, incoming_crs):
    """Merge group CRs by id; prefer entries with studentId assigned.
    When both entries have a student, the one with the newer updated_at/elected_on wins
    so that an admin-approved CR switch is never overwritten by a stale peer sync."""
    merged = {}
    for arr in [existing_crs or [], incoming_crs or []]:
        for cr in arr:
            cid = int(cr.get('id') or 0)
            if not cid:
                continue
            prev = merged.get(cid)
            if not prev:
                merged[cid] = dict(cr)
                continue
            # Never overwrite a populated entry with an empty one
            if prev.get('studentId') and not cr.get('studentId'):
                continue
            # When both have a student, prefer the newer entry by updated_at / elected_on
            if prev.get('studentId') and cr.get('studentId'):
                prev_stamp = max(
                    _parse_sync_stamp(prev.get('updated_at')),
                    _parse_sync_stamp(prev.get('elected_on'))
                )
                cr_stamp = max(
                    _parse_sync_stamp(cr.get('updated_at')),
                    _parse_sync_stamp(cr.get('elected_on'))
                )
                if prev_stamp > cr_stamp:
                    # prev is newer — keep it, only fill any missing fields from cr
                    new_entry = {**cr, **prev}
                    merged[cid] = new_entry
                    continue
            new_entry = {**prev, **cr}
            # Preserve 'ended' against stale clients that still have 'active' for the same student.
            prev_status = str(prev.get('status') or '').strip().lower()
            new_status = str(cr.get('status') or '').strip().lower()
            if (prev_status == 'ended' and new_status not in ('ended',)
                    and prev.get('studentId') and cr.get('studentId')
                    and int(prev.get('studentId') or 0) == int(cr.get('studentId') or 0)):
                new_entry['status'] = 'ended'
            merged[cid] = new_entry
    return list(merged.values())


def _merge_class_reps_superset(existing_reps, incoming_reps):
    """Merge class reps by id; prefer entries with studentId assigned.
    An 'ended' status is preserved unless the incoming entry assigns a *different* student."""
    merged = {}
    for arr in [existing_reps or [], incoming_reps or []]:
        for rep in arr:
            rid = int(rep.get('id') or 0)
            if not rid:
                continue
            prev = merged.get(rid)
            if not prev:
                merged[rid] = dict(rep)
                continue
            if prev.get('studentId') and not rep.get('studentId'):
                continue
            new_entry = {**prev, **rep}
            prev_status = str(prev.get('status') or '').strip().lower()
            new_status = str(rep.get('status') or '').strip().lower()
            if (prev_status == 'ended' and new_status not in ('ended',)
                    and prev.get('studentId') and rep.get('studentId')
                    and int(prev.get('studentId') or 0) == int(rep.get('studentId') or 0)):
                new_entry['status'] = 'ended'
            merged[rid] = new_entry
    return list(merged.values())


def _merge_party_members_superset(prev_members, new_members):
    """Merge two party member arrays into a union by member id — never lose a member.

    Prior behaviour overwrote the whole members array with whichever party record was
    processed last (last-writer-wins). That dropped members added on a client whose
    snapshot had not yet synced everywhere: an incoming snapshot from another client with
    a stale (older, non-empty) members array would silently overwrite the newer one, so
    members added earlier would vanish on the next sync. Same class of bug as the
    star-merge regression. This merges by member id so every member is preserved; on a
    conflicting id the record with the later elected_on / updated_at wins, preferring the
    one with more populated fields when timestamps are equal or absent."""
    merged = {}
    for arr in [prev_members or [], new_members or []]:
        for member in arr:
            if not isinstance(member, dict):
                continue
            mid = member.get('id')
            if mid is None or mid == '':
                # No id — key by studentId so we still de-dupe rather than duplicating.
                sid = member.get('studentId')
                mid = f"sid-{sid}" if sid is not None else None
                if mid is None:
                    continue
            key = str(mid)
            existing = merged.get(key)
            if not existing:
                merged[key] = dict(member)
                continue
            # Prefer the member record that looks newer/richer.
            def _stamp(m):
                return _parse_sync_stamp(m.get('updated_at') or m.get('elected_on') or m.get('created_at'))
            if _stamp(member) > _stamp(existing):
                merged[key] = {**existing, **member}
            elif _stamp(existing) > _stamp(member):
                merged[key] = {**member, **existing}
            else:
                # Equal/unknown timestamps — keep whichever has more non-empty fields.
                def _richness(m):
                    return sum(1 for v in m.values() if v not in (None, '', 0))
                merged[key] = {**existing, **member} if _richness(member) >= _richness(existing) else {**member, **existing}
    return list(merged.values())


def _merge_parties_superset(existing_parties, incoming_parties):
    """Merge parties by code; never overwrite non-empty members with empty, and never
    lose individual members to a stale snapshot."""
    merged = {}
    for arr in [existing_parties or [], incoming_parties or []]:
        for party in arr:
            key = str(party.get('code') or '').strip().upper() or str(party.get('id') or '')
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = dict(party)
                continue
            prev_members = prev.get('members') or []
            new_members = party.get('members') or []
            # Union the members by id so members added on either side are preserved;
            # then merge the rest of the party metadata with the incoming record winning.
            merged_members = _merge_party_members_superset(prev_members, new_members)
            merged[key] = {**prev, **party, 'members': merged_members}
    return list(merged.values())


def _merge_postholder_tickets(existing_tickets, incoming_tickets):
    """Merge postholder ticket balances keyed by studentId.
    Take the max of each ticket type so grants are never lost."""
    merged = {}
    local = existing_tickets if isinstance(existing_tickets, dict) else {}
    remote = incoming_tickets if isinstance(incoming_tickets, dict) else {}
    all_keys = set(list(local.keys()) + list(remote.keys()))
    for key in all_keys:
        lt = local.get(key) or {}
        rt = remote.get(key) or {}
        half = max(_parse_int_safe(lt.get('half_tickets'), 0), _parse_int_safe(rt.get('half_tickets'), 0))
        erase = max(_parse_int_safe(lt.get('erase_tickets'), 0), _parse_int_safe(rt.get('erase_tickets'), 0))
        if half > 0 or erase > 0:
            merged[key] = {'half_tickets': half, 'erase_tickets': erase}
    return merged


def _merge_postholder_ticket_log(existing_log, incoming_log):
    """Merge postholder ticket log entries by id; never drop entries, prefer newer by timestamp."""
    merged = {}
    for arr in [existing_log or [], incoming_log or []]:
        for entry in arr:
            if not isinstance(entry, dict):
                continue
            eid = int(entry.get('id') or 0)
            if not eid:
                continue
            prev = merged.get(eid)
            if not prev:
                merged[eid] = dict(entry)
                continue
            prev_stamp = _parse_sync_stamp(prev.get('timestamp'))
            new_stamp = _parse_sync_stamp(entry.get('timestamp'))
            if prev_stamp > new_stamp:
                merged[eid] = {**entry, **prev}
            else:
                merged[eid] = {**prev, **entry}
    return list(merged.values())


def _merge_pending_cr_requests_superset(existing_reqs, incoming_reqs):
    """Merge pending CR requests by id.
    Never downgrade a resolved (approved/rejected) request back to pending.
    """
    merged = {}
    for arr in [existing_reqs or [], incoming_reqs or []]:
        for req in arr:
            if not isinstance(req, dict):
                continue
            rid = int(req.get('id') or 0)
            if not rid:
                continue
            prev = merged.get(rid)
            if not prev:
                merged[rid] = dict(req)
                continue
            prev_status = str(prev.get('status') or '').strip().lower()
            new_status = str(req.get('status') or '').strip().lower()
            # Keep the resolved state if it has already been acted on
            if prev_status in ('approved', 'rejected') and new_status == 'pending':
                continue
            merged[rid] = {**prev, **req}
    return list(merged.values())


def _merge_pending_cr_requests_teacher(existing_reqs, incoming_reqs, teacher_login_id):
    """Teacher-safe merge for pending CR requests.
    Teachers can only create new pending requests; they cannot resolve existing ones.
    """
    now_iso = datetime.now(timezone.utc).isoformat(timespec='milliseconds').replace('+00:00', 'Z')
    existing = existing_reqs if isinstance(existing_reqs, list) else []
    by_id = {
        _parse_int_safe(req.get('id'), 0): dict(req)
        for req in existing
        if isinstance(req, dict) and _parse_int_safe(req.get('id'), 0) > 0
    }

    sanitized_new = []
    if not isinstance(incoming_reqs, list):
        incoming_reqs = []
    for raw in incoming_reqs:
        if not isinstance(raw, dict):
            continue
        rid = _parse_int_safe(raw.get('id'), 0)
        if rid <= 0 or rid in by_id:
            continue
        group = str(raw.get('group') or '').strip().upper()
        student_id = _parse_int_safe(raw.get('studentId'), 0)
        elected_on = str(raw.get('elected_on') or '').strip()[:10]
        if not group or student_id <= 0:
            continue
        post = str(raw.get('post') or f'GR - Group {group}').strip() or f'GR - Group {group}'
        note = str(raw.get('note') or '').strip()[:250]
        sanitized_new.append({
            'id': rid,
            'group': group,
            'post': post,
            'studentId': student_id,
            'elected_on': elected_on,
            'note': note,
            'requested_by': teacher_login_id or 'Teacher',
            'requested_at': str(raw.get('requested_at') or now_iso).strip() or now_iso,
            'status': 'pending',
            'updated_at': now_iso,
        })

    return _merge_pending_cr_requests_superset(existing, sanitized_new)


def _merge_resource_transactions_superset(existing_rows, incoming_rows):
    """Merge resource transactions by id."""
    merged = {}

    def _normalize(item):
        if not isinstance(item, dict):
            return None, None
        tid = _parse_int_safe(item.get('id'), 0)
        if tid <= 0:
            return None, None
        normalized = dict(item)
        normalized['id'] = tid
        sid = _parse_int_safe(item.get('studentId'), 0)
        if sid > 0:
            normalized['studentId'] = sid
        normalized['month'] = str(item.get('month') or '').strip()
        normalized['updated_at'] = str(item.get('updated_at') or '').strip()
        normalized['created_at'] = str(item.get('created_at') or normalized['updated_at'] or '').strip()
        return str(tid), normalized

    for src in (existing_rows or []), (incoming_rows or []):
        if not isinstance(src, list):
            continue
        for item in src:
            key, normalized = _normalize(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            prev_stamp = _parse_sync_stamp(prev.get('updated_at') or prev.get('created_at'))
            next_stamp = _parse_sync_stamp(normalized.get('updated_at') or normalized.get('created_at'))
            merged[key] = normalized if next_stamp >= prev_stamp else prev

    return list(merged.values())




def _merge_resource_advantage_deductions_superset(existing_rows, incoming_rows):
    """
    Merge resource_advantage_deductions by id.
    Safety rules:
    - Never delete a deduction record.
    - Once reversed=True, never revert back to False.
    - On conflict, keep the record that has reversed=True; otherwise keep the newer one.
    """
    merged = {}

    def _normalize(item):
        if not isinstance(item, dict):
            return None, None
        did = _parse_int_safe(item.get('id'), 0)
        if did <= 0:
            return None, None
        normalized = dict(item)
        normalized['id'] = did
        sid = _parse_int_safe(item.get('studentId'), 0)
        if sid > 0:
            normalized['studentId'] = sid
        normalized['month'] = str(item.get('month') or '').strip()
        normalized['points_deducted'] = max(0, _parse_int_safe(item.get('points_deducted'), 0))
        normalized['transaction_id'] = _parse_int_safe(item.get('transaction_id'), 0)
        normalized['reversed'] = bool(item.get('reversed'))
        normalized['created_at'] = str(item.get('created_at') or '').strip()
        return str(did), normalized

    for src in (existing_rows or []), (incoming_rows or []):
        if not isinstance(src, list):
            continue
        for item in src:
            key, normalized = _normalize(item)
            if not key:
                continue
            prev = merged.get(key)
            if not prev:
                merged[key] = normalized
                continue
            # Reversal is permanent — once reversed, keep it reversed regardless of timestamp.
            if prev.get('reversed') and not normalized.get('reversed'):
                continue  # keep prev (already reversed)
            if normalized.get('reversed') and not prev.get('reversed'):
                merged[key] = normalized  # incoming has reversal, take it
                continue
            # Both same reversal state — use the newer record.
            prev_stamp = _parse_sync_stamp(prev.get('created_at') or '')
            next_stamp = _parse_sync_stamp(normalized.get('created_at') or '')
            merged[key] = normalized if next_stamp >= prev_stamp else prev

    return list(merged.values())


def _build_teacher_resource_request(existing_payload, raw, teacher_login_id, month_key, now_iso):
    """Sanitize a teacher-created resource request row."""
    if not isinstance(existing_payload, dict) or not isinstance(raw, dict):
        return None

    rid = _parse_int_safe(raw.get('id'), 0)
    if rid <= 0:
        return None

    mode = str(raw.get('type') or '').strip().lower()
    if mode not in {'redeem_points', 'cash_purchase'}:
        return None

    month = str(raw.get('month') or month_key or '').strip()
    if month != str(month_key or '').strip():
        return None

    student_id = _parse_int_safe(raw.get('studentId'), 0)
    student_roll = _normalize_roll_value(raw.get('student_roll') or raw.get('studentRoll') or '')
    if student_id <= 0 and student_roll:
        student_id = _find_student_id_by_roll(existing_payload, student_roll)
    if student_id <= 0:
        return None

    students = existing_payload.get('students', []) if isinstance(existing_payload.get('students'), list) else []
    student_obj = next((s for s in students if _parse_int_safe((s or {}).get('id'), 0) == student_id), None)
    if not student_roll and isinstance(student_obj, dict):
        student_roll = _normalize_roll_value(student_obj.get('roll'))

    item_id = _parse_int_safe(raw.get('item_id') or raw.get('itemId'), 0)
    if item_id <= 0:
        return None

    cabinet = existing_payload.get('resource_cabinet', []) if isinstance(existing_payload.get('resource_cabinet'), list) else []
    cabinet_item = next((it for it in cabinet if _parse_int_safe((it or {}).get('id'), 0) == item_id), None)
    if not isinstance(cabinet_item, dict):
        return None

    qty = max(1, _parse_int_safe(raw.get('qty'), 1))

    def _float_or(value, default_value):
        try:
            return float(value)
        except Exception:
            return float(default_value)

    unit_price = max(0.0, _float_or(raw.get('price_per_unit'), cabinet_item.get('price_per_unit') or 0))
    total_cost = max(0.0, _float_or(raw.get('total_cost'), unit_price * qty))
    points_cost = max(0, _parse_int_safe(raw.get('points_cost'), 0))
    cash_paid = max(0, _parse_int_safe(raw.get('cash_paid'), 0))

    return {
        'id': rid,
        'type': mode,
        'studentId': student_id,
        'student_roll': student_roll,
        'month': month,
        'item_id': item_id,
        'item_name': str(raw.get('item_name') or cabinet_item.get('name') or '').strip(),
        'unit': str(raw.get('unit') or cabinet_item.get('unit') or '').strip(),
        'qty': qty,
        'price_per_unit': unit_price,
        'total_cost': total_cost,
        'points_cost': points_cost if mode == 'redeem_points' else 0,
        'cash_paid': cash_paid if mode == 'cash_purchase' else 0,
        'urgent': False,
        'admin_veto': False,
        'status': 'pending_teacher',
        'teacher_decision': '',
        'teacher_remark': '',
        'created_by_login_id': teacher_login_id,
        'created_at': now_iso,
        'updated_at': now_iso,
    }


def _merge_resource_requests_teacher(existing_payload, incoming_requests, teacher_login_id, month_key):
    """
    Teachers can only:
    - create new pending_teacher requests for the current server month
    - recommend / not recommend existing requests for the current month
    - add teacher remark
    """
    now_iso = datetime.now(timezone.utc).isoformat(timespec='milliseconds').replace('+00:00', 'Z')
    existing_obj = existing_payload if isinstance(existing_payload, dict) else {}
    existing_requests = existing_obj.get('resource_requests', []) if isinstance(existing_obj.get('resource_requests'), list) else []
    by_id = {}
    for item in existing_requests:
        if isinstance(item, dict):
            rid = _parse_int_safe(item.get('id'), 0)
            if rid > 0 and rid not in by_id:
                by_id[rid] = dict(item)

    updated = []
    if not isinstance(incoming_requests, list):
        incoming_requests = []
    for raw in incoming_requests:
        if not isinstance(raw, dict):
            continue
        rid = _parse_int_safe(raw.get('id'), 0)
        if rid <= 0:
            continue

        if rid not in by_id:
            created = _build_teacher_resource_request(existing_obj, raw, teacher_login_id, month_key, now_iso)
            if created:
                updated.append(created)
            continue

        current = by_id[rid]
        if str(current.get('month') or '').strip() != str(month_key or '').strip():
            continue
        status = str(current.get('status') or '').strip().lower()
        if status not in {'pending_teacher', 'recommended', 'not_recommended', 'pending_admin'}:
            continue

        decision = str(raw.get('teacher_decision') or raw.get('decision') or '').strip().lower()
        remark = str(raw.get('teacher_remark') or raw.get('remark') or '').strip()
        if decision not in {'recommended', 'not_recommended'} and not remark:
            continue

        next_row = dict(current)
        if remark:
            next_row['teacher_remark'] = remark
        if decision in {'recommended', 'not_recommended'}:
            next_row['teacher_decision'] = decision
            if decision == 'recommended':
                next_row['status'] = 'pending_admin'
            else:
                next_row['status'] = 'not_recommended'
        next_row['teacher_login_id'] = teacher_login_id
        next_row['teacher_updated_at'] = now_iso
        next_row['updated_at'] = now_iso
        updated.append(next_row)

    return _merge_resource_requests_superset(existing_requests, updated)


def _find_student_id_by_roll(payload, roll_value):
    roll = _normalize_roll_value(roll_value)
    if not roll:
        return 0
    students = payload.get('students', []) if isinstance(payload, dict) else []
    if not isinstance(students, list):
        return 0
    for s in students:
        if not isinstance(s, dict):
            continue
        if _normalize_roll_value(s.get('roll')) == roll:
            return _parse_int_safe(s.get('id'), 0)
    return 0


def _student_resource_request_patch(existing_payload, actor_login_id, incoming_payload):
    """
    Build a safe patch from a student submission.
    Only allows creating a new resource request for the logged-in student.
    """
    current_month = _server_now_iso()[:7]
    student_id = _find_student_id_by_roll(existing_payload or {}, actor_login_id)
    if student_id <= 0:
        return None, "Student roll not found on server roster"

    incoming_requests = []
    if isinstance(incoming_payload, dict) and isinstance(incoming_payload.get('resource_requests'), list):
        incoming_requests = incoming_payload.get('resource_requests') or []
    if not incoming_requests:
        return None, "No resource request provided"

    raw = incoming_requests[0] if isinstance(incoming_requests[0], dict) else None
    if not raw:
        return None, "Invalid resource request"

    now_iso = datetime.now(timezone.utc).isoformat(timespec='milliseconds').replace('+00:00', 'Z')
    rid = _parse_int_safe(raw.get('id'), 0)
    if rid <= 0:
        rid = int(datetime.now().timestamp() * 1000)

    req_type = str(raw.get('type') or '').strip().lower()
    if req_type not in {'redeem_points', 'cash_purchase'}:
        return None, "Invalid request type"

    item_id = _parse_int_safe(raw.get('item_id') or raw.get('itemId'), 0)
    qty = max(1, _parse_int_safe(raw.get('qty') or raw.get('quantity'), 0))
    if item_id <= 0:
        return None, "Item not selected"

    # Validate item exists.
    cabinet = (existing_payload or {}).get('resource_cabinet', []) or []
    if not isinstance(cabinet, list):
        cabinet = []
    item = next((it for it in cabinet if isinstance(it, dict) and _parse_int_safe(it.get('id'), 0) == item_id), None)
    if not item:
        return None, "Item not found in cabinet"

    price = float(item.get('price_per_unit') or 0)
    total_cost = max(0.0, price * float(qty))

    cash_paid = 0.0
    if req_type == 'cash_purchase':
        try:
            cash_paid = float(raw.get('cash_paid') or raw.get('paid_amount') or 0)
        except Exception:
            cash_paid = 0.0
        cash_paid = max(0.0, cash_paid)

    request_row = {
        'id': rid,
        'type': req_type,
        'studentId': student_id,
        'student_roll': _normalize_roll_value(actor_login_id),
        'month': current_month,
        'item_id': item_id,
        'item_name': str(item.get('name') or '').strip(),
        'unit': str(item.get('unit') or '').strip(),
        'qty': qty,
        'price_per_unit': price,
        'total_cost': total_cost,
        'cash_paid': cash_paid,
        'urgent': False,
        'status': 'pending_teacher',
        'created_at': now_iso,
        'updated_at': now_iso
    }
    return {'resource_requests': [request_row]}, ""


def _student_profile_change_appeal_patch(existing_payload, actor_login_id, incoming_payload):
    """
    Build a safe patch from a student submission for profile-change requests.
    Students can only create a new appeal of type=profile_change for themselves (append-only).
    """
    student_id = _find_student_id_by_roll(existing_payload or {}, actor_login_id)
    if student_id <= 0:
        return None, "Student roll not found on server roster"

    incoming_appeals = []
    if isinstance(incoming_payload, dict) and isinstance(incoming_payload.get('appeals'), list):
        incoming_appeals = incoming_payload.get('appeals') or []
    if not incoming_appeals:
        return None, "No appeal provided"

    raw = incoming_appeals[0] if isinstance(incoming_appeals[0], dict) else None
    if not raw:
        return None, "Invalid appeal"

    appeal_type = str(raw.get('type') or '').strip().lower()
    if appeal_type != 'profile_change':
        return None, "Invalid appeal type"

    existing_ids = set()
    for item in (existing_payload or {}).get('appeals', []) or []:
        if not isinstance(item, dict):
            continue
        aid = item.get('id')
        if aid is None:
            continue
        existing_ids.add(str(aid).strip())

    now_iso = datetime.now(timezone.utc).isoformat(timespec='milliseconds').replace('+00:00', 'Z')
    aid = _parse_int_safe(raw.get('id'), 0)
    if aid <= 0:
        aid = int(datetime.now().timestamp() * 1000)
    if str(aid).strip() in existing_ids:
        return None, "Duplicate appeal id"

    requested_name = str(raw.get('requested_name') or raw.get('requestedName') or '').strip()
    requested_profile = raw.get('requested_profile_data') if isinstance(raw.get('requested_profile_data'), dict) else {}
    allowed_keys = {
        'fatherName', 'motherName', 'dateOfBirth', 'bloodGroup', 'aadhar',
        'phone', 'email', 'parentPhone', 'admissionDate', 'academicYear', 'address'
    }
    sanitized_profile = {k: requested_profile.get(k) for k in requested_profile.keys() if k in allowed_keys}
    if not requested_name and not sanitized_profile:
        return None, "No profile changes provided"

    # Resolve canonical student roll + name from the server roster.
    roll_norm = _normalize_roll_value(actor_login_id)
    student_name = ''
    for s in (existing_payload or {}).get('students', []) or []:
        if not isinstance(s, dict):
            continue
        if _parse_int_safe(s.get('id'), 0) == student_id:
            student_name = str(s.get('base_name') or s.get('name') or '').strip()
            break

    appeal_row = {
        'id': aid,
        'type': 'profile_change',
        'subject': 'Profile Change Request',
        'message': f"Student requested profile update for {student_name or roll_norm}.",
        'from_role': 'student',
        'created_by': roll_norm,
        'target_role': 'admin',
        'forwarded_to': 'admin',
        'status': 'pending_admin',
        'recommendation': '',
        'student_id': student_id,
        'student_roll': roll_norm,
        'student_name': student_name,
        'requested_name': requested_name,
        'requested_profile_data': sanitized_profile,
        'created_at': now_iso,
        'updated_at': now_iso,
    }
    return {'appeals': [appeal_row]}, ""


def _extract_month_roster_rolls(payload, month_key):
    rolls = set()
    if not isinstance(payload, dict):
        return rolls
    month_students = payload.get('month_students', {}) or {}
    month_profiles = payload.get('month_roster_profiles', {}) or {}

    for value in month_students.get(month_key, []) or []:
        roll = _normalize_roll_value(value)
        if roll.startswith('EA'):
            rolls.add(roll)

    for profile in month_profiles.get(month_key, []) or []:
        if not isinstance(profile, dict):
            continue
        roll = _normalize_roll_value(profile.get('roll'))
        if roll.startswith('EA'):
            rolls.add(roll)

    return rolls


def _enforce_current_month_roster_integrity(incoming_data, existing_data):
    """
    Prevent stale client snapshots from shrinking current-month roster visibility.
    Keeps/repairs student entries for current roster rolls (e.g. Feb 2026 46-student roster).
    """
    if not isinstance(incoming_data, dict):
        return incoming_data

    current_month = _server_now_iso()[:7]
    month_key = current_month
    incoming_rolls = _extract_month_roster_rolls(incoming_data, month_key)
    existing_rolls = _extract_month_roster_rolls(existing_data or {}, month_key)

    if not incoming_rolls and '2026-02' != month_key:
        month_key = '2026-02'
        incoming_rolls = _extract_month_roster_rolls(incoming_data, month_key)
        existing_rolls = _extract_month_roster_rolls(existing_data or {}, month_key)

    roster_rolls = incoming_rolls or existing_rolls
    if not roster_rolls:
        return incoming_data

    incoming_students = list(incoming_data.get('students', []) or [])
    existing_students = list((existing_data or {}).get('students', []) or [])

    by_roll_incoming = {}
    for student in incoming_students:
        if not isinstance(student, dict):
            continue
        roll = _normalize_roll_value(student.get('roll'))
        if roll and roll not in by_roll_incoming:
            by_roll_incoming[roll] = student

    by_roll_existing = {}
    for student in existing_students:
        if not isinstance(student, dict):
            continue
        roll = _normalize_roll_value(student.get('roll'))
        if roll and roll not in by_roll_existing:
            by_roll_existing[roll] = student

    profile_by_roll = {}
    month_profiles = (incoming_data.get('month_roster_profiles', {}) or {}).get(month_key, []) or []
    for profile in month_profiles:
        if not isinstance(profile, dict):
            continue
        roll = _normalize_roll_value(profile.get('roll'))
        if roll:
            profile_by_roll[roll] = profile

    next_id = max([_parse_int_safe(student.get('id'), 0) for student in incoming_students if isinstance(student, dict)] + [0])

    # Ensure each roster roll has a student record in incoming payload.
    for roll in sorted(roster_rolls):
        target = by_roll_incoming.get(roll)
        if target is None and roll in by_roll_existing:
            source = dict(by_roll_existing[roll])
            incoming_students.append(source)
            by_roll_incoming[roll] = source
            target = source
        if target is None:
            next_id += 1
            profile = profile_by_roll.get(roll, {})
            name = str(profile.get('base_name') or profile.get('name') or roll).strip() or roll
            raw_name = str(profile.get('name') or name).strip() or name
            class_value = profile.get('class')
            target = {
                'id': next_id,
                'roll': roll,
                'name': name,
                'base_name': name,
                'raw_name': raw_name,
                'class': _parse_int_safe(class_value, None) if class_value not in (None, '') else None,
                'fees': 0,
                'vote_power': None,
                'stars': 0,
                'veto_count': 0,
                'active': True
            }
            incoming_students.append(target)
            by_roll_incoming[roll] = target

        # Do not force-active here: admins may intentionally deactivate duplicates.
        # We only guarantee the roster roll has a record (to prevent "missing students" issues).
        if 'active' not in target:
            target['active'] = True

        # Align display identity with month profile if available.
        profile = profile_by_roll.get(roll)
        if isinstance(profile, dict):
            base_name = str(profile.get('base_name') or profile.get('name') or '').strip()
            raw_name = str(profile.get('name') or '').strip()
            if base_name:
                target['base_name'] = base_name
                target['name'] = base_name
            if raw_name:
                target['raw_name'] = raw_name
            class_value = profile.get('class')
            if class_value not in (None, ''):
                target['class'] = _parse_int_safe(class_value, target.get('class'))

    incoming_data['students'] = incoming_students

    # Normalize month_students list for target month to canonical roll values.
    month_students = incoming_data.setdefault('month_students', {})
    month_students[month_key] = sorted(roster_rolls)

    return incoming_data

def _normalize_parties(parties):
    normalized = []
    for idx, party in enumerate(parties, start=1):
        code = str(party.get('code', '')).strip()
        if not code:
            continue
        try:
            power = int(party.get('power') or 0)
        except (TypeError, ValueError):
            power = 0
        party_id = party.get('id') or idx
        normalized.append({'id': int(party_id), 'code': code, 'power': power})
    return normalized


def _normalize_leadership(posts):
    normalized = []
    for idx, post in enumerate(posts, start=1):
        title = str(post.get('post', '')).strip()
        if not title:
            continue
        holder = str(post.get('holder', '')).strip()
        post_id = post.get('id') or idx
        normalized.append({'id': int(post_id), 'post': title, 'holder': holder})
    return normalized


def _load_politics_data():
    path = _politics_file_path()
    if not os.path.exists(path):
        return {
            'parties': DEFAULT_PARTIES.copy(),
            'leadership': DEFAULT_LEADERSHIP.copy()
        }
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        data['parties'] = _normalize_parties(data.get('parties', []))
        data['leadership'] = _normalize_leadership(data.get('leadership', []))
        return data
    except Exception:
        return {
            'parties': DEFAULT_PARTIES.copy(),
            'leadership': DEFAULT_LEADERSHIP.copy()
        }


def _save_politics_data(data):
    os.makedirs(_storage_root_path(), exist_ok=True)
    payload = {
        'parties': _normalize_parties(data.get('parties', [])),
        'leadership': _normalize_leadership(data.get('leadership', []))
    }
    _shared_atomic_write_json(_politics_file_path(), payload, indent=2)
    return payload


def _extract_party_and_leadership(ws):
    parties = []
    leadership = []

    party_row = None
    party_col = None
    power_col = None
    post_row = None
    post_col = None
    holder_col = None

    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if isinstance(cell.value, str):
                label = cell.value.strip().lower()
                if label == 'party':
                    party_row = cell.row
                    party_col = cell.column
                if label == 'power':
                    power_col = cell.column
                if label == 'post':
                    post_row = cell.row
                    post_col = cell.column
                if label == 'post holders':
                    holder_col = cell.column

    if party_row and party_col and power_col:
        for r in range(party_row + 1, party_row + 20):
            code = ws.cell(r, party_col).value
            power = ws.cell(r, power_col).value
            if not code:
                continue
            if str(code).strip().upper() == 'TOTAL':
                continue
            if isinstance(power, (int, float)):
                parties.append({'code': str(code).strip(), 'power': int(power)})

    if post_row and post_col and holder_col:
        for r in range(post_row + 1, post_row + 30):
            post = ws.cell(r, post_col).value
            holder = ws.cell(r, holder_col).value
            if post:
                leadership.append({'post': str(post).strip(), 'holder': str(holder).strip() if holder else ''})

    return parties, leadership


_seed_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'seed_data.json')
with open(_seed_path, 'r', encoding='utf-8') as _seed_file:
    FEB26_SEED = json.load(_seed_file)

# ============== ROUTES ==============

@points_bp.route('/balances', methods=['GET'])
@login_required
def get_student_balances():
    """Return authoritative star and VETO balances for all students for a month.

    This is the single source of truth that the client should use to confirm
    available counts before applying star/VETO actions. Uses the same data
    store and the same formula as the server-side reconciliation functions so
    the numbers are always consistent with what the server persists.

    Query params:
        month  YYYY-MM (defaults to current month)
    """
    month_key = str(request.args.get('month') or '').strip()
    if not re.match(r'^\d{4}-\d{2}$', month_key):
        month_key = _server_now_iso()[:7]

    data = _load_offline_data() or {}
    students = data.get('students', []) or []

    balances = []
    for student in students:
        sid = _parse_int_safe(student.get('id'), 0)
        if sid <= 0:
            continue
        balances.append({
            'id': sid,
            'roll': student.get('roll', ''),
            'name': student.get('name', ''),
            'available_stars': _compute_student_star_balance(data, sid, month_key),
            'available_vetos': _compute_student_veto_balance(data, sid, month_key),
            'individual_veto_count': max(0, _parse_int_safe(student.get('veto_count'), 0)),
            'role_veto_count': max(0, _parse_int_safe(student.get('role_veto_count'), 0)),
            'global_stars': max(0, _parse_int_safe(student.get('stars'), 0)),
        })

    resp = jsonify({'success': True, 'month': month_key, 'balances': balances})
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@points_bp.route('/validate-action', methods=['POST'])
@csrf.exempt  # JSON API endpoint — secured by @login_required
@login_required
def validate_action():
    """Server-side validation of a proposed star/VETO action.

    The client should call this before applying any star or VETO delta so the
    server can confirm the action is valid given the authoritative balance.
    Returns the available counts and whether the proposed delta is allowed.

    Body JSON:
        student_id   int
        month        YYYY-MM
        delta_stars  int  (negative = spending, positive = awarding)
        delta_vetos  int  (negative = spending, positive = awarding)
    """
    payload = request.get_json(silent=True) or {}
    student_id = _parse_int_safe(payload.get('student_id'), 0)
    month_key = str(payload.get('month') or '').strip()
    delta_stars = _parse_int_safe(payload.get('delta_stars'), 0)
    delta_vetos = _parse_int_safe(payload.get('delta_vetos'), 0)

    if student_id <= 0:
        return jsonify({'success': False, 'error': 'student_id required'}), 400
    if not re.match(r'^\d{4}-\d{2}$', month_key):
        month_key = _server_now_iso()[:7]

    data = _load_offline_data() or {}
    available_stars = _compute_student_star_balance(data, student_id, month_key)
    available_vetos = _compute_student_veto_balance(data, student_id, month_key)

    star_valid = not (delta_stars < 0 and available_stars < abs(delta_stars))
    veto_valid = not (delta_vetos < 0 and available_vetos < abs(delta_vetos))

    resp = jsonify({
        'success': True,
        'student_id': student_id,
        'month': month_key,
        'available_stars': available_stars,
        'available_vetos': available_vetos,
        'star_action': {
            'valid': star_valid,
            'delta': delta_stars,
            'error': None if star_valid else (
                f'Insufficient stars: {available_stars} available, {abs(delta_stars)} requested'
            ),
        },
        'veto_action': {
            'valid': veto_valid,
            'delta': delta_vetos,
            'error': None if veto_valid else (
                f'Insufficient VETOs: {available_vetos} available, {abs(delta_vetos)} requested'
            ),
        },
        'overall_valid': star_valid and veto_valid,
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@points_bp.route('/record-roll-change', methods=['POST'])
@csrf.exempt  # Admin-only; CSRF exempt to allow in-app fetch with credentials
@login_required
@_ledger_write_guard
def record_roll_change():
    """Record a student roll-number change and propagate it forward through history.

    When a student is promoted (e.g. EA24A01 → EA25A01 from September 2025),
    admin calls this endpoint. It:
      1. Appends an entry to data['roll_history'] with old/new roll and effective month
      2. Updates the student's current roll in data['students']
      3. Copies the month_roster_profile from old_roll to new_roll for every
         month >= effective_month (so carry-forward still works)
      4. Returns the new roll history for this student

    Body JSON:
        student_id      int
        new_roll        str    e.g. "EA25A01"
        effective_month str    YYYY-MM — month from which new_roll applies
        reason          str    optional, e.g. "annual_promotion"
    """
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    payload = request.get_json(silent=True) or {}
    student_id = _parse_int_safe(payload.get('student_id'), 0)
    new_roll = _normalize_roll_value(payload.get('new_roll') or '')
    effective_month = str(payload.get('effective_month') or '').strip()
    reason = str(payload.get('reason') or 'manual_update')[:200]

    if student_id <= 0:
        return jsonify({'success': False, 'error': 'student_id required'}), 400
    if not new_roll:
        return jsonify({'success': False, 'error': 'new_roll required'}), 400
    if not re.match(r'^\d{4}-\d{2}$', effective_month):
        return jsonify({'success': False, 'error': 'effective_month must be YYYY-MM'}), 400

    data = _load_offline_data() or {}
    students = data.get('students', []) or []

    student = next((s for s in students if _parse_int_safe(s.get('id'), 0) == student_id), None)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404

    old_roll = _normalize_roll_value(student.get('roll'))
    if old_roll == new_roll:
        return jsonify({'success': False, 'error': 'new_roll is the same as current roll'}), 400

    # 1. Record in roll_history
    if not isinstance(data.get('roll_history'), list):
        data['roll_history'] = []

    now_iso = _server_now_iso()
    history_entry = {
        'id': int(time.time() * 1000),
        'student_id': student_id,
        'old_roll': old_roll,
        'new_roll': new_roll,
        'effective_month': effective_month,
        'changed_by': str(current_user.login_id or '').strip(),
        'changed_at': now_iso,
        'reason': reason,
    }
    data['roll_history'].append(history_entry)

    # 2. Update the student's current roll
    student['roll'] = new_roll
    student['updated_at'] = now_iso
    student['roll_updated_at'] = now_iso
    group_match = re.match(r'^EA\d{2}([A-Z])', new_roll)
    if group_match:
        student['group'] = group_match.group(1)

    # Retire any leftover active records that still carry the old roll or
    # already collide with the new roll. These are stale duplicates from
    # previous syncs/imports and should not stay visible once the roll has
    # been reassigned.
    retired_duplicate_ids = []
    for other in students:
        if not isinstance(other, dict):
            continue
        other_id = _parse_int_safe(other.get('id'), 0)
        if other is student:
            continue
        other_roll = _normalize_roll_value(other.get('roll'))
        if other_id != student_id and other_roll not in {old_roll, new_roll}:
            continue
        other['active'] = False
        other['retired_at'] = now_iso
        other['retired_reason'] = f'roll_changed_to:{new_roll}'
        if other_id == student_id:
            other['roll'] = new_roll
        retired_duplicate_ids.append(other_id or other_roll)

    # 3. Update month_roster_profiles for months >= effective_month:
    #    rename the old_roll entry to new_roll so carry-forward still works.
    #    month_roster_profiles[month] is a LIST of profile dicts (not a dict keyed by roll).
    month_profiles = data.get('month_roster_profiles', {}) or {}
    month_students = data.get('month_students', {}) or {}
    copied_months = []
    for month_key, profiles in month_profiles.items():
        month_key_str = str(month_key or '').strip()
        if not isinstance(profiles, list):
            continue
        if month_key_str < effective_month:
            continue  # historical months keep the old roll — correct by design
        changed_profiles = False
        next_by_roll = {}
        for profile in profiles:
            if not isinstance(profile, dict):
                continue
            roll_val = _normalize_roll_value(profile.get('roll'))
            # remove any lingering old roll; replace with new unless new already exists
            if roll_val == old_roll:
                roll_val = new_roll
                profile['roll'] = new_roll
                changed_profiles = True
            if not roll_val:
                continue
            # keep only one entry per roll (last one wins)
            next_by_roll[roll_val] = profile
        # if new roll already existed, the old one is simply dropped
        if changed_profiles:
            copied_months.append(month_key)
        month_profiles[month_key] = list(next_by_roll.values())
    data['month_roster_profiles'] = month_profiles

    # 3b. Update month_students for months >= effective_month (roster list of rolls)
    for month_key, rolls in list(month_students.items()):
        month_key_str = str(month_key or '').strip()
        if month_key_str < effective_month:
            continue
        if not isinstance(rolls, list):
            continue
        changed_rolls = False
        next_rolls = []
        seen = set()
        for value in rolls:
            roll_val = _normalize_roll_value(value)
            if roll_val == old_roll:
                roll_val = new_roll
                changed_rolls = True
            if roll_val and roll_val not in seen:
                seen.add(roll_val)
                next_rolls.append(roll_val)
        # also remove any lingering old_roll even if new_roll not added
        if old_roll in seen:
            changed_rolls = True
            next_rolls = [r for r in next_rolls if r != old_roll]
            seen.discard(old_roll)
        if changed_rolls:
            month_students[month_key] = next_rolls
    data['month_students'] = month_students

    # 3c. Update month_student_extras roll keys for months >= effective_month.
    month_student_extras = data.get('month_student_extras', {}) or {}
    if isinstance(month_student_extras, dict):
        for month_key, month_extras in list(month_student_extras.items()):
            month_key_str = str(month_key or '').strip()
            if month_key_str < effective_month:
                continue
            if not isinstance(month_extras, dict):
                continue
            next_extras = {}
            local_changed = False
            for key, value in month_extras.items():
                normalized_key = _normalize_roll_value(key)
                if normalized_key == old_roll:
                    next_extras[new_roll] = value
                    local_changed = True
                else:
                    next_extras[key] = value
            if local_changed:
                month_student_extras[month_key] = next_extras
        data['month_student_extras'] = month_student_extras

    # 3d. Update roll fields in leadership / CR tables to keep labels in sync
    def _swap_roll_field(items, field):
        if not isinstance(items, list):
            return
        for item in items:
            try:
                if _normalize_roll_value(item.get(field)) == old_roll:
                    item[field] = new_roll
            except Exception:
                continue

    _swap_roll_field(data.get('leadership', []), 'roll')
    _swap_roll_field(data.get('class_reps', []), 'roll')
    _swap_roll_field(data.get('group_crs', []), 'roll')
    _swap_roll_field(data.get('post_holder_history', []), 'roll')
    _swap_roll_field(data.get('attendance', []), 'roll')
    _swap_roll_field(data.get('appeals', []), 'student_roll')
    _swap_roll_field(data.get('appeals', []), 'roll')
    _swap_roll_field(data.get('resource_requests', []), 'student_roll')
    _swap_roll_field(data.get('resource_requests', []), 'roll')
    _swap_roll_field(data.get('resource_transactions', []), 'student_roll')
    _swap_roll_field(data.get('resource_transactions', []), 'roll')
    _swap_roll_field(data.get('resource_advantage_deductions', []), 'student_roll')
    _swap_roll_field(data.get('resource_advantage_deductions', []), 'roll')

    # 3e. Update party member rolls (if present)
    if isinstance(data.get('parties'), list):
        for party in data['parties']:
            members = party.get('members') if isinstance(party, dict) else None
            if not isinstance(members, list):
                continue
            for member in members:
                try:
                    if _normalize_roll_value(member.get('roll')) == old_roll:
                        member['roll'] = new_roll
                except Exception:
                    continue

    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='roll-change')

    student_history = [
        e for e in data['roll_history']
        if _parse_int_safe(e.get('student_id'), 0) == student_id
    ]

    resp = jsonify({
        'success': True,
        'student_id': student_id,
        'old_roll': old_roll,
        'new_roll': new_roll,
        'effective_month': effective_month,
        'profiles_updated': copied_months,
        'retired_duplicate_ids': retired_duplicate_ids,
        'history': student_history,
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@points_bp.route('/roll-history', methods=['GET'])
@login_required
def get_roll_history():
    """Return roll number change history.

    Query params:
        student_id   int  optional — filter to one student
    """
    if current_user.role not in ('admin', 'teacher'):
        return jsonify({'success': False, 'error': 'Admin or teacher access required'}), 403

    sid_filter = _parse_int_safe(request.args.get('student_id'), 0)
    data = _load_offline_data() or {}
    history = data.get('roll_history', []) or []

    if sid_filter > 0:
        history = [e for e in history if _parse_int_safe(e.get('student_id'), 0) == sid_filter]

    # Enrich with current student name for readability
    students = {_parse_int_safe(s.get('id'), 0): s for s in (data.get('students', []) or [])}
    enriched = []
    for entry in sorted(history, key=lambda e: str(e.get('changed_at') or ''), reverse=True):
        sid = _parse_int_safe(entry.get('student_id'), 0)
        s = students.get(sid, {})
        enriched.append({
            **entry,
            'student_name': s.get('name', ''),
            'current_roll': _normalize_roll_value(s.get('roll')),
        })

    resp = jsonify({'success': True, 'count': len(enriched), 'history': enriched})
    resp.headers['Cache-Control'] = 'no-store'
    return resp


# Content hash of offline_scoreboard.html — recomputed on first request after server start.
_offline_html_hash = None


def _get_offline_html_hash():
    """Compute a short content hash for cache-busting the 2.4 MB HTML payload."""
    global _offline_html_hash
    if _offline_html_hash is not None:
        return _offline_html_hash
    try:
        html_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'offline_scoreboard.html')
        import hashlib
        h = hashlib.md5(open(html_path, 'rb').read()).hexdigest()[:10]
        _offline_html_hash = h
    except Exception:
        _offline_html_hash = '0'
    return _offline_html_hash


@points_bp.route('/offline')
def offline_scoreboard():
    """Serve offline scoreboard HTML — ETag-revalidated every load, never stale."""
    response = send_file('static/offline_scoreboard.html')
    # no-cache (NOT no-store): the browser keeps a copy but MUST revalidate on
    # every load. send_file emits an mtime/size ETag, so an unchanged file gets
    # a 0-byte 304 instead of re-downloading ~3.1 MB (~550 KB gzipped); any edit
    # bumps mtime+size → new ETag → fresh 200. Freshness guarantee identical to
    # no-store (every load still hits the server); repeat loads are near-instant.
    response.headers['Cache-Control'] = 'no-cache, must-revalidate'
    return response


@points_bp.route('/public')
def public_scoreboard():
    """Public live scoreboard (read-only, no login required)."""
    response = current_app.make_response(render_template('scoreboard/public_live.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@points_bp.route('/seed-data', methods=['GET'])
def seed_data():
    """Provide seed data lazily so the main HTML stays lightweight."""
    payload = FEB26_SEED
    response = jsonify(payload)
    response.headers['Cache-Control'] = 'public, max-age=86400'
    return response


def _is_party_president_designation(designation):
    if not designation:
        return False
    text = str(designation).strip().lower()
    return text == 'party president' or text == 'pp'


def _resolve_leadership_student(post, students):
    if not post:
        return None
    # studentId check
    student_id = post.get('studentId')
    if student_id is not None:
        try:
            sid = int(student_id)
            for s in students:
                if int(s.get('id', 0)) == sid:
                    return s
        except ValueError:
            pass
    # roll check
    roll = post.get('roll')
    if roll:
        roll_norm = str(roll).strip().upper()
        for s in students:
            if str(s.get('roll', '')).strip().upper() == roll_norm:
                return s
    # holder name check
    holder = post.get('holder')
    if holder:
        holder_norm = str(holder).strip().upper()
        for s in students:
            name = s.get('name')
            base_name = s.get('base_name')
            if (name and str(name).strip().upper() == holder_norm) or (base_name and str(base_name).strip().upper() == holder_norm):
                return s
    return None


def _get_student_group_jurisdiction(student_id, data):
    try:
        sid = int(student_id)
    except (ValueError, TypeError):
        return set()
    students = data.get('students', [])
    jurisdiction = set()
    
    # 1. Leadership
    major_posts = ['leader', 'co-leader', 'leader of opposition', 'lop']
    for post in data.get('leadership', []):
        if post.get('status') == 'active':
            resolved = _resolve_leadership_student(post, students)
            if resolved and int(resolved.get('id', 0)) == sid:
                p_name = str(post.get('post', '')).lower()
                if any(mp in p_name for mp in major_posts):
                    return {'A', 'B', 'C', 'D', 'H'}
                    
    # 2. Group CRs
    for rep in data.get('group_crs', []):
        try:
            rep_sid = int(rep.get('studentId', 0))
        except (ValueError, TypeError):
            continue
        if rep.get('status') == 'active' and rep_sid == sid and rep.get('group'):
            jurisdiction.add(str(rep.get('group')).strip().upper())
            
    # 3. Class Reps
    for rep in data.get('class_reps', []):
        try:
            rep_sid = int(rep.get('studentId', 0))
        except (ValueError, TypeError):
            continue
        if rep.get('status') == 'active' and rep_sid == sid:
            return {'A', 'B', 'C', 'D', 'H'}
            
    return jurisdiction


def _validate_veto_jurisdictions(data):
    # Retrieve students roster
    students = data.get('students', [])
    scores = data.get('scores', [])
    
    # Map student ID to student dict for convenience
    student_map = {}
    for s in students:
        if 'id' in s:
            try:
                student_map[int(s['id'])] = s
            except (ValueError, TypeError):
                pass
    
    # We will search for all postholder veto usages.
    # A postholder veto usage is a score entry where notes indicate a postholder veto action.
    # Note formats:
    # 1) "[VETO-Postholder] Removed penalty for Group X on current day"
    # 2) "[VETO-Shield-Group X] Activated VETO shield for Group X postholders"
    # Where X is the group letter (A, B, C, D, H)
    
    import re
    penalty_regex = re.compile(r'\[VETO-Postholder\] Removed penalty for Group ([A-Z])', re.IGNORECASE)
    shield_regex = re.compile(r'\[VETO-Shield-Group ([A-Z])\]', re.IGNORECASE)
    
    # Keep track of valid postholder actions by (date, group) to validate targets
    valid_penalty_removals = set() # elements: (date, group)
    valid_shields = set() # elements: (date, group)
    
    for score in scores:
        notes = score.get('notes', '')
        if not notes:
            continue
        
        sid = score.get('studentId')
        if sid is None:
            continue
        try:
            sid = int(sid)
        except (ValueError, TypeError):
            continue
            
        date = score.get('date', '')
        
        # Check for postholder penalty removal
        penalty_match = penalty_regex.search(notes)
        if penalty_match:
            group = penalty_match.group(1).upper()
            jurisdiction = _get_student_group_jurisdiction(sid, data)
            if group not in jurisdiction:
                current_app.logger.warning(
                    f"Validation failed: Student {sid} saved postholder veto for Group {group} "
                    f"but jurisdiction is {jurisdiction}."
                )
                return False
            valid_penalty_removals.add((date, group))
            
        # Check for postholder shield activation
        shield_match = shield_regex.search(notes)
        if shield_match:
            group = shield_match.group(1).upper()
            jurisdiction = _get_student_group_jurisdiction(sid, data)
            if group not in jurisdiction:
                current_app.logger.warning(
                    f"Validation failed: Student {sid} activated veto shield for Group {group} "
                    f"but jurisdiction is {jurisdiction}."
                )
                return False
            valid_shields.add((date, group))

    # Validate that no student has an unauthorized penalty removal
    # A penalty removal note: "[VETO-Postholder-Penalty-Group G] Penalty removed"
    # on date D.
    target_penalty_regex = re.compile(r'\[VETO-Postholder-Penalty-Group ([A-Z])\] Penalty removed', re.IGNORECASE)
    for score in scores:
        notes = score.get('notes', '')
        if not notes:
            continue
        target_match = target_penalty_regex.search(notes)
        if target_match:
            group = target_match.group(1).upper()
            date = score.get('date', '')
            if (date, group) not in valid_penalty_removals:
                current_app.logger.warning(
                    f"Validation failed: Unauthorized penalty removal for Group {group} on {date} "
                    f"without a valid postholder veto usage."
                )
                return False

    # Validate that no active shield in leadership/group_crs/class_reps/parties is active without a valid shield record
    # Let's check leadership
    for entry in data.get('leadership', []):
        if entry.get('status') == 'active' and entry.get('veto_shield_active'):
            date = entry.get('veto_shield_used_on', '')
            resolved = _resolve_leadership_student(entry, students)
            if resolved:
                group = str(resolved.get('group', '')).strip().upper()
                if group and (date, group) not in valid_shields:
                    current_app.logger.warning(
                        f"Validation failed: Active leadership shield for Group {group} on {date} "
                        f"without a valid postholder veto shield activation."
                    )
                    return False
                    
    # Let's check group_crs
    for entry in data.get('group_crs', []):
        if entry.get('status') == 'active' and entry.get('veto_shield_active'):
            date = entry.get('veto_shield_used_on', '')
            sid = entry.get('studentId')
            if sid is not None:
                try:
                    sid = int(sid)
                    if sid in student_map:
                        student = student_map[sid]
                        group = str(student.get('group', '')).strip().upper()
                        if group and (date, group) not in valid_shields:
                            current_app.logger.warning(
                                f"Validation failed: Active group CR shield for Group {group} on {date} "
                                f"without a valid postholder veto shield activation."
                            )
                            return False
                except (ValueError, TypeError):
                    pass
                    
    # Let's check class_reps
    for entry in data.get('class_reps', []):
        if entry.get('status') == 'active' and entry.get('veto_shield_active'):
            date = entry.get('veto_shield_used_on', '')
            sid = entry.get('studentId')
            if sid is not None:
                try:
                    sid = int(sid)
                    if sid in student_map:
                        student = student_map[sid]
                        group = str(student.get('group', '')).strip().upper()
                        if group and (date, group) not in valid_shields:
                            current_app.logger.warning(
                                f"Validation failed: Active class rep shield for Group {group} on {date} "
                                f"without a valid postholder veto shield activation."
                            )
                            return False
                except (ValueError, TypeError):
                    pass

    # Let's check parties president
    for party in data.get('parties', []):
        for member in party.get('members', []):
            if member.get('status') == 'active' and member.get('veto_shield_active'):
                # Check designation is party president
                designation = str(member.get('designation', '')).lower()
                if _is_party_president_designation(designation):
                    date = member.get('veto_shield_used_on', '')
                    sid = member.get('studentId')
                    if sid is not None:
                        try:
                            sid = int(sid)
                            if sid in student_map:
                                student = student_map[sid]
                                group = str(student.get('group', '')).strip().upper()
                                if group and (date, group) not in valid_shields:
                                    current_app.logger.warning(
                                        f"Validation failed: Active party president shield for Group {group} on {date} "
                                        f"without a valid postholder veto shield activation."
                                    )
                                    return False
                        except (ValueError, TypeError):
                            pass

    return True


def _is_valid_replication_request():
    """
    Validate peer replication requests with secure key comparison.

    Security improvements:
    - Requires SYNC_SHARED_KEY to be set (minimum 16 chars recommended)
    - Uses HMAC comparison to prevent timing attacks
    - Validates required headers are present
    """
    import hmac

    if request.headers.get('X-EA-Replicated') != '1':
        return False

    expected_key = resolve_sync_shared_key()
    provided_key = request.headers.get('X-EA-Sync-Key', '').strip()

    # Fail if no replication key source is configured.
    if not expected_key:
        current_app.logger.warning("Replication key missing (set SYNC_SHARED_KEY or SECRET_KEY) - rejecting replication request")
        return False

    # Keep a minimum length guard while allowing common SECRET_KEY fallbacks.
    if len(expected_key) < 8:
        current_app.logger.error("Replication key too short (minimum 8 characters required)")
        return False

    # Security: Use HMAC comparison to prevent timing attacks
    return hmac.compare_digest(expected_key, provided_key)


@points_bp.route('/offline-data', methods=['GET', 'POST'])
@csrf.exempt  # Required for peer-to-peer sync, but secured with sync key validation
@limiter.limit("2000 per hour")  # LAN mode: allow frequent sync while preventing runaway abuse
@_ledger_write_guard
def offline_data():
    replicated_auth = _is_valid_replication_request()
    is_replicated = request.headers.get('X-EA-Replicated') == '1'
    user, actor_role, actor_login_id = _get_request_user()

    if request.method == 'POST' and not user and not replicated_auth:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    if request.method == 'POST' and str(os.getenv('EA_RESTORE_LOCK', '')).strip() == '1':
        return jsonify({'success': False, 'error': 'Restore lock enabled'}), 423

    if request.method == 'GET':
        data = _load_offline_data()
        if not data:
            return ('', 204)
        min_students = _min_safe_student_roster()
        if _is_tiny_roster(data, min_students):
            data, src = _recover_tiny_roster_if_needed(data, min_students=min_students)
            if _is_tiny_roster(data, min_students):
                # Hard safety: never serve known-corrupt tiny rosters (prevents old clients from applying them).
                current_app.logger.error(
                    "Refusing to serve tiny roster snapshot (%s students). Recovery source=%s",
                    _student_count(data),
                    src or 'none',
                )
                resp = jsonify({'success': False, 'error': 'Roster snapshot incomplete. Recovery required.'})
                resp.headers['Cache-Control'] = 'no-store'
                return resp, 503
        data, _ = _recover_stale_snapshot_if_needed(data, min_students=min_students)
        if _ensure_score_timestamps(data):
            # Defer the heavy write to a background thread — the in-memory
            # cache already has the fix, so subsequent reads are correct.
            # Blocking the GET response for a 1+ second write is unnecessary.
            def _deferred_timestamp_save():
                try:
                    with app.app_context():
                        _save_offline_data(data)
                except Exception:
                    try:
                        current_app.logger.exception("Failed to persist score timestamp normalization (deferred)")
                    except Exception:
                        pass
            try:
                app = current_app._get_current_object()
                threading.Thread(target=_deferred_timestamp_save, daemon=True).start()
            except Exception:
                try:
                    _save_offline_data(data)
                except Exception:
                    current_app.logger.exception("Failed to persist score timestamp normalization on GET")

        is_authenticated = (user is not None)
        is_admin = is_authenticated and str(actor_role or '').strip().lower() == 'admin'
        anon_full_allowed = replicated_auth or str(os.getenv('EA_ALLOW_ANON_FULL_SYNC', '')).strip() == '1'
        if not is_authenticated and not anon_full_allowed:
            # Unauthenticated viewers (wall displays, logged-out SPA tabs) get a
            # sanitized recent-months snapshot — never fees/appeals/logs/profile
            # data. Replication peers authenticate via X-EA-Replicated +
            # X-EA-Sync-Key and still receive the full ledger. Set
            # EA_ALLOW_ANON_FULL_SYNC=1 to restore the legacy open behavior.
            data = _sanitize_anonymous_snapshot(data)
        elif is_authenticated and not is_admin:
            allowed_months = _allowed_months_for_user(data, user)
            data = _clip_payload_to_allowed_months(data, allowed_months)

        updated_at = data.get('server_updated_at') or data.get('updated_at')
        since = request.args.get('since') if hasattr(request, 'args') else None
        if since and updated_at:
            server_stamp = _parse_sync_stamp(updated_at)
            since_stamp = _parse_sync_stamp(since)
            if server_stamp and since_stamp and server_stamp <= since_stamp:
                # Bandwidth optimization for WAN: if client is already at/above this server stamp,
                # avoid sending the full payload. SSE (offline-events) still provides realtime updates.
                return ('', 204)

            # Delta sync: send only scores added/modified since the client's last sync.
            # This reduces a typical 34 MB full dump to <200 KB for incremental updates.
            if server_stamp and since_stamp and server_stamp > since_stamp:
                # Filter by updated_at FIRST so edits to old score rows are
                # included in the delta (filtering by score date alone silently
                # missed edits to historical rows — clients then kept stale values).
                delta_scores = [
                    s for s in data.get('scores', [])
                    if isinstance(s, dict) and _parse_sync_stamp(s.get('updated_at') or s.get('created_at') or s.get('date') or '') and
                       _parse_sync_stamp(s.get('updated_at') or s.get('created_at') or s.get('date') or '') >= since_stamp
                ]
                # If delta is small enough (<50% of total), send delta payload.
                # Otherwise fall through to full sync for consistency.
                total_scores = len(data.get('scores', []))
                if 0 < len(delta_scores) < total_scores * 0.5:
                    delta_out = {
                        'delta': True,
                        'since': since,
                        'scores': delta_scores,
                        'students': data.get('students', []),  # students array is small (~85 entries)
                        'parties': data.get('parties', []),
                        'chess_champion': data.get('chess_champion', {}),
                        'leadership': data.get('leadership', []),
                        'post_holder_history': data.get('post_holder_history', []),
                        'veto_tracking': data.get('veto_tracking', {}),
                        'server_updated_at': updated_at,
                        'updated_at': updated_at,
                        'server_version': data.get('server_version', 0),
                        '_cache_bust_version': data.get('_cache_bust_version', ''),
                    }
                    # Include month_roster_profiles if changed (small dict)
                    if data.get('month_roster_profiles'):
                        delta_out['month_roster_profiles'] = data['month_roster_profiles']
                    resp = jsonify({'data': delta_out, 'updated_at': updated_at})
                    resp.headers['Cache-Control'] = 'no-store'
                    return resp
        
        # ── Serialized-response cache (admin full-sync only) ──────────────
        # For admin users requesting a full sync (no ?since=), the response
        # body is identical across consecutive GETs until a write changes the
        # data.  Caching the pre-serialized bytes saves ~124ms per request.
        cache_mtime = 0
        cache_size = -1
        cache_ver = -1
        if is_admin and not since:
            try:
                path = _offline_data_path()
                st = os.stat(path)
                cache_mtime = getattr(st, 'st_mtime_ns', int(st.st_mtime * 1e9))
                cache_size = st.st_size
                cache_ver = _parse_int_safe(data.get('server_version'), 0)
                expected_etag = f'W/"{cache_ver}-{cache_size}"'
                # Conditional GET: if the client already has this version,
                # return 304 Not Modified — saves ~18 MB of bandwidth per pull.
                if_none_match = (request.headers.get('If-None-Match') or '').strip()
                if if_none_match and if_none_match == expected_etag:
                    resp = Response(status=304)
                    resp.headers['ETag'] = expected_etag
                    resp.headers['Cache-Control'] = 'no-store'
                    return resp
                cached_body, cached_etag = _get_serialized_response(cache_mtime, cache_size, cache_ver)
                if cached_body is not None:
                    resp = Response(cached_body, mimetype='application/json')
                    resp.headers['Cache-Control'] = 'no-store'
                    if cached_etag:
                        resp.headers['ETag'] = cached_etag
                    return resp
            except Exception:
                pass  # Fall through to normal serialization

        # DIAGNOSTIC FIX: Ensure attendance records are present in GET response
        attendance_records = data.get('attendance', [])
        if not attendance_records:
            current_app.logger.debug(f"GET /offline-data: No attendance records in snapshot (students: {len(data.get('students', []))}, scores: {len(data.get('scores', []))})")
        else:
            current_app.logger.debug(f"GET /offline-data: Returning {len(attendance_records)} attendance records")
        
        # Keep sync snapshots lean for reliability/performance. Large collections are fetched
        # via dedicated endpoints and should not inflate every pull payload to prevent
        # "QuotaExceeded" errors in browsers with strict storage limits.
        data_out = dict(data)
        data_out.pop('activity_log', None)
        data_out['notification_history'] = (data_out.get('notification_history') or [])[-50:]  # Keep only recent
        data_out['proposal_messages'] = (data_out.get('proposal_messages') or [])[-30:]  # Keep only recent
        data_out['_sync_ops'] = []  # Clear pending sync ops
        resp = jsonify({'data': data_out, 'updated_at': updated_at})
        resp.headers['Cache-Control'] = 'no-store'

        # Store the serialized response for future admin full-sync cache hits.
        if is_admin and not since:
            try:
                body = resp.get_data()
                etag = f'W/"{cache_ver}-{cache_size}"'
                _store_serialized_response(cache_mtime, cache_size, cache_ver, body, etag)
            except Exception:
                pass

        return resp

    payload = request.get_json(silent=True) or {}
    data = payload.get('data', payload)
    historical_score_ops = payload.get('historical_score_ops', []) if isinstance(payload, dict) else []
    request_peers = payload.get('peers', []) if isinstance(payload, dict) else []
    request_op_id = str(payload.get('op_id') or '').strip() if isinstance(payload, dict) else ''
    request_base_version = _parse_int_safe(payload.get('base_version'), 0) if isinstance(payload, dict) else 0
    if request_op_id:
        g.ea_ledger_op_id = request_op_id
    if not isinstance(data, dict):
        return jsonify({'success': False, 'error': 'Invalid payload'}), 400

    actor_login_id = actor_login_id if user else ''
    if user:
        actor_role = actor_role
    elif replicated_auth:
        declared_role = payload.get('actor_role') if isinstance(payload, dict) else ''
        if not declared_role and isinstance(data, dict):
            declared_role = data.get('actor_role', '')
        actor_role = str(declared_role or 'admin').strip().lower()
    else:
        actor_role = 'admin'

    actor_role = str(actor_role or 'admin').strip().lower()
    if actor_role not in ['admin', 'teacher', 'student']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    replica_purpose = ''
    if isinstance(payload, dict):
        replica_purpose = str(payload.get('replica_purpose') or '').strip().lower()
    if not replica_purpose and isinstance(data, dict):
        replica_purpose = str(data.get('replica_purpose') or '').strip().lower()

    # Master safety: In master mode, reject peer replication except for narrow teacher patches.
    if request.method == 'POST' and is_replicated and str(os.getenv('EA_MASTER_MODE', '')).strip() == '1':
        if not (actor_role == 'teacher' and replica_purpose == 'teacher_patch'):
            return jsonify({'success': False, 'error': 'Peer replication disabled on master mode'}), 409

    existing = _load_offline_data() or {}
    existing_version = _parse_int_safe(existing.get('server_version'), 0)
    if request_op_id and _is_duplicate_sync_op(existing, request_op_id):
        current_stamp = existing.get('server_updated_at') or existing.get('updated_at')
        return jsonify({
            'success': True,
            'dedup': True,
            'updated_at': current_stamp,
            'server_version': existing_version
        })

    if (
        request_base_version > 0 and
        existing_version > 0 and
        request_base_version != existing_version and
        not replicated_auth and
        not bool(payload.get('force_replace')) and
        not bool(payload.get('authoritative_master_push'))
    ):
        current_stamp = existing.get('server_updated_at') or existing.get('updated_at')
        return jsonify({
            'success': False,
            'error': 'Version conflict',
            'code': 'stale_base_version',
            'updated_at': current_stamp,
            'server_version': existing_version
        }), 409
    min_students = _min_safe_student_roster()
    if _is_tiny_roster(existing, min_students):
        existing, _ = _recover_tiny_roster_if_needed(existing, min_students=min_students)
    # Skip stale-recovery peer fetch on POST — it adds network latency (Render cold-start can be
    # 30+ seconds) and is unnecessary when we're already receiving a push from an authenticated client.
    if request.method == 'GET':
        existing, _ = _recover_stale_snapshot_if_needed(existing, min_students=min_students)
    authoritative_master_push = bool(payload.get('authoritative_master_push')) if isinstance(payload, dict) else False
    force_replace = bool(payload.get('force_replace')) if isinstance(payload, dict) else False
    if authoritative_master_push and replicated_auth:
        force_replace = True
    if replicated_auth and authoritative_master_push:
        data['chess_champion'] = _normalize_chess_champion(data.get('chess_champion'))
    else:
        data['chess_champion'] = _merge_chess_champion_superset(
            existing.get('chess_champion', {}),
            data.get('chess_champion', {})
        )
    incoming_count = _student_count(data)
    existing_count = _student_count(existing)
    if actor_role not in ['teacher', 'student'] and not force_replace and incoming_count > 0 and incoming_count < min_students:
        current_stamp = existing.get('server_updated_at') or existing.get('updated_at')
        return jsonify({
            'success': False,
            'error': f'Incoming roster snapshot too small ({incoming_count} students). Upload rejected.',
            'updated_at': current_stamp
        }), 409
    if actor_role not in ['teacher', 'student'] and not force_replace and _is_suspicious_student_shrink(existing, data):
        current_app.logger.warning(
            "Suspicious student shrink detected and rejected. "
            "Source: %s, Existing Count: %s, Incoming Count: %s, Existing Stamp: %s, Incoming Stamp: %s",
            request.remote_addr,
            _student_count(existing),
            _student_count(data),
            existing.get('server_updated_at') or existing.get('updated_at'),
            data.get('server_updated_at') or data.get('updated_at')
        )
        current_stamp = existing.get('server_updated_at') or existing.get('updated_at')
        return jsonify({
            'success': False,
            'error': 'Incoming snapshot would shrink student master data. Upload rejected.',
            'updated_at': current_stamp
        }), 409
    incoming_stamp = _payload_sync_stamp(data)
    existing_stamp = _payload_sync_stamp(existing)
    if actor_role == 'teacher':
        data = _filter_teacher_payload_to_edit_window(data, actor_login_id or 'Teacher')
        if user:
            data = _clip_payload_to_allowed_months(
                data,
                _allowed_months_for_user(existing, user)
            )
        merged = existing if existing else {}
        merged['chess_champion'] = _normalize_chess_champion(existing.get('chess_champion', {}))
        merged.setdefault('students', existing.get('students', []))
        merged.setdefault('month_students', existing.get('month_students', {}))
        merged.setdefault('month_roster_profiles', existing.get('month_roster_profiles', {}))
        merged.setdefault('parties', existing.get('parties', []))
        merged.setdefault('leadership', existing.get('leadership', []))
        merged.setdefault('class_reps', existing.get('class_reps', []))
        merged.setdefault('group_crs', existing.get('group_crs', []))
        merged.setdefault('election_candidates', existing.get('election_candidates', []))
        merged.setdefault('election_votes', existing.get('election_votes', []))
        merged.setdefault('election_individual_votes', existing.get('election_individual_votes', []))
        merged.setdefault('election_teacher_votes', existing.get('election_teacher_votes', []))
        merged.setdefault('pending_election_results', existing.get('pending_election_results', []))
        merged.setdefault('appeals', existing.get('appeals', []))
        merged.setdefault('attendance', existing.get('attendance', []))
        merged.setdefault('notification_history', existing.get('notification_history', []))
        merged.setdefault('resource_cabinet', existing.get('resource_cabinet', []))
        merged.setdefault('resource_requests', existing.get('resource_requests', []))
        merged.setdefault('resource_transactions', existing.get('resource_transactions', []))
        merged.setdefault('syllabus_catalog', existing.get('syllabus_catalog', {}))
        merged.setdefault('syllabus_tracking', existing.get('syllabus_tracking', []))
        if isinstance(data.get('syllabus_catalog'), dict):
            merged['syllabus_catalog'] = merge_syllabus_catalog_superset(
                existing.get('syllabus_catalog', {}),
                data.get('syllabus_catalog', {}),
                _parse_int_safe
            )
        merged['scores'] = _merge_teacher_scores(existing, data)
        if isinstance(data.get('appeals'), list):
            merged['appeals'] = _merge_appeals_superset(existing.get('appeals', []), data.get('appeals', []))
        if isinstance(data.get('attendance'), list):
            incoming_attendance_count = len(data.get('attendance', []))
            existing_attendance_count = len(existing.get('attendance', []))
            merged['attendance'] = _merge_attendance_superset(existing, data)
            merged_attendance_count = len(merged.get('attendance', []))
            current_app.logger.info(
                f"[TEACHER SYNC] Attendance merged | "
                f"incoming: {incoming_attendance_count}, existing: {existing_attendance_count}, result: {merged_attendance_count} | "
                f"teacher: {actor_login_id or 'Teacher'}"
            )
        if isinstance(data.get('election_teacher_votes'), list):
            merged['election_teacher_votes'] = _merge_election_votes_superset(
                existing.get('election_teacher_votes', []),
                data.get('election_teacher_votes', []),
                mode='teacher'
            )
        if isinstance(data.get('pending_election_results'), list):
            merged['pending_election_results'] = _merge_pending_results_superset(
                existing.get('pending_election_results', []),
                data.get('pending_election_results', [])
            )
        if isinstance(data.get('notification_history'), list):
            merged['notification_history'] = _merge_notification_history(
                existing.get('notification_history', []),
                data.get('notification_history', [])
            )
        if isinstance(data.get('resource_requests'), list):
            merged['resource_requests'] = _merge_resource_requests_teacher(
                existing,
                data.get('resource_requests', []),
                actor_login_id or 'Teacher',
                _server_now_iso()[:7]
            )
        if isinstance(data.get('syllabus_tracking'), list):
            merged['syllabus_tracking'] = merge_syllabus_tracking_superset(
                existing.get('syllabus_tracking', []),
                data.get('syllabus_tracking', []),
                _parse_int_safe,
                _parse_sync_stamp
            )
        # Teachers cannot directly modify post-holder source tables.
        # They must submit approval requests that Admin applies.
        if isinstance(data.get('pending_cr_requests'), list):
            merged['pending_cr_requests'] = _merge_pending_cr_requests_teacher(
                existing.get('pending_cr_requests', []),
                data.get('pending_cr_requests', []),
                actor_login_id or 'Teacher'
            )
        merged = _enforce_current_month_roster_integrity(merged, existing)
        _reconcile_role_veto_monthly(merged)
        _reconcile_veto_counters_from_scores(merged)
        _ensure_score_timestamps(merged)
        if not _validate_veto_jurisdictions(merged):
            return jsonify({'success': False, 'error': 'Veto jurisdiction validation failed'}), 400
        merged['updated_at'] = data.get('updated_at', existing.get('updated_at'))
        merged['server_updated_at'] = _server_now_iso()
        _record_sync_op(merged, request_op_id, actor_login_id or 'Teacher')
        _save_offline_data(merged)
        _broadcast_sync_event(merged['server_updated_at'], source='teacher')
        if not is_replicated:
            if str(os.getenv('EA_MASTER_MODE', '')).strip() == '1':
                _forward_offline_data_to_peers_async(merged, request_peers)
            else:
                patch = _build_teacher_replication_patch(merged, actor_login_id or 'Teacher')
                _forward_offline_data_to_peers_async(patch, request_peers)
        return jsonify({'success': True, 'updated_at': merged['server_updated_at'], 'server_version': _parse_int_safe(merged.get('server_version'), 0)})

    if actor_role == 'student':
        if user:
            data = _clip_payload_to_allowed_months(
                data,
                _allowed_months_for_user(existing, user)
            )
        # Students can only:
        # - create resource requests for themselves (append-only, server builds canonical row)
        # - submit profile-change requests to admin via appeals (append-only, sanitized)
        patch = {}
        if isinstance(data, dict) and isinstance(data.get('resource_requests'), list) and data.get('resource_requests'):
            req_patch, err = _student_resource_request_patch(existing, actor_login_id, data)
            if not req_patch:
                return jsonify({'success': False, 'error': err or 'Invalid student request'}), 400
            patch.update(req_patch)
        if isinstance(data, dict) and isinstance(data.get('appeals'), list) and data.get('appeals'):
            appeal_patch, err = _student_profile_change_appeal_patch(existing, actor_login_id, data)
            if not appeal_patch:
                return jsonify({'success': False, 'error': err or 'Invalid student appeal'}), 400
            patch.update(appeal_patch)
        if not patch:
            return jsonify({'success': False, 'error': 'No valid student update provided'}), 400
        existing_obj = existing if isinstance(existing, dict) else {}
        merged = existing_obj
        merged['chess_champion'] = _normalize_chess_champion(existing_obj.get('chess_champion', {}))
        merged.setdefault('resource_cabinet', existing_obj.get('resource_cabinet', []))
        merged.setdefault('resource_requests', existing_obj.get('resource_requests', []))
        merged.setdefault('resource_transactions', existing_obj.get('resource_transactions', []))
        merged.setdefault('appeals', existing_obj.get('appeals', []))
        if isinstance(patch.get('resource_requests'), list):
            merged['resource_requests'] = _merge_resource_requests_superset(
                merged.get('resource_requests', []),
                patch.get('resource_requests', [])
            )
        if isinstance(patch.get('appeals'), list):
            merged['appeals'] = _merge_appeals_superset(existing_obj.get('appeals', []), patch.get('appeals', []))
        _ensure_score_timestamps(merged)
        if not _validate_veto_jurisdictions(merged):
            return jsonify({'success': False, 'error': 'Veto jurisdiction validation failed'}), 400
        merged['server_updated_at'] = _server_now_iso()
        _record_sync_op(merged, request_op_id, actor_login_id or 'Student')
        _save_offline_data(merged)
        _broadcast_sync_event(merged['server_updated_at'], source='student')
        if not is_replicated:
            _forward_offline_data_to_peers_async(merged, request_peers)
        return jsonify({'success': True, 'updated_at': merged['server_updated_at'], 'server_version': _parse_int_safe(merged.get('server_version'), 0)})

    if existing and incoming_stamp and existing_stamp and incoming_stamp < existing_stamp:
        # Trusted master push can override stale timestamp drift on peers.
        if replicated_auth and authoritative_master_push and incoming_count >= min_students:
            current_app.logger.warning(
                "Accepting authoritative master snapshot despite older stamp. incoming=%s existing=%s",
                incoming_count,
                existing_count
            )
        # Case 1: existing is a tiny/corrupt roster — accept any healthy incoming snapshot.
        elif _is_tiny_roster(existing, min_students) and incoming_count >= min_students:
            current_app.logger.warning(
                "Accepting healthy snapshot (%s students) over tiny-roster data (%s students) despite older stamp.",
                incoming_count,
                existing_count
            )
        # Case 2: trusted peer push arrives with significantly MORE students than existing.
        # This heals the common Render-restart scenario where the dyno reseeds from FEB26_SEED
        # (46 students, datetime.now() stamp) and then the local master tries to push the real
        # 91-student snapshot whose stamp is genuinely older than the freshly-seeded timestamp.
        elif replicated_auth and _is_suspicious_student_shrink(data, existing) and incoming_count >= min_students:
            current_app.logger.warning(
                "Accepting trusted peer push (%s students) over apparent seed/shrunk data (%s students) despite older stamp.",
                incoming_count,
                existing_count
            )
        else:
            current_stamp = existing.get('server_updated_at') or existing.get('updated_at')
            return jsonify({
                'success': False,
                'error': 'Server has newer data',
                'updated_at': current_stamp
            }), 409

    # Patch-safety: if a client sends a partial payload (e.g. only fee/resource updates),
    # ensure we don't accidentally overwrite core tables like students/month roster.
    if isinstance(existing, dict):
        # Preserve veto_tracking from existing if missing, or merge usage logs if present
        if 'veto_tracking' not in data or not data.get('veto_tracking'):
            data['veto_tracking'] = existing.get('veto_tracking', {})
        elif 'veto_tracking' in existing and existing.get('veto_tracking'):
            exist_vt = existing['veto_tracking']
            incoming_vt = data['veto_tracking']
            merged_log = list(exist_vt.get('usage_log', []))
            exist_timestamps = {entry.get('timestamp') for entry in merged_log if entry.get('timestamp')}
            for entry in incoming_vt.get('usage_log', []):
                ts = entry.get('timestamp')
                if ts and ts not in exist_timestamps:
                    merged_log.append(entry)
                    exist_timestamps.add(ts)
            data['veto_tracking'] = dict(exist_vt)
            data['veto_tracking']['usage_log'] = merged_log

        if 'students' not in data:
            data['students'] = existing.get('students', [])
        elif isinstance(data.get('students'), list) and isinstance(existing.get('students'), list):
            # Merge students: never downgrade active status without a genuinely newer timestamp.
            data['students'] = _merge_students_preserve_active(
                existing.get('students', []),
                data.get('students', [])
            )
        # Always superset-merge month rosters so a partial client payload never shrinks the server roster.
        data['month_students'] = _merge_month_students_superset(
            existing.get('month_students', {}),
            data.get('month_students', {})
        )
        data['month_roster_profiles'] = _merge_month_roster_profiles_superset(
            existing.get('month_roster_profiles', {}),
            data.get('month_roster_profiles', {})
        )
        # Guard against accidental UI/import payloads that clear office-holder tables.
        # Preserve existing non-empty structures unless caller explicitly uses force_replace.
        if not force_replace:
            protected_list_tables = [
                'leadership',
                'group_crs',
                'class_reps',
                'parties',
                'post_holder_history',
                'postholder_ticket_log'
            ]
            for key in protected_list_tables:
                incoming = data.get(key)
                existing_val = existing.get(key)
                if isinstance(existing_val, list) and existing_val and (not isinstance(incoming, list) or len(incoming) == 0):
                    data[key] = existing_val
            protected_object_tables = [
                'postholder_tickets'
            ]
            for key in protected_object_tables:
                incoming = data.get(key)
                existing_val = existing.get(key)
                if isinstance(existing_val, dict) and existing_val and (not isinstance(incoming, dict) or len(incoming.keys()) == 0):
                    data[key] = existing_val

            # Specific protection for leadership selection state:
            # If the server has a populated candidates list, but incoming is empty or missing, retain existing server state.
            existing_ls = existing.get('leadership_selection_state')
            incoming_ls = data.get('leadership_selection_state')
            if isinstance(existing_ls, dict):
                existing_candidates = existing_ls.get('candidates', [])
                if isinstance(existing_candidates, list) and len(existing_candidates) > 0:
                    if not isinstance(incoming_ls, dict) or not isinstance(incoming_ls.get('candidates'), list) or len(incoming_ls.get('candidates')) == 0:
                        data['leadership_selection_state'] = existing_ls
                        current_app.logger.warning("[SYNC] Retained existing leadership_selection_state because incoming was empty/missing candidates.")

    if isinstance(existing, dict):
        data['scores'] = _merge_scores_superset(existing.get('scores', []), data.get('scores', []))
        data['attendance'] = _merge_attendance_superset(existing, data)
        data['appeals'] = _merge_appeals_superset(existing.get('appeals', []), data.get('appeals', []))
        data['election_votes'] = _merge_election_votes_superset(
            existing.get('election_votes', []),
            data.get('election_votes', []),
            mode='party'
        )
        data['election_individual_votes'] = _merge_election_votes_superset(
            existing.get('election_individual_votes', []),
            data.get('election_individual_votes', []),
            mode='individual'
        )
        data['election_teacher_votes'] = _merge_election_votes_superset(
            existing.get('election_teacher_votes', []),
            data.get('election_teacher_votes', []),
            mode='teacher'
        )
        data['pending_election_results'] = _merge_pending_results_superset(
            existing.get('pending_election_results', []),
            data.get('pending_election_results', [])
        )
        data['fee_records'] = _merge_fee_records_superset(
            existing.get('fee_records', []),
            data.get('fee_records', [])
        )
        data['resource_cabinet'] = _merge_resource_cabinet_superset(
            existing.get('resource_cabinet', []),
            data.get('resource_cabinet', [])
        )
        data['resource_requests'] = _merge_resource_requests_superset(
            existing.get('resource_requests', []),
            data.get('resource_requests', [])
        )
        data['resource_transactions'] = _merge_resource_transactions_superset(
            existing.get('resource_transactions', []),
            data.get('resource_transactions', [])
        )
        data['resource_advantage_deductions'] = _merge_resource_advantage_deductions_superset(
            existing.get('resource_advantage_deductions', []),
            data.get('resource_advantage_deductions', [])
        )
        data['notification_history'] = _merge_notification_history(
            existing.get('notification_history', []),
            data.get('notification_history', [])
        )
        data['proposals'] = _merge_records_superset(
            existing.get('proposals', []),
            data.get('proposals', []),
            key_fields=('id',),
            ts_fields=('updated_at', 'created_at', 'open_at')
        )
        data['proposal_votes'] = _merge_records_superset(
            existing.get('proposal_votes', []),
            data.get('proposal_votes', []),
            key_fields=('proposal_id', 'voter_login_id'),
            ts_fields=('updated_at', 'created_at')
        )
        data['proposal_messages'] = _merge_records_superset(
            existing.get('proposal_messages', []),
            data.get('proposal_messages', []),
            key_fields=('id',),
            ts_fields=('created_at',)
        )
        data['score_adjustment_actions'] = _merge_records_superset(
            existing.get('score_adjustment_actions', []),
            data.get('score_adjustment_actions', []),
            key_fields=('id',),
            ts_fields=('created_at',)
        )
        data['leadership'] = _merge_leadership_superset(
            existing.get('leadership', []),
            data.get('leadership', [])
        )
        data['group_crs'] = _merge_group_crs_superset(
            existing.get('group_crs', []),
            data.get('group_crs', [])
        )
        data['class_reps'] = _merge_class_reps_superset(
            existing.get('class_reps', []),
            data.get('class_reps', [])
        )
        data['parties'] = _merge_parties_superset(
            existing.get('parties', []),
            data.get('parties', [])
        )
        data['pending_cr_requests'] = _merge_pending_cr_requests_superset(
            existing.get('pending_cr_requests', []),
            data.get('pending_cr_requests', [])
        )
        data['syllabus_catalog'] = merge_syllabus_catalog_superset(
            existing.get('syllabus_catalog', {}),
            data.get('syllabus_catalog', {}),
            _parse_int_safe
        )
        data['syllabus_tracking'] = merge_syllabus_tracking_superset(
            existing.get('syllabus_tracking', []),
            data.get('syllabus_tracking', []),
            _parse_int_safe,
            _parse_sync_stamp
        )
        data['post_holder_history'] = _merge_records_superset(
            existing.get('post_holder_history', []),
            data.get('post_holder_history', []),
            key_fields=('id',),
            ts_fields=('updated_at', 'created_at')
        )
        data['postholder_tickets'] = _merge_postholder_tickets(
            existing.get('postholder_tickets', {}),
            data.get('postholder_tickets', {})
        )
        data['postholder_ticket_log'] = _merge_postholder_ticket_log(
            existing.get('postholder_ticket_log', []),
            data.get('postholder_ticket_log', [])
        )

    data = _preserve_locked_historical_window(existing, data)
    admin_historical_ops_allowed = (
        not is_replicated and
        current_user.is_authenticated and
        str(current_user.role or '').strip().lower() == 'admin' and
        actor_role == 'admin'
    )
    if admin_historical_ops_allowed and isinstance(historical_score_ops, list) and historical_score_ops:
        data = _apply_admin_historical_score_ops(data, historical_score_ops, actor_login_id or 'Admin')

    data = _enforce_current_month_roster_integrity(data, existing)
    _reconcile_role_veto_monthly(data)
    _reconcile_veto_counters_from_scores(data)
    _ensure_score_timestamps(data)
    if not _validate_veto_jurisdictions(data):
        return jsonify({'success': False, 'error': 'Veto jurisdiction validation failed'}), 400
    if is_replicated:
        # Preserve the source timestamp so the receiver never appears artificially
        # newer than the sender.  Generating a fresh _server_now_iso() here would
        # make Render's stamp >30 s ahead of local's (due to cold-start latency),
        # causing the next background-sync cycle to pull from Render and silently
        # revert any local changes (deactivations, edits) made after the push.
        data['server_updated_at'] = data.get('server_updated_at') or _server_now_iso()
    else:
        data['server_updated_at'] = _server_now_iso()
    _record_sync_op(data, request_op_id, actor_login_id or ('replica' if is_replicated else 'Admin'))
    _save_offline_data(data)
    _broadcast_sync_event(data['server_updated_at'], source='replica' if is_replicated else 'client')
    if not is_replicated:
        _forward_offline_data_to_peers_async(data, request_peers)
    return jsonify({'success': True, 'updated_at': data['server_updated_at'], 'server_version': _parse_int_safe(data.get('server_version'), 0)})


@points_bp.route('/offline-events')
@login_required
def offline_events():
    subscriber = _subscribe_sync_events()

    def generate():
        try:
            yield 'retry: 2500\n\n'
            existing = _load_offline_data() or {}
            stamp = existing.get('server_updated_at') or existing.get('updated_at') or ''
            yield f"event: sync\ndata: {json.dumps({'updated_at': stamp, 'source': 'init'})}\n\n"
            while True:
                try:
                    payload = subscriber.get(timeout=20)
                    yield f"event: sync\ndata: {payload}\n\n"
                except Empty:
                    yield ': keepalive\n\n'
        finally:
            _unsubscribe_sync_events(subscriber)

    headers = {
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    }
    return Response(stream_with_context(generate()), mimetype='text/event-stream', headers=headers)


@points_bp.route('/offline-server-health', methods=['POST'])
@csrf.exempt
@login_required
def offline_server_health():
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    requested_urls = payload.get('urls', []) if isinstance(payload, dict) else []
    urls = _normalize_peer_list(requested_urls)
    if not urls:
        urls = get_sync_peers()
    current_base = (request.host_url or '').rstrip('/')
    if current_base and current_base not in urls:
        urls.insert(0, current_base)
    urls = list(dict.fromkeys(urls))

    items = []
    for base in urls:
        endpoint = f"{base}/scoreboard/offline-data"
        status = 'offline'
        error = ''
        data_stamp = ''
        students = None
        scores = None
        try:
            req = urllib.request.Request(endpoint, method='GET')
            with urllib.request.urlopen(req, timeout=4) as resp:
                raw = resp.read()
                parsed = json.loads(raw.decode('utf-8'))
                payload_data = parsed.get('data', {}) if isinstance(parsed, dict) else {}
                status = 'online'
                data_stamp = parsed.get('updated_at') or payload_data.get('server_updated_at') or payload_data.get('updated_at') or ''
                if isinstance(payload_data, dict):
                    if isinstance(payload_data.get('students'), list):
                        students = len(payload_data.get('students'))
                    if isinstance(payload_data.get('scores'), list):
                        scores = len(payload_data.get('scores'))
                if not data_stamp:
                    status = 'degraded'
                    error = 'No data stamp'
        except Exception as exc:
            status = 'offline'
            error = str(exc)

        items.append({
            'base_url': base,
            'status': status,
            'data_stamp': data_stamp,
            'students': students,
            'scores': scores,
            'error': error
        })

    return jsonify({'success': True, 'items': items, 'checked_at': _server_now_iso()})


@points_bp.route('/supabase-health', methods=['GET'])
@login_required
def supabase_health():
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    cfg = _supabase_snapshot_config()
    enabled = bool(cfg.get('enabled_read') or cfg.get('enabled_write'))
    response = {
        'success': True,
        'enabled': enabled,
        'read_enabled': bool(cfg.get('enabled_read')),
        'write_enabled': bool(cfg.get('enabled_write')),
        'table': cfg.get('table'),
        'row_id': cfg.get('row_id'),
        'checked_at': _server_now_iso(),
    }
    if not enabled:
        response.update({'status': 'disabled', 'error': 'Supabase env not configured'})
        return jsonify(response)

    endpoint = f"{cfg.get('url', '').rstrip('/')}/rest/v1/{cfg.get('table', 'offline_snapshots')}"
    params = urllib.parse.urlencode({
        'select': 'id,updated_at,source',
        'id': f"eq.{cfg.get('row_id', 'main')}",
        'limit': 1
    })
    started = time.time()
    try:
        req = urllib.request.Request(
            f"{endpoint}?{params}",
            headers=_supabase_headers(cfg, key_name='read_key', for_write=False),
            method='GET'
        )
        with urllib.request.urlopen(req, timeout=12) as resp:
            body = resp.read()
        elapsed_ms = int((time.time() - started) * 1000)
        rows = json.loads(body.decode('utf-8', errors='replace')) if body else []
        row = rows[0] if isinstance(rows, list) and rows and isinstance(rows[0], dict) else {}
        response.update({
            'status': 'online',
            'latency_ms': elapsed_ms,
            'snapshot_updated_at': row.get('updated_at') or '',
            'snapshot_source': row.get('source') or '',
        })
    except Exception as exc:
        elapsed_ms = int((time.time() - started) * 1000)
        degraded_hint = ''
        degraded_latency = None
        try:
            root_started = time.time()
            root_req = urllib.request.Request(
                f"{cfg.get('url', '').rstrip('/')}/rest/v1/",
                headers=_supabase_headers(cfg, key_name='read_key', for_write=False),
                method='GET'
            )
            with urllib.request.urlopen(root_req, timeout=8) as root_resp:
                _ = root_resp.read()
            degraded_latency = int((time.time() - root_started) * 1000)
            degraded_hint = 'Project reachable; table read endpoint is blocked/slow'
        except Exception:
            degraded_hint = ''
        response.update({
            'status': 'degraded' if degraded_hint else 'offline',
            'latency_ms': elapsed_ms,
            'probe_latency_ms': degraded_latency,
            'probe_note': degraded_hint,
            'error': str(exc),
        })
    return jsonify(response)


@points_bp.route('/offline-force-publish', methods=['POST'])
@csrf.exempt
@login_required
@_ledger_write_guard
def offline_force_publish():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    data = payload.get('data') if isinstance(payload, dict) else None
    public_snapshot = payload.get('public_snapshot') if isinstance(payload, dict) else None
    request_peers = payload.get('peers', []) if isinstance(payload, dict) else []
    wait_for_results = bool(payload.get('wait_for_results')) if isinstance(payload, dict) else False
    cloudflare_only = bool(payload.get('cloudflare_only')) if isinstance(payload, dict) else False
    used_fallback_snapshot = False
    if not isinstance(data, dict):
        data = _load_offline_data() or {}
        used_fallback_snapshot = True
    if not isinstance(data, dict):
        data = {}
    existing = _load_offline_data() or {}
    data = _preserve_locked_historical_window(existing, data)
    data['chess_champion'] = _merge_chess_champion_superset(
        existing.get('chess_champion', {}),
        data.get('chess_champion', {})
    )

    data['server_updated_at'] = _server_now_iso()
    _save_offline_data(data)
    _broadcast_sync_event(data['server_updated_at'], source='force-publish')
    auto_push_public_site = _auto_push_public_site_enabled()

    if cloudflare_only:
        try:
            public_site_result = _publish_public_site_snapshot(data, push=auto_push_public_site, public_snapshot=public_snapshot)
        except Exception as exc:
            current_app.logger.exception("Force publish cloudflare-only failed")
            return jsonify({
                'success': False,
                'replication_ok': False,
                'mode': 'cloudflare_only',
                'used_fallback_snapshot': used_fallback_snapshot,
                'updated_at': data.get('server_updated_at'),
                'gist': {'status': 'skipped'},
                'peers': [],
                'public_site': {'status': 'failed', 'error': str(exc)},
                'error': f'Public site publish failed: {exc}'
            })
        public_site_ok = str(public_site_result.get('status') or '').lower() in {'pushed', 'up_to_date', 'written'}
        return jsonify({
            'success': True,
            'replication_ok': public_site_ok,
            'mode': 'cloudflare_only',
            'used_fallback_snapshot': used_fallback_snapshot,
            'updated_at': data.get('server_updated_at'),
            'gist': {'status': 'skipped'},
            'peers': [],
            'public_site': public_site_result
        })

    # Default mode: queue remote replication in background and return fast.
    # This avoids user-visible publish failures caused by WAN/Supabase timeouts.
    if not wait_for_results:
        public_site_result = _publish_public_site_snapshot(data, push=False)
        if auto_push_public_site:
            threading.Thread(
                target=_publish_public_site_snapshot,
                args=(dict(data), True),
                daemon=True,
                name='ea-public-site-push'
            ).start()
            public_site_result['status'] = 'queued'
        _gist_push_snapshot_async(data, reason='force_publish')
        _forward_offline_data_to_peers_async(data, request_peers)
        return jsonify({
            'success': True,
            'replication_ok': True,
            'mode': 'queued',
            'used_fallback_snapshot': used_fallback_snapshot,
            'updated_at': data.get('server_updated_at'),
            'gist': {'status': 'queued'},
            'peers': [],
            'public_site': public_site_result
        })

    gist_result = {'status': 'skipped'}
    public_site_result = _publish_public_site_snapshot(data, push=auto_push_public_site)

    peers = get_sync_peers() + _normalize_peer_list(request_peers)
    peers = list(dict.fromkeys(peers))
    current_origin = (request.host_url or '').rstrip('/')
    shared_key = resolve_sync_shared_key()
    is_master = str(os.getenv('EA_MASTER_MODE', '')).strip() == '1'
    peer_body_payload = {'data': payload_for_external_replication(data)}
    if is_master:
        peer_body_payload['authoritative_master_push'] = True
        peer_body_payload['force_replace'] = True
    peer_body = json.dumps(peer_body_payload).encode('utf-8')

    peer_targets = []
    for peer in peers:
        base = str(peer or '').rstrip('/')
        if not base or (current_origin and base == current_origin):
            continue
        peer_targets.append(base)

    peer_results = []
    result_lock = threading.Lock()
    threads = []

    def _push_gist_worker():
        nonlocal gist_result
        started = time.time()
        try:
            ok = _gist_push_snapshot(data, reason='force_publish', timeout_sec=15)
            elapsed_ms = int((time.time() - started) * 1000)
            gist_result = {
                'status': 'ok' if ok else 'failed',
                'latency_ms': elapsed_ms,
            }
        except Exception as exc:
            elapsed_ms = int((time.time() - started) * 1000)
            gist_result = {
                'status': 'failed',
                'latency_ms': elapsed_ms,
                'error': str(exc),
            }

    def _push_peer_worker(base):
        target_url = f'{base}/scoreboard/offline-data'
        started = time.time()
        result = {'base_url': base, 'status': 'failed'}
        try:
            req = urllib.request.Request(
                target_url,
                data=peer_body,
                method='POST',
                headers={
                    'Content-Type': 'application/json',
                    'X-EA-Replicated': '1',
                    'X-EA-Sync-Key': shared_key
                }
            )
            with urllib.request.urlopen(req, timeout=12) as resp:
                raw = resp.read()
            updated_at = ''
            try:
                parsed = json.loads(raw.decode('utf-8', errors='replace')) if raw else {}
                if isinstance(parsed, dict):
                    updated_at = parsed.get('updated_at') or ''
            except Exception:
                updated_at = ''
            result['status'] = 'ok'
            result['updated_at'] = updated_at
        except Exception as exc:
            result['error'] = str(exc)
        result['latency_ms'] = int((time.time() - started) * 1000)
        with result_lock:
            peer_results.append(result)

    supa_thread = threading.Thread(target=_push_gist_worker, daemon=True)
    supa_thread.start()
    threads.append(supa_thread)
    for base in peer_targets:
        t = threading.Thread(target=_push_peer_worker, args=(base,), daemon=True)
        t.start()
        threads.append(t)

    deadline = time.time() + 16.0
    for t in threads:
        remaining = deadline - time.time()
        if remaining <= 0:
            break
        t.join(timeout=remaining)

    for t in threads:
        if t.is_alive():
            if t is supa_thread:
                gist_result = {'status': 'timeout', 'error': 'Timed out'}
            else:
                # Best-effort: workers append own results; on timeout add synthetic entry.
                pass

    seen_bases = {str(item.get('base_url') or '').rstrip('/') for item in peer_results}
    for base in peer_targets:
        if base not in seen_bases:
            peer_results.append({'base_url': base, 'status': 'timeout', 'error': 'Timed out', 'latency_ms': 16000})

    peer_ok = any(item.get('status') == 'ok' for item in peer_results) if peer_targets else False
    public_site_ok = str(public_site_result.get('status') or '').lower() in {'pushed', 'up_to_date', 'written'}
    replication_ok = bool(peer_ok or gist_result.get('status') == 'ok' or public_site_ok)
    return jsonify({
        'success': True,
        'replication_ok': replication_ok,
        'used_fallback_snapshot': used_fallback_snapshot,
        'updated_at': data.get('server_updated_at'),
        'gist': gist_result,
        'peers': peer_results,
        'public_site': public_site_result
    })


@points_bp.route('/offline-backup', methods=['GET'])
@login_required
def offline_backup():
    data = _load_offline_data()
    if not data:
        return jsonify({'success': False, 'error': 'No data to backup'}), 404

    fd, temp_path = tempfile.mkstemp(prefix='offline_backup_', suffix='.json')
    with os.fdopen(fd, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    @after_this_request
    def _cleanup(response):
        try:
            os.remove(temp_path)
        except Exception:
            pass
        return response

    filename = f'offline_scoreboard_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    return send_file(temp_path, as_attachment=True, download_name=filename)


def _validate_and_sanitize_fee_updates(updates):
    import re
    if not isinstance(updates, dict):
        return {}
    
    sanitized = {}
    
    # 1. Base fee amount
    if 'amount' in updates:
        val = updates['amount']
        if val is not None:
            try:
                sanitized['amount'] = float(val)
            except (ValueError, TypeError):
                sanitized['amount'] = 0.0
        else:
            sanitized['amount'] = None
            
    # 2. Structural fields
    for field in ['start_date', 'last_paid_date', 'due_date']:
        if field in updates:
            val = str(updates[field] or '').strip()
            if re.match(r'^\d{4}-\d{2}-\d{2}$', val) or not val or val == '-':
                sanitized[field] = val if val != '-' else None
                
    # 3. Pending amount
    if 'pending_amount' in updates:
        val = updates['pending_amount']
        try:
            sanitized['pending_amount'] = float(val)
        except (ValueError, TypeError):
            sanitized['pending_amount'] = 0.0
            
    # 4. Payment history
    if 'payment_history' in updates:
        hist = updates['payment_history']
        if isinstance(hist, list):
            sanitized_hist = []
            for item in hist:
                if not isinstance(item, dict):
                    continue
                # Sanitize single transaction entry
                sanitized_item = {
                    'txn_id': str(item.get('txn_id') or '').strip(),
                    'date': str(item.get('date') or item.get('paid_on') or '').strip(),
                    'mode': str(item.get('mode') or 'cash').strip().lower(),
                    'category': str(item.get('category') or 'tuition').strip().lower(),
                    'note': str(item.get('note') or '').strip()[:500],
                    'recorded_by': str(item.get('recorded_by') or '').strip()[:100],
                    'status': str(item.get('status') or 'confirmed').strip().lower()
                }
                
                # Check numeric amount
                try:
                    sanitized_item['amount'] = float(item.get('amount', 0))
                except (ValueError, TypeError):
                    sanitized_item['amount'] = 0.0
                    
                # Check reversal flags
                if 'is_reversal' in item:
                    sanitized_item['is_reversal'] = bool(item['is_reversal'])
                if 'original_txn_id' in item:
                    sanitized_item['original_txn_id'] = str(item['original_txn_id'] or '').strip()
                    
                sanitized_hist.append(sanitized_item)
            sanitized['payment_history'] = sanitized_hist
            
    return sanitized


def _sync_fee_records_to_sqlite(student_id, roll_number, student_name, payment_history):
    from app.models.fees import FeeTransaction
    from app import db
    
    if not isinstance(payment_history, list):
        return
        
    for txn in payment_history:
        if not isinstance(txn, dict):
            continue
        txn_id = txn.get('txn_id')
        if not txn_id:
            continue # Skip legacy records without unique transaction IDs
            
        existing = FeeTransaction.query.filter_by(txn_id=txn_id).first()
        
        # Get operator/recorder
        recorded_by = txn.get('recorded_by') or txn.get('recordedBy')
        if not recorded_by:
            try:
                recorded_by = current_user.login_id
            except Exception:
                recorded_by = 'Admin'
                
        # Handle decimal amount
        try:
            amount = float(txn.get('amount', 0))
        except (ValueError, TypeError):
            amount = 0.0
            
        is_reversal = bool(txn.get('is_reversal', False) or txn.get('status') == 'reversed' or 'reversal' in str(txn.get('note', '')).lower())
        
        if existing:
            # Update fields in case they changed (e.g. status reversed)
            existing.status = txn.get('status', 'confirmed')
            existing.is_reversal = is_reversal
            existing.note = txn.get('note')
        else:
            new_txn = FeeTransaction(
                student_id=student_id,
                roll_number=roll_number or f"ST{student_id}",
                student_name=student_name or f"Student #{student_id}",
                txn_id=txn_id,
                date=txn.get('date') or txn.get('paid_on') or datetime.now().strftime('%Y-%m-%d'),
                amount=amount,
                mode=txn.get('mode') or 'cash',
                category=txn.get('category') or 'tuition',
                ref_no=txn.get('ref_no') or txn.get('refNo'),
                note=txn.get('note'),
                recorded_by=recorded_by,
                status=txn.get('status', 'confirmed'),
                is_reversal=is_reversal,
                original_txn_id=txn.get('original_txn_id')
            )
            db.session.add(new_txn)
            
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.error("Failed to commit fee transaction to SQLite: %s", exc)


def _auto_generate_monthly_pdf_if_needed(month_key, db_data=None):
    if not month_key:
        return
    from app.utils.data_paths import get_storage_root
    pdf_dir = os.path.join(get_storage_root(), 'fee_reports')
    os.makedirs(pdf_dir, exist_ok=True)
    
    pdf_path = os.path.join(pdf_dir, f"fee_report_{month_key}.pdf")
    
    if db_data is None:
        db_data = _load_offline_data() or {}
        
    from app.utils.fee_pdf import generate_monthly_fee_report
    try:
        generate_monthly_fee_report(month_key, db_data, pdf_path)
    except Exception as exc:
        try:
            current_app.logger.error("Auto PDF generation failed for month %s: %s", month_key, exc)
        except Exception:
            pass


@points_bp.route('/fee-update', methods=['POST'])
@csrf.exempt
@login_required
@limiter.limit("500 per hour")
@_ledger_write_guard
def fee_update():
    """Directly update a student fee record, commit to SQLite database, and auto-generate PDF report.

    Body JSON: { "studentId": <int>, "updates": { ... fee record fields ... } }
    """
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    req_data = request.get_json(force=True, silent=True) or {}
    raw_sid = req_data.get('studentId')
    updates = req_data.get('updates')

    if raw_sid is None or not isinstance(updates, dict):
        return jsonify({'success': False, 'error': 'Missing studentId or updates'}), 400

    try:
        sid = int(raw_sid)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid studentId'}), 400

    db_data = _load_offline_data()
    if not db_data:
        return jsonify({'success': False, 'error': 'Database not available'}), 503

    # Verify student exists
    students = db_data.get('students') or []
    student = next((s for s in students if s.get('id') == sid), None)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404

    # Sanitize and validate incoming updates
    sanitized_updates = _validate_and_sanitize_fee_updates(updates)

    fee_records = db_data.get('fee_records') or []
    now_iso = _server_now_iso()

    existing = next((r for r in fee_records if r.get('studentId') == sid), None)
    if existing:
        existing.update(sanitized_updates)
        existing['studentId'] = sid
        existing['updated_at'] = now_iso
    else:
        new_record = {'studentId': sid, 'created_at': now_iso, 'updated_at': now_iso}
        new_record.update(sanitized_updates)
        fee_records.append(new_record)

    # Keep student.fees in sync for convenience
    amount = sanitized_updates.get('amount')
    if amount is not None:
        student['fees'] = int(amount)

    db_data['fee_records'] = fee_records
    db_data['server_updated_at'] = now_iso
    _save_offline_data(db_data)

    # ACID Double-Write: Sync transaction history into local SQLite table
    student_name = student.get('name') or student.get('base_name')
    roll_number = student.get('roll')
    history = sanitized_updates.get('payment_history') or (existing.get('payment_history') if existing else [])
    if history:
        _sync_fee_records_to_sqlite(sid, roll_number, student_name, history)

    # Auto-generate Monthly PDF Report on payments/reversals
    try:
        if history:
            # Detect month of latest transaction to update the corresponding report
            sorted_hist = sorted(history, key=lambda h: str(h.get('date', '')))
            if sorted_hist:
                latest_date = sorted_hist[-1].get('date') or ''
                if len(latest_date) >= 7:
                    month_key = latest_date[:7]
                    _auto_generate_monthly_pdf_if_needed(month_key, db_data)
    except Exception:
        pass

    return jsonify({'success': True, 'studentId': sid})


@points_bp.route('/api/fees/reconstruct', methods=['POST'])
@csrf.exempt
@login_required
@_ledger_write_guard
def reconstruct_fees_from_sqlite():
    """Rebuild JSON fee records from the durable SQLite transactions ledger in case of system crash."""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    from app.models.fees import FeeTransaction
    db_data = _load_offline_data()
    if not db_data:
        return jsonify({'success': False, 'error': 'Database not available'}), 503
        
    try:
        txns = FeeTransaction.query.all()
    except Exception as exc:
        return jsonify({'success': False, 'error': f"Failed to read from SQLite database: {exc}"}), 500
        
    # Group transactions by student ID
    by_student = {}
    for t in txns:
        sid = t.student_id
        if sid not in by_student:
            by_student[sid] = []
        by_student[sid].append(t)
        
    fee_records = db_data.get('fee_records') or []
    fee_map = {r.get('studentId'): r for r in fee_records}
    
    rebuilt_records = []
    students = db_data.get('students') or []
    
    for s in students:
        sid = s.get('id')
        if not sid:
            continue
            
        student_txns = by_student.get(sid, [])
        if not student_txns and sid not in fee_map:
            continue
            
        # Compile payment history from transactions
        history = []
        for t in student_txns:
            history.append({
                'txn_id': t.txn_id,
                'date': t.date,
                'amount': t.amount,
                'mode': t.mode,
                'category': t.category,
                'ref_no': t.ref_no,
                'note': t.note,
                'recorded_by': t.recorded_by,
                'status': t.status,
                'is_reversal': t.is_reversal,
                'original_txn_id': t.original_txn_id
            })
            
        # Reconstruct base record
        existing = fee_map.get(sid) or {}
        
        # Sort history by date to find last paid date and next due date
        sorted_hist = sorted(history, key=lambda h: str(h.get('date', '')))
        last_paid = None
        for h in reversed(sorted_hist):
            if not h.get('is_reversal') and h.get('status') != 'reversed':
                # Skipped cycles (mode='skipped') are non-payments — don't treat
                # them as the last paid date.
                if str(h.get('mode', '')).lower() == 'skipped':
                    continue
                last_paid = h.get('date')
                break
                
        # Keep existing structural fields like start_date, last_paid_date, due_date
        rebuilt = {
            'studentId': sid,
            'amount': existing.get('amount') or s.get('fees', 0),
            'start_date': existing.get('start_date') or s.get('created_at', '')[:10] or '2026-01-01',
            'last_paid_date': last_paid or existing.get('last_paid_date'),
            'due_date': existing.get('due_date'),
            'pending_amount': existing.get('pending_amount') or 0.0,
            'payment_history': history,
            'created_at': existing.get('created_at') or _server_now_iso(),
            'updated_at': _server_now_iso()
        }
        rebuilt_records.append(rebuilt)
        
    db_data['fee_records'] = rebuilt_records
    db_data['server_updated_at'] = _server_now_iso()
    _save_offline_data(db_data)
    
    return jsonify({
        'success': True,
        'message': f"Successfully reconstructed fee ledger for {len(rebuilt_records)} students from SQLite DB"
    })


@points_bp.route('/api/fees/reports', methods=['GET'])
@login_required
def list_fee_reports():
    """List all generated monthly PDF fee reports from persistent storage."""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    from app.utils.data_paths import get_storage_root
    pdf_dir = os.path.join(get_storage_root(), 'fee_reports')
    os.makedirs(pdf_dir, exist_ok=True)
    
    files = []
    for f in os.listdir(pdf_dir):
        if f.startswith('fee_report_') and f.endswith('.pdf'):
            full_path = os.path.join(pdf_dir, f)
            st = os.stat(full_path)
            month = f.replace('fee_report_', '').replace('.pdf', '')
            files.append({
                'filename': f,
                'month': month,
                'size_kb': round(st.st_size / 1024, 1),
                'created_at': datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            })
            
    files.sort(key=lambda item: item['month'], reverse=True)
    return jsonify({'success': True, 'reports': files})


@points_bp.route('/api/fees/reports/generate', methods=['POST'])
@csrf.exempt
@login_required
def trigger_fee_report_generation():
    """Trigger manual generation or regeneration of a monthly PDF report."""
    import re
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    req_data = request.get_json(force=True, silent=True) or {}
    month = req_data.get('month', '').strip()
    if not month or not re.match(r'^\d{4}-\d{2}$', month):
        return jsonify({'success': False, 'error': 'Invalid month format (YYYY-MM)'}), 400
        
    db_data = _load_offline_data()
    if not db_data:
        return jsonify({'success': False, 'error': 'Database not available'}), 503
        
    try:
        _auto_generate_monthly_pdf_if_needed(month, db_data)
        return jsonify({'success': True, 'month': month, 'filename': f"fee_report_{month}.pdf"})
    except Exception as exc:
        return jsonify({'success': False, 'error': f"Generation failed: {exc}"}), 500


@points_bp.route('/api/fees/reports/download/<filename>', methods=['GET'])
@login_required
def download_fee_report(filename):
    """Download a specific generated monthly PDF report."""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    filename = os.path.basename(filename)
    if not filename.startswith('fee_report_') or not filename.endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Invalid filename'}), 400
        
    from app.utils.data_paths import get_storage_root
    pdf_dir = os.path.join(get_storage_root(), 'fee_reports')
    pdf_path = os.path.join(pdf_dir, filename)
    
    if not os.path.exists(pdf_path):
        return jsonify({'success': False, 'error': 'Report not found'}), 404
        
    return send_file(pdf_path, as_attachment=True, download_name=filename, mimetype='application/pdf')


@points_bp.route('/offline-restore-points', methods=['GET'])
@login_required
def offline_restore_points():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    storage_root = os.path.normpath(_storage_root_path())
    candidates = []
    roots = [
        ('live', _offline_data_path()),
        ('rolling', _offline_backup_dir()),
        ('hourly', _offline_hourly_backup_dir()),
        ('startup', _offline_startup_restore_dir()),
        ('storage', storage_root),
        ('legacy-instance', current_app.instance_path)
    ]

    seen = set()
    for source, root in roots:
        if source == 'live':
            path = root
            if os.path.isfile(path):
                rel = os.path.relpath(path, storage_root)
                key = rel.replace('\\', '/')
                if key not in seen:
                    seen.add(key)
                    candidates.append((source, path, key))
            continue
        if not os.path.isdir(root):
            continue
        for name in os.listdir(root):
            if not name.endswith('.json'):
                continue
            if source == 'legacy-instance' and not name.startswith('offline_scoreboard_data'):
                continue
            path = os.path.join(root, name)
            if not os.path.isfile(path):
                continue
            if source == 'legacy-instance':
                rel = f"legacy/{name}"
            else:
                rel = os.path.relpath(path, storage_root)
            key = rel.replace('\\', '/')
            if key in seen:
                continue
            seen.add(key)
            candidates.append((source, path, key))

    meta = _load_restore_points_meta()
    items = []
    for source, path, key in candidates:
        try:
            stat = os.stat(path)
            key_meta = meta.get(key, {}) if isinstance(meta.get(key), dict) else {}
            items.append({
                'id': key,
                'source': source,
                'name': os.path.basename(path),
                'path': key,
                'modified_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'size': stat.st_size,
                'locked': bool(key_meta.get('locked')),
                'label': str(key_meta.get('label') or '').strip()
            })
        except Exception:
            continue

    items.sort(key=lambda item: item.get('modified_at', ''), reverse=True)
    return jsonify({'success': True, 'items': items})


@points_bp.route('/offline-restore-point-lock', methods=['POST'])
@csrf.exempt  # JSON API endpoint — secured by @login_required + admin role check
@login_required
def offline_restore_point_lock():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    payload = request.get_json(silent=True) or {}
    restore_id = str(payload.get('id') or '').strip().replace('\\', '/')
    if not restore_id or '..' in restore_id:
        return jsonify({'success': False, 'error': 'Invalid restore id'}), 400
    lock_state = bool(payload.get('locked'))
    label = str(payload.get('label') or '').strip()
    storage_root = os.path.normpath(_storage_root_path())
    if restore_id.startswith('legacy/'):
        source_path = os.path.normpath(os.path.join(current_app.instance_path, restore_id.split('/', 1)[1]))
        allowed_root = os.path.normpath(current_app.instance_path)
    else:
        source_path = os.path.normpath(os.path.join(storage_root, restore_id))
        allowed_root = storage_root
    if not source_path.startswith(allowed_root):
        return jsonify({'success': False, 'error': 'Invalid restore path'}), 400
    if not os.path.isfile(source_path):
        return jsonify({'success': False, 'error': 'Restore file not found'}), 404

    meta = _load_restore_points_meta()
    entry = meta.get(restore_id, {}) if isinstance(meta.get(restore_id), dict) else {}
    entry['locked'] = lock_state
    entry['label'] = label[:80]
    entry['updated_at'] = _server_now_iso()
    meta[restore_id] = entry
    _save_restore_points_meta(meta)
    return jsonify({'success': True, 'id': restore_id, 'locked': lock_state, 'label': entry['label']})


@points_bp.route('/offline-restore', methods=['POST'])
@csrf.exempt  # JSON API endpoint — secured by @login_required + admin role check
@login_required
@_ledger_write_guard
def offline_restore():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    restore_id = str(payload.get('id') or '').strip().replace('\\', '/')
    if not restore_id:
        return jsonify({'success': False, 'error': 'Missing restore id'}), 400
    if '..' in restore_id:
        return jsonify({'success': False, 'error': 'Invalid restore id'}), 400

    storage_root = os.path.normpath(_storage_root_path())
    if restore_id.startswith('legacy/'):
        source_path = os.path.normpath(os.path.join(current_app.instance_path, restore_id.split('/', 1)[1]))
        allowed_root = os.path.normpath(current_app.instance_path)
    else:
        source_path = os.path.normpath(os.path.join(storage_root, restore_id))
        allowed_root = storage_root
    if not source_path.startswith(allowed_root):
        return jsonify({'success': False, 'error': 'Invalid restore path'}), 400
    if not os.path.isfile(source_path):
        return jsonify({'success': False, 'error': 'Restore file not found'}), 404

    try:
        with open(source_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        return jsonify({'success': False, 'error': 'Restore file is not valid JSON'}), 400
    if not isinstance(data, dict):
        return jsonify({'success': False, 'error': 'Restore payload invalid'}), 400

    # Always create a backup of current live state before restore.
    current = _load_offline_data()
    if isinstance(current, dict):
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safety = os.path.join(storage_root, f'offline_scoreboard_data.pre_ui_restore_{stamp}.json')
        _atomic_write_json(safety, current)

    data['server_updated_at'] = _server_now_iso()
    data['updated_at'] = data.get('updated_at') or data['server_updated_at']
    _save_offline_data(data)
    _broadcast_sync_event(data['server_updated_at'], source='admin-restore')
    _forward_offline_data_to_peers_async(data, [])

    return jsonify({
        'success': True,
        'updated_at': data['server_updated_at'],
        'students': len(data.get('students', []) or []),
        'scores': len(data.get('scores', []) or [])
    })


@points_bp.route('/manifest.webmanifest')
@login_required
def offline_manifest():
    return send_file('static/offline_manifest.webmanifest', mimetype='application/manifest+json')


@points_bp.route('/sw.js')
def offline_sw():
    """Serve the service worker with a dynamically injected cache version.

    The cache name is derived from offline_scoreboard.html's mtime so it
    auto-bumps on every deployment — no manual version bump required.
    The SW file itself is served with no-store so browsers always re-fetch it
    and pick up updates immediately (within the SW 24-hour update window).
    """
    sw_path = os.path.join(current_app.root_path, 'static', 'offline_sw.js')
    html_path = os.path.join(current_app.root_path, 'static', 'offline_scoreboard.html')

    try:
        version = int(os.path.getmtime(html_path))
    except OSError:
        version = 0

    with open(sw_path, 'r', encoding='utf-8') as f:
        sw_content = f.read()

    # Replace the placeholder cache name with the mtime-derived version.
    sw_content = re.sub(
        r"const CACHE_NAME = 'ea-offline-v\d+';.*",
        f"const CACHE_NAME = 'ea-offline-v{version}';",
        sw_content
    )

    response = Response(sw_content, mimetype='application/javascript')
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


def _get_request_user():
    """
    Resolve the user making the request from either the Flask-Login session
    or custom headers: X-EA-Login-ID and X-EA-Login-Code.
    Returns: (user_obj, role, login_id) or (None, None, None)
    """
    if current_user.is_authenticated:
        return current_user, current_user.role, current_user.login_id
    
    login_id = request.headers.get('X-EA-Login-ID')
    login_code = request.headers.get('X-EA-Login-Code')
    
    if login_id and login_code:
        user = User.query.filter_by(login_id=login_id).first()
        if user:
            # Check login code validity
            if user.login_code:
                if user.login_code_expires_at and datetime.utcnow() > user.login_code_expires_at:
                    return None, None, None
                if str(login_code).strip().upper() == str(user.login_code).strip().upper():
                    return user, user.role, user.login_id
            else:
                if user.check_password(login_code):
                    return user, user.role, user.login_id
    return None, None, None


@points_bp.route('/session')
@csrf.exempt
def scoreboard_session():
    user, role, login_id = _get_request_user()
    if not user:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
    role = (role or 'student').strip().lower()
    if role not in {'admin', 'teacher', 'student'}:
        role = 'student'

    response_data = {
        'login_id': login_id,
        'role': role,
        'server_timezone': _get_server_timezone(),
        'server_time': _server_now_iso(),
        'fees_enabled': _fees_module_enabled(),
    }

    # For students, add their roll number to enable personalized filtering
    if role == 'student':
        response_data['student_roll'] = login_id

    return jsonify(response_data)


# ─── Device Monitoring ────────────────────────────────────────────────────────
_DEVICE_LOG_MAX = 2000


@points_bp.route('/device-checkin', methods=['POST', 'GET'])
@csrf.exempt  # JSON API endpoint — secured by @login_required + role check
@login_required
def device_checkin():
    """POST: record a device check-in. GET (admin-only): return full log."""
    log_path = _device_log_path()

    if request.method == 'GET':
        if current_user.role != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                log = json.load(f)
            if not isinstance(log, list):
                log = []
        except Exception:
            log = []
        return jsonify({'log': log, 'count': len(log)})

    # POST — record check-in from any logged-in user
    data = request.get_json(silent=True) or {}

    def _s(val, maxlen=100):
        return str(val or '')[:maxlen].strip()

    entry = {
        'ts': datetime.now(timezone.utc).isoformat(),
        'login_id': current_user.login_id,
        'role': current_user.role or 'student',
        'ip': (_s(request.headers.get('X-Forwarded-For', '') or request.remote_addr or '', 90)
               .split(',')[0].strip())[:45],
        'device_id': _s(data.get('device_id'), 64),
        'device_name': _s(data.get('device_name'), 100),
        'brand': _s(data.get('brand'), 50),
        'os': _s(data.get('os'), 80),
        'browser': _s(data.get('browser'), 80),
        'screen': _s(data.get('screen'), 20),
        'event': _s(data.get('event', 'login'), 20),
    }

    try:
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                log = json.load(f)
            if not isinstance(log, list):
                log = []
        except Exception:
            log = []

        log.append(entry)
        if len(log) > _DEVICE_LOG_MAX:
            log = log[-_DEVICE_LOG_MAX:]

        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump(log, f)
    except Exception as e:
        current_app.logger.error('device_checkin write error: %s', e)

    # Also upsert DeviceSession row so the Control Panel reflects live status
    try:
        dev_id = _s(data.get('device_id'), 128) or None
        q = DeviceSession.query.filter_by(user_id=current_user.id, login_id=current_user.login_id)
        if dev_id:
            q = q.filter_by(device_id=dev_id)
        row = q.order_by(DeviceSession.last_seen.desc(), DeviceSession.id.desc()).first()
        is_new = row is None
        if is_new:
            row = DeviceSession(
                user_id=current_user.id,
                login_id=current_user.login_id,
                role=current_user.role or 'student',
                device_id=dev_id,
            )
            db.session.add(row)
        row.device_id = dev_id or row.device_id
        _dn = _s(data.get('device_name'), 120)
        _br = _s(data.get('brand'), 50)
        # Combine brand + model for a readable device_name (e.g. "Samsung Galaxy S24")
        if _dn and _br and _br.lower() not in _dn.lower():
            row.device_name = f"{_br} {_dn}"[:120]
        elif _dn:
            row.device_name = _dn
        row.os = _s(data.get('os'), 80) or row.os
        row.browser = _s(data.get('browser'), 80) or row.browser
        row.ip = entry['ip'] or row.ip
        row.last_seen = datetime.utcnow()
        row.status = 'online'
        # Set login_at only on new sessions or when transitioning from offline→online
        if is_new or str(row.status or '').lower() != 'online':
            row.login_at = row.last_seen
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'ok': True})


@points_bp.route('/online-sessions', methods=['GET'])
@csrf.exempt
@login_required
def online_sessions():
    """Return currently-online sessions with login time and live duration.
    Admin-only. Sessions not seen in 15 min are marked offline first."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        cutoff = datetime.utcnow() - timedelta(minutes=15)
        stale = DeviceSession.query.filter(
            DeviceSession.status == 'online',
            DeviceSession.last_seen < cutoff
        ).all()
        for s in stale:
            s.status = 'offline'
        if stale:
            db.session.commit()
    except Exception:
        db.session.rollback()

    now = datetime.utcnow()
    rows = (
        DeviceSession.query
        .filter_by(status='online')
        .order_by(DeviceSession.login_at.desc(), DeviceSession.last_seen.desc())
        .all()
    )
    items = []
    for s in rows:
        login_at = s.login_at or s.last_seen
        duration_sec = int((now - login_at).total_seconds()) if login_at else 0
        items.append({
            'login_id': s.login_id,
            'role': s.role,
            'device_name': s.device_name or '',
            'os': s.os or '',
            'browser': s.browser or '',
            'ip': s.ip or '',
            'login_at': login_at.isoformat() if login_at else '',
            'last_seen': s.last_seen.isoformat() if s.last_seen else '',
            'duration_sec': max(0, duration_sec),
        })
    return jsonify({'success': True, 'count': len(items), 'sessions': items})


@points_bp.route('/device-log/clear', methods=['POST'])
@csrf.exempt  # JSON API endpoint — secured by @login_required + role check
@login_required
def device_log_clear():
    """Admin: wipe the device connection log."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    log_path = _device_log_path()
    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump([], f)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True})


@points_bp.route('/')
@login_required
def scoreboard_home():
    """Main scoreboard page"""
    return render_template('scoreboard/index.html')


@points_bp.route('/party-data', methods=['GET', 'POST'])
@login_required
@_ledger_write_guard
def party_data():
    """Get or update party system data"""
    data = _load_politics_data()
    if request.method == 'POST':
        if current_user.role != 'admin':
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        payload = request.get_json()
        if not isinstance(payload, dict):
            return jsonify({'success': False, 'error': 'Invalid request data'}), 400

        parties = payload.get('parties', [])
        if not isinstance(parties, list):
            return jsonify({'success': False, 'error': 'Parties must be a list'}), 400

        # Validate each party entry
        for idx, party in enumerate(parties):
            if not isinstance(party, dict):
                return jsonify({'success': False, 'error': f'Party at index {idx} is invalid'}), 400

            # Required fields validation
            if 'id' not in party or not isinstance(party['id'], int):
                return jsonify({'success': False, 'error': f'Party at index {idx} missing valid id'}), 400
            if 'code' not in party or not isinstance(party['code'], str) or len(party['code']) > 10:
                return jsonify({'success': False, 'error': f'Party at index {idx} has invalid code'}), 400
            if 'name' not in party or not isinstance(party['name'], str) or len(party['name']) > 100:
                return jsonify({'success': False, 'error': f'Party at index {idx} has invalid name'}), 400
            if 'power' not in party or not isinstance(party['power'], int) or party['power'] < 0 or party['power'] > 1000:
                return jsonify({'success': False, 'error': f'Party at index {idx} has invalid power (0-1000)'}), 400

        data['parties'] = parties
        saved = _save_politics_data(data)
        return jsonify({'success': True, 'parties': saved['parties']})
    return jsonify({'success': True, 'parties': data['parties']})


@points_bp.route('/leadership-data', methods=['GET', 'POST'])
@login_required
@_ledger_write_guard
def leadership_data():
    """Get or update leadership posts"""
    data = _load_politics_data()
    if request.method == 'POST':
        if current_user.role != 'admin':
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        payload = request.get_json()
        if not isinstance(payload, dict):
            return jsonify({'success': False, 'error': 'Invalid request data'}), 400

        leadership = payload.get('leadership', [])
        if not isinstance(leadership, list):
            return jsonify({'success': False, 'error': 'Leadership must be a list'}), 400

        # Validate each leadership entry
        valid_statuses = {'active', 'suspended', 'vacant'}
        for idx, post in enumerate(leadership):
            if not isinstance(post, dict):
                return jsonify({'success': False, 'error': f'Leadership post at index {idx} is invalid'}), 400

            # Required fields validation
            if 'id' not in post or not isinstance(post['id'], int):
                return jsonify({'success': False, 'error': f'Leadership post at index {idx} missing valid id'}), 400
            if 'post' not in post or not isinstance(post['post'], str) or len(post['post']) > 100:
                return jsonify({'success': False, 'error': f'Leadership post at index {idx} has invalid post name'}), 400

            # Optional fields validation
            if 'holder' in post and post['holder'] and not isinstance(post['holder'], str):
                return jsonify({'success': False, 'error': f'Leadership post at index {idx} has invalid holder'}), 400
            if 'status' in post and post['status'] not in valid_statuses:
                return jsonify({'success': False, 'error': f'Leadership post at index {idx} has invalid status (must be: {", ".join(valid_statuses)})'}), 400
            if 'vetoQuota' in post:
                veto_quota = post['vetoQuota']
                if not isinstance(veto_quota, int) or veto_quota < 0 or veto_quota > 20:
                    return jsonify({'success': False, 'error': f'Leadership post at index {idx} has invalid vetoQuota (0-20)'}), 400

        data['leadership'] = leadership
        saved = _save_politics_data(data)
        return jsonify({'success': True, 'leadership': saved['leadership']})
    return jsonify({'success': True, 'leadership': data['leadership']})


@points_bp.route('/election-results/<post>')
@login_required
def election_results(post):
    """Get election results for a specific post."""
    data = _load_offline_data() or {}
    student_votes = data.get('election_individual_votes', [])
    teacher_votes = data.get('election_teacher_votes', [])

    # Filter votes for the specific post
    student_votes = [v for v in student_votes if v.get('post') == post]
    teacher_votes = [v for v in teacher_votes if v.get('post') == post]

    results = _calculate_election_results(data, student_votes, teacher_votes)
    return jsonify({'success': True, 'results': results})



@points_bp.route('/data')
@login_required
def get_scoreboard_data():
    """Get scoreboard data with filters"""
    try:
        # Get filters from request
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        class_filter = request.args.get('class', 'All')
        group_filter = request.args.get('group', 'All')
        search = request.args.get('search', '').lower()
        month_key = request.args.get('month')
        
        # Parse dates
        if not date_from or not date_to:
            today = date.today()
            date_from = today.replace(day=1)
            date_to = today
        else:
            date_from = datetime.fromisoformat(date_from).date()
            date_to = datetime.fromisoformat(date_to).date()
        
        # Query students - filter by active users only
        query = StudentProfile.query.join(User).filter(User.is_active == True).all()
        
        # Apply filters
        filtered_students = []
        for student in query:
            # Class filter
            if class_filter != 'All' and str(student.class_name) != str(class_filter):
                continue
            # Group filter
            if group_filter != 'All' and str(student.group) != str(group_filter):
                continue
            # Search filter
            if search and search not in student.full_name.lower() and search not in student.roll_number.lower():
                continue
            filtered_students.append(student)

        student_ids = [student.id for student in filtered_students]
        points_by_student = {}
        if student_ids:
            points_records = StudentPoints.query.filter(
                StudentPoints.student_id.in_(student_ids),
                StudentPoints.date_recorded >= date_from,
                StudentPoints.date_recorded <= date_to
            ).all()
            for record in points_records:
                points_by_student.setdefault(record.student_id, []).append(record)

        students_data = []
        for student in filtered_students:
            points_records = points_by_student.get(student.id, [])
            total_points = sum(p.points for p in points_records)
            total_stars = sum(p.stars for p in points_records)
            total_vetos = sum(p.vetos for p in points_records)
            
            # Create daily breakdown
            daily_data = {}
            for record in points_records:
                date_key = record.date_recorded.isoformat()
                daily_data[date_key] = {
                    'points': record.points,
                    'stars': record.stars,
                    'vetos': record.vetos,
                    'notes': record.notes
                }
            
            profile_data = student.profile_data or {}
            students_data.append({
                'id': student.id,
                'roll_number': student.roll_number,
                'full_name': student.full_name,
                'class': student.class_name,
                'group': student.group,
                'fees': profile_data.get('fees'),
                'vote_power': profile_data.get('vote_power'),
                'sheet_total_score': profile_data.get('total_score'),
                'sheet_rank': profile_data.get('rank'),
                'total_points': total_points,
                'total_stars': total_stars,
                'total_vetos': total_vetos,
                'daily_data': daily_data,
                'net_score': total_points + (total_stars * 10) - (total_vetos * 5)
            })
        
        # Sort by net score descending
        students_data.sort(key=lambda x: x['net_score'], reverse=True)
        
        # Add ranks
        for idx, student in enumerate(students_data, 1):
            student['rank'] = idx
        
        # Get date columns for table headers
        date_range = []
        current = date_from
        while current <= date_to:
            date_range.append(current.isoformat())
            current += timedelta(days=1)
        
        return jsonify({
            'success': True,
            'students': students_data,
            'date_range': date_range,
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat()
        })
    except Exception as e:
        current_app.logger.error(f"Error in get_scoreboard_data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/proposals', methods=['GET', 'POST'])
@csrf.exempt  # JSON API — secured by @login_required
@login_required
def proposals_data():
    data = _load_offline_data() or {}
    proposals = list(data.get('proposals') or [])
    votes = list(data.get('proposal_votes') or [])
    messages = list(data.get('proposal_messages') or [])

    if request.method == 'GET':
        allowed_months = _allowed_months_for_user(data, current_user)
        visible = []
        for p in proposals:
            if not isinstance(p, dict):
                continue
            if allowed_months is not None:
                pm = _month_key_from_date_like(p.get('month') or p.get('created_at') or p.get('open_at'))
                if pm not in allowed_months:
                    continue
            if _is_proposal_stakeholder(data, p, current_user):
                pid = _parse_int_safe(p.get('id'), 0)
                vote_counts = {'support': 0, 'oppose': 0, 'abstain': 0}
                for v in votes:
                    if _parse_int_safe(v.get('proposal_id'), 0) != pid:
                        continue
                    ch = str(v.get('choice') or '').strip().lower()
                    if ch in vote_counts:
                        vote_counts[ch] += 1
                msg_count = sum(1 for m in messages if _parse_int_safe(m.get('proposal_id'), 0) == pid)
                row = dict(p)
                row['vote_counts'] = vote_counts
                row['message_count'] = msg_count
                visible.append(row)
        visible.sort(key=lambda x: _parse_sync_stamp(x.get('created_at')), reverse=True)
        return jsonify({'success': True, 'proposals': visible})

    # POST: create proposal (admin/teacher)
    role = str(current_user.role or '').strip().lower()
    if role not in ('admin', 'teacher'):
        return jsonify({'success': False, 'error': 'Only Admin/Teacher can create proposals'}), 403
    if not _is_month_allowed_for_user(data, current_user, _server_now_iso()[:7]):
        return jsonify({'success': False, 'error': 'No month access to create proposal'}), 403
    payload = request.get_json(silent=True) or {}
    title = str(payload.get('title') or '').strip()[:200]
    body = str(payload.get('body') or '').strip()[:4000]
    scope = str(payload.get('scope') or 'student_council').strip().lower()
    if scope not in ('student_council', 'all_students'):
        scope = 'student_council'
    if not title or not body:
        return jsonify({'success': False, 'error': 'title and body are required'}), 400

    now_iso = _server_now_iso()
    next_id = max([_parse_int_safe(p.get('id'), 0) for p in proposals if isinstance(p, dict)] + [0]) + 1
    proposal = {
        'id': next_id,
        'title': title,
        'body': body,
        'scope': scope,
        'month': _server_now_iso()[:7],
        'status': 'open',
        'created_by': str(current_user.login_id or ''),
        'created_by_role': role,
        'open_at': now_iso,
        'created_at': now_iso,
        'updated_at': now_iso,
    }
    proposals.append(proposal)
    data['proposals'] = proposals
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='proposal')
    _forward_offline_data_to_peers_async(data, [])
    return jsonify({'success': True, 'proposal': proposal, 'updated_at': now_iso})


@points_bp.route('/proposals/<int:proposal_id>/vote', methods=['POST'])
@csrf.exempt  # JSON API — secured by @login_required
@login_required
def proposal_vote(proposal_id):
    data = _load_offline_data() or {}
    proposals = list(data.get('proposals') or [])
    proposal = next((p for p in proposals if _parse_int_safe(p.get('id'), 0) == proposal_id), None)
    if not proposal:
        return jsonify({'success': False, 'error': 'Proposal not found'}), 404
    if str(proposal.get('status') or 'open').strip().lower() != 'open':
        return jsonify({'success': False, 'error': 'Proposal is closed'}), 409
    if not _is_proposal_stakeholder(data, proposal, current_user):
        return jsonify({'success': False, 'error': 'Not a stakeholder for this proposal'}), 403
    if not _is_month_allowed_for_user(data, current_user, proposal.get('month') or proposal.get('created_at')):
        return jsonify({'success': False, 'error': 'No month access for this proposal'}), 403

    payload = request.get_json(silent=True) or {}
    choice = str(payload.get('choice') or '').strip().lower()
    if choice not in ('support', 'oppose', 'abstain'):
        return jsonify({'success': False, 'error': 'Invalid vote choice'}), 400

    votes = list(data.get('proposal_votes') or [])
    now_iso = _server_now_iso()
    voter_login = str(current_user.login_id or '').strip()
    replaced = False
    for row in votes:
        if not isinstance(row, dict):
            continue
        if _parse_int_safe(row.get('proposal_id'), 0) != proposal_id:
            continue
        if str(row.get('voter_login_id') or '').strip() != voter_login:
            continue
        row['choice'] = choice
        row['updated_at'] = now_iso
        replaced = True
        break
    if not replaced:
        votes.append({
            'id': int(time.time() * 1000),
            'proposal_id': proposal_id,
            'voter_login_id': voter_login,
            'voter_role': str(current_user.role or ''),
            'choice': choice,
            'created_at': now_iso,
            'updated_at': now_iso,
        })
    data['proposal_votes'] = votes
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='proposal-vote')
    _forward_offline_data_to_peers_async(data, [])
    return jsonify({'success': True, 'updated_at': now_iso})


@points_bp.route('/proposals/<int:proposal_id>/messages', methods=['GET', 'POST'])
@csrf.exempt  # JSON API — secured by @login_required
@login_required
def proposal_messages(proposal_id):
    data = _load_offline_data() or {}
    proposals = list(data.get('proposals') or [])
    proposal = next((p for p in proposals if _parse_int_safe(p.get('id'), 0) == proposal_id), None)
    if not proposal:
        return jsonify({'success': False, 'error': 'Proposal not found'}), 404
    if not _is_proposal_stakeholder(data, proposal, current_user):
        return jsonify({'success': False, 'error': 'Not a stakeholder for this proposal'}), 403
    if not _is_month_allowed_for_user(data, current_user, proposal.get('month') or proposal.get('created_at')):
        return jsonify({'success': False, 'error': 'No month access for this proposal'}), 403

    messages = list(data.get('proposal_messages') or [])
    if request.method == 'GET':
        rows = [m for m in messages if _parse_int_safe(m.get('proposal_id'), 0) == proposal_id]
        rows.sort(key=lambda x: _parse_sync_stamp(x.get('created_at')))
        return jsonify({'success': True, 'messages': rows[-500:]})

    payload = request.get_json(silent=True) or {}
    text = str(payload.get('message') or '').strip()
    if not text:
        return jsonify({'success': False, 'error': 'Message is required'}), 400
    if len(text) > 2000:
        return jsonify({'success': False, 'error': 'Message too long'}), 400
    now_iso = _server_now_iso()
    msg_row = {
        'id': int(time.time() * 1000),
        'proposal_id': proposal_id,
        'login_id': str(current_user.login_id or ''),
        'role': str(current_user.role or ''),
        'message': text,
        'created_at': now_iso,
    }
    messages.append(msg_row)
    data['proposal_messages'] = messages[-20000:]
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='proposal-chat')
    _forward_offline_data_to_peers_async(data, [])
    return jsonify({'success': True, 'message': msg_row, 'updated_at': now_iso})


@points_bp.route('/allowed-months', methods=['GET'])
@login_required
def allowed_months():
    data = _load_offline_data() or {}
    allowed = _allowed_months_for_user(data, current_user)
    if allowed is None:
        # Unrestricted users receive known month keys for UI filters.
        month_keys = set()
        for key in (data.get('month_students') or {}).keys():
            mk = _month_key_from_date_like(key)
            if mk:
                month_keys.add(mk)
        for key in (data.get('month_roster_profiles') or {}).keys():
            mk = _month_key_from_date_like(key)
            if mk:
                month_keys.add(mk)
        if not month_keys:
            month_keys.add(_server_now_iso()[:7])
        allowed = month_keys
    return jsonify({'success': True, 'months': sorted(list(allowed))})


@points_bp.route('/admin/control-panel-data', methods=['GET'])
@login_required
def admin_control_panel_data():
    if str(current_user.role or '').strip().lower() != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    # Mark stale device sessions as offline (not seen in 15 min)
    try:
        cutoff = datetime.utcnow() - timedelta(minutes=15)
        stale = DeviceSession.query.filter(
            DeviceSession.status == 'online',
            DeviceSession.last_seen < cutoff
        ).all()
        for s in stale:
            s.status = 'offline'
        if stale:
            db.session.commit()
    except Exception:
        db.session.rollback()

    users = User.query.order_by(User.role.asc(), User.login_id.asc()).all()
    user_rows = []
    for user in users:
        last_action = (
            AccountAction.query
            .filter_by(target_user_id=user.id)
            .order_by(AccountAction.created_at.desc(), AccountAction.id.desc())
            .first()
        )
        status = 'active' if bool(user.is_active) else 'inactive'
        if last_action and str(last_action.action or '').strip().lower() in ('hold', 'delete'):
            status = str(last_action.action or '').strip().lower()
        user_rows.append({
            'id': user.id,
            'login_id': user.login_id,
            'role': user.role,
            'status': status,
            'is_active': bool(user.is_active),
            'created_at': user.created_at.isoformat() if user.created_at else '',
            'last_login': user.last_login.isoformat() if user.last_login else '',
            'last_login_ip': str(user.last_login_ip or ''),
        })

    sessions = (
        DeviceSession.query
        .order_by(DeviceSession.last_seen.desc(), DeviceSession.id.desc())
        .limit(1000)
        .all()
    )
    session_rows = [{
        'id': s.id,
        'user_id': s.user_id,
        'login_id': s.login_id,
        'role': s.role,
        'device_id': s.device_id,
        'device_name': s.device_name,
        'os': s.os,
        'browser': s.browser,
        'ip': s.ip,
        'login_at': s.login_at.isoformat() if s.login_at else '',
        'last_seen': s.last_seen.isoformat() if s.last_seen else '',
        'status': s.status,
    } for s in sessions]

    windows = (
        UserAccessWindow.query
        .order_by(UserAccessWindow.updated_at.desc(), UserAccessWindow.id.desc())
        .all()
    )
    window_rows = [{
        'id': w.id,
        'user_id': w.user_id,
        'month_from': str(w.month_from or ''),
        'month_to': str(w.month_to or ''),
        'set_by': w.set_by,
        'updated_at': w.updated_at.isoformat() if w.updated_at else '',
    } for w in windows]

    join_codes = (
        JoinCode.query
        .order_by(JoinCode.created_at.desc(), JoinCode.id.desc())
        .limit(20)
        .all()
    )
    code_rows = [{
        'id': c.id,
        'active': bool(c.active),
        'expires_at': c.expires_at.isoformat() if c.expires_at else '',
        'created_by': c.created_by,
        'created_at': c.created_at.isoformat() if c.created_at else '',
    } for c in join_codes]

    # Public site (Cloudflare) login credentials — admin-managed, published on
    # Force Publish. Only metadata is exposed here (never salts/hashes).
    pub_creds = []
    try:
        pub_rows = (
            PublicSiteCredential.query
            .order_by(PublicSiteCredential.roll.asc())
            .all()
        )
        admin_id_map = {u.id: u.login_id for u in User.query.all()}
        for r in pub_rows:
            pub_creds.append({
                'id': r.id,
                'roll': str(r.roll or '').upper(),
                'active': bool(r.active),
                'set_by_login': admin_id_map.get(r.set_by, '') if r.set_by else '',
                'set_at': r.set_at.isoformat() if r.set_at else '',
            })
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Failed to load public site credentials')

    return jsonify({
        'success': True,
        'users': user_rows,
        'device_sessions': session_rows,
        'access_windows': window_rows,
        'join_codes': code_rows,
        'public_credentials': pub_creds,
        'student_rolls': _student_rolls_from_ledger(),
    })


@points_bp.route('/admin/join-code', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_rotate_join_code():
    if str(current_user.role or '').strip().lower() != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    code = str(payload.get('code') or '').strip()
    if len(code) < 4:
        return jsonify({'success': False, 'error': 'Join code must be at least 4 characters'}), 400
    expires_at = None
    expires_raw = str(payload.get('expires_at') or '').strip()
    if expires_raw:
        try:
            expires_at = datetime.fromisoformat(expires_raw.replace('Z', '+00:00'))
        except Exception:
            return jsonify({'success': False, 'error': 'Invalid expires_at format'}), 400

    try:
        JoinCode.query.update({'active': False})
        row = JoinCode(
            code_hash=generate_password_hash(code),
            active=True,
            expires_at=expires_at,
            created_by=_parse_int_safe(current_user.id, 0) or None,
            created_at=datetime.utcnow()
        )
        db.session.add(row)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    return jsonify({'success': True, 'message': 'Join code rotated'})


@points_bp.route('/admin/account-action', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_account_action():
    if str(current_user.role or '').strip().lower() != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    login_id = str(payload.get('login_id') or '').strip()
    action = str(payload.get('action') or '').strip().lower()
    reason = str(payload.get('reason') or '').strip()[:500]
    if action not in ('hold', 'resume', 'delete'):
        return jsonify({'success': False, 'error': 'Invalid action'}), 400
    if not login_id:
        return jsonify({'success': False, 'error': 'login_id required'}), 400

    user = User.query.filter_by(login_id=login_id).first()
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404
    if user.login_id == 'Admin' and action in ('hold', 'delete'):
        return jsonify({'success': False, 'error': 'Cannot disable Admin account'}), 400

    try:
        if action == 'resume':
            user.is_active = True
        else:
            user.is_active = False
        db.session.add(AccountAction(
            target_user_id=user.id,
            action=action,
            reason=reason,
            by_user_id=_parse_int_safe(current_user.id, 0),
            created_at=datetime.utcnow()
        ))
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500

    return jsonify({'success': True, 'message': f'Action {action} applied for {login_id}'})


@points_bp.route('/admin/access-window', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_set_access_window():
    if str(current_user.role or '').strip().lower() != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    login_id = str(payload.get('login_id') or '').strip()
    month_from = _month_key_from_date_like(payload.get('month_from'))
    month_to = _month_key_from_date_like(payload.get('month_to'))
    if not login_id or not month_from or not month_to:
        return jsonify({'success': False, 'error': 'login_id, month_from, month_to are required'}), 400

    user = User.query.filter_by(login_id=login_id).first()
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404
    if str(user.role or '').strip().lower() != 'teacher':
        return jsonify({'success': False, 'error': 'Access windows can be configured for teachers only'}), 400

    try:
        UserAccessWindow.query.filter_by(user_id=user.id).delete()
        db.session.add(UserAccessWindow(
            user_id=user.id,
            month_from=month_from,
            month_to=month_to,
            set_by=_parse_int_safe(current_user.id, 0),
            updated_at=datetime.utcnow()
        ))
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    return jsonify({'success': True, 'message': f'Window saved for {login_id}'})


@points_bp.route('/admin/reset-login-codes', methods=['POST'])
@csrf.exempt
@login_required
def admin_reset_login_codes():
    if str(current_user.role or '').strip().lower() != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    import random
    def _generate_random_code(length=6):
        chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
        return ''.join(random.choice(chars) for _ in range(length))

    def _get_current_month_end():
        now = datetime.utcnow()
        if now.month == 12:
            next_month = datetime(now.year + 1, 1, 1)
        else:
            next_month = datetime(now.year, now.month + 1, 1)
        return next_month - timedelta(seconds=1)

    try:
        users = User.query.filter(User.login_id != 'Admin').all()
        expires_at = _get_current_month_end()
        generated = []

        for user in users:
            code = _generate_random_code()
            user.login_code = code
            user.login_code_expires_at = expires_at
            generated.append({
                'login_id': user.login_id,
                'role': user.role,
                'code': code
            })
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully regenerated codes for {len(users)} users expiring at {expires_at.isoformat()}',
            'expires_at': expires_at.isoformat(),
            'codes': generated
        })
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500


def _require_admin():
    """Returns (ok, error_response). ok=True means caller is admin."""
    if str(current_user.role or '').strip().lower() != 'admin':
        return False, (jsonify({'success': False, 'error': 'Unauthorized'}), 403)
    return True, None


def _upsert_public_credential(roll, password, actor_id):
    """Create or update a PublicSiteCredential row with a fresh salt+hash."""
    roll = _normalize_roll(roll)
    if not roll:
        raise ValueError('roll is required')
    if not str(password or ''):
        raise ValueError('password is required')
    salt, digest = _hash_public_credential(password)
    row = PublicSiteCredential.query.filter_by(roll=roll).first()
    if row:
        row.salt = salt
        row.hash = digest
        row.active = True
        row.set_by = actor_id
        row.set_at = datetime.utcnow()
    else:
        row = PublicSiteCredential(
            roll=roll, salt=salt, hash=digest, active=True,
            set_by=actor_id, set_at=datetime.utcnow(),
        )
        db.session.add(row)
    return row


@points_bp.route('/admin/public-credential', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_set_public_credential():
    ok, err = _require_admin()
    if not ok:
        return err
    payload = request.get_json(silent=True) or {}
    roll = _normalize_roll(payload.get('roll'))
    password = str(payload.get('password') or '')
    if not roll:
        return jsonify({'success': False, 'error': 'roll is required'}), 400
    if not password:
        return jsonify({'success': False, 'error': 'password is required'}), 400
    if len(password) < 4:
        return jsonify({'success': False, 'error': 'password must be at least 4 characters'}), 400
    try:
        _upsert_public_credential(roll, password, _parse_int_safe(current_user.id, 0) or None)
        db.session.commit()
    except ValueError as ve:
        return jsonify({'success': False, 'error': str(ve)}), 400
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('Failed to set public credential')
        return jsonify({'success': False, 'error': str(exc)}), 500
    return jsonify({'success': True, 'message': f'Credential set for {roll}'})


@points_bp.route('/admin/public-credential/bulk', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_bulk_set_public_credentials():
    ok, err = _require_admin()
    if not ok:
        return err
    payload = request.get_json(silent=True) or {}
    password = str(payload.get('password') or '')
    rolls_raw = payload.get('rolls')
    if not password:
        return jsonify({'success': False, 'error': 'password is required'}), 400
    if len(password) < 4:
        return jsonify({'success': False, 'error': 'password must be at least 4 characters'}), 400
    if rolls_raw is None or (isinstance(rolls_raw, str) and not rolls_raw.strip()):
        rolls = _student_rolls_from_ledger()
    elif isinstance(rolls_raw, list):
        rolls = [_normalize_roll(r) for r in rolls_raw if str(r or '').strip()]
    elif isinstance(rolls_raw, str):
        rolls = [_normalize_roll(r) for r in re.split(r'[,\s]+', rolls_raw) if str(r or '').strip()]
    else:
        rolls = []
    rolls = list(dict.fromkeys(rolls))  # de-dup, preserve order
    if not rolls:
        return jsonify({'success': False, 'error': 'No rolls provided and no students found in ledger'}), 400
    actor_id = _parse_int_safe(current_user.id, 0) or None
    saved = 0
    failed = []
    try:
        for roll in rolls:
            try:
                _upsert_public_credential(roll, password, actor_id)
                saved += 1
            except Exception as row_err:
                failed.append({'roll': roll, 'error': str(row_err)})
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('Bulk public credential set failed')
        return jsonify({'success': False, 'error': str(exc)}), 500
    return jsonify({
        'success': True,
        'message': f'Credentials set for {saved} student(s)',
        'saved': saved,
        'failed': failed,
    })


@points_bp.route('/admin/public-credential/toggle', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_toggle_public_credential():
    ok, err = _require_admin()
    if not ok:
        return err
    payload = request.get_json(silent=True) or {}
    roll = _normalize_roll(payload.get('roll'))
    active = bool(payload.get('active'))
    if not roll:
        return jsonify({'success': False, 'error': 'roll is required'}), 400
    row = PublicSiteCredential.query.filter_by(roll=roll).first()
    if not row:
        return jsonify({'success': False, 'error': f'No credential for {roll}'}), 404
    try:
        row.active = active
        row.set_at = datetime.utcnow()
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    state = 'enabled' if active else 'disabled'
    return jsonify({'success': True, 'message': f'{roll} {state}'})


@points_bp.route('/admin/public-credential/delete', methods=['POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + admin role check
@login_required
def admin_delete_public_credential():
    ok, err = _require_admin()
    if not ok:
        return err
    payload = request.get_json(silent=True) or {}
    roll = _normalize_roll(payload.get('roll'))
    if not roll:
        return jsonify({'success': False, 'error': 'roll is required'}), 400
    row = PublicSiteCredential.query.filter_by(roll=roll).first()
    if not row:
        return jsonify({'success': False, 'error': f'No credential for {roll}'}), 404
    try:
        db.session.delete(row)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    return jsonify({'success': True, 'message': f'Credential revoked for {roll}'})


@points_bp.route('/auth/check-updates', methods=['GET', 'POST'])
@csrf.exempt
def auth_check_updates():
    """
    Background API to check for new scoreboard updates.
    Accepts X-EA-Login-ID and X-EA-Login-Code in headers or JSON request for secure authentication.
    """
    login_id = request.headers.get('X-EA-Login-ID') or request.values.get('login_id')
    login_code = request.headers.get('X-EA-Login-Code') or request.values.get('login_code')

    if not login_id or not login_code:
        if current_user.is_authenticated:
            login_id = current_user.login_id
        else:
            return jsonify({'success': False, 'error': 'Credentials required'}), 401

    user = User.query.filter_by(login_id=login_id).first()
    if not user:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    if user.login_code:
        if user.login_code_expires_at and datetime.utcnow() > user.login_code_expires_at:
            return jsonify({'success': False, 'error': 'Login code expired'}), 401
        if str(login_code).strip().upper() != str(user.login_code).strip().upper():
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    else:
        if not user.check_password(login_code):
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    data = _load_offline_data() or {}
    server_updated_at = data.get('server_updated_at') or data.get('updated_at') or _server_now_iso()

    return jsonify({
        'success': True,
        'server_updated_at': server_updated_at
    })


@points_bp.route('/student-transfers', methods=['GET', 'POST'])
@csrf.exempt  # JSON API endpoint secured by @login_required + student role check
@login_required
@_ledger_write_guard
def student_transfers():
    if str(current_user.role or '').strip().lower() != 'student':
        return jsonify({'success': False, 'error': 'Student role required'}), 403

    now = datetime.utcnow()
    now_iso = _server_now_iso()
    now_date = _parse_date_key(now_iso) or date.today()
    month_key = now_date.strftime('%Y-%m')
    data = _load_offline_data() or {}

    sender_id = _student_id_by_login(data, current_user.login_id)
    if sender_id <= 0:
        return jsonify({'success': False, 'error': 'Student roster mapping not found'}), 404
    if not _is_month_allowed_for_user(data, current_user, month_key):
        return jsonify({'success': False, 'error': 'No month access for transfer operation'}), 403

    if request.method == 'GET':
        rows = (
            StudentTransfer.query
            .filter(
                (StudentTransfer.from_student_id == sender_id) |
                (StudentTransfer.to_student_id == sender_id)
            )
            .order_by(StudentTransfer.created_at.desc(), StudentTransfer.id.desc())
            .limit(200)
            .all()
        )
        out = [{
            'id': r.id,
            'from_student_id': r.from_student_id,
            'to_student_id': r.to_student_id,
            'transfer_type': r.transfer_type,
            'amount': r.amount,
            'created_at': r.created_at.isoformat() if r.created_at else '',
            'lock_until': r.lock_until.isoformat() if r.lock_until else '',
        } for r in rows]
        return jsonify({'success': True, 'transfers': out})

    payload = request.get_json(silent=True) or {}
    transfer_type = str(payload.get('transfer_type') or '').strip().lower()
    amount = _parse_int_safe(payload.get('amount'), 0)
    to_login_id = str(payload.get('to_login_id') or '').strip().upper()
    if transfer_type not in ('points', 'stars'):
        return jsonify({'success': False, 'error': 'transfer_type must be points or stars'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'error': 'amount must be > 0'}), 400
    if transfer_type == 'points' and amount > 50:
        return jsonify({'success': False, 'error': 'Maximum point transfer is 50 per transaction'}), 400
    if transfer_type == 'stars' and amount > 3:
        return jsonify({'success': False, 'error': 'Maximum star transfer is 3 in a transfer'}), 400
    if not to_login_id or not to_login_id.startswith('EA'):
        return jsonify({'success': False, 'error': 'to_login_id must be a valid student roll'}), 400

    receiver_id = _student_id_by_login(data, to_login_id)
    if receiver_id <= 0:
        return jsonify({'success': False, 'error': 'Recipient not found'}), 404
    if receiver_id == sender_id:
        return jsonify({'success': False, 'error': 'Cannot transfer to self'}), 400

    lock_window = now - timedelta(hours=24)
    existing_pair = (
        StudentTransfer.query
        .filter_by(from_student_id=sender_id, to_student_id=receiver_id)
        .filter(StudentTransfer.created_at >= lock_window)
        .order_by(StudentTransfer.created_at.desc())
        .first()
    )
    if existing_pair:
        return jsonify({'success': False, 'error': 'You can transfer to this student only once in 24 hours'}), 409
    reverse_pair = (
        StudentTransfer.query
        .filter_by(from_student_id=receiver_id, to_student_id=sender_id)
        .filter(StudentTransfer.created_at >= lock_window)
        .order_by(StudentTransfer.created_at.desc())
        .first()
    )
    if reverse_pair:
        return jsonify({'success': False, 'error': 'Reverse transfer lock active for this pair (24h)'}), 409

    sender = next((s for s in (data.get('students') or []) if _parse_int_safe(s.get('id'), 0) == sender_id), None)
    receiver = next((s for s in (data.get('students') or []) if _parse_int_safe(s.get('id'), 0) == receiver_id), None)
    if not sender or not receiver:
        return jsonify({'success': False, 'error': 'Sender/receiver records unavailable'}), 409

    if transfer_type == 'points':
        sender_points = _sum_points_for_student_month(data, sender_id, month_key)
        if sender_points <= 0:
            return jsonify({'success': False, 'error': 'Sender has no transferable points'}), 409
        if amount > sender_points:
            return jsonify({'success': False, 'error': 'Transfer exceeds sender available points'}), 409
        _upsert_score_delta(
            data, sender_id, now_date.isoformat(), month_key,
            delta_points=-amount, delta_stars=0,
            note=f'[TRANSFER OUT points:{amount} to {to_login_id}]'
        )
        _upsert_score_delta(
            data, receiver_id, now_date.isoformat(), month_key,
            delta_points=amount, delta_stars=0,
            note=f'[TRANSFER IN points:{amount} from {current_user.login_id}]'
        )
    else:
        stars_24h = (
            db.session.query(db.func.coalesce(db.func.sum(StudentTransfer.amount), 0))
            .filter_by(from_student_id=sender_id, transfer_type='stars')
            .filter(StudentTransfer.created_at >= lock_window)
            .scalar() or 0
        )
        if _parse_int_safe(stars_24h, 0) + amount > 3:
            return jsonify({'success': False, 'error': 'Star transfer exceeds 24h cap (3)'}), 409
        sender_stars = _parse_int_safe(sender.get('stars'), 0)
        if sender_stars <= 0 or amount > sender_stars:
            return jsonify({'success': False, 'error': 'Sender has insufficient stars'}), 409
        sender['stars'] = max(0, sender_stars - amount)
        receiver['stars'] = max(0, _parse_int_safe(receiver.get('stars'), 0) + amount)
        _upsert_score_delta(
            data, sender_id, now_date.isoformat(), month_key,
            delta_points=0, delta_stars=-amount,
            note=f'[TRANSFER OUT stars:{amount} to {to_login_id}]'
        )
        _upsert_score_delta(
            data, receiver_id, now_date.isoformat(), month_key,
            delta_points=0, delta_stars=amount,
            note=f'[TRANSFER IN stars:{amount} from {current_user.login_id}]'
        )

    transfer_row = StudentTransfer(
        from_student_id=sender_id,
        to_student_id=receiver_id,
        transfer_type=transfer_type,
        amount=amount,
        created_at=now,
        lock_until=now + timedelta(hours=24)
    )
    try:
        db.session.add(transfer_row)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    _ensure_score_timestamps(data)
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='student-transfer')
    _forward_offline_data_to_peers_async(data, [])

    return jsonify({
        'success': True,
        'message': f'{transfer_type.capitalize()} transfer completed',
        'updated_at': now_iso
    })


@points_bp.route('/add-points', methods=['POST'])
@login_required
@_ledger_write_guard
def add_points():
    """Add points for a student"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()

        # Security: Validate incoming data
        if not isinstance(data, dict):
            return jsonify({'success': False, 'error': 'Invalid request data'}), 400

        # Validate student_id
        student_id = data.get('student_id')
        if not student_id or not isinstance(student_id, int):
            return jsonify({'success': False, 'error': 'Invalid student ID'}), 400

        # Validate student exists and user account is active
        student = StudentProfile.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404

        # Check if associated user account is active
        if not student.user or not student.user.is_active:
            return jsonify({'success': False, 'error': 'Student account is inactive'}), 403

        # Validate and parse date
        try:
            date_str = data.get('date')
            if not date_str:
                return jsonify({'success': False, 'error': 'Date is required'}), 400
            date_recorded = datetime.fromisoformat(date_str).date()
        except (ValueError, TypeError) as e:
            return jsonify({'success': False, 'error': 'Invalid date format (use YYYY-MM-DD)'}), 400

        # Validate date is not in future
        if date_recorded > datetime.now().date():
            return jsonify({'success': False, 'error': 'Cannot record points for future dates'}), 400

        # Validate date is not too far in past (within current academic year)
        from datetime import date
        current_year = date.today().year
        if date_recorded.year < (current_year - 1):
            return jsonify({'success': False, 'error': 'Date is too far in the past'}), 400

        # Validate and sanitize numeric values
        try:
            points = int(data.get('points', 0))
            stars = int(data.get('stars', 0))
            vetos = int(data.get('vetos', 0))
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Points, stars, and vetos must be integers'}), 400

        # Security: Validate numeric ranges (prevent data corruption)
        if not (0 <= points <= 1000):
            return jsonify({'success': False, 'error': 'Points must be between 0 and 1000'}), 400
        if not (0 <= stars <= 100):
            return jsonify({'success': False, 'error': 'Stars must be between 0 and 100'}), 400
        if not (0 <= vetos <= 50):
            return jsonify({'success': False, 'error': 'Vetos must be between 0 and 50'}), 400

        # Sanitize notes (prevent XSS)
        notes = str(data.get('notes', ''))[:500]  # Limit to 500 chars
        
        # Check if record exists (only manual entries; notebook entries have their own entry_type)
        record = StudentPoints.query.filter_by(
            student_id=student_id,
            date_recorded=date_recorded,
            entry_type='manual',
        ).first()
        
        if record:
            # Update existing
            record.points = points
            record.stars = stars
            record.vetos = vetos
            record.notes = notes
            record.recorded_by = current_user.login_id
        else:
            # Create new
            record = StudentPoints(
                student_id=student_id,
                date_recorded=date_recorded,
                points=points,
                stars=stars,
                vetos=vetos,
                notes=notes,
                recorded_by=current_user.login_id,
                entry_type='manual',
            )
            db.session.add(record)
        
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Points added successfully',
            'data': record.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in add_points: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/leader-adjust-score', methods=['POST'])
@csrf.exempt  # JSON API — secured by @login_required + role check
@login_required
@_ledger_write_guard
def leader_adjust_score():
    """
    Leader rule:
    - can neutralize negative score to zero for only 1 student per day (today only).
    Co-Leader rule:
    - can reduce a negative score by at most 30 points (today only).
    """
    if str(current_user.role or '').strip().lower() != 'student':
        return jsonify({'success': False, 'error': 'Only student post-holders can use this action'}), 403

    payload = request.get_json(silent=True) or {}
    student_id = _parse_int_safe(payload.get('student_id'), 0)
    date_key = str(payload.get('date') or '').strip()
    reason = str(payload.get('reason') or '').strip()[:300]
    if student_id <= 0 or not date_key:
        return jsonify({'success': False, 'error': 'student_id and date are required'}), 400

    today = _parse_date_key(_server_now_iso()) or date.today()
    target_date = _parse_date_key(date_key)
    if not target_date or target_date != today:
        return jsonify({'success': False, 'error': 'Action allowed only for current server date'}), 400

    data = _load_offline_data() or {}
    role_type = _active_leadership_role_for_login(data, current_user.login_id)
    if role_type not in ('leader', 'co_leader'):
        return jsonify({'success': False, 'error': 'Active Leader/Co-Leader role required'}), 403

    score_row = _get_score_row_for_student_date(data, student_id, date_key)
    if not score_row:
        return jsonify({'success': False, 'error': 'Score row not found for target student/date'}), 404

    current_points = _parse_int_safe(score_row.get('points'), 0)
    if current_points >= 0:
        return jsonify({'success': False, 'error': 'Only negative score can be adjusted'}), 400

    actions = list(data.get('score_adjustment_actions') or [])
    actor_login = str(current_user.login_id or '').strip()
    today_str = today.isoformat()

    if role_type == 'leader':
        # Leader can affect only 1 student/day.
        prior = [
            a for a in actions
            if isinstance(a, dict)
            and str(a.get('actor_login_id') or '').strip() == actor_login
            and str(a.get('action_date') or '').strip() == today_str
            and str(a.get('mode') or '').strip() == 'leader_zero'
        ]
        if prior:
            first_sid = _parse_int_safe(prior[0].get('target_student_id'), 0)
            if first_sid != student_id:
                return jsonify({'success': False, 'error': 'Leader can neutralize only one student per day'}), 409
        delta = abs(current_points)
        new_points = 0
        mode = 'leader_zero'
    else:
        # Co-Leader can reduce penalty up to 30 points.
        relief = min(30, abs(current_points))
        if relief <= 0:
            return jsonify({'success': False, 'error': 'No reducible penalty found'}), 400
        delta = relief
        new_points = current_points + relief
        if new_points > 0:
            new_points = 0
            delta = abs(current_points)
        mode = 'co_leader_reduce'

    now_iso = _server_now_iso()
    score_row['points'] = new_points
    score_row['updated_at'] = now_iso
    note = str(score_row.get('notes') or '').strip()
    suffix = (
        f"[{mode}] {actor_login} adjusted {current_points} -> {new_points}"
        + (f" | reason: {reason}" if reason else '')
    )
    score_row['notes'] = f"{note} | {suffix}" if note else suffix
    # Append to inline score history for audit trail
    if not isinstance(score_row.get('history'), list):
        score_row['history'] = []
    score_row['history'].append({
        'delta': delta,
        'total': new_points,
        'actor': actor_login,
        'role': role_type,
        'timestamp': now_iso,
        'note': f"[{mode}] {reason}" if reason else f"[{mode}]"
    })

    action_row = {
        'id': int(time.time() * 1000),
        'actor_login_id': actor_login,
        'actor_user_id': _parse_int_safe(current_user.id, 0),
        'actor_role': role_type,
        'target_student_id': student_id,
        'target_date': date_key,
        'action_date': today_str,
        'delta_points': delta,
        'mode': mode,
        'reason': reason,
        'created_at': now_iso,
    }
    actions.append(action_row)
    data['score_adjustment_actions'] = actions[-5000:]
    try:
        db_action = ScoreAdjustmentAction(
            actor_user_id=_parse_int_safe(current_user.id, 0),
            actor_login_id=actor_login,
            actor_role=role_type,
            target_student_id=student_id,
            target_date=target_date,
            delta_points=delta,
            mode=mode,
            reason=reason
        )
        db.session.add(db_action)
        db.session.commit()
    except Exception:
        db.session.rollback()

    _ensure_score_timestamps(data)
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(data['server_updated_at'], source='leader-adjust')
    _forward_offline_data_to_peers_async(data, [])
    return jsonify({
        'success': True,
        'mode': mode,
        'old_points': current_points,
        'new_points': new_points,
        'updated_at': data['server_updated_at']
    })


@points_bp.route('/award-gcb', methods=['POST'])
@csrf.exempt
@login_required
@_ledger_write_guard
def award_gcb():
    """Award or revoke GCB (Good Conduct Badge) immunity for a student. Admin only.

    GCB immunity means:
    - Absence penalty is not applied to this student.
    - Score entries cannot be saved below -20 for this student.
    """
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    payload = request.get_json(silent=True) or {}
    student_id = _parse_int_safe(payload.get('student_id'), 0)
    awarded = bool(payload.get('awarded', False))

    if student_id <= 0:
        return jsonify({'success': False, 'error': 'student_id required'}), 400

    data = _load_offline_data() or {}
    students = data.get('students', []) or []
    target = None
    for s in students:
        if isinstance(s, dict) and _parse_int_safe(s.get('id'), 0) == student_id:
            target = s
            break
    if target is None:
        return jsonify({'success': False, 'error': 'Student not found'}), 404

    target['gcb'] = awarded
    now_iso = _server_now_iso()
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='award-gcb')
    _forward_offline_data_to_peers_async(data, [])
    return jsonify({
        'success': True,
        'student_id': student_id,
        'roll': target.get('roll', ''),
        'gcb': awarded,
        'updated_at': now_iso,
    })


@points_bp.route('/add-student', methods=['POST'])
@login_required
@_ledger_write_guard
def add_student():
    """Add a new student"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        # Check if student already exists
        existing = StudentProfile.query.filter_by(roll_number=data.get('roll_number')).first()
        if existing:
            return jsonify({'success': False, 'error': 'Student with this roll number already exists'}), 400
        
        student = StudentProfile(
            roll_number=data.get('roll_number'),
            full_name=data.get('full_name'),
            class_name=data.get('class'),
            group=data.get('group', 'A'),
            user_id=None  # Not linked to user account
        )

        profile_data = student.profile_data or {}
        if data.get('fees') is not None:
            profile_data['fees'] = data.get('fees')
        if data.get('vote_power') is not None:
            profile_data['vote_power'] = data.get('vote_power')
        student.profile_data = profile_data
        
        db.session.add(student)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Student added successfully',
            'student': {
                'id': student.id,
                'roll_number': student.roll_number,
                'full_name': student.full_name,
                'class': student.class_name,
                'group': student.group,
                'fees': profile_data.get('fees'),
                'vote_power': profile_data.get('vote_power')
            }
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in add_student: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/delete-student/<int:student_id>', methods=['DELETE'])
@login_required
@_ledger_write_guard
def delete_student(student_id):
    """Delete a student"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        student = StudentProfile.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        db.session.delete(student)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Student deleted successfully'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in delete_student: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/update-profile/<int:student_id>', methods=['POST'])
@login_required
@_ledger_write_guard
def update_profile(student_id):
    """Update student profile with extended fields"""
    # Security: Only admin and teacher can update profiles
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin or Teacher access required'}), 403

    try:
        student = StudentProfile.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404

        # Check if associated user account is active
        if not student.user or not student.user.is_active:
            return jsonify({'success': False, 'error': 'Student account is inactive'}), 403

        data = request.get_json()

        # Security: Validate incoming data is a dictionary
        if not isinstance(data, dict):
            return jsonify({'success': False, 'error': 'Invalid request data'}), 400
        
        # Update basic fields
        if 'full_name' in data:
            student.full_name = data['full_name']
        if 'class' in data:
            student.class_name = data['class']
        if 'group' in data:
            student.group = data['group']
        
        # Update extended profile fields
        if not student.profile_data:
            student.profile_data = {}
        
        profile_updates = {
            'fatherName', 'motherName', 'dateOfBirth', 'bloodGroup', 'aadhar',
            'phone', 'email', 'address', 'parentPhone', 'admissionDate', 'academicYear',
            'fees', 'vote_power', 'total_score', 'rank'
        }
        
        for field in profile_updates:
            if field in data:
                student.profile_data[field] = data[field]
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Profile updated successfully',
            'student': {
                'id': student.id,
                'roll_number': student.roll_number,
                'full_name': student.full_name,
                'class': student.class_name,
                'group': student.group,
                'profile_data': student.profile_data
            }
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in update_profile: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


# ── Student portrait photo upload ──────────────────────────────────────────
# Saves the full-resolution image to app/static/uploads/students/{id}.{ext}
# and records the relative URL path in the student's profile_data.photo_path
# in the offline ledger. On Force Publish, the uploads dir is mirrored to
# public_site/static/uploads/students/ so portraits appear on Cloudflare too.
_PHOTO_ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}
_PHOTO_ALLOWED_MIMETYPES = {
    'image/jpeg', 'image/jpg', 'image/png', 'image/webp',
    'application/octet-stream',  # browsers sometimes send this for webp
}
_PHOTO_MAX_SIZE = 5 * 1024 * 1024  # 5 MB


def _student_uploads_dir():
    """Absolute path to the student portrait uploads directory."""
    return os.path.join(current_app.root_path, 'static', 'uploads', 'students')


def _student_photo_url(student_id, ext):
    """Relative URL path for a student's portrait (works on LAN + Cloudflare)."""
    return f'/static/uploads/students/{student_id}.{ext}'


@points_bp.route('/upload-student-photo', methods=['POST'])
@csrf.exempt  # Multipart API — secured by login + role check
@_ledger_write_guard
def upload_student_photo():
    """Upload a portrait photo for a student (admin/teacher only).

    Accepts multipart form data:
        file        - the image file (jpg/png/webp, max 5 MB)
        student_id  - int form field

    Saves to app/static/uploads/students/{student_id}.{ext}, updates the
    student's profile_data.photo_path in the offline ledger, and returns
    {success, photo_path}.

    Auth: supports both Flask-Login sessions (LAN) and X-EA-Login-ID /
    X-EA-Login-Code headers (cross-origin / Cloudflare tunnel mode).
    """
    # Resolve user from session or cross-origin headers
    user, role, login_id = _get_request_user()
    if not user:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    if str(role or '').strip().lower() not in ('admin', 'teacher'):
        return jsonify({'success': False, 'error': 'Admin or Teacher access required'}), 403

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    student_id = _parse_int_safe(request.form.get('student_id'), 0)
    if student_id <= 0:
        return jsonify({'success': False, 'error': 'Valid student_id required'}), 400

    # Validate extension
    original_name = str(file.filename or '')
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in _PHOTO_ALLOWED_EXTENSIONS:
        return jsonify({'success': False, 'error': 'Only JPG, PNG, and WebP images are allowed'}), 400

    # Validate MIME type (lenient — accept octet-stream for webp)
    if file.content_type and file.content_type not in _PHOTO_ALLOWED_MIMETYPES:
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400

    # Validate file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    if file_size > _PHOTO_MAX_SIZE:
        return jsonify({'success': False, 'error': f'File too large. Maximum size: {_PHOTO_MAX_SIZE // (1024*1024)}MB'}), 400
    if file_size == 0:
        return jsonify({'success': False, 'error': 'Empty file'}), 400

    # Verify student exists in the offline ledger
    data = _load_offline_data() or {}
    students = data.get('students', []) or []
    student = next((s for s in students if _parse_int_safe(s.get('id'), 0) == student_id), None)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404

    # Save the file
    uploads_dir = _student_uploads_dir()
    os.makedirs(uploads_dir, exist_ok=True)

    # Remove any existing photo for this student (different extensions)
    for old_ext in _PHOTO_ALLOWED_EXTENSIONS:
        old_path = os.path.join(uploads_dir, f'{student_id}{old_ext}')
        if os.path.exists(old_path):
            try:
                os.unlink(old_path)
            except Exception:
                current_app.logger.warning(f'Could not remove old photo {old_path}')

    filename = f'{student_id}{ext}'
    save_path = os.path.join(uploads_dir, filename)
    try:
        file.save(save_path)
    except Exception as e:
        current_app.logger.error(f'Failed to save student photo: {e}')
        return jsonify({'success': False, 'error': 'Failed to save file'}), 500

    photo_url = _student_photo_url(student_id, ext.lstrip('.'))

    # Update the student's profile_data in the offline ledger
    profile = student.get('profile_data')
    if not isinstance(profile, dict):
        profile = {}
    profile['photo_path'] = photo_url
    student['profile_data'] = profile
    student['updated_at'] = _server_now_iso()

    _save_offline_data(data)

    current_app.logger.info(f'Student photo uploaded for student_id={student_id} by {login_id} -> {photo_url}')
    return jsonify({'success': True, 'photo_path': photo_url})


@points_bp.route('/remove-student-photo', methods=['POST'])
@csrf.exempt
@_ledger_write_guard
def remove_student_photo():
    """Remove a student's portrait photo (admin/teacher only).

    Body JSON: { student_id: int }
    Deletes the file from disk and clears profile_data.photo_path.
    """
    user, role, login_id = _get_request_user()
    if not user:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    if str(role or '').strip().lower() not in ('admin', 'teacher'):
        return jsonify({'success': False, 'error': 'Admin or Teacher access required'}), 403

    payload = request.get_json(silent=True) or {}
    student_id = _parse_int_safe(payload.get('student_id'), 0)
    if student_id <= 0:
        return jsonify({'success': False, 'error': 'Valid student_id required'}), 400

    data = _load_offline_data() or {}
    students = data.get('students', []) or []
    student = next((s for s in students if _parse_int_safe(s.get('id'), 0) == student_id), None)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404

    uploads_dir = _student_uploads_dir()
    removed_files = []
    for old_ext in _PHOTO_ALLOWED_EXTENSIONS:
        old_path = os.path.join(uploads_dir, f'{student_id}{old_ext}')
        if os.path.exists(old_path):
            try:
                os.unlink(old_path)
                removed_files.append(old_path)
            except Exception:
                current_app.logger.warning(f'Could not remove photo {old_path}')

    profile = student.get('profile_data')
    if isinstance(profile, dict) and profile.get('photo_path'):
        profile.pop('photo_path', None)
        student['profile_data'] = profile
        student['updated_at'] = _server_now_iso()
        _save_offline_data(data)

    return jsonify({'success': True, 'removed_files': len(removed_files)})


@points_bp.route('/import-excel', methods=['POST'])
@login_required
@_ledger_write_guard
def import_excel():
    """Import student data and scores from Excel file"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
            return jsonify({'success': False, 'error': 'Only Excel files (.xlsx, .xls, .xlsm) are supported'}), 400

        # Validate MIME type for additional security
        allowed_mimetypes = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
            'application/vnd.ms-excel.sheet.macroEnabled.12'  # .xlsm
        ]
        if file.content_type not in allowed_mimetypes and file.content_type != 'application/octet-stream':
            return jsonify({'success': False, 'error': 'Invalid file type'}), 400

        # Check file size (max 50MB as configured in app config)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        max_size = current_app.config.get('MAX_CONTENT_LENGTH', 52428800)
        if file_size > max_size:
            return jsonify({'success': False, 'error': f'File too large. Maximum size: {max_size // (1024*1024)}MB'}), 400

        # Security: Use unique temporary file to prevent race conditions and path traversal
        import uuid
        temp_suffix = file.filename.split('.')[-1] if '.' in file.filename else 'xlsx'
        temp_file = tempfile.NamedTemporaryFile(
            mode='w+b',
            suffix=f'.{temp_suffix}',
            prefix=f'ea_import_{uuid.uuid4().hex}_',
            delete=False
        )
        temp_path = temp_file.name
        temp_file.close()

        try:
            # Save uploaded file to unique temporary path
            file.save(temp_path)

            # Load workbook with security settings (read_only to prevent formula execution)
            import openpyxl
            wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True, keep_vba=False)
        except Exception as e:
            # Security: Always cleanup temp file on error
            if os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except:
                    pass
            return jsonify({'success': False, 'error': f'Failed to read Excel file: {str(e)}'}), 400
        sheet_name = request.form.get('sheet') or request.args.get('sheet')
        if not sheet_name:
            for name in wb.sheetnames:
                if name.strip().lower() == 'feb 26':
                    sheet_name = name
                    break
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        
        imported_count = 0
        errors = []
        
        header_row = [cell.value for cell in ws[1]]
        header_map = {}
        for idx, value in enumerate(header_row, start=1):
            if isinstance(value, str):
                header_map[value.strip().lower()] = idx

        def find_header(candidates):
            for key in candidates:
                for header, idx in header_map.items():
                    if key in header:
                        return idx
            return None

        roll_col = find_header(['roll'])
        name_col = find_header(['student name', 'name'])
        class_col = find_header(['class'])
        fees_col = find_header(['fees'])
        total_col = find_header(['total score'])
        rank_col = find_header(['rank'])
        vote_col = find_header(['vote power', 'votepower'])

        if not roll_col or not name_col:
            return jsonify({'success': False, 'error': 'Missing roll or student name column'}), 400

        date_columns = []
        for idx, header in enumerate(header_row, start=1):
            if isinstance(header, (datetime, date)):
                date_columns.append((idx, header.date()))
            elif isinstance(header, str):
                try:
                    parsed = datetime.fromisoformat(header.strip())
                    date_columns.append((idx, parsed.date()))
                except (ValueError, TypeError):
                    pass
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                roll_number = ws.cell(row_idx, roll_col).value
                full_name = ws.cell(row_idx, name_col).value
                class_name = ws.cell(row_idx, class_col).value if class_col else 'Unknown'
                fees_value = ws.cell(row_idx, fees_col).value if fees_col else None
                total_value = ws.cell(row_idx, total_col).value if total_col else None
                rank_value = ws.cell(row_idx, rank_col).value if rank_col else None
                vote_value = ws.cell(row_idx, vote_col).value if vote_col else None
                
                if not roll_number or not full_name:
                    continue

                roll_str = str(roll_number).strip()
                group_match = re.search(r'^EA\\d{2}([A-Z])', roll_str)
                group = group_match.group(1) if group_match else 'A'

                # Get or create student
                student = StudentProfile.query.filter_by(roll_number=roll_str).first()
                if not student:
                    student = StudentProfile(
                        roll_number=roll_str,
                        full_name=str(full_name),
                        class_name=str(class_name),
                        group=group
                    )
                    db.session.add(student)
                    db.session.flush()
                else:
                    student.full_name = str(full_name)
                    student.class_name = str(class_name)
                    student.group = group

                profile_data = student.profile_data or {}
                if isinstance(fees_value, (int, float)):
                    profile_data['fees'] = int(fees_value)
                if isinstance(vote_value, (int, float)):
                    profile_data['vote_power'] = int(vote_value)
                if isinstance(total_value, (int, float)):
                    profile_data['total_score'] = int(total_value)
                if isinstance(rank_value, (int, float)):
                    profile_data['rank'] = int(rank_value)
                student.profile_data = profile_data
                
                # Process date columns (scores)
                for col_idx, date_recorded in date_columns:
                    score_value = ws.cell(row_idx, col_idx).value
                    if score_value is None or score_value == '':
                        continue
                    if not isinstance(score_value, (int, float)):
                        continue

                    record = StudentPoints.query.filter_by(
                        student_id=student.id,
                        date_recorded=date_recorded
                    ).first()

                    if record:
                        record.points = int(score_value)
                        record.recorded_by = current_user.login_id
                    else:
                        record = StudentPoints(
                            student_id=student.id,
                            date_recorded=date_recorded,
                            points=int(score_value),
                            recorded_by=current_user.login_id
                        )
                        db.session.add(record)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        # Extract party and leadership data if present
        parties, leadership = _extract_party_and_leadership(ws)
        if parties or leadership:
            _save_politics_data({
                'parties': parties or DEFAULT_PARTIES,
                'leadership': leadership or DEFAULT_LEADERSHIP
            })

        db.session.commit()
        os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'message': f'Imported {imported_count} records successfully',
            'imported_count': imported_count,
            'errors': errors
        })
    except Exception as e:
        current_app.logger.error(f"Error in import_excel: {str(e)}")
        # Cleanup temp file if it exists
        try:
            if 'temp_path' in locals() and os.path.exists(temp_path):
                os.unlink(temp_path)
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/seed-feb26', methods=['POST'])
@login_required
@_ledger_write_guard
def seed_feb26():
    """Seed database with Feb 26 sheet data"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    mode = request.args.get('mode', 'replace_unlinked')
    try:
        rolls = {s['roll'] for s in FEB26_SEED['students']}

        if mode == 'replace_unlinked':
            StudentProfile.query.filter(
                StudentProfile.user_id.is_(None),
                StudentProfile.roll_number.notin_(list(rolls))
            ).delete(synchronize_session=False)

        # Upsert students
        for student_data in FEB26_SEED['students']:
            roll = student_data['roll']
            student = StudentProfile.query.filter_by(roll_number=roll).first()
            group_match = re.search(r'^EA\\d{2}([A-Z])', roll)
            group = group_match.group(1) if group_match else 'A'
            if not student:
                student = StudentProfile(
                    roll_number=roll,
                    full_name=student_data.get('name'),
                    class_name=str(student_data.get('class')),
                    group=group
                )
                db.session.add(student)
                db.session.flush()
            else:
                student.full_name = student_data.get('name')
                student.class_name = str(student_data.get('class'))
                student.group = group

            profile_data = student.profile_data or {}
            profile_data['fees'] = student_data.get('fees', 0)
            profile_data['vote_power'] = student_data.get('vote_power')
            profile_data['total_score'] = student_data.get('total_score')
            profile_data['rank'] = student_data.get('rank')
            student.profile_data = profile_data

        db.session.flush()

        seed_id_map = {s['id']: s['roll'] for s in FEB26_SEED['students']}
        imported_scores = 0
        for score in FEB26_SEED['scores']:
            roll = seed_id_map.get(score['studentId'])
            if not roll:
                continue
            student = StudentProfile.query.filter_by(roll_number=roll).first()
            if not student:
                continue
            date_recorded = datetime.fromisoformat(score['date']).date()
            record = StudentPoints.query.filter_by(
                student_id=student.id,
                date_recorded=date_recorded
            ).first()
            if record:
                record.points = int(score['points'])
                record.recorded_by = current_user.login_id
            else:
                record = StudentPoints(
                    student_id=student.id,
                    date_recorded=date_recorded,
                    points=int(score['points']),
                    recorded_by=current_user.username
                )
                db.session.add(record)
            imported_scores += 1

        _save_politics_data({
            'parties': FEB26_SEED.get('parties', DEFAULT_PARTIES),
            'leadership': FEB26_SEED.get('leadership', DEFAULT_LEADERSHIP)
        })

        db.session.commit()
        return jsonify({
            'success': True,
            'students': len(FEB26_SEED['students']),
            'scores': imported_scores
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in seed_feb26: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/leaderboard')
@login_required
def get_leaderboard():
    """Get top students leaderboard"""
    try:
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        limit = request.args.get('limit', 10, type=int)
        month_start = date(year, month, 1)
        month_end = month_start + relativedelta(months=1)
        
        # Build query
        leaderboard = db.session.query(
            StudentProfile.id,
            StudentProfile.full_name,
            StudentProfile.class_name,
            db.func.sum(StudentPoints.points).label('total_points'),
            db.func.sum(StudentPoints.stars).label('total_stars'),
            db.func.sum(StudentPoints.vetos).label('total_vetos')
        ).join(
            StudentPoints, StudentProfile.id == StudentPoints.student_id
        ).filter(
            StudentPoints.date_recorded >= month_start,
            StudentPoints.date_recorded < month_end,
        ).group_by(
            StudentProfile.id,
            StudentProfile.full_name,
            StudentProfile.class_name
        ).order_by(
            (db.func.sum(StudentPoints.points) + 
             db.func.sum(StudentPoints.stars) * 10 - 
             db.func.sum(StudentPoints.vetos) * 5).desc()
        ).limit(limit).all()
        
        result = []
        for idx, record in enumerate(leaderboard, 1):
            net_score = (record.total_points or 0) + ((record.total_stars or 0) * 10) - ((record.total_vetos or 0) * 5)
            result.append({
                'rank': idx,
                'name': record.full_name,
                'class': record.class_name,
                'total_points': record.total_points or 0,
                'total_stars': record.total_stars or 0,
                'total_vetos': record.total_vetos or 0,
                'net_score': net_score
            })
        
        return jsonify({'success': True, 'leaderboard': result})
    except Exception as e:
        current_app.logger.error(f"Error in get_leaderboard: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@points_bp.route('/month-summary')
@login_required
def get_month_summary():
    """Get previous months for tab navigation"""
    try:
        today = date.today()
        months_data = []
        
        for i in range(4):  # Current + 3 previous months
            check_date = today - relativedelta(months=i)
            month_key = f"{check_date.year}-{check_date.month:02d}"
            month_name = calendar.month_name[check_date.month]
            
            months_data.append({
                'key': month_key,
                'name': f"{month_name} {check_date.year}",
                'year': check_date.year,
                'month': check_date.month,
                'is_current': i == 0
            })
        
        return jsonify({
            'success': True,
            'months': months_data,
            'current_month': today.month,
            'current_year': today.year,
            'today': today.isoformat()
        })
    except Exception as e:
        current_app.logger.error(f"Error in get_month_summary: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


# ============== EXCEL IMPORT HELPERS ==============

# Column header patterns that must NEVER be treated as daily score columns.
# These appear in all three Excel format variants as text-string headers.
_EXCEL_EXCLUDED_COL_KEYWORDS = frozenset([
    'previous advantage', 'prev advantage', 'previous adv',
    'cumulative', 'cumul point', 'running total',
    'final score', 'total score', 'grand total', 'overall score',
    'prize money', 'prize used', 'combined score',
    'project', 'activity point', 'activity score',
    'vote power', 'votepower',
    'rank', 'ranking',
    'awf', 'buffer',
    'roll', 'roll no', 'student name', 'name',
    'class', 'group', 'fees', 'fee',
    'bonus', 'advantage',
])


def _excel_col_excluded(header_val):
    """Return True if this column should never be imported as a daily score."""
    if not isinstance(header_val, str):
        return False
    s = header_val.strip().lower()
    for kw in _EXCEL_EXCLUDED_COL_KEYWORDS:
        if kw in s:
            return True
    return False


def _sheet_name_to_month_key(sheet_name):
    """
    Parse a sheet name like 'Jan 26', 'Feb 26', 'October 2025' → '2026-01', '2026-02', '2025-10'.
    Returns None if the name is not a recognisable month pattern.
    """
    import calendar as _cal
    abbr_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}
    full_map = {m.lower(): i for i, m in enumerate(_cal.month_name) if m}
    parts = str(sheet_name or '').strip().split()
    if len(parts) != 2:
        return None
    mon_str = parts[0].strip().lower()
    yr_str = parts[1].strip()
    mon_num = abbr_map.get(mon_str) or full_map.get(mon_str)
    if not mon_num:
        return None
    try:
        yr = int(yr_str)
        if yr < 100:
            yr += 2000
        return f"{yr:04d}-{mon_num:02d}"
    except (ValueError, TypeError):
        return None


def _detect_excel_header_row(ws):
    """
    Return (header_row_number, header_values_list).
    Tries row 1 first; if no 'roll' column is found, tries row 2.
    This handles Format B sheets (Oct 24–Aug 25) where row 1 contains
    'EXCEL ACADEMY LEADERSHIP BOARD' and the real headers are in row 2.
    """
    for row_num in (1, 2):
        try:
            row_vals = [cell.value for cell in ws[row_num]]
        except Exception:
            continue
        has_roll = any(
            isinstance(v, str) and 'roll' in v.strip().lower()
            for v in row_vals
        )
        if has_roll:
            return row_num, row_vals
    # Fallback
    return 1, [cell.value for cell in ws[1]]


# ============== REFINED IMPORT ENDPOINTS ==============

@points_bp.route('/import-historical-data', methods=['POST'])
@csrf.exempt  # JSON API — secured by @login_required + admin role check
@login_required
@_ledger_write_guard
def import_historical_data():
    """
    Historical Excel import for months before Feb 2026.

    For each historical month sheet (Aug 24 – Jan 26 only), this importer:
    1) Reads monthly totals from "Total Score" (fallback: "Final Score"), and
    2) Imports per-date star usage markers from real date columns only.

    Feb 2026 and newer months are never touched.
    """
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    # Hard limit: never touch Feb 26 onwards
    HISTORY_CUTOFF = '2026-02'

    temp_path = None
    try:
        # ── File validation ──────────────────────────────────────────────────
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
            return jsonify({'success': False, 'error': 'Only Excel files (.xlsx, .xls, .xlsm) are supported'}), 400

        import uuid
        temp_suffix = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'xlsx'
        tmp = tempfile.NamedTemporaryFile(
            mode='w+b', suffix=f'.{temp_suffix}',
            prefix=f'ea_hist_{uuid.uuid4().hex}_', delete=False
        )
        temp_path = tmp.name
        tmp.close()

        try:
            file.save(temp_path)
            import openpyxl
            wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True, keep_vba=False)
        except Exception as e:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
            return jsonify({'success': False, 'error': f'Cannot read Excel file: {e}'}), 400

        # ── Load live snapshot and build student lookups ────────────────
        snapshot = _load_offline_data() or {}
        roll_to_sid = {}
        name_to_sid = {}
        for s in (snapshot.get('students') or []):
            if not isinstance(s, dict):
                continue
            roll = str(s.get('roll') or '').strip().upper()
            sid = s.get('id')
            if roll and sid is not None and roll not in roll_to_sid:
                roll_to_sid[roll] = sid
            name = _name_key(s.get('base_name') or s.get('name') or s.get('raw_name') or '')
            if name and sid is not None and name not in name_to_sid:
                name_to_sid[name] = sid

        if not roll_to_sid and not name_to_sid:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
            return jsonify({
                'success': False,
                'error': 'No students found in offline snapshot. Force-publish data from admin panel first.'
            }), 400

        # ── Determine which sheets to process (historical only) ───────────
        sheet_filter = request.form.get('sheet', '').strip()
        target_sheets = []
        if sheet_filter and sheet_filter in wb.sheetnames:
            mk = _sheet_name_to_month_key(sheet_filter)
            if mk and mk < HISTORY_CUTOFF:
                target_sheets = [(sheet_filter, mk)]
        if not target_sheets:
            for name in wb.sheetnames:
                mk = _sheet_name_to_month_key(name)
                if mk and mk < HISTORY_CUTOFF:
                    target_sheets.append((name, mk))

        if not target_sheets:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
            return jsonify({
                'success': False,
                'error': (
                    'No historical month sheets found (looking for months before Feb 2026). '
                    'Sheet names must be like "Jan 26", "Oct 24" etc. '
                    f'Sheets in workbook: {", ".join(wb.sheetnames[:10])}'
                )
            }), 400

        # ── Helpers ─────────────────────────────────────────────────────────
        def _to_date(v, month_hint=None):
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            # Day-only numeric headers (1..31) are common in workbook variants.
            if isinstance(v, (int, float)) and month_hint:
                day_num = int(v)
                if 1 <= day_num <= 31:
                    try:
                        yy, mm = [int(x) for x in month_hint.split('-')]
                        return date(yy, mm, day_num)
                    except Exception:
                        return None
            if isinstance(v, str):
                text = v.strip()
                if not text:
                    return None
                if month_hint and re.fullmatch(r'\d{1,2}', text):
                    try:
                        yy, mm = [int(x) for x in month_hint.split('-')]
                        return date(yy, mm, int(text))
                    except Exception:
                        return None
                # Common textual forms seen in sheets
                for fmt in (
                    '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y',
                    '%d-%b-%Y', '%d-%B-%Y', '%d-%b-%y', '%d-%B-%y',
                    '%b %d %Y', '%B %d %Y'
                ):
                    try:
                        return datetime.strptime(text, fmt).date()
                    except Exception:
                        pass
                # Last resort: ISO parser
                try:
                    return datetime.fromisoformat(text).date()
                except Exception:
                    return None
            return None

        def _to_int_score(v):
            """Parse a numeric score from Excel cell values safely."""
            if v is None:
                return None
            if isinstance(v, bool):
                return int(v)
            if isinstance(v, (int, float)):
                return int(round(v))
            text = str(v).strip()
            if not text:
                return None
            m = re.search(r'-?\d+(?:\.\d+)?', text.replace(',', ''))
            if not m:
                return None
            try:
                return int(round(float(m.group(0))))
            except Exception:
                return None

        def _extract_star_usage(v):
            """
            Return star usage count from a cell.
            Supports '*', '**', '***', and compact '*xN' style.
            """
            if v is None:
                return 0
            if isinstance(v, (int, float)):
                return 0
            text = str(v).strip()
            if not text:
                return 0
            compact = re.search(r'\*\s*[xX]\s*(\d+)', text)
            if compact:
                return max(0, _parse_int_safe(compact.group(1), 0))
            return text.count('*')

        def _extract_cell_points(v):
            """
            Extract daily points from a date cell.
            Accepts numeric cells and mixed text like '20*' or '-15 V'.
            """
            if v is None:
                return None
            if isinstance(v, bool):
                return int(v)
            if isinstance(v, (int, float)):
                return int(round(v))
            text = str(v).strip()
            if not text:
                return None
            m = re.search(r'-?\d+(?:\.\d+)?', text.replace(',', ''))
            if not m:
                return None
            try:
                return int(round(float(m.group(0))))
            except Exception:
                return None

        # ── Process each sheet ─────────────────────────────────────────────
        now_iso = _server_now_iso()
        existing_scores = list(snapshot.get('scores') or [])
        max_id = max((_parse_int_safe(sc.get('id')) for sc in existing_scores), default=0)

        months_cleared = []
        months_processed = []
        month_report = {}
        total_imported = 0
        total_star_markers = 0

        for sheet_name, month_key in sorted(target_sheets, key=lambda x: x[1]):
            ws = wb[sheet_name]
            hdr_row_num, header_row = _detect_excel_header_row(ws)

            # Find roll column from the header row
            roll_col = None
            name_col = None
            total_col = None
            final_score_col = None
            for idx, val in enumerate(header_row, start=1):
                if not isinstance(val, str):
                    continue
                s = val.strip().lower()
                if roll_col is None and ('roll' in s):
                    roll_col = idx
                if name_col is None and (('student' in s and 'name' in s) or s in {'name', 'candidate'}):
                    name_col = idx
                if total_col is None and 'total score' in s:
                    total_col = idx
                if final_score_col is None and 'final score' in s:
                    final_score_col = idx

            if total_col is None and final_score_col is not None:
                total_col = final_score_col

            if not roll_col:
                month_report[month_key] = {
                    'status': 'skipped', 'sheet': sheet_name,
                    'reason': 'roll column not found'
                }
                continue

            # Strictly include only date headers that belong to this month.
            date_cols = []
            for idx, val in enumerate(header_row, start=1):
                if _excel_col_excluded(val):
                    continue
                d = _to_date(val, month_hint=month_key)
                if not d:
                    continue
                mk = d.strftime('%Y-%m')
                if mk == month_key:
                    date_cols.append((idx, d))

            if not date_cols and not total_col:
                month_report[month_key] = {
                    'status': 'skipped', 'sheet': sheet_name,
                    'reason': 'no date-header columns and no total score column found'
                }
                continue

            # Clear all previously admin-imported scores for this month
            if month_key not in months_cleared:
                existing_scores = [
                    sc for sc in existing_scores
                    if not (
                        str(sc.get('month') or sc.get('date', '')[:7]) == month_key
                        # Historical months are rebuilt fully from Excel date columns.
                        # This avoids stale/duplicate rows causing inflated totals.
                    )
                ]
                months_cleared.append(month_key)

            # Rebuild month rows: one total row per student + per-date star usage rows.
            month_count = 0
            month_daily_score_rows = 0
            month_star_count = 0
            unknown_roll_rows = 0
            for row_idx in range(hdr_row_num + 1, (ws.max_row or 0) + 1):
                raw_roll = ws.cell(row_idx, roll_col).value
                if raw_roll is None:
                    continue
                roll = str(raw_roll).strip().upper()
                if not roll or roll.startswith('ROLL'):
                    continue
                full_name = ws.cell(row_idx, name_col).value if name_col else None
                sid = name_to_sid.get(_name_key(full_name)) if full_name else None
                if sid is None:
                    sid = roll_to_sid.get(roll)
                if sid is None:
                    unknown_roll_rows += 1
                    continue

                pts = None
                if total_col:
                    pts = _to_int_score(ws.cell(row_idx, total_col).value)
                if pts is None and date_cols:
                    # Fallback for sheets where total column is not available.
                    run_sum = 0
                    has_numeric = False
                    for col_idx, _d in date_cols:
                        score_val = ws.cell(row_idx, col_idx).value
                        if isinstance(score_val, (int, float)):
                            run_sum += int(round(score_val))
                            has_numeric = True
                    if has_numeric:
                        pts = run_sum

                # Keep workbook total alignment (including explicit zero totals).
                if pts is not None:
                    max_id += 1
                    existing_scores.append({
                        'id': max_id,
                        'studentId': sid,
                        'month': month_key,
                        'date': month_key + '-15',   # mid-month placeholder date
                        'points': pts,
                        'stars': 0,
                        'vetos': 0,
                        'notes': 'excel_total_score',
                        'recordedBy': 'admin',
                        'created_at': now_iso,
                        'updated_at': now_iso,
                    })
                    month_count += 1
                    total_imported += 1

                # Import historical daily score points from date columns.
                for col_idx, d in date_cols:
                    day_points = _extract_cell_points(ws.cell(row_idx, col_idx).value)
                    if day_points is None:
                        continue
                    # Skip no-op zeros to keep dataset compact.
                    if day_points == 0:
                        continue
                    max_id += 1
                    existing_scores.append({
                        'id': max_id,
                        'studentId': sid,
                        'month': month_key,
                        'date': d.isoformat(),
                        'points': day_points,
                        'stars': 0,
                        'vetos': 0,
                        'notes': 'excel_daily_score',
                        'recordedBy': 'admin',
                        'created_at': now_iso,
                        'updated_at': now_iso,
                    })
                    month_daily_score_rows += 1

                # Import historical star usage markers (for red '*' cell display).
                for col_idx, d in date_cols:
                    uses = _extract_star_usage(ws.cell(row_idx, col_idx).value)
                    if uses <= 0:
                        continue
                    max_id += 1
                    existing_scores.append({
                        'id': max_id,
                        'studentId': sid,
                        'month': month_key,
                        'date': d.isoformat(),
                        'points': 0,
                        'stars': -uses,
                        'vetos': 0,
                        'notes': f'excel_star_usage:{uses}',
                        'recordedBy': 'admin',
                        'created_at': now_iso,
                        'updated_at': now_iso,
                    })
                    month_star_count += 1
                    total_star_markers += 1

            months_processed.append(month_key)
            month_report[month_key] = {
                'status': 'ok', 'sheet': sheet_name,
                'students_imported': month_count,
                'daily_score_rows': month_daily_score_rows,
                'star_usage_rows': month_star_count,
                'header_row': hdr_row_num,
                'date_columns': len(date_cols),
                'total_col': total_col,
                'unknown_roll_rows': unknown_roll_rows,
            }

        # ── Persist updated snapshot ───────────────────────────────────────
        if months_processed:
            snapshot['scores'] = existing_scores
            snapshot['server_updated_at'] = now_iso
            snapshot['updated_at'] = now_iso
            _save_offline_data(snapshot)

        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception:
                pass

        return jsonify({
            'success': True,
            'message': (
                f'Imported totals for {total_imported} student-months across '
                f'{len(months_processed)} month(s) and '
                f'imported {total_star_markers} historical star-usage marker row(s). '
                f'Skipped {len(target_sheets) - len(months_processed)} sheet(s).'
            ),
            'total_imported': total_imported,
            'total_star_markers': total_star_markers,
            'months_processed': sorted(months_processed),
            'month_report': month_report,
        })

    except Exception as e:
        current_app.logger.error(f"Historical import error: {e}")
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception:
                pass
        return jsonify({'success': False, 'error': str(e)}), 500


@points_bp.route('/import-latest-roster', methods=['POST'])
@login_required
@_ledger_write_guard
def import_latest_roster():
    """
    Import latest roster for CURRENT MONTH ONLY
    Filters out any previous month data even if present in Excel
    Preserves student active/inactive status from system
    Only updates current month scoreboard scores
    """
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    try:
        from datetime import datetime as dt
        from dateutil.relativedelta import relativedelta

        current_month_start = dt.now().replace(day=1).date()
        next_month_start = (dt.now().replace(day=1) + relativedelta(months=1)).date()

        # File validation
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
            return jsonify({'success': False, 'error': 'Only Excel files are supported'}), 400

        # Parse Excel and filter to CURRENT MONTH ONLY
        import uuid
        temp_suffix = file.filename.split('.')[-1] if '.' in file.filename else 'xlsx'
        temp_file = tempfile.NamedTemporaryFile(
            mode='w+b',
            suffix=f'.{temp_suffix}',
            prefix=f'ea_roster_{uuid.uuid4().hex}_',
            delete=False
        )
        temp_path = temp_file.name
        temp_file.close()

        try:
            file.save(temp_path)
            import openpyxl
            wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True, keep_vba=False)
        except Exception as e:
            if os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except:
                    pass
            return jsonify({'success': False, 'error': f'Failed to read Excel: {str(e)}'}), 400

        ws = wb.active
        header_row = [cell.value for cell in ws[1]]

        header_map = {}
        for idx, value in enumerate(header_row, start=1):
            if isinstance(value, str):
                header_map[value.strip().lower()] = idx

        def find_header(candidates):
            for key in candidates:
                for header, idx in header_map.items():
                    if key in header:
                        return idx
            return None

        roll_col = find_header(['roll'])
        if not roll_col:
            os.remove(temp_path)
            return jsonify({'success': False, 'error': 'Roll column not found'}), 400

        # Find date columns - ONLY CURRENT MONTH
        date_columns = []
        excluded_dates = []

        for idx, header in enumerate(header_row, start=1):
            parsed_date = None
            if isinstance(header, (datetime, date)):
                parsed_date = header.date() if isinstance(header, datetime) else header
            elif isinstance(header, str):
                try:
                    parsed_date = dt.fromisoformat(header.strip()).date()
                except:
                    pass

            if parsed_date:
                if current_month_start <= parsed_date < next_month_start:
                    date_columns.append((idx, parsed_date))
                else:
                    excluded_dates.append(parsed_date.isoformat())

        if not date_columns:
            os.remove(temp_path)
            return jsonify({
                'success': False,
                'error': f'No current month dates found. All {len(excluded_dates)} dates are from previous months. Use "Historical Data" import instead.'
            }), 400

        # Count imported scores
        imported = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            roll_number = ws.cell(row_idx, roll_col).value
            if not roll_number:
                continue

            for col_idx, date_recorded in date_columns:
                score_value = ws.cell(row_idx, col_idx).value
                if score_value is not None and isinstance(score_value, (int, float)):
                    imported += 1

        os.remove(temp_path)

        return jsonify({
            'success': True,
            'message': f'Latest roster import: {imported} scores for current month',
            'imported_scores': imported,
            'date_columns': len(date_columns),
            'excluded_historical_dates': len(excluded_dates),
            'info': 'Historical data excluded. System settings preserved.'
        })

    except Exception as e:
        current_app.logger.error(f"Latest roster import error: {str(e)}")
        try:
            if 'temp_path' in locals() and os.path.exists(temp_path):
                os.unlink(temp_path)
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Points Transfer (admin/teacher-initiated) ──────────────────────────────
@points_bp.route('/transfer-points', methods=['POST'])
@csrf.exempt
@login_required
@_ledger_write_guard
def transfer_points():
    """Admin/teacher-initiated point transfer between two students.

    Role-based limits:
      Leader (L)          – up to 50 per transfer
      Co-Leader (CoL)     – up to 40
      Leadership posts    – up to 30
      All CRs             – up to 30
      All PPs / DPPs      – no hard cap (use available balance)
      Normal students     – can only donate to their own PP/DPP
    """
    role = str(current_user.role or '').strip().lower()
    if role not in ('admin', 'teacher'):
        return jsonify({'success': False, 'error': 'Admin or teacher role required'}), 403

    payload = request.get_json(silent=True) or {}
    from_student_id = _parse_int_safe(payload.get('from_student_id'), 0)
    to_student_id = _parse_int_safe(payload.get('to_student_id'), 0)
    amount = _parse_int_safe(payload.get('amount'), 0)
    reason = str(payload.get('reason') or '').strip()[:200]
    transfer_date = str(payload.get('date') or '').strip()

    if from_student_id <= 0 or to_student_id <= 0:
        return jsonify({'success': False, 'error': 'Valid from/to student IDs required'}), 400
    if from_student_id == to_student_id:
        return jsonify({'success': False, 'error': 'Cannot transfer to self'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'error': 'Amount must be positive'}), 400

    now = datetime.utcnow()
    now_iso = _server_now_iso()
    now_date = _parse_date_key(now_iso) or date.today()
    month_key = now_date.strftime('%Y-%m')

    # Resolve transfer date (default to today)
    t_date = _parse_date_key(transfer_date) if transfer_date else now_date
    if not t_date:
        t_date = now_date
    t_date_iso = t_date.isoformat()
    t_month = t_date.strftime('%Y-%m')

    data = _load_offline_data() or {}
    if not _is_month_allowed_for_user(data, current_user, t_month):
        return jsonify({'success': False, 'error': 'No access to target month'}), 403

    # Lookup students
    students = data.get('students', []) or []
    sender = next((s for s in students if _parse_int_safe(s.get('id'), 0) == from_student_id), None)
    receiver = next((s for s in students if _parse_int_safe(s.get('id'), 0) == to_student_id), None)
    if not sender:
        return jsonify({'success': False, 'error': 'Sender not found'}), 404
    if not receiver:
        return jsonify({'success': False, 'error': 'Receiver not found'}), 404

    # ── Determine sender's role for limit enforcement ─────────────────────
    sender_limit = 0  # 0 means "no cap" (PP/DPP); will be set for others
    sender_is_post_holder = False
    sender_is_cr = False
    sender_is_pp = False
    sender_is_dpp = False
    sender_party_code = ''
    sender_leadership_type = ''

    check_date = t_date

    # Check leadership posts
    for post in data.get('leadership', []) or []:
        if _parse_int_safe(post.get('studentId'), 0) != from_student_id:
            continue
        if not _is_active_assignment(post, 'leadership', check_date):
            continue
        sender_is_post_holder = True
        role_type = _leadership_role_type(post.get('post'))
        if role_type == 'leader':
            sender_leadership_type = 'leader'
            sender_limit = 50
        elif role_type == 'co_leader':
            sender_leadership_type = 'co_leader'
            sender_limit = 40
        elif role_type == 'lop':
            sender_leadership_type = 'lop'
            sender_limit = 30
        else:
            # Other leadership posts (captains, in-charges)
            sender_limit = 30
        break

    # Check CR posts
    if not sender_is_post_holder:
        for post in data.get('class_reps', []) or []:
            if _parse_int_safe(post.get('studentId'), 0) == from_student_id and _is_active_assignment(post, 'class_rep', check_date):
                sender_is_cr = True
                sender_limit = 30
                break
        if not sender_is_cr:
            for post in data.get('group_crs', []) or []:
                if _parse_int_safe(post.get('studentId'), 0) == from_student_id and _is_active_assignment(post, 'group_cr', check_date):
                    sender_is_cr = True
                    sender_limit = 30
                    break

    # Check party membership (PP/DPP)
    for party in data.get('parties', []) or []:
        for member in party.get('members', []) or []:
            if _parse_int_safe(member.get('studentId'), 0) != from_student_id:
                continue
            m_status = str(member.get('status') or 'active').strip().lower()
            if m_status not in ('active', ''):
                continue
            sender_party_code = str(party.get('code') or '').strip().upper()
            designation = str(member.get('designation') or '').strip().lower()
            if designation in ('party president', 'pp'):
                sender_is_pp = True
                sender_limit = 0  # no cap
            elif designation in ('deputy party president', 'dpp'):
                sender_is_dpp = True
                sender_limit = 0  # no cap
            break
        if sender_party_code:
            break

    is_sender_privileged = sender_is_post_holder or sender_is_cr or sender_is_pp or sender_is_dpp

    # ── Permission checks ─────────────────────────────────────────────────
    if not is_sender_privileged:
        # Normal student: can only donate to their own PP/DPP
        if not sender_party_code:
            return jsonify({'success': False, 'error': 'Normal students can only transfer to their PP/DPP'}), 403
        # Find sender's party PP/DPP
        sender_party = next((p for p in data.get('parties', []) or [] if str(p.get('code', '')).strip().upper() == sender_party_code), None)
        if not sender_party:
            return jsonify({'success': False, 'error': 'Sender party not found'}), 404
        pp_dpp_ids = set()
        for member in sender_party.get('members', []) or []:
            designation = str(member.get('designation') or '').strip().lower()
            if designation in ('party president', 'pp', 'deputy party president', 'dpp'):
                mid = _parse_int_safe(member.get('studentId'), 0)
                if mid > 0:
                    pp_dpp_ids.add(mid)
        if to_student_id not in pp_dpp_ids:
            return jsonify({'success': False, 'error': 'Normal students can only transfer to their PP/DPP'}), 403
        sender_limit = 30  # normal student donation cap

    elif sender_is_pp or sender_is_dpp:
        # PP/DPP: can transfer between party members only
        if not sender_party_code:
            return jsonify({'success': False, 'error': 'PP/DPP party not found'}), 409
        sender_party = next((p for p in data.get('parties', []) or [] if str(p.get('code', '')).strip().upper() == sender_party_code), None)
        if not sender_party:
            return jsonify({'success': False, 'error': 'Sender party not found'}), 404
        party_member_ids = set()
        for member in sender_party.get('members', []) or []:
            mid = _parse_int_safe(member.get('studentId'), 0)
            if mid > 0:
                party_member_ids.add(mid)
        if to_student_id not in party_member_ids:
            return jsonify({'success': False, 'error': 'PP/DPP can only transfer within party members'}), 403
    # else: post holder / CR / leader / co-leader can transfer to any student

    # ── Enforce transfer limit ─────────────────────────────────────────────
    if sender_limit > 0 and amount > sender_limit:
        return jsonify({'success': False, 'error': f'Transfer exceeds limit of {sender_limit} for this role'}), 400

    # ── Check sender balance ──────────────────────────────────────────────
    sender_balance = _sum_points_for_student_month(data, from_student_id, t_month)
    if sender_balance <= 0:
        return jsonify({'success': False, 'error': 'Sender has no transferable points this month'}), 409
    if amount > sender_balance:
        return jsonify({'success': False, 'error': f'Transfer exceeds sender available balance ({sender_balance})'}), 409

    # ── Execute transfer ──────────────────────────────────────────────────
    from_roll = str(sender.get('roll') or '').strip()
    to_roll = str(receiver.get('roll') or '').strip()
    reason_part = f' — {reason}' if reason else ''
    note_out = f'[PTS TRANSFER OUT -{amount} to {to_roll}]{reason_part}'
    note_in = f'[PTS TRANSFER IN +{amount} from {from_roll}]{reason_part}'

    _upsert_score_delta(data, from_student_id, t_date_iso, t_month,
                        delta_points=-amount, delta_stars=0, note=note_out)
    _upsert_score_delta(data, to_student_id, t_date_iso, t_month,
                        delta_points=amount, delta_stars=0, note=note_in)

    # Record in StudentTransfer table
    transfer_row = StudentTransfer(
        from_student_id=from_student_id,
        to_student_id=to_student_id,
        transfer_type='points',
        amount=amount,
        created_at=now,
        lock_until=now + timedelta(hours=24),
    )
    try:
        db.session.add(transfer_row)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 500
    _ensure_score_timestamps(data)
    data['server_updated_at'] = now_iso
    _save_offline_data(data)
    _broadcast_sync_event(now_iso, source='points-transfer')
    _forward_offline_data_to_peers_async(data, [])

    return jsonify({
        'success': True,
        'message': f'Transferred {amount} point(s): {from_roll} -> {to_roll}',
        'from_roll': from_roll,
        'to_roll': to_roll,
        'amount': amount,
        'date': t_date_iso,
        'updated_at': now_iso,
    })


# ── DEBUG: Temporary cache verification endpoint ────────────────────────────
@points_bp.route('/__debug_cache_status')
@login_required
def _debug_cache_status():
    """Debug endpoint: verify caching is working. Admin only."""
    if current_user.role != 'admin':
        abort(403)
    from app.utils.data_paths import _data_cache
    import time

    # Measure load time
    t1 = time.perf_counter()
    data = _load_offline_data()
    elapsed_ms = (time.perf_counter() - t1) * 1000

    return jsonify({
        'cache_status': {
            'path': _data_cache.get('path'),
            'mtime_ns': _data_cache.get('mtime_ns'),
            'size_bytes': _data_cache.get('size'),
            'has_data': _data_cache.get('data') is not None,
        },
        'last_load_ms': round(elapsed_ms, 2),
        'students_count': len(data.get('students', [])),
        'scores_count': len(data.get('scores', [])),
        'active_file_path': _offline_data_path(),
    })
