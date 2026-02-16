#!/usr/bin/env python
"""
Smart Excel extractor for EA scoreboard workbooks.

Features:
- Reads only visible sheets, rows, and columns
- Detects student header row heuristically
- Extracts known fields (roll, name, class, fees, totals, stars, veto)
- Extracts day-wise score columns
- Extracts trailing non-date columns as month-wise extra columns
- Outputs normalized JSON for auditing/import workflows
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


MONTH_LOOKUP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


KNOWN_KEYWORDS = {
    "roll": ("roll",),
    "name": ("student", "name", "candidate"),
    "class": ("class",),
    "fees": ("fees",),
    "vote_power": ("vote power", "votepower", "vote"),
    "total": ("total", "final score", "final"),
    "rank": ("rank",),
    "stars": ("stars", "star"),
    "veto": ("veto", "vetoes", "vetos"),
}


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_key(value: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return key[:48] if key else ""


def infer_month_from_sheet_name(sheet_name: str) -> Tuple[Optional[int], Optional[int]]:
    text = sheet_name.lower()
    month = None
    for token, number in MONTH_LOOKUP.items():
        if token in text:
            month = number
            break
    year = None
    year_match = re.search(r"(20\d{2}|\d{2})", text)
    if year_match:
        raw = year_match.group(1)
        year = int(raw) if len(raw) == 4 else 2000 + int(raw)
    return month, year


def is_relevant_extra_header(header: str) -> bool:
    if not header:
        return False
    if re.fullmatch(r"column\d*", header, re.I):
        return False
    if re.fullmatch(r"unnamed.*", header, re.I):
        return False
    if re.fullmatch(r"\d+(\.\d+)?", header):
        return False
    return True


def is_date_header(value: object, month: Optional[int], year: Optional[int]) -> bool:
    if value is None:
        return False
    if isinstance(value, datetime):
        return True
    text = normalize_text(value)
    if not text:
        return False
    if re.fullmatch(r"\d{1,2}", text):
        return bool(month and year)
    if re.search(r"\b\d{1,2}\s*[A-Za-z]{3,}\b", text):
        return True
    try:
        datetime.fromisoformat(text)
        return True
    except Exception:
        return False


def parse_score_cell(value: object) -> Dict[str, int | float | None]:
    if value is None or normalize_text(value) == "":
        return {"points": None, "stars": 0, "vetos": 0}
    if isinstance(value, (int, float)):
        return {"points": value, "stars": 0, "vetos": 0}
    text = normalize_text(value)
    stars = len(re.findall(r"\*", text))
    vetos = sum(len(x) for x in re.findall(r"v+", text, flags=re.I))
    match = re.search(r"-?\d+(\.\d+)?", text)
    points = float(match.group(0)) if match else None
    if points is not None and points.is_integer():
        points = int(points)
    return {"points": points, "stars": stars, "vetos": vetos}


def detect_header_row(rows: List[List[object]]) -> int:
    for idx, row in enumerate(rows):
        lowered = [normalize_text(v).lower() for v in row]
        has_roll = any("roll" in x for x in lowered)
        has_name = any(("name" in x or "student" in x or "candidate" in x) for x in lowered)
        if has_roll and has_name:
            return idx
    return -1


def detect_column_index(header_row: List[object], candidates: Tuple[str, ...]) -> int:
    lowered = [normalize_text(v).lower() for v in header_row]
    for i, value in enumerate(lowered):
        if any(token in value for token in candidates):
            return i
    return -1


@dataclass
class ExtractResult:
    students: List[dict]
    scores: List[dict]
    month_students: Dict[str, List[str]]
    month_extra_columns: Dict[str, List[dict]]
    month_student_extras: Dict[str, Dict[str, dict]]
    scanned_sheets: List[str]


def extract_workbook(path: Path) -> ExtractResult:
    wb = load_workbook(path, data_only=True, keep_vba=True)
    students_map: Dict[str, dict] = {}
    scores: List[dict] = []
    month_students: Dict[str, set] = {}
    month_extra_columns: Dict[str, List[dict]] = {}
    month_student_extras: Dict[str, Dict[str, dict]] = {}
    scanned_sheets: List[str] = []
    student_id = 1
    score_id = 1

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        scanned_sheets.append(ws.title)
        month, year = infer_month_from_sheet_name(ws.title)

        hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
        hidden_cols = {
            idx - 1
            for idx in range(1, ws.max_column + 1)
            if ws.column_dimensions.get(get_column_letter(idx)) and ws.column_dimensions[get_column_letter(idx)].hidden
        }

        rows: List[List[object]] = []
        for r in range(1, ws.max_row + 1):
            if r in hidden_rows:
                continue
            row_values: List[object] = []
            for c in range(1, ws.max_column + 1):
                if (c - 1) in hidden_cols:
                    continue
                row_values.append(ws.cell(r, c).value)
            rows.append(row_values)

        if not rows:
            continue

        header_idx = detect_header_row(rows)
        if header_idx < 0:
            continue
        header = rows[header_idx]

        roll_col = detect_column_index(header, KNOWN_KEYWORDS["roll"])
        name_col = detect_column_index(header, KNOWN_KEYWORDS["name"])
        class_col = detect_column_index(header, KNOWN_KEYWORDS["class"])
        fees_col = detect_column_index(header, KNOWN_KEYWORDS["fees"])
        vote_col = detect_column_index(header, KNOWN_KEYWORDS["vote_power"])
        total_col = detect_column_index(header, KNOWN_KEYWORDS["total"])
        rank_col = detect_column_index(header, KNOWN_KEYWORDS["rank"])
        stars_col = detect_column_index(header, KNOWN_KEYWORDS["stars"])
        veto_col = detect_column_index(header, KNOWN_KEYWORDS["veto"])

        known_cols = {
            i
            for i in [roll_col, name_col, class_col, fees_col, vote_col, total_col, rank_col, stars_col, veto_col]
            if i >= 0
        }
        date_cols = [i for i, cell in enumerate(header) if i not in known_cols and is_date_header(cell, month, year)]
        month_key = f"{year:04d}-{month:02d}" if month and year else None
        if not month_key and date_cols:
            month_key = datetime.now().strftime("%Y-%m")

        if month_key and month_key not in month_students:
            month_students[month_key] = set()
        if month_key and month_key not in month_extra_columns:
            month_extra_columns[month_key] = []
        if month_key and month_key not in month_student_extras:
            month_student_extras[month_key] = {}

        trailing_start = (max(date_cols) + 1) if date_cols else len(header)
        seen_extra = {col["key"] for col in month_extra_columns.get(month_key or "", [])}
        extra_cols: List[Tuple[int, str, str]] = []
        for idx in range(trailing_start, len(header)):
            if idx in known_cols or idx in date_cols:
                continue
            label = normalize_text(header[idx])
            if not is_relevant_extra_header(label):
                continue
            base_key = normalize_key(label) or f"extra_{idx + 1}"
            key = base_key
            suffix = 2
            while key in seen_extra:
                key = f"{base_key}_{suffix}"
                suffix += 1
            seen_extra.add(key)
            extra_cols.append((idx, key, label))
        if month_key and extra_cols:
            month_extra_columns[month_key].extend({"key": k, "label": l} for _, k, l in extra_cols)

        for row in rows[header_idx + 1 :]:
            roll = normalize_text(row[roll_col]) if roll_col >= 0 and roll_col < len(row) else ""
            name = normalize_text(row[name_col]) if name_col >= 0 and name_col < len(row) else ""
            if not roll or not name:
                continue
            if not re.fullmatch(r"EA[A-Z0-9]{5}", roll.upper()):
                continue

            row_key = roll.upper()
            if row_key not in students_map:
                students_map[row_key] = {
                    "id": student_id,
                    "roll": roll,
                    "name": name,
                    "class": None,
                    "fees": None,
                    "total_score": None,
                    "rank": None,
                    "vote_power": None,
                }
                student_id += 1

            student = students_map[row_key]
            if month_key:
                month_students[month_key].add(row_key)

            def parse_int(col_idx: int) -> Optional[int]:
                if col_idx < 0 or col_idx >= len(row):
                    return None
                value = row[col_idx]
                if value is None or normalize_text(value) == "":
                    return None
                try:
                    return int(float(value))
                except Exception:
                    return None

            class_val = parse_int(class_col)
            fees_val = parse_int(fees_col)
            total_val = parse_int(total_col)
            rank_val = parse_int(rank_col)
            vote_val = parse_int(vote_col)
            if class_val is not None:
                student["class"] = class_val
            if fees_val is not None:
                student["fees"] = fees_val
            if total_val is not None:
                student["total_score"] = total_val
            if rank_val is not None:
                student["rank"] = rank_val
            if vote_val is not None:
                student["vote_power"] = vote_val

            if month_key and extra_cols:
                slot = month_student_extras[month_key].setdefault(row_key, {})
                for idx, key, _ in extra_cols:
                    if idx >= len(row):
                        continue
                    value = row[idx]
                    if value is None or normalize_text(value) == "":
                        continue
                    slot[key] = value

            for d_idx in date_cols:
                if d_idx >= len(row):
                    continue
                parsed = parse_score_cell(row[d_idx])
                if parsed["points"] is None:
                    continue
                if month_key:
                    day_header = normalize_text(header[d_idx])
                    if re.fullmatch(r"\d{1,2}", day_header):
                        date_str = f"{month_key}-{int(day_header):02d}"
                    else:
                        date_str = f"{month_key}-01"
                else:
                    date_str = datetime.now().strftime("%Y-%m-%d")
                scores.append(
                    {
                        "id": score_id,
                        "studentId": student["id"],
                        "date": date_str,
                        "month": date_str[:7],
                        "points": parsed["points"],
                        "stars": parsed["stars"],
                        "vetos": parsed["vetos"],
                        "notes": "",
                    }
                )
                score_id += 1

    return ExtractResult(
        students=list(students_map.values()),
        scores=scores,
        month_students={k: sorted(list(v)) for k, v in month_students.items()},
        month_extra_columns=month_extra_columns,
        month_student_extras=month_student_extras,
        scanned_sheets=scanned_sheets,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract visible scoreboard data from EA Excel workbook.")
    parser.add_argument("excel_file", help="Path to .xlsx/.xlsm workbook")
    parser.add_argument("-o", "--out", help="Output JSON file path", default="smart_extract_output.json")
    args = parser.parse_args()

    workbook_path = Path(args.excel_file)
    if not workbook_path.exists():
        raise SystemExit(f"File not found: {workbook_path}")

    result = extract_workbook(workbook_path)
    output = {
        "students": result.students,
        "scores": result.scores,
        "month_students": result.month_students,
        "month_extra_columns": result.month_extra_columns,
        "month_student_extras": result.month_student_extras,
        "scanned_sheets": result.scanned_sheets,
        "generated_at": datetime.now().isoformat(),
    }

    out_path = Path(args.out)
    out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Extraction complete: {out_path}")
    print(f"Visible sheets scanned: {len(result.scanned_sheets)}")
    print(f"Students: {len(result.students)}")
    print(f"Scores: {len(result.scores)}")
    print(f"Months: {len(result.month_students)}")


if __name__ == "__main__":
    main()

