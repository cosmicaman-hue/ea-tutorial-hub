#!/usr/bin/env python
"""
Reconcile one active month from workbook into offline scoreboard JSON.

Use this to keep historical data intact while forcing current-month roster/scores
to exactly match the workbook sheet (e.g. "Feb 26").
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


MONTH_LOOKUP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def normalize_roll(value: object) -> str:
    return str(value or "").strip().upper()


def normalize_name_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\*+", "", text)
    text = re.sub(r"\((v+)\)", "", text, flags=re.I)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def parse_name_meta(raw_name: object) -> Tuple[str, int, int]:
    text = str(raw_name or "").strip()
    stars = len(re.findall(r"\*", text))
    vetos = sum(len(match) for match in re.findall(r"v+", text, flags=re.I))
    base = re.sub(r"\*+", "", text)
    base = re.sub(r"\((v+)\)", "", base, flags=re.I)
    base = re.sub(r"\s{2,}", " ", base).strip()
    return (base or text, stars, vetos)


def parse_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return int(round(value))
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"-?\d+", text)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def infer_month_year(sheet_name: str) -> Tuple[Optional[int], Optional[int]]:
    text = sheet_name.lower()
    month = None
    for token, number in MONTH_LOOKUP.items():
        if token in text:
            month = number
            break
    year = None
    year_match = re.search(r"(20\d{2}|\d{2})", text)
    if year_match:
        raw = year_match.group(1)
        year = int(raw) if len(raw) == 4 else (2000 + int(raw))
    return month, year


def visible_matrix(ws) -> List[List[object]]:
    hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
    hidden_cols = {
        idx - 1
        for idx in range(1, ws.max_column + 1)
        if ws.column_dimensions.get(get_column_letter(idx)) and ws.column_dimensions[get_column_letter(idx)].hidden
    }
    rows: List[List[object]] = []
    for r in range(1, ws.max_row + 1):
        if r in hidden_rows:
            continue
        row_values = []
        for c in range(1, ws.max_column + 1):
            if (c - 1) in hidden_cols:
                continue
            row_values.append(ws.cell(r, c).value)
        rows.append(row_values)
    return rows


def find_header_row(rows: List[List[object]]) -> int:
    for idx, row in enumerate(rows):
        lowered = [str(v or "").strip().lower() for v in row]
        has_roll = any("roll" in cell for cell in lowered)
        has_name = any(("student" in cell and "name" in cell) or cell == "name" for cell in lowered)
        if has_roll and has_name:
            return idx
    return -1


def detect_col(header: List[object], keywords: Tuple[str, ...]) -> int:
    lowered = [str(v or "").strip().lower() for v in header]
    for idx, cell in enumerate(lowered):
        if any(k in cell for k in keywords):
            return idx
    return -1


def parse_date_header(value: object, month: Optional[int], year: Optional[int]) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    day_num = parse_int(value)
    if day_num is not None and month and year and 1 <= day_num <= 31:
        try:
            return date(year, month, day_num).isoformat()
        except ValueError:
            return None

    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except Exception:
        return None


def derive_group(roll: str) -> str:
    match = re.match(r"^EA\d{2}([A-Z])", normalize_roll(roll))
    return match.group(1) if match else ""


def load_sheet_profiles(workbook_path: Path, sheet_name: str) -> Tuple[List[dict], List[dict], str]:
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    month, year = infer_month_year(ws.title)
    month_key = f"{year:04d}-{month:02d}" if month and year else datetime.now().strftime("%Y-%m")

    rows = visible_matrix(ws)
    if not rows:
        raise RuntimeError("No visible rows found in selected worksheet")

    header_idx = find_header_row(rows)
    if header_idx < 0:
        raise RuntimeError("Could not locate header row with Roll and Student Name")
    header = rows[header_idx]

    roll_col = detect_col(header, ("roll",))
    name_col = detect_col(header, ("student name", "name"))
    class_col = detect_col(header, ("class",))
    fees_col = detect_col(header, ("fees",))
    total_col = detect_col(header, ("total score", "total"))
    rank_col = detect_col(header, ("rank",))
    vote_col = detect_col(header, ("vote power", "votepower", "vote"))

    if roll_col < 0 or name_col < 0:
        raise RuntimeError("Missing roll/name columns")

    known_cols = {idx for idx in [roll_col, name_col, class_col, fees_col, total_col, rank_col, vote_col] if idx >= 0}
    date_cols: List[Tuple[int, str]] = []
    for idx, cell in enumerate(header):
        if idx in known_cols:
            continue
        parsed = parse_date_header(cell, month, year)
        if parsed:
            date_cols.append((idx, parsed))
    if not date_cols:
        raise RuntimeError("No date columns detected on sheet")

    profiles: List[dict] = []
    month_scores: List[dict] = []
    for row in rows[header_idx + 1 :]:
        if roll_col >= len(row) or name_col >= len(row):
            continue
        roll = normalize_roll(row[roll_col])
        raw_name = str(row[name_col] or "").strip()
        if not roll or not raw_name:
            continue
        if not re.fullmatch(r"EA[A-Z0-9]{5}", roll):
            continue
        base_name, stars, vetos = parse_name_meta(raw_name)
        if not base_name:
            continue

        class_val = parse_int(row[class_col]) if class_col >= 0 and class_col < len(row) else None
        fees_val = parse_int(row[fees_col]) if fees_col >= 0 and fees_col < len(row) else None
        total_val = parse_int(row[total_col]) if total_col >= 0 and total_col < len(row) else None
        rank_val = parse_int(row[rank_col]) if rank_col >= 0 and rank_col < len(row) else None
        vote_val = parse_int(row[vote_col]) if vote_col >= 0 and vote_col < len(row) else None

        profiles.append({
            "roll": roll,
            "name": raw_name,
            "base_name": base_name,
            "class": class_val,
            "fees": fees_val,
            "total_score": total_val,
            "rank": rank_val,
            "vote_power": vote_val,
            "stars": stars,
            "veto_count": vetos,
        })

        for col_idx, score_date in date_cols:
            if col_idx >= len(row):
                continue
            points = parse_int(row[col_idx])
            if points is None:
                continue
            month_scores.append({
                "roll": roll,
                "date": score_date,
                "month": score_date[:7],
                "points": points,
                "stars": 0,
                "vetos": 0,
                "notes": "",
            })

    if not profiles:
        raise RuntimeError("No student rows parsed from selected worksheet")
    return profiles, month_scores, month_key


def choose_student_for_roll(students: List[dict], roll: str, base_name: str) -> Optional[dict]:
    candidates = [student for student in students if normalize_roll(student.get("roll")) == roll]
    if not candidates:
        return None
    target_name = normalize_name_key(base_name)
    for student in candidates:
        if normalize_name_key(student.get("base_name") or student.get("name")) == target_name:
            return student
    for student in candidates:
        if student.get("active") is not False:
            return student
    return candidates[0]


def reconcile(data_path: Path, workbook_path: Path, sheet_name: str, explicit_month: Optional[str]) -> None:
    profiles, month_scores, detected_month = load_sheet_profiles(workbook_path, sheet_name)
    month_key = explicit_month or detected_month

    if not data_path.exists():
        raise RuntimeError(f"Data file not found: {data_path}")
    data = json.loads(data_path.read_text(encoding="utf-8"))
    students = list(data.get("students", []) or [])
    scores = list(data.get("scores", []) or [])
    month_students = dict(data.get("month_students", {}) or {})
    month_roster_profiles = dict(data.get("month_roster_profiles", {}) or {})

    max_student_id = 0
    for student in students:
        try:
            max_student_id = max(max_student_id, int(student.get("id") or 0))
        except Exception:
            continue

    roll_to_student: Dict[str, dict] = {}
    active_rolls = set()

    for profile in profiles:
        roll = profile["roll"]
        chosen = choose_student_for_roll(students, roll, profile["base_name"])
        if not chosen:
            max_student_id += 1
            chosen = {
                "id": max_student_id,
                "roll": roll,
                "name": profile["name"],
                "base_name": profile["base_name"],
                "raw_name": profile["name"],
                "class": profile["class"] if profile["class"] is not None else 0,
                "fees": profile["fees"] if profile["fees"] is not None else 0,
                "total_score": profile["total_score"],
                "rank": profile["rank"],
                "vote_power": profile["vote_power"],
                "stars": profile["stars"],
                "veto_count": profile["veto_count"],
                "group": derive_group(roll),
                "active": True,
            }
            students.append(chosen)

        chosen["roll"] = roll
        chosen["name"] = profile["name"]
        chosen["base_name"] = profile["base_name"]
        chosen["raw_name"] = profile["name"]
        if profile["class"] is not None:
            chosen["class"] = profile["class"]
        if profile["fees"] is not None:
            chosen["fees"] = profile["fees"]
        chosen["total_score"] = profile["total_score"]
        chosen["rank"] = profile["rank"]
        if profile["vote_power"] is not None:
            chosen["vote_power"] = profile["vote_power"]
        chosen["stars"] = max(0, int(profile["stars"] or 0))
        chosen["veto_count"] = max(0, int(profile["veto_count"] or 0))
        chosen["group"] = chosen.get("group") or derive_group(roll)
        chosen["active"] = True

        active_rolls.add(roll)
        roll_to_student[roll] = chosen

    for student in students:
        if normalize_roll(student.get("roll")) not in active_rolls:
            student["active"] = False

    roster_profiles = []
    for profile in profiles:
        roster_profiles.append({
            "roll": profile["roll"],
            "name": profile["name"],
            "base_name": profile["base_name"],
            "class": profile["class"],
        })
    month_roster_profiles[month_key] = roster_profiles
    month_students[month_key] = sorted(active_rolls)

    kept_scores = []
    max_score_id = 0
    for score in scores:
        if str(score.get("month") or "") == month_key:
            continue
        kept_scores.append(score)
        try:
            max_score_id = max(max_score_id, int(score.get("id") or 0))
        except Exception:
            continue

    for score in month_scores:
        roll = score["roll"]
        student = roll_to_student.get(roll)
        if not student:
            continue
        max_score_id += 1
        kept_scores.append({
            "id": max_score_id,
            "studentId": student["id"],
            "date": score["date"],
            "points": int(score["points"] or 0),
            "stars": int(score.get("stars") or 0),
            "vetos": int(score.get("vetos") or 0),
            "month": score["month"],
            "notes": score.get("notes") or "",
            "recordedBy": "admin"
        })

    now_iso = datetime.now().isoformat()
    data["students"] = students
    data["scores"] = kept_scores
    data["month_students"] = month_students
    data["month_roster_profiles"] = month_roster_profiles
    data["updated_at"] = now_iso
    data["server_updated_at"] = now_iso

    backup_dir = data_path.parent / "offline_scoreboard_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"offline_scoreboard_pre_reconcile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    backup_path.write_text(json.dumps(json.loads(data_path.read_text(encoding="utf-8")), indent=2, ensure_ascii=False), encoding="utf-8")

    data_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    totals = {}
    for score in data["scores"]:
        if str(score.get("month")) != month_key:
            continue
        sid = score.get("studentId")
        totals[sid] = totals.get(sid, 0) + int(score.get("points") or 0)
    student_by_id = {student.get("id"): student for student in students}
    top = sorted(
        [
            (points, student_by_id.get(sid, {}).get("roll"), student_by_id.get(sid, {}).get("base_name") or student_by_id.get(sid, {}).get("name"))
            for sid, points in totals.items()
        ],
        key=lambda item: (-item[0], str(item[1] or ""))
    )

    print(f"Workbook sheet reconciled: {sheet_name}")
    print(f"Target month: {month_key}")
    print(f"Active students set from sheet: {len(active_rolls)}")
    print(f"Month scores rewritten: {len([s for s in data['scores'] if str(s.get('month')) == month_key])}")
    print("Top 10 after reconcile:")
    for idx, row in enumerate(top[:10], start=1):
        print(f"{idx:02d}. {row[1]} | {row[2]} | {row[0]}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile current month from workbook into offline JSON.")
    parser.add_argument("--xlsx", required=True, help="Workbook path (xlsx/xlsm)")
    parser.add_argument("--sheet", default="Feb 26", help="Sheet name to read")
    parser.add_argument("--month", default=None, help="Force month key (YYYY-MM)")
    parser.add_argument("--data", default="instance/offline_scoreboard_data.json", help="Offline JSON data path")
    args = parser.parse_args()

    reconcile(
        data_path=Path(args.data),
        workbook_path=Path(args.xlsx),
        sheet_name=args.sheet,
        explicit_month=args.month
    )


if __name__ == "__main__":
    main()
