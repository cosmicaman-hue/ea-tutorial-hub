#!/usr/bin/env python
"""
Harden offline scoreboard data from authoritative workbook.

This script:
- reads only visible sheets/rows/columns from workbook
- extracts month-wise rosters and date-wise scores
- avoids non-month sheets (Ranking/Pivot)
- deduplicates students by roll globally
- rebuilds scoreboard-critical sections in offline_scoreboard_data.json
- preserves app-only sections (appeals, attendance, parties, leadership, etc.)
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def normalize_roll(value: object) -> str:
    return str(value or "").strip().upper()


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_name_key(value: object) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"\*+", "", text)
    text = re.sub(r"\((v+)\)", "", text, flags=re.I)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def parse_name_meta(raw_name: object) -> Tuple[str, int, int]:
    text = normalize_text(raw_name)
    stars = len(re.findall(r"\*", text))
    # Count letters v/V occurrences; many cells use V / (vv) forms.
    vetos = len(re.findall(r"v", text, flags=re.I))
    base = re.sub(r"\*+", "", text)
    base = re.sub(r"\((v+)\)", "", base, flags=re.I)
    base = re.sub(r"\s{2,}", " ", base).strip()
    return (base or text, stars, vetos)


def parse_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(\.\d+)?", text)
    if not match:
        return None
    try:
        return int(round(float(match.group(0))))
    except ValueError:
        return None


def infer_sheet_month_year(sheet_name: str) -> Tuple[Optional[int], Optional[int]]:
    lowered = sheet_name.lower()
    month = None
    for token, number in MONTH_MAP.items():
        if token in lowered:
            month = number
            break
    year = None
    match = re.search(r"(20\d{2}|\b\d{2}\b)", lowered)
    if match:
        token = match.group(1)
        year = int(token) if len(token) == 4 else 2000 + int(token)
    return month, year


def is_month_sheet(sheet_name: str) -> bool:
    lowered = sheet_name.lower()
    if "ranking" in lowered or "pivot" in lowered:
        return False
    month, year = infer_sheet_month_year(sheet_name)
    return month is not None and year is not None


def visible_rows_matrix(ws) -> List[List[object]]:
    hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
    hidden_cols = {
        idx - 1
        for idx in range(1, ws.max_column + 1)
        if ws.column_dimensions.get(get_column_letter(idx)) and ws.column_dimensions[get_column_letter(idx)].hidden
    }
    rows: List[List[object]] = []
    for r in range(1, ws.max_row + 1):
        if r in hidden_rows:
            continue
        row_values = []
        for c in range(1, ws.max_column + 1):
            if (c - 1) in hidden_cols:
                continue
            row_values.append(ws.cell(r, c).value)
        rows.append(row_values)
    return rows


def find_header_row(rows: List[List[object]]) -> int:
    for idx, row in enumerate(rows):
        lowered = [normalize_text(cell).lower() for cell in row]
        has_roll = any("roll" in value for value in lowered)
        has_name = any(("student" in value and "name" in value) or value == "candidate" or value == "name" for value in lowered)
        if has_roll and has_name:
            return idx
    return -1


def detect_col(header: List[object], keywords: Tuple[str, ...]) -> int:
    lowered = [normalize_text(cell).lower() for cell in header]
    for idx, value in enumerate(lowered):
        if any(keyword in value for keyword in keywords):
            return idx
    return -1


def parse_date_header(value: object, sheet_month: Optional[int], sheet_year: Optional[int]) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return None
    numeric_day = parse_int(text)
    if numeric_day is not None and sheet_month and sheet_year and 1 <= numeric_day <= 31:
        try:
            return date(sheet_year, sheet_month, numeric_day).isoformat()
        except ValueError:
            return None
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except Exception:
        return None


def parse_score_cell(value: object) -> Tuple[Optional[int], int, int]:
    if value is None:
        return (None, 0, 0)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (int(round(float(value))), 0, 0)
    text = normalize_text(value)
    if not text:
        return (None, 0, 0)
    stars = len(re.findall(r"\*", text))
    vetos = len(re.findall(r"v", text, flags=re.I))
    points = parse_int(text)
    if points is None and (stars or vetos):
        points = 0
    return (points, stars, vetos)


def is_relevant_extra_header(label: str) -> bool:
    if not label:
        return False
    lowered = label.strip().lower()
    if not lowered:
        return False
    if re.fullmatch(r"column\d*", lowered):
        return False
    return True


def to_extra_key(label: str, fallback_idx: int) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", label.strip().lower()).strip("_")
    return key or f"extra_{fallback_idx}"


def month_sort_key(month_key: str) -> Tuple[int, int]:
    year, month = month_key.split("-")
    return (int(year), int(month))


def derive_group_from_roll(roll: str) -> str:
    roll_norm = normalize_roll(roll)
    match = re.match(r"^EA\d{2}([A-Z])", roll_norm)
    return match.group(1) if match else ""


def choose_existing_student(candidates: List[dict], preferred_base_name: str) -> Optional[dict]:
    if not candidates:
        return None
    target_key = normalize_name_key(preferred_base_name)
    for student in candidates:
        key = normalize_name_key(student.get("base_name") or student.get("name"))
        if key and key == target_key:
            return student
    for student in candidates:
        if student.get("active") is not False:
            return student
    return candidates[0]


def extract_workbook(workbook_path: Path) -> Dict[str, object]:
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)

    month_students: Dict[str, set] = defaultdict(set)
    month_roster_profiles: Dict[str, Dict[str, dict]] = defaultdict(dict)
    month_extra_columns: Dict[str, Dict[str, dict]] = defaultdict(dict)
    month_student_extras: Dict[str, Dict[str, dict]] = defaultdict(dict)
    score_points: Dict[Tuple[str, str, str], dict] = {}
    roll_profiles_by_month: Dict[str, Dict[str, dict]] = defaultdict(dict)
    scanned_sheets: List[str] = []

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if not is_month_sheet(ws.title):
            continue

        rows = visible_rows_matrix(ws)
        if not rows:
            continue

        header_idx = find_header_row(rows)
        if header_idx < 0:
            continue

        scanned_sheets.append(ws.title)
        sheet_month, sheet_year = infer_sheet_month_year(ws.title)
        header = rows[header_idx]

        roll_col = detect_col(header, ("roll",))
        name_col = detect_col(header, ("student name", "student's name", "candidate", "name"))
        class_col = detect_col(header, ("class", "grade"))
        fees_col = detect_col(header, ("fees",))
        vote_col = detect_col(header, ("vote power", "votepower", "vote"))
        total_col = detect_col(header, ("total score", "final score", "final"))
        rank_col = detect_col(header, ("rank",))

        if roll_col < 0 or name_col < 0:
            continue

        known_cols = {idx for idx in [roll_col, name_col, class_col, fees_col, vote_col, total_col, rank_col] if idx >= 0}

        date_cols: List[Tuple[int, str]] = []
        for idx, cell in enumerate(header):
            if idx in known_cols:
                continue
            parsed_date = parse_date_header(cell, sheet_month, sheet_year)
            if parsed_date:
                date_cols.append((idx, parsed_date))

        if date_cols:
            month_key = sorted(date_cols, key=lambda item: item[1])[0][1][:7]
        else:
            if not (sheet_month and sheet_year):
                continue
            month_key = f"{sheet_year:04d}-{sheet_month:02d}"

        last_date_idx = max((idx for idx, _ in date_cols), default=-1)
        for idx in range(last_date_idx + 1, len(header)):
            if idx in known_cols:
                continue
            label = normalize_text(header[idx])
            if not is_relevant_extra_header(label):
                continue
            key = to_extra_key(label, idx + 1)
            if key not in month_extra_columns[month_key]:
                month_extra_columns[month_key][key] = {"key": key, "label": label}

        for row in rows[header_idx + 1 :]:
            if roll_col >= len(row) or name_col >= len(row):
                continue

            roll = normalize_roll(row[roll_col])
            raw_name = normalize_text(row[name_col])
            if not roll or not raw_name:
                continue
            if not re.fullmatch(r"EA[A-Z0-9]{5}", roll):
                continue

            base_name, name_stars, name_vetos = parse_name_meta(raw_name)
            class_val = parse_int(row[class_col]) if class_col >= 0 and class_col < len(row) else None
            fees_val = parse_int(row[fees_col]) if fees_col >= 0 and fees_col < len(row) else None
            vote_val = parse_int(row[vote_col]) if vote_col >= 0 and vote_col < len(row) else None
            total_val = parse_int(row[total_col]) if total_col >= 0 and total_col < len(row) else None
            rank_val = parse_int(row[rank_col]) if rank_col >= 0 and rank_col < len(row) else None

            profile = {
                "roll": roll,
                "name": raw_name,
                "base_name": base_name,
                "class": class_val,
                "fees": fees_val,
                "vote_power": vote_val,
                "total_score": total_val,
                "rank": rank_val,
                "stars": max(0, name_stars),
                "veto_count": max(0, name_vetos),
            }

            roll_profiles_by_month[month_key][roll] = profile
            month_roster_profiles[month_key][roll] = {
                "roll": roll,
                "name": raw_name,
                "base_name": base_name,
                "class": class_val,
            }
            month_students[month_key].add(roll)

            if last_date_idx >= 0:
                row_extra = month_student_extras[month_key].setdefault(roll, {})
                for col_key, col_data in month_extra_columns[month_key].items():
                    # Recover original index by matching labels from header.
                    matches = [i for i, h in enumerate(header) if normalize_text(h) == col_data["label"]]
                    if not matches:
                        continue
                    col_idx = matches[0]
                    if col_idx >= len(row):
                        continue
                    cell = row[col_idx]
                    if cell is None:
                        continue
                    text = normalize_text(cell)
                    if not text:
                        continue
                    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                        row_extra[col_key] = int(round(float(cell)))
                    else:
                        row_extra[col_key] = text
                if not row_extra:
                    month_student_extras[month_key].pop(roll, None)

            for col_idx, date_iso in date_cols:
                if col_idx >= len(row):
                    continue
                points, stars, vetos = parse_score_cell(row[col_idx])
                if points is None and stars == 0 and vetos == 0:
                    continue
                score_key = (month_key, roll, date_iso)
                if score_key not in score_points:
                    score_points[score_key] = {
                        "month": month_key,
                        "roll": roll,
                        "date": date_iso,
                        "points": int(points or 0),
                        "stars": int(stars),
                        "vetos": int(vetos),
                        "notes": "",
                    }
                else:
                    # If duplicate date columns/rows appear, accumulate instead of overriding.
                    score_points[score_key]["points"] += int(points or 0)
                    score_points[score_key]["stars"] += int(stars)
                    score_points[score_key]["vetos"] += int(vetos)

    if not roll_profiles_by_month:
        raise RuntimeError("No valid month sheets were extracted from workbook.")

    month_keys = sorted(roll_profiles_by_month.keys(), key=month_sort_key)
    latest_month = month_keys[-1]

    latest_profile_by_roll: Dict[str, dict] = {}
    for month_key in month_keys:
        for roll, profile in roll_profiles_by_month[month_key].items():
            latest_profile_by_roll[roll] = profile

    return {
        "latest_month": latest_month,
        "month_keys": month_keys,
        "latest_profile_by_roll": latest_profile_by_roll,
        "month_students": {k: sorted(v) for k, v in month_students.items()},
        "month_roster_profiles": {
            k: [month_roster_profiles[k][roll] for roll in sorted(month_roster_profiles[k].keys())]
            for k in sorted(month_roster_profiles.keys(), key=month_sort_key)
        },
        "month_extra_columns": {
            k: list(month_extra_columns[k].values())
            for k in sorted(month_extra_columns.keys(), key=month_sort_key)
        },
        "month_student_extras": {
            k: month_student_extras[k]
            for k in sorted(month_student_extras.keys(), key=month_sort_key)
        },
        "score_points": list(score_points.values()),
        "scanned_sheets": scanned_sheets,
    }


def apply_hardening(existing_data: dict, extracted: dict) -> dict:
    latest_month = extracted["latest_month"]
    latest_rolls = set(extracted["month_students"].get(latest_month, []))

    existing_students = existing_data.get("students", []) or []
    existing_by_roll: Dict[str, List[dict]] = defaultdict(list)
    max_id = 0
    for student in existing_students:
        roll = normalize_roll(student.get("roll"))
        if roll:
            existing_by_roll[roll].append(student)
        try:
            max_id = max(max_id, int(student.get("id") or 0))
        except Exception:
            continue

    canonical_students: List[dict] = []
    roll_to_student_id: Dict[str, int] = {}
    used_ids = set()

    for roll in sorted(extracted["latest_profile_by_roll"].keys()):
        profile = extracted["latest_profile_by_roll"][roll]
        candidates = existing_by_roll.get(roll, [])
        chosen = choose_existing_student(candidates, profile.get("base_name") or profile.get("name") or "")
        if chosen and int(chosen.get("id") or 0) in used_ids:
            chosen = None

        if chosen is None:
            max_id += 1
            chosen = {
                "id": max_id,
                "profile_data": {},
            }

        student_id = int(chosen.get("id") or 0)
        used_ids.add(student_id)
        roll_to_student_id[roll] = student_id

        merged = dict(chosen)
        merged["id"] = student_id
        merged["roll"] = roll
        merged["name"] = profile.get("name") or profile.get("base_name") or merged.get("name") or ""
        merged["base_name"] = profile.get("base_name") or parse_name_meta(merged["name"])[0]
        merged["raw_name"] = profile.get("name") or merged["name"]
        if profile.get("class") is not None:
            merged["class"] = profile.get("class")
        if profile.get("fees") is not None:
            merged["fees"] = profile.get("fees")
        if profile.get("vote_power") is not None:
            merged["vote_power"] = profile.get("vote_power")
        if profile.get("total_score") is not None:
            merged["total_score"] = profile.get("total_score")
        if profile.get("rank") is not None:
            merged["rank"] = profile.get("rank")
        merged["stars"] = max(0, int(profile.get("stars") or 0))
        merged["veto_count"] = max(0, int(profile.get("veto_count") or 0))
        merged["group"] = derive_group_from_roll(roll)
        merged["active"] = roll in latest_rolls
        canonical_students.append(merged)

    # Preserve non-workbook rolls as inactive archival records.
    for student in existing_students:
        roll = normalize_roll(student.get("roll"))
        sid = int(student.get("id") or 0)
        if not roll or roll in roll_to_student_id:
            continue
        if sid in used_ids:
            continue
        archived = dict(student)
        archived["active"] = False
        used_ids.add(sid)
        canonical_students.append(archived)

    score_records: List[dict] = []
    score_id = 1
    for item in sorted(extracted["score_points"], key=lambda x: (x["month"], x["date"], x["roll"])):
        roll = item["roll"]
        student_id = roll_to_student_id.get(roll)
        if not student_id:
            continue
        score_records.append({
            "id": score_id,
            "studentId": student_id,
            "date": item["date"],
            "month": item["month"],
            "points": int(item["points"] or 0),
            "stars": int(item["stars"] or 0),
            "vetos": int(item["vetos"] or 0),
            "notes": item.get("notes", ""),
            "recordedBy": "admin",
        })
        score_id += 1

    hardened = dict(existing_data)
    hardened["students"] = canonical_students
    hardened["scores"] = score_records
    hardened["month_students"] = extracted["month_students"]
    hardened["month_roster_profiles"] = extracted["month_roster_profiles"]
    hardened["month_extra_columns"] = extracted["month_extra_columns"]
    hardened["month_student_extras"] = extracted["month_student_extras"]
    hardened["months"] = []
    now_iso = datetime.now().isoformat()
    hardened["updated_at"] = now_iso
    hardened["server_updated_at"] = now_iso

    return hardened


def main() -> None:
    parser = argparse.ArgumentParser(description="Harden offline scoreboard data from workbook.")
    parser.add_argument("--workbook", required=True, help="Path to authoritative .xlsm/.xlsx workbook")
    parser.add_argument("--data", default="instance/offline_scoreboard_data.json", help="Path to offline JSON data")
    parser.add_argument("--backup-dir", default="instance/offline_scoreboard_backups", help="Backup directory")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    data_path = Path(args.data)
    backup_dir = Path(args.backup_dir)

    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not data_path.exists():
        raise SystemExit(f"Data file not found: {data_path}")

    extracted = extract_workbook(workbook_path)
    existing_data = json.loads(data_path.read_text(encoding="utf-8"))
    hardened = apply_hardening(existing_data, extracted)

    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"offline_scoreboard_pre_harden_{timestamp}.json"
    backup_path.write_text(json.dumps(existing_data, ensure_ascii=False, indent=2), encoding="utf-8")

    temp_path = data_path.with_suffix(".tmp.json")
    temp_path.write_text(json.dumps(hardened, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(data_path)

    # Quick summary
    month_key = extracted["latest_month"]
    totals = defaultdict(int)
    student_lookup = {s["id"]: s for s in hardened["students"]}
    for score in hardened["scores"]:
        if score["month"] != month_key:
            continue
        totals[score["studentId"]] += int(score["points"] or 0)
    top_rows = sorted(
        [
            (
                points,
                student_lookup.get(sid, {}).get("roll", ""),
                student_lookup.get(sid, {}).get("base_name", "") or student_lookup.get(sid, {}).get("name", "")
            )
            for sid, points in totals.items()
        ],
        key=lambda item: (-item[0], item[1])
    )[:10]

    print(f"Hardened from workbook: {workbook_path}")
    print(f"Scanned month sheets: {len(extracted['scanned_sheets'])}")
    print(f"Latest month: {month_key}")
    print(f"Students in latest month: {len(extracted['month_students'].get(month_key, []))}")
    print(f"Total students (canonical + archived): {len(hardened['students'])}")
    print(f"Total score records: {len(hardened['scores'])}")
    print("Top 10 latest month:")
    for idx, row in enumerate(top_rows, start=1):
        print(f"{idx:02d}. {row[1]} | {row[2]} | {row[0]}")
    print(f"Backup written: {backup_path}")


if __name__ == "__main__":
    main()
